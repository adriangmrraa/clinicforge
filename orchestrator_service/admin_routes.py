import os
import uuid
import json
import csv
import io
import html as html_module
import asyncio
import asyncpg
import httpx
import logging
import re
from pathlib import Path
from datetime import datetime, timedelta, date, time, timezone
from typing import Any, Dict, List, Literal, Optional, Union
from decimal import Decimal
from fastapi import (
    APIRouter,
    HTTPException,
    Depends,
    Query,
    Request,
    status,
    Header,
    BackgroundTasks,
    UploadFile,
    File,
    Form,
)
from slowapi import Limiter
from slowapi.util import get_remote_address
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, Response
from pydantic import BaseModel, Field
from db import db
from gcal_service import gcal_service
from analytics_service import analytics_service
from services.liquidation_service import liquidation_service
from services.financial_dashboard_service import financial_dashboard_service
from email_service import (
    email_service,
    send_welcome_email,
    send_plan_payment_confirmation_email,
)

logger = logging.getLogger(__name__)
limiter = Limiter(key_func=get_remote_address)

# Configuración
from core.credentials import (
    get_tenant_credential,
    save_tenant_credential,
    CREDENTIALS_FERNET_KEY,
    encrypt_value,
    decrypt_value,
    TELEGRAM_BOT_TOKEN,
    TELEGRAM_WEBHOOK_SECRET,
    TELEGRAM_WEBHOOK_ACCESS_TOKEN,
)
from core.auth import verify_admin_token, get_resolved_tenant_id, ADMIN_TOKEN
from core.security_utils import generate_signed_url

INTERNAL_API_TOKEN = os.getenv("INTERNAL_API_TOKEN", "")
WHATSAPP_SERVICE_URL = os.getenv("WHATSAPP_SERVICE_URL", "http://whatsapp:8002")

# Usar encrypt_value y decrypt_value importados de core.credentials

# Legacy fixed-offset fallback. New code paths in this module SHOULD resolve the
# tenant timezone via services.tz_resolver.get_tenant_tz(tenant_id) — kept here only
# for safety on jobs/utilities that don't yet have a tenant context.
ARG_TZ = timezone(timedelta(hours=-3))

from services.whisper_service import transcribe_audio_url
from shared.odontogram_utils import normalize_to_v3
import glob

# Treatment Plan Schemas
from schemas.treatment_plan import (
    AddPlanItemBody,
    UpdatePlanItemBody,
    TreatmentPlanItemResponse,
    TreatmentPlanDetailResponse,
    ItemStatus,
    PlanStatus,
    RegisterPaymentBody,
    LinkPlanItemBody,
)


class AppointmentResponse(BaseModel):
    id: Union[str, uuid.UUID]  # UUID en appointments (create manual usa uuid.uuid4)
    patient_id: int
    appointment_datetime: datetime
    duration_minutes: int
    status: str
    urgency_level: Optional[str] = None
    source: Optional[str] = None
    appointment_type: Optional[str] = None
    notes: Optional[str] = None
    end_datetime: Optional[datetime] = None  # Para FullCalendar (start + duration)
    patient_name: str
    patient_phone: Optional[str] = None
    professional_name: Optional[str] = None
    professional_id: Optional[int] = None
    appointment_name: str


class PreInstructions(BaseModel):
    """Structured pre-treatment instructions (migration 036).

    Stored as JSONB in treatment_types.pre_instructions. Legacy free text
    rows are wrapped as {"general_notes": "<text>"} by the migration so
    they remain readable through this shape.
    """

    preparation_days_before: Optional[int] = None
    fasting_required: Optional[bool] = None
    fasting_hours: Optional[int] = None
    medications_to_avoid: Optional[List[str]] = Field(default_factory=list)
    medications_to_take: Optional[List[str]] = Field(default_factory=list)
    what_to_bring: Optional[List[str]] = Field(default_factory=list)
    general_notes: Optional[str] = None


class PostInstructions(BaseModel):
    """Structured post-treatment recovery protocol (migration 036).

    Stored as JSONB in treatment_types.post_instructions. Coexists with
    the legacy list-of-timed-entries shape (wrapped by migration 036 as
    {"general_notes": "<serialized list>"}). The get_treatment_instructions
    tool reads both shapes via _format_post_instructions_dict and the
    legacy list renderer.
    """

    care_duration_days: Optional[int] = None
    dietary_restrictions: Optional[List[str]] = Field(default_factory=list)
    activity_restrictions: Optional[List[str]] = Field(default_factory=list)
    allowed_medications: Optional[List[str]] = Field(default_factory=list)
    prohibited_medications: Optional[List[str]] = Field(default_factory=list)
    sutures_removal_day: Optional[int] = None
    normal_symptoms: Optional[List[str]] = Field(default_factory=list)
    alarm_symptoms: Optional[List[str]] = Field(default_factory=list)
    escalation_message: Optional[str] = None


class TreatmentTypeResponse(BaseModel):
    id: int
    code: str
    name: str
    description: Optional[str] = None
    default_duration_minutes: int
    min_duration_minutes: Optional[int] = None
    max_duration_minutes: Optional[int] = None
    complexity_level: Optional[str] = None
    category: Optional[str] = None
    requires_multiple_sessions: bool
    session_gap_days: Optional[int] = None
    is_active: bool
    is_available_for_booking: bool
    internal_notes: Optional[str] = None
    base_price: Optional[float] = 0
    priority: Optional[str] = "medium"
    professional_ids: Optional[List[int]] = []
    # Migration 036: pre_instructions is JSONB now (was Text). Response shape
    # is whatever is stored — dict or null. Frontend handles parsing.
    pre_instructions: Optional[Any] = None
    post_instructions: Optional[Any] = None
    followup_template: Optional[Any] = None


router = APIRouter(prefix="/admin", tags=["Dental Admin"])

# ============================================
# MODELS - Treatment Plan Billing (Tarea 2.1)
# ============================================


class TreatmentPlanItemCreate(BaseModel):
    """Modelo para crear un ítem dentro de un plan de tratamiento."""

    treatment_type_code: Optional[str] = None
    custom_description: Optional[str] = None
    estimated_price: float = 0
    approved_price: Optional[float] = None
    status: Optional[str] = "pending"


class CreateTreatmentPlanBody(BaseModel):
    """Body para crear un nuevo plan de tratamiento."""

    name: str
    professional_id: Optional[int] = None
    notes: Optional[str] = None
    items: Optional[List[TreatmentPlanItemCreate]] = None  # Items iniciales opcionales


class UpdateTreatmentPlanBody(BaseModel):
    """Body para actualizar un plan de tratamiento."""

    name: Optional[str] = None
    professional_id: Optional[int] = None
    status: Optional[str] = None  # draft, approved, in_progress, completed, cancelled
    approved_total: Optional[float] = None
    notes: Optional[str] = None
    # Campos de configuración presupuesto (opcionales)
    payment_conditions: Optional[str] = None  # e.g. "Válido por 30 días"
    discount_pct: Optional[float] = None  # porcentaje de descuento 0-100
    discount_amount: Optional[float] = None  # descuento fijo en $
    installments: Optional[int] = None  # cantidad de cuotas
    installments_amount: Optional[float] = None  # monto por cuota


class TreatmentPlanItemResponse(BaseModel):
    """Response item de plan de tratamiento."""

    id: str
    plan_id: str
    tenant_id: int
    treatment_type_code: Optional[str]
    custom_description: Optional[str]
    estimated_price: float
    approved_price: Optional[float]
    status: str
    sort_order: int
    created_at: datetime
    updated_at: datetime
    # Joined fields
    treatment_name: Optional[str] = None
    appointments_count: Optional[int] = 0


class TreatmentPlanPaymentResponse(BaseModel):
    """Response de pago de plan."""

    id: str
    plan_id: str
    tenant_id: int
    amount: float
    payment_method: str
    payment_date: date
    recorded_by: Optional[str]
    appointment_id: Optional[str]
    receipt_data: Optional[Dict[str, Any]]
    notes: Optional[str]
    created_at: datetime


class TreatmentPlanResponse(BaseModel):
    """Response básico de plan (para listas)."""

    id: str
    tenant_id: int
    patient_id: int
    professional_id: Optional[int]
    name: str
    status: str
    estimated_total: float
    approved_total: Optional[float]
    approved_by: Optional[str]
    approved_at: Optional[datetime]
    notes: Optional[str]
    created_at: datetime
    updated_at: datetime
    # Aggregated fields
    items_count: int = 0
    paid_total: float = 0
    pending_total: float = 0
    # Joined fields
    professional_name: Optional[str] = None


class TreatmentPlanDetailResponse(BaseModel):
    """Response completo de plan con arrays de items y payments."""

    id: str
    tenant_id: int
    patient_id: int
    professional_id: Optional[int]
    name: str
    status: str
    estimated_total: float
    approved_total: Optional[float]
    approved_by: Optional[str]
    approved_at: Optional[datetime]
    notes: Optional[str]
    created_at: datetime
    updated_at: datetime
    # Arrays
    items: List[TreatmentPlanItemResponse] = []
    payments: List[TreatmentPlanPaymentResponse] = []
    # Joined fields
    professional_name: Optional[str] = None
    patient_name: Optional[str] = None
    # Calculated totals
    paid_total: float = 0
    pending_total: float = 0


# ============================================
# DEBUG ENDPOINTS - Para diagnóstico en producción
# ============================================


@router.get("/debug/auth-test", dependencies=[Depends(verify_admin_token)])
async def debug_auth_test(request: Request, x_admin_token: str = Header(None)):
    """
    Endpoint de debug para verificar problemas de autenticación 401.
    Público - no requiere auth.
    """
    headers = dict(request.headers)

    # Obtener X-Admin-Token de headers (case-insensitive)
    received_admin_token = None
    for key, value in headers.items():
        if key.lower() == "x-admin-token":
            received_admin_token = value
            break

    debug_info = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "received_admin_token_present": received_admin_token is not None,
        "received_admin_token_preview": received_admin_token[:10] + "..."
        if received_admin_token
        else None,
        "expected_admin_token_defined": bool(ADMIN_TOKEN),
        "expected_admin_token_length": len(ADMIN_TOKEN) if ADMIN_TOKEN else 0,
        "tokens_match": received_admin_token == ADMIN_TOKEN
        if received_admin_token and ADMIN_TOKEN
        else False,
        "client_ip": request.client.host if request.client else "unknown",
        "user_agent": headers.get("User-Agent", "unknown")[:100],
        "cors_origin": headers.get("Origin", "none"),
        "authorization_header_present": "Authorization" in headers,
        "cookie_header_present": "Cookie" in headers,
    }

    logger.info(f"DEBUG_AUTH_TEST: {debug_info}")

    return debug_info


@router.get("/debug/health", dependencies=[Depends(verify_admin_token)])
async def debug_health():
    """
    Health check público.
    """
    return {
        "status": "healthy",
        "service": "orchestrator",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": os.getenv("API_VERSION", "1.0.0"),
    }


@router.get("/debug/env-safe", dependencies=[Depends(verify_admin_token)])
async def debug_env_safe():
    """
    Variables de entorno seguras (sin exponer valores sensibles).
    """
    return {
        "ADMIN_TOKEN_defined": bool(ADMIN_TOKEN),
        "ADMIN_TOKEN_length": len(ADMIN_TOKEN) if ADMIN_TOKEN else 0,
        "CORS_ALLOWED_ORIGINS": os.getenv("CORS_ALLOWED_ORIGINS", "not_set"),
        "LOG_LEVEL": os.getenv("LOG_LEVEL", "INFO"),
        "backend_version": os.getenv("API_VERSION", "1.0.0"),
        "environment": os.getenv("NODE_ENV", "production"),
    }


def normalize_phone(phone: str) -> str:
    """Asegura que el número tenga el formato +123456789 (E.164)"""
    clean = re.sub(r"\D", "", phone)
    if not phone.startswith("+"):
        return "+" + clean
    return "+" + clean


# --- Helper para emitir eventos de Socket.IO ---
async def emit_appointment_event(
    event_type: str, data: Dict[str, Any], request: Request
):
    """Emit appointment events via Socket.IO through the app state."""
    if hasattr(request.app.state, "emit_appointment_event"):
        await request.app.state.emit_appointment_event(event_type, data)


# --- Background Task para envío a WhatsApp ---
async def send_to_whatsapp_task(phone: str, message: str, business_number: str):
    """Tarea en segundo plano para no bloquear la UI mientras se envía el mensaje."""
    normalized = normalize_phone(phone)
    logger.info(
        f"📤 Intentando envío manual a WhatsApp: {normalized} via {WHATSAPP_SERVICE_URL}"
    )
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            ws_resp = await client.post(
                f"{WHATSAPP_SERVICE_URL}/send",
                json={"to": normalized, "message": message},
                headers={
                    "X-Internal-Token": INTERNAL_API_TOKEN,
                    "X-Correlation-Id": str(uuid.uuid4()),
                },
                params={"from_number": business_number},
            )
            if ws_resp.status_code != 200:
                logger.error(
                    f"❌ Error en WhatsApp Service ({ws_resp.status_code}): {ws_resp.text}"
                )
            else:
                logger.info(f"✅ WhatsApp background send success for {normalized}")
    except Exception as e:
        logger.error(
            f"❌ WhatsApp background send CRITICAL failed for {normalized}: {str(e)}"
        )


# --- Dependencias de Seguridad (Triple Capa Nexus v7.6) ---
from core.auth import verify_admin_token, verify_ceo_token, get_resolved_tenant_id

_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _strip_html(text: str) -> str:
    """Elimina tags HTML/JS de un string para prevenir stored XSS."""
    return _HTML_TAG_RE.sub("", text).strip()


async def get_allowed_tenant_ids(user_data=Depends(verify_admin_token)) -> List[int]:
    """
    Lista de tenant_id que el usuario puede ver (chats, sesiones).
    CEO: tenants donde tiene professional vinculado + tenant del JWT.
    Secretary/Professional: solo su clínica resuelta.
    """
    try:
        if user_data.role == "ceo":
            # Solo tenants donde el CEO tiene un professional vinculado
            rows = await db.pool.fetch(
                "SELECT DISTINCT tenant_id FROM professionals WHERE user_id = $1 ORDER BY tenant_id ASC",
                uuid.UUID(user_data.user_id),
            )
            allowed = [int(r["tenant_id"]) for r in rows] if rows else []
            # Siempre incluir el tenant del JWT como fallback
            jwt_tid = getattr(user_data, "tenant_id", 1)
            if isinstance(jwt_tid, int) and jwt_tid not in allowed:
                allowed.append(jwt_tid)
            return allowed if allowed else [1]
        try:
            tid = await db.pool.fetchval(
                "SELECT tenant_id FROM professionals WHERE user_id = $1",
                uuid.UUID(user_data.user_id),
            )
            if tid is not None:
                return [int(tid)]
        except (ValueError, TypeError):
            pass
        first = await db.pool.fetchval("SELECT id FROM tenants ORDER BY id ASC LIMIT 1")
        return [int(first)] if first is not None else [1]
    except Exception:
        return [1]  # Fallback para no devolver 500


@router.get("/patients/phone/{phone}/context", tags=["Pacientes"])
async def get_patient_clinical_context(
    phone: str,
    request: Request,
    tenant_id_override: Optional[int] = None,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Retorna el contexto clínico completo de un paciente por su teléfono.
    Si se pasa tenant_id_override (ej. CEO eligió una clínica en Chats), se usa si el usuario tiene acceso.
    """
    tenant_id = (
        tenant_id_override
        if (tenant_id_override is not None and tenant_id_override in allowed_ids)
        else resolved_tenant_id
    )

    # 1. Resolver paciente (por Teléfono o por ID de Plataforma)
    # Si el "phone" no empieza con + o no son solo números, lo tratamos como plataforma_id
    is_pure_platform_id = not (phone.startswith("+") or phone.isdigit())

    patient = None
    if not is_pure_platform_id:
        normalized_phone = normalize_phone(phone)
        patient = await db.pool.fetchrow(
            """
            SELECT id, first_name, last_name, phone_number, status, urgency_level, urgency_reason, preferred_schedule,
                   acquisition_source, meta_ad_id, meta_ad_headline, meta_ad_body, external_ids, medical_history
            FROM patients 
            WHERE tenant_id = $1 AND (phone_number = $2 OR phone_number = $3)
            AND status != 'deleted'
        """,
            tenant_id,
            normalized_phone,
            phone,
        )

    if not patient:
        # Buscar por external_ids (IG/FB)
        # Probamos el ID contra todas las claves del JSONB
        patient = await db.pool.fetchrow(
            """
            SELECT id, first_name, last_name, phone_number, status, urgency_level, urgency_reason, preferred_schedule,
                   acquisition_source, meta_ad_id, meta_ad_headline, meta_ad_body, external_ids, medical_history
            FROM patients 
            WHERE tenant_id = $1 
            AND (
                external_ids->>'instagram' = $2 OR 
                external_ids->>'facebook' = $2 OR 
                external_ids->>'chatwoot' = $2
            )
            AND status != 'deleted'
        """,
            tenant_id,
            phone,
        )

    if patient:
        from core.auth import log_pii_access

        await log_pii_access(
            request, user_data, patient["id"], action="read_clinical_context"
        )

        # 2. Última cita (pasada) - Mapeo a 'date' para el frontend
        last_apt = await db.pool.fetchrow(
            """
            SELECT a.id, a.appointment_datetime AS date, a.appointment_type AS type, a.status, 
                   a.duration_minutes, p.first_name as professional_name
            FROM appointments a
            LEFT JOIN professionals p ON a.professional_id = p.id
            WHERE a.tenant_id = $1 AND a.patient_id = $2 AND a.appointment_datetime < NOW()
            ORDER BY a.appointment_datetime DESC LIMIT 1
        """,
            tenant_id,
            patient["id"],
        )

        # 3. Próxima cita (futura) - Mapeo a 'date' para el frontend
        upcoming_apt = await db.pool.fetchrow(
            """
            SELECT a.id, a.appointment_datetime AS date, a.appointment_type AS type, a.status,
                   a.duration_minutes, p.first_name as professional_name
            FROM appointments a
            LEFT JOIN professionals p ON a.professional_id = p.id
            WHERE a.tenant_id = $1 AND a.patient_id = $2 AND a.appointment_datetime >= NOW()
            AND a.status IN ('scheduled', 'confirmed')
            ORDER BY a.appointment_datetime ASC LIMIT 1
        """,
            tenant_id,
            patient["id"],
        )

        # 4. Plan de tratamiento (del último registro clínico)
        clinical_record = await db.pool.fetchrow(
            """
            SELECT treatment_plan, diagnosis, created_at as record_date
            FROM clinical_records
            WHERE tenant_id = $1 AND patient_id = $2
            ORDER BY created_at DESC LIMIT 1
        """,
            tenant_id,
            patient["id"],
        )

        # 5. Respuesta combinada
        return {
            "patient": dict(patient),
            "last_appointment": dict(last_apt) if last_apt else None,
            "upcoming_appointment": dict(upcoming_apt) if upcoming_apt else None,
            "treatment_plan": clinical_record["treatment_plan"]
            if clinical_record
            else None,
            "diagnosis": clinical_record["diagnosis"] if clinical_record else None,
            "is_guest": patient["status"] == "guest",
        }

    # Si no existe, es un lead puro sin registro previo
    return {
        "patient": None,
        "last_appointment": None,
        "upcoming_appointment": None,
        "treatment_plan": None,
        "is_guest": True,
    }


class StatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None  # active, suspended, pending


# --- RUTAS DE ADMINISTRACIÓN DE USUARIOS ---


@router.get(
    "/users/pending",
    tags=["Usuarios"],
    summary="Listar usuarios pendientes de aprobación",
)
async def get_pending_users(
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Retorna la lista de usuarios con estado 'pending' (Solo CEO/Secretary)"""
    if user_data.role not in ["ceo", "secretary"]:
        raise HTTPException(
            status_code=403,
            detail="Solo el personal administrador puede ver usuarios pendientes.",
        )

    users = await db.fetch(
        """
        SELECT id, email, role, status, created_at, first_name, last_name
        FROM users
        WHERE status = 'pending' AND tenant_id = $1
        ORDER BY created_at DESC
    """,
        tenant_id,
    )
    return [dict(u) for u in users]


@router.get(
    "/users", tags=["Usuarios"], summary="Listar todos los usuarios del sistema"
)
async def get_all_users(
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Retorna la lista de todos los usuarios de la clínica (Solo CEO/Secretary)"""
    if user_data.role not in ["ceo", "secretary"]:
        raise HTTPException(
            status_code=403,
            detail="Solo el personal administrador puede listar usuarios.",
        )

    users = await db.fetch(
        """
        SELECT DISTINCT u.id, u.email, u.role, u.status, u.created_at, u.updated_at, u.first_name, u.last_name
        FROM users u
        LEFT JOIN professionals p ON p.user_id = u.id
        WHERE p.tenant_id = ANY($1::int[]) OR p.user_id IS NULL
        ORDER BY u.status ASC, u.created_at DESC
    """,
        allowed_ids,
    )
    return [dict(u) for u in users]


@router.post(
    "/users/{user_id}/status",
    tags=["Usuarios"],
    summary="Aprobar o suspender un usuario",
)
async def update_user_status(
    user_id: str, payload: StatusUpdate, user_data=Depends(verify_ceo_token)
):
    """Actualiza el estado de un usuario (Aprobación/Suspensión) - Solo CEO.
    Al aprobar (active): si es professional/secretary y no tiene fila en professionals, se crea una para la primera sede."""

    target_user = await db.fetchrow(
        "SELECT id, email, role, COALESCE(first_name, '') as first_name, COALESCE(last_name, '') as last_name FROM users WHERE id = $1",
        user_id,
    )
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    await db.execute(
        "UPDATE users SET status = $1, updated_at = NOW() WHERE id = $2",
        payload.status,
        user_id,
    )

    uid = uuid.UUID(user_id)
    if target_user["role"] in ("professional", "secretary"):
        is_active = payload.status == "active"
        # Sincronizar is_active en professionals si ya tiene fila(s)
        await db.execute(
            "UPDATE professionals SET is_active = $1 WHERE user_id = $2", is_active, uid
        )
        # Al aprobar: si no tiene ninguna fila en professionals, crear una para la primera sede (puede usar la plataforma)
        if payload.status == "active":
            has_row = await db.pool.fetchval(
                "SELECT 1 FROM professionals WHERE user_id = $1", uid
            )
            if not has_row:
                first_tenant = await db.pool.fetchval(
                    "SELECT id FROM tenants ORDER BY id ASC LIMIT 1"
                )
                tenant_id = int(first_tenant) if first_tenant is not None else 1
                first_name = (target_user["first_name"] or "").strip() or "Profesional"
                last_name = (target_user["last_name"] or "").strip() or " "
                email = (target_user["email"] or "").strip()
                wh_json = json.dumps(generate_default_working_hours())
                try:
                    await db.pool.execute(
                        """
                        INSERT INTO professionals (tenant_id, user_id, first_name, last_name, email, phone_number,
                        specialty, registration_id, is_active, working_hours, created_at, updated_at)
                        VALUES ($1, $2, $3, $4, $5, NULL, NULL, NULL, TRUE, $6::jsonb, NOW(), NOW())
                    """,
                        tenant_id,
                        uid,
                        first_name,
                        last_name,
                        email,
                        wh_json,
                    )
                except asyncpg.UndefinedColumnError as e:
                    err_str = str(e).lower()
                    if "phone_number" in err_str:
                        await db.pool.execute(
                            """
                            INSERT INTO professionals (tenant_id, user_id, first_name, last_name, email,
                            specialty, registration_id, is_active, working_hours, created_at, updated_at)
                            VALUES ($1, $2, $3, $4, $5, NULL, NULL, TRUE, $6::jsonb, NOW(), NOW())
                        """,
                            tenant_id,
                            uid,
                            first_name,
                            last_name,
                            email,
                            wh_json,
                        )
                    elif "updated_at" in err_str:
                        await db.pool.execute(
                            """
                            INSERT INTO professionals (tenant_id, user_id, first_name, last_name, email, phone_number,
                            specialty, registration_id, is_active, working_hours, created_at)
                            VALUES ($1, $2, $3, $4, $5, NULL, NULL, NULL, TRUE, $6::jsonb, NOW())
                        """,
                            tenant_id,
                            uid,
                            first_name,
                            last_name,
                            email,
                            wh_json,
                        )
                    else:
                        raise

    # Enviar email de bienvenida al aprobar (activar) un profesional o secretaria
    if payload.status == "active" and target_user["role"] in (
        "professional",
        "secretary",
    ):
        try:
            # Resolver tenant_id: buscar la sede vinculada al profesional, o la primera sede
            _wel_tenant = await db.pool.fetchval(
                "SELECT tenant_id FROM professionals WHERE user_id = $1 LIMIT 1", uid
            )
            if not _wel_tenant:
                _wel_tenant = await db.pool.fetchval(
                    "SELECT id FROM tenants ORDER BY id ASC LIMIT 1"
                )
            _wel_tid = int(_wel_tenant) if _wel_tenant is not None else 1
            asyncio.create_task(
                send_welcome_email(_wel_tid, str(uid), target_user["role"], db.pool)
            )
            logger.info(
                f"📧 Email de bienvenida programado para usuario aprobado {uid}"
            )
        except Exception as e:
            logger.warning(f"Welcome email failed for approved user {uid}: {e}")

    return {
        "message": f"Usuario {target_user['email']} actualizado a {payload.status}."
    }


class PatientCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = ""
    phone_number: str
    email: Optional[str] = None
    dni: Optional[str] = None
    insurance: Optional[str] = None  # Obra Social
    city: Optional[str] = None
    birth_date: Optional[date] = None


class AppointmentCreate(BaseModel):
    patient_id: Optional[int] = None
    patient_phone: Optional[str] = None  # Si es paciente nuevo rápido
    professional_id: int
    appointment_datetime: datetime
    duration_minutes: Optional[int] = 30  # Por defecto 30 min
    appointment_type: str = "checkup"
    notes: Optional[str] = None
    check_collisions: bool = True  # Por defecto verificar colisiones


class GCalendarBlockCreate(BaseModel):
    google_event_id: str
    title: str
    description: Optional[str] = None
    start_datetime: datetime
    end_datetime: datetime
    all_day: bool = False
    professional_id: Optional[int] = None


class ClinicalNote(BaseModel):
    content: str
    odontogram_data: Optional[Dict] = None


# Modelos para Ficha Médica - Documentos
class OdontogramUpdate(BaseModel):
    odontogram_data: Dict[str, Any]


class PatientDocumentCreate(BaseModel):
    document_type: str = "clinical"
    filename: str


class PatientDocument(BaseModel):
    id: int
    tenant_id: int
    patient_id: int
    filename: str
    file_path: str
    file_size: Optional[int]
    mime_type: Optional[str]
    document_type: str
    uploaded_by: Optional[int]
    created_at: datetime


class ProfessionalCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    specialty: Optional[str] = None
    license_number: Optional[str] = None
    address: Optional[str] = None
    is_active: bool = True
    availability: Dict[str, Any] = {}
    working_hours: Optional[Dict[str, Any]] = None
    tenant_id: Optional[int] = (
        None  # Clínica a la que se vincula; si viene y está permitido, se usa; si no, contexto
    )


class ProfessionalUpdate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    specialty: Optional[str] = None
    license_number: Optional[str] = None
    is_active: bool
    availability: Dict[str, Any]
    working_hours: Optional[Dict[str, Any]] = None
    google_calendar_id: Optional[str] = None
    consultation_price: Optional[float] = None
    is_priority_professional: Optional[bool] = False


def generate_default_working_hours():
    """Genera el JSON de horarios por defecto (Mon-Sat, 09:00-18:00)"""
    start = os.getenv("CLINIC_START_TIME", "09:00")
    end = os.getenv("CLINIC_END_TIME", "18:00")
    slot = {"start": start, "end": end}
    wh = {}
    for day in [
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
        "sunday",
    ]:
        is_working_day = day != "sunday"
        wh[day] = {"enabled": is_working_day, "slots": [slot] if is_working_day else []}
    return wh


class TreatmentTypeCreate(BaseModel):
    code: str
    name: str
    description: Optional[str] = ""
    default_duration_minutes: int = 30
    min_duration_minutes: int = 15
    max_duration_minutes: int = 60
    complexity_level: str = "medium"
    category: str = "restorative"
    requires_multiple_sessions: bool = False
    session_gap_days: int = 0
    is_active: bool = True
    is_available_for_booking: bool = True
    internal_notes: Optional[str] = ""
    base_price: Optional[float] = 0
    priority: Optional[str] = "medium"
    professional_ids: Optional[List[int]] = None
    # Migration 036: pre_instructions accepts dict (PreInstructions),
    # legacy plain string (wrapped on save), or raw dict for flexibility.
    pre_instructions: Optional[Union[PreInstructions, dict, str]] = None
    # post_instructions accepts dict (PostInstructions), legacy timed-list,
    # raw dict, or string (rare).
    post_instructions: Optional[Union[PostInstructions, list, dict, str]] = None
    followup_template: Optional[Any] = None
    confirm_unusual_price: bool = False


class TreatmentTypeUpdate(BaseModel):
    name: str
    description: Optional[str] = ""
    default_duration_minutes: int
    min_duration_minutes: int
    max_duration_minutes: int
    complexity_level: str
    category: str
    requires_multiple_sessions: bool
    session_gap_days: int
    is_active: bool
    is_available_for_booking: bool
    internal_notes: Optional[str] = ""
    base_price: Optional[float] = 0
    priority: Optional[str] = "medium"
    # Migration 036: same Union shape as TreatmentTypeCreate.
    pre_instructions: Optional[Union[PreInstructions, dict, str]] = None
    post_instructions: Optional[Union[PostInstructions, list, dict, str]] = None
    followup_template: Optional[Any] = None
    confirm_unusual_price: bool = False


class ChatSendMessage(BaseModel):
    phone: str
    tenant_id: int
    message: str


class HumanInterventionToggle(BaseModel):
    phone: str
    tenant_id: int  # Clínica: override/silencio es por (tenant_id, phone), independiente por clínica
    activate: bool
    duration: Optional[int] = 86400000  # 24 horas en ms


class ConnectSovereignPayload(BaseModel):
    """Token de Auth0 para conectar Google Calendar de forma soberana por clínica."""

    access_token: str  # Token de Auth0 (se guarda cifrado en credentials)
    tenant_id: Optional[int] = (
        None  # Solo CEO puede especificar; si no, se usa la clínica resuelta
    )


class CredentialPayload(BaseModel):
    id: Optional[int] = None
    name: str
    value: str
    category: str
    description: Optional[str] = ""
    scope: str = "global"  # global | tenant
    tenant_id: Optional[int] = None


# ==================== INTEGRATIONS WIZARD ENDPOINTS (Spec 16) ====================


class IntegrationConfig(BaseModel):
    provider: str  # ycloud | chatwoot
    # Chatwoot fields
    chatwoot_base_url: Optional[str] = None
    chatwoot_api_token: Optional[str] = None
    chatwoot_account_id: Optional[str] = (
        None  # String to handle potential non-numeric IDs in future or masks
    )
    # YCloud fields
    ycloud_api_key: Optional[str] = None
    ycloud_webhook_secret: Optional[str] = None

    tenant_id: Optional[int] = None  # Optional override for multi-tenant setup


class HighRiskProtocol(BaseModel):
    """Validación estricta de cada entrada del dict high_risk_protocols.

    Cada condición (key del dict) tiene una política con estos campos.
    `extra=forbid` garantiza que keys desconocidas tiren 422 y no se cuelen
    en el JSONB.
    """

    requires_medical_clearance: bool = False
    requires_pre_appointment_call: bool = False
    restricted_treatments: list[str] = []
    notes: str = ""

    model_config = {"extra": "forbid"}


@router.get("/integrations/{provider}/config", tags=["Integraciones"])
async def get_integration_config(
    provider: str,
    request: Request,
    tenant_id: Optional[int] = None,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Obtiene la configuración enmascarada para un proveedor específico.
    Provider: 'chatwoot' | 'ycloud'
    Query param 'tenant_id' permite a CEO ver config de otra sede.
    """
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Acceso denegado.")

    # Use query param if provided (CEO override), otherwise current context
    target_tenant_id = tenant_id if tenant_id else resolved_tenant_id

    config = {}

    if provider == "chatwoot":
        # Retrieve Chatwoot credentials
        rows = await db.fetch(
            """
            SELECT name, value FROM credentials 
            WHERE tenant_id = $1 AND name IN ('CHATWOOT_BASE_URL', 'CHATWOOT_API_TOKEN', 'CHATWOOT_ACCOUNT_ID')
        """,
            target_tenant_id,
        )

        data = {r["name"]: r["value"] for r in rows}

        # Base URL (Visible)
        config["chatwoot_base_url"] = data.get(
            "CHATWOOT_BASE_URL", "https://app.chatwoot.com"
        )

        # API Token (Masked)
        token = data.get("CHATWOOT_API_TOKEN")
        if token:
            decrypted = decrypt_value(token) or token
            config["chatwoot_api_token"] = (
                f"••••••••{decrypted[-4:]}" if len(decrypted) > 4 else "••••••••"
            )

        # Account ID (Visible)
        config["chatwoot_account_id"] = data.get("CHATWOOT_ACCOUNT_ID", "")

        # --- Generate Webhook URL (Moved from chat_api.py) ---
        # 1. Get/Create Webhook Token
        webhook_token = await db.pool.fetchval(
            "SELECT value FROM credentials WHERE tenant_id = $1 AND name = 'WEBHOOK_ACCESS_TOKEN'",
            target_tenant_id,
        )
        if not webhook_token:
            webhook_token = uuid.uuid4().hex + uuid.uuid4().hex
            await db.pool.execute(
                """
                INSERT INTO credentials (tenant_id, name, value, category, description, scope, created_at, updated_at)
                VALUES ($1, 'WEBHOOK_ACCESS_TOKEN', $2, 'system', 'Token para webhook de Chatwoot', 'tenant', NOW(), NOW())
                ON CONFLICT (tenant_id, name) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """,
                target_tenant_id,
                webhook_token,
            )

        config["access_token"] = webhook_token

        # 2. Construct URL
        api_base = os.getenv("BASE_URL", "").rstrip("/")
        if not api_base:
            # Fallback using request
            scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
            host = request.headers.get("x-forwarded-host", request.url.netloc)
            api_base = f"{scheme}://{host}"

        config["api_base"] = api_base
        config["webhook_path"] = "/admin/chatwoot/webhook"
        config["full_webhook_url"] = (
            f"{api_base}/admin/chatwoot/webhook?access_token={webhook_token}"
        )

    elif provider == "ycloud":
        rows = await db.fetch(
            """
            SELECT name, value FROM credentials 
            WHERE tenant_id = $1 AND name IN ('YCLOUD_API_KEY', 'YCLOUD_WEBHOOK_SECRET')
        """,
            target_tenant_id,
        )
        data = {r["name"]: r["value"] for r in rows}

        # API Key (Masked)
        key = data.get("YCLOUD_API_KEY")
        if key:
            decrypted = decrypt_value(key) or key
            config["ycloud_api_key"] = (
                f"••••••••{decrypted[-4:]}" if len(decrypted) > 4 else "••••••••"
            )

        # Webhook Secret (Masked)
        secret = data.get("YCLOUD_WEBHOOK_SECRET")
        if secret:
            decrypted = decrypt_value(secret) or secret
            config["ycloud_webhook_secret"] = (
                f"••••••••{decrypted[-4:]}" if len(decrypted) > 4 else "••••••••"
            )

        # --- Generate Webhook URL for YCloud ---
        webhook_token = await db.pool.fetchval(
            "SELECT value FROM credentials WHERE tenant_id = $1 AND name = 'WEBHOOK_ACCESS_TOKEN'",
            target_tenant_id,
        )
        if not webhook_token:
            webhook_token = uuid.uuid4().hex + uuid.uuid4().hex
            await db.pool.execute(
                """
                INSERT INTO credentials (tenant_id, name, value, category, description, scope, created_at, updated_at)
                VALUES ($1, 'WEBHOOK_ACCESS_TOKEN', $2, 'system', 'Token para webhook de Chatwoot', 'tenant', NOW(), NOW())
                ON CONFLICT (tenant_id, name) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """,
                target_tenant_id,
                webhook_token,
            )

        config["access_token"] = webhook_token

        api_base = os.getenv("BASE_URL", "").rstrip("/")
        if not api_base:
            scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
            host = request.headers.get("x-forwarded-host", request.url.netloc)
            api_base = f"{scheme}://{host}"

        config["api_base"] = api_base
        config["webhook_path"] = "/admin/ycloud/webhook"
        config["ycloud_webhook_url"] = (
            f"{api_base}/admin/ycloud/webhook?access_token={webhook_token}"
        )

    else:
        raise HTTPException(
            status_code=400, detail=f"Proveedor no soportado: {provider}"
        )

    return config


@router.post("/integrations/{provider}/config", tags=["Integraciones"])
async def save_integration_config(
    provider: str,
    payload: IntegrationConfig,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Guarda la configuración de integración para YCloud o Chatwoot.
    Maneja encriptación y upsert de múltiples credenciales.
    """
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Acceso denegado.")

    target_tenant_id = payload.tenant_id if payload.tenant_id else resolved_tenant_id

    creds_to_save = []

    if provider == "chatwoot":
        if payload.chatwoot_base_url:
            creds_to_save.append(
                ("CHATWOOT_BASE_URL", payload.chatwoot_base_url, "Chatwoot URL", False)
            )

        if payload.chatwoot_api_token and not payload.chatwoot_api_token.startswith(
            "••••"
        ):
            creds_to_save.append(
                (
                    "CHATWOOT_API_TOKEN",
                    payload.chatwoot_api_token,
                    "Chatwoot API Token",
                    True,
                )
            )

        if payload.chatwoot_account_id:
            creds_to_save.append(
                (
                    "CHATWOOT_ACCOUNT_ID",
                    payload.chatwoot_account_id,
                    "Chatwoot Account ID",
                    False,
                )
            )

    elif provider == "ycloud":
        if payload.ycloud_api_key and not payload.ycloud_api_key.startswith("••••"):
            creds_to_save.append(
                ("YCLOUD_API_KEY", payload.ycloud_api_key, "YCloud API Key", True)
            )

        if (
            payload.ycloud_webhook_secret
            and not payload.ycloud_webhook_secret.startswith("••••")
        ):
            creds_to_save.append(
                (
                    "YCLOUD_WEBHOOK_SECRET",
                    payload.ycloud_webhook_secret,
                    "YCloud Webhook Secret",
                    True,
                )
            )

    else:
        raise HTTPException(status_code=400, detail="Proveedor no reconocido")

    # Execute Upserts
    for key, value, desc, encrypt in creds_to_save:
        final_val = value.strip()
        if encrypt:
            encrypted = encrypt_value(final_val)
            if not encrypted:
                raise HTTPException(status_code=500, detail="Error de encriptación")
            final_val = encrypted

        # Check integrity: if exists update, else insert
        existing = await db.fetchval(
            "SELECT id FROM credentials WHERE tenant_id = $1 AND name = $2",
            target_tenant_id,
            key,
        )

        if existing:
            await db.execute(
                """
                UPDATE credentials SET value = $1, updated_at = NOW()
                WHERE id = $2
            """,
                final_val,
                existing,
            )
        else:
            await db.execute(
                """
                INSERT INTO credentials (name, value, category, description, scope, tenant_id, created_at, updated_at)
                VALUES ($1, $2, $3, $4, 'tenant', $5, NOW(), NOW())
            """,
                key,
                final_val,
                provider,
                desc,
                target_tenant_id,
            )

    return {"message": f"Integración {provider} actualizada correctamente."}
    """Lista simple de tenants para selectores (id, name)."""
    # Permitir a CEO y Secretary (para selectores de contexto)
    rows = await db.fetch("SELECT id, clinic_name FROM tenants ORDER BY id ASC")
    return [{"id": r["id"], "name": r["clinic_name"]} for r in rows]


# ==================== ENDPOINTS CHAT MANAGEMENT ====================


@router.get("/chat/tenants", dependencies=[Depends(verify_admin_token)], tags=["Chat"])
async def get_chat_tenants(allowed_ids: List[int] = Depends(get_allowed_tenant_ids)):
    """
    Lista de clínicas que el usuario puede ver en Chats.
    CEO: todas. Secretary/Professional: una sola (su clínica).
    Usado por el selector de Clínicas en ChatsView.
    """
    if not allowed_ids:
        return []
    rows = await db.pool.fetch(
        "SELECT id, clinic_name FROM tenants WHERE id = ANY($1::int[]) ORDER BY id ASC",
        allowed_ids,
    )
    return [{"id": r["id"], "clinic_name": r["clinic_name"]} for r in rows]


@router.get("/chat/sessions", dependencies=[Depends(verify_admin_token)], tags=["Chat"])
async def get_chat_sessions(
    tenant_id: int,
    request: Request,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Sesiones de chat activas para la clínica indicada.
    El usuario solo ve sesiones cuyo tenant_id coincide con su clínica (o cualquiera si es CEO).
    Human Override y ventana de 24h son por (tenant_id, phone): independientes por clínica.
    """
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    # Sesiones = pacientes de esta clínica que tienen al menos un mensaje en esta clínica
    has_tenant_in_cm = await db.pool.fetchval(
        "SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public' AND table_name='chat_messages' AND column_name='tenant_id')"
    )
    if not has_tenant_in_cm:
        # Fallback: DB sin parche 15, filtrar solo por patients.tenant_id (mensajes sin tenant)
        rows = await db.pool.fetch(
            """
            SELECT * FROM (
                SELECT DISTINCT ON (p.phone_number)
                    p.phone_number,
                    p.id as patient_id,
                    p.first_name || ' ' || COALESCE(p.last_name, '') as patient_name,
                    cm.content as last_message,
                    cm.created_at as last_message_time,
                    p.human_handoff_requested,
                    p.human_override_until,
                    p.last_derivhumano_at,
                    CASE 
                        WHEN p.human_handoff_requested AND p.human_override_until > NOW() THEN 'human_handling'
                        WHEN p.human_override_until > NOW() THEN 'silenced'
                        ELSE 'active'
                    END as status,
                    urgency.urgency_level
                FROM patients p
                LEFT JOIN LATERAL (
                    SELECT content, created_at FROM chat_messages
                    WHERE from_number = p.phone_number
                    ORDER BY created_at DESC LIMIT 1
                ) cm ON true
                LEFT JOIN LATERAL (
                    SELECT urgency_level FROM appointments
                    WHERE tenant_id = $1 AND patient_id = p.id
                    ORDER BY created_at DESC LIMIT 1
                ) urgency ON true
                WHERE p.tenant_id = $1
                AND EXISTS (SELECT 1 FROM chat_messages WHERE from_number = p.phone_number)
                ORDER BY p.phone_number, cm.created_at DESC
            ) sub
            ORDER BY last_message_time DESC NULLS LAST
        """,
            tenant_id,
        )
    else:
        rows = await db.pool.fetch(
            """
            SELECT * FROM (
                SELECT DISTINCT ON (p.phone_number)
                    p.phone_number,
                    p.id as patient_id,
                    p.first_name || ' ' || COALESCE(p.last_name, '') as patient_name,
                    cm.content as last_message,
                    cm.created_at as last_message_time,
                    p.human_handoff_requested,
                    p.human_override_until,
                    p.last_derivhumano_at,
                    CASE 
                        WHEN p.human_handoff_requested AND p.human_override_until > NOW() THEN 'human_handling'
                        WHEN p.human_override_until > NOW() THEN 'silenced'
                        ELSE 'active'
                    END as status,
                    urgency.urgency_level,
                    cc.last_read_at,
                    cc.last_user_message_at as conv_last_user_msg_at,
                    cc.id as conversation_id
                FROM patients p
                LEFT JOIN chat_conversations cc ON cc.tenant_id = p.tenant_id AND cc.channel = 'whatsapp' AND cc.external_user_id = p.phone_number
                LEFT JOIN LATERAL (
                    SELECT content, created_at FROM chat_messages
                    WHERE tenant_id = $1 AND from_number = p.phone_number AND conversation_id = cc.id
                    ORDER BY created_at DESC LIMIT 1
                ) cm ON true
                LEFT JOIN LATERAL (
                    SELECT urgency_level FROM appointments
                    WHERE tenant_id = $1 AND patient_id = p.id
                    ORDER BY created_at DESC LIMIT 1
                ) urgency ON true
                WHERE p.tenant_id = $1
                AND EXISTS (SELECT 1 FROM chat_messages WHERE tenant_id = $1 AND from_number = p.phone_number)
                ORDER BY p.phone_number, cm.created_at DESC
            ) sub
            ORDER BY last_message_time DESC NULLS LAST
        """,
            tenant_id,
        )
    sessions = []
    for row in rows:
        unread_sql = """
            SELECT COUNT(*) FROM chat_messages
            WHERE correlation_id = (SELECT id FROM chat_conversations WHERE tenant_id = $1 AND id = $2) -- Just a safety check, or use conversation_id directly
            OR conversation_id = $2
            AND role = 'user'
            AND tenant_id = $1
            AND created_at > GREATEST(
                COALESCE(
                    (SELECT created_at FROM chat_messages
                     WHERE (conversation_id = $2 OR (from_number = $3 AND tenant_id = $1)) AND role = 'assistant'
                     ORDER BY created_at DESC LIMIT 1),
                    '1970-01-01'::timestamptz
                ),
                COALESCE($4, '1970-01-01'::timestamptz)
            )
        """
        # unread_sql simplificado con ID de conversación (Spec 14 isolation)
        unread_sql = """
            SELECT COUNT(*) FROM chat_messages
            WHERE (conversation_id = $1 OR (from_number = $2 AND tenant_id = $3))
            AND role = 'user'
            AND tenant_id = $3
            AND created_at > GREATEST(
                COALESCE(
                    (SELECT created_at FROM chat_messages
                     WHERE (conversation_id = $1 OR (from_number = $2 AND tenant_id = $3)) 
                     AND role = 'assistant' AND tenant_id = $3
                     ORDER BY created_at DESC LIMIT 1),
                    '1970-01-01'::timestamptz
                ),
                COALESCE($4, '1970-01-01'::timestamptz)
            )
        """
        unread = await db.pool.fetchval(
            unread_sql,
            row.get("conversation_id"),
            row["phone_number"],
            tenant_id,
            row.get("last_read_at"),
        )

        last_user_msg = row.get("conv_last_user_msg_at")
        is_window_open = False
        if last_user_msg:
            # Ventana estricta de 24h (Spec 14 / Clarificación)
            now_utc = datetime.now(timezone.utc)
            msg_utc = (
                last_user_msg
                if last_user_msg.tzinfo
                else last_user_msg.replace(tzinfo=timezone.utc)
            )
            is_window_open = (now_utc - msg_utc) < timedelta(hours=24)
        sessions.append(
            {
                "phone_number": row["phone_number"],
                "patient_id": row["patient_id"],
                "patient_name": row["patient_name"],
                "tenant_id": tenant_id,
                "last_message": row["last_message"] or "",
                "last_message_time": row["last_message_time"].isoformat()
                if row["last_message_time"]
                else None,
                "unread_count": unread or 0,
                "status": row["status"],
                "human_override_until": row["human_override_until"].isoformat()
                if row["human_override_until"]
                else None,
                "urgency_level": row["urgency_level"],
                "last_derivhumano_at": row["last_derivhumano_at"].isoformat()
                if row["last_derivhumano_at"]
                else None,
                "is_window_open": is_window_open,
                "last_user_message_time": last_user_msg.isoformat()
                if last_user_msg
                else None,
            }
        )
    if sessions:
        from core.auth import log_pii_access

        await log_pii_access(
            request, user_data, f"Tenant:{tenant_id}", action="list_chat_sessions"
        )
    return sessions


@router.get(
    "/chat/messages/{phone}", dependencies=[Depends(verify_admin_token)], tags=["Chat"]
)
async def get_chat_messages(
    phone: str,
    tenant_id: int,
    request: Request,
    limit: int = 50,
    offset: int = 0,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Historial de mensajes para un número en la clínica indicada. Aislado por tenant_id."""
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    has_tenant = await db.pool.fetchval(
        "SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public' AND table_name='chat_messages' AND column_name='tenant_id')"
    )
    if has_tenant:
        rows = await db.pool.fetch(
            """
            SELECT id, from_number, role, content, created_at, correlation_id, content_attributes
            FROM chat_messages
            WHERE from_number = $1 AND tenant_id = $2
            ORDER BY created_at DESC
            LIMIT $3 OFFSET $4
        """,
            phone,
            tenant_id,
            limit,
            offset,
        )
    else:
        rows = await db.pool.fetch(
            """
            SELECT id, from_number, role, content, created_at, correlation_id, content_attributes
            FROM chat_messages
            WHERE from_number = $1
            ORDER BY created_at DESC
            LIMIT $2 OFFSET $3
        """,
            phone,
            limit,
            offset,
        )

    # Invertir para que lleguen en orden cronológico al frontend
    rows = sorted(rows, key=lambda x: x["created_at"])

    messages = []
    for row in rows:
        # Detectar si es un mensaje de derivhumano (sistema indica handoff)
        is_derivhumano = (
            row["role"] == "assistant"
            and "representante humano" in row["content"].lower()
        )

        # Parse content_attributes if it's a JSON string
        attachments = []
        raw_attrs = row.get("content_attributes")
        if raw_attrs:
            if isinstance(raw_attrs, str):
                try:
                    attachments = json.loads(raw_attrs)
                except (json.JSONDecodeError, TypeError):
                    attachments = []
            elif isinstance(raw_attrs, list):
                attachments = raw_attrs

        messages.append(
            {
                "id": row["id"],
                "from_number": row["from_number"],
                "role": row["role"],
                "content": row["content"],
                "created_at": row["created_at"].isoformat(),
                "is_derivhumano": is_derivhumano,
                "attachments": attachments,
            }
        )

    logger.info(
        f"📤 Devueltos {len(messages)} mensajes para {phone} (attachments check: {[len(m.get('attachments', []) or []) for m in messages]})"
    )
    if messages:
        from core.auth import log_pii_access

        await log_pii_access(
            request, user_data, f"Phone:{phone}", action="read_chat_messages"
        )
    return messages


@router.put(
    "/chat/sessions/{phone}/read",
    dependencies=[Depends(verify_admin_token)],
    tags=["Chat"],
)
async def mark_chat_session_read(
    phone: str,
    tenant_id: int,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Marca la conversación (tenant_id, phone) como leída. Por clínica."""
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")

    # Persistir estado de lectura (Spec 14 / Fase 3)
    await db.execute(
        """
        UPDATE chat_conversations 
        SET last_read_at = NOW(), updated_at = NOW()
        WHERE tenant_id = $1 AND channel = 'whatsapp' AND external_user_id = $2
    """,
        tenant_id,
        phone,
    )

    return {"status": "ok", "phone": phone, "tenant_id": tenant_id}


@router.post(
    "/chat/human-intervention",
    dependencies=[Depends(verify_admin_token)],
    tags=["Chat"],
)
async def toggle_human_intervention(
    payload: HumanInterventionToggle,
    request: Request,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Activa o desactiva la intervención humana para un chat (tenant_id + phone).
    Independiente por clínica: intervención en Clínica A no afecta Clínica B.
    """
    if payload.tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    if payload.activate:
        try:
            from services.tz_resolver import get_tenant_tz

            _tz = await get_tenant_tz(payload.tenant_id)
        except Exception:
            _tz = ARG_TZ
        override_until = datetime.now(_tz) + timedelta(milliseconds=payload.duration)
        await db.pool.execute(
            """
            UPDATE patients
            SET human_handoff_requested = TRUE,
                human_override_until = $1,
                last_derivhumano_at = NULL,
                updated_at = NOW()
            WHERE tenant_id = $2 AND phone_number = $3
        """,
            override_until,
            payload.tenant_id,
            payload.phone,
        )
        # Spec 24 / Phase 5: Sync with chat_conversations for omnichannel buffering parity
        await db.pool.execute(
            """
            UPDATE chat_conversations
            SET human_override_until = $1,
                updated_at = NOW()
            WHERE tenant_id = $2 AND channel = 'whatsapp' AND external_user_id = $3
        """,
            override_until,
            payload.tenant_id,
            payload.phone,
        )

        logger.info(
            f"👤 Intervención humana activada para {payload.phone} (tenant={payload.tenant_id}) hasta {override_until}"
        )
        await emit_appointment_event(
            "HUMAN_OVERRIDE_CHANGED",
            {
                "phone_number": payload.phone,
                "tenant_id": payload.tenant_id,
                "enabled": True,
                "until": override_until.isoformat(),
            },
            request,
        )
        return {
            "status": "activated",
            "phone": payload.phone,
            "tenant_id": payload.tenant_id,
            "until": override_until.isoformat(),
        }
    else:
        await db.pool.execute(
            """
            UPDATE patients
            SET human_handoff_requested = FALSE,
                human_override_until = NULL,
                updated_at = NOW()
            WHERE tenant_id = $1 AND phone_number = $2
        """,
            payload.tenant_id,
            payload.phone,
        )
        # Spec 24 / Phase 5: Sync with chat_conversations for omnichannel buffering parity
        await db.pool.execute(
            """
            UPDATE chat_conversations
            SET human_override_until = NULL,
                updated_at = NOW()
            WHERE tenant_id = $1 AND channel = 'whatsapp' AND external_user_id = $2
        """,
            payload.tenant_id,
            payload.phone,
        )

        logger.info(
            f"🤖 IA reactivada para {payload.phone} (tenant={payload.tenant_id})"
        )
        await emit_appointment_event(
            "HUMAN_OVERRIDE_CHANGED",
            {
                "phone_number": payload.phone,
                "tenant_id": payload.tenant_id,
                "enabled": False,
            },
            request,
        )
        return {
            "status": "deactivated",
            "phone": payload.phone,
            "tenant_id": payload.tenant_id,
        }


@router.post(
    "/chat/remove-silence", dependencies=[Depends(verify_admin_token)], tags=["Chat"]
)
async def remove_silence(
    payload: dict,
    request: Request,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Remueve el silencio de la IA para (tenant_id, phone). Por clínica."""
    phone = payload.get("phone")
    tenant_id = payload.get("tenant_id")
    if not phone or tenant_id is None:
        raise HTTPException(status_code=400, detail="phone y tenant_id requeridos")
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    await db.pool.execute(
        """
        UPDATE patients
        SET human_handoff_requested = FALSE,
            human_override_until = NULL,
            updated_at = NOW()
        WHERE tenant_id = $1 AND phone_number = $2
    """,
        tenant_id,
        phone,
    )
    await emit_appointment_event(
        "HUMAN_OVERRIDE_CHANGED",
        {
            "phone_number": phone,
            "tenant_id": tenant_id,
            "enabled": False,
        },
        request,
    )
    return {"status": "removed", "phone": phone, "tenant_id": tenant_id}


@router.get("/internal/credentials/{name}", tags=["Internal"])
async def get_internal_credential(name: str, x_internal_token: str = Header(None)):
    """Permite a servicios internos obtener credenciales de forma segura."""
    if not INTERNAL_API_TOKEN:
        raise HTTPException(
            status_code=503, detail="Internal credentials endpoint disabled"
        )
    if x_internal_token != INTERNAL_API_TOKEN:
        raise HTTPException(status_code=401, detail="Internal token invalid")

    # Mapeo de nombres a variables de entorno o valores de BD
    creds = {
        "YCLOUD_API_KEY": os.getenv("YCLOUD_API_KEY"),
        "YCLOUD_WEBHOOK_SECRET": os.getenv("YCLOUD_WEBHOOK_SECRET"),
        "OPENAI_API_KEY": os.getenv("OPENAI_API_KEY"),
        "YCLOUD_Phone_Number_ID": os.getenv("YCLOUD_Phone_Number_ID")
        or os.getenv("BOT_PHONE_NUMBER"),
    }

    val = creds.get(name)
    if val is None:  # Solo 404 si la clave no existe en nuestro mapeo
        raise HTTPException(
            status_code=404,
            detail=f"Credential '{name}' not supported by this endpoint",
        )

    return {"name": name, "value": val}


# ==================== ENDPOINTS CREDENCIALES (SEGURIDAD) ====================

# --- SPEC 19: MEDIA MANAGEMENT & MAINTENANCE ---


@router.get("/chat/media/proxy", tags=["Chat"])
async def media_proxy(url: str, user_data=Depends(verify_admin_token)):
    """
    Proxy seguro para servir archivos multimedia (YCloud, Local, Chatwoot).
    Permite visualizar contenido sin exponer URLs directas o lidiar con CORS/Expiración.
    """
    logger.info(f"📥 Proxy media request: url={url[:100]}...")

    # SSRF protection: validate URL before proxying
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Only HTTP/HTTPS URLs allowed")
    hostname = parsed.hostname or ""
    if (
        hostname in ("localhost", "127.0.0.1", "0.0.0.0")
        or hostname.startswith("10.")
        or hostname.startswith("192.168.")
        or hostname.startswith("172.")
        or hostname == "169.254.169.254"
    ):
        raise HTTPException(status_code=400, detail="Internal URLs not allowed")

    # Headers con autenticación para YCloud si disponible
    headers = {}
    ycloud_key = os.getenv("YCLOUD_API_KEY")
    if ycloud_key and "ycloud" in url.lower():
        headers["X-API-Key"] = ycloud_key

    # Timeout aumentado a 60 segundos para archivos grandes (imágenes HD, videos, audios)
    async with httpx.AsyncClient(
        timeout=60.0, follow_redirects=True, headers=headers
    ) as client:
        try:
            resp = await client.get(url)
            logger.info(
                f"📥 Proxy response status: {resp.status_code}, content-type: {resp.headers.get('Content-Type')}"
            )
            if resp.status_code != 200:
                raise HTTPException(
                    status_code=resp.status_code, detail="Remote media unreachable"
                )

            return StreamingResponse(
                resp.iter_bytes(),
                media_type=resp.headers.get("Content-Type", "application/octet-stream"),
            )
        except httpx.TimeoutException:
            logger.error(f"⏱️ Proxy timeout para URL: {url[:100]}")
            raise HTTPException(
                status_code=504,
                detail="Timeout descargando media (archivo muy grande o conexión lenta)",
            )
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Proxy error: {e}")
            raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.post("/chat/transcribe-again", tags=["Chat"])
async def transcribe_again(
    payload: Dict[str, Any],
    background_tasks: BackgroundTasks,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Manual re-trigger of Whisper transcription for a given audio message.
    Payload: { 'url': str, 'tenant_id': int, 'conversation_id': str, 'external_user_id': str }
    """
    tid = payload.get("tenant_id")
    if tid not in allowed_ids:
        raise HTTPException(status_code=403, detail="Sin permiso para este tenant.")

    background_tasks.add_task(
        transcribe_audio_url,
        url=payload.get("url"),
        tenant_id=tid,
        conversation_id=payload.get("conversation_id"),
        external_user_id=payload.get("external_user_id"),
    )
    return {"message": "Transcription re-queued successfully."}


@router.post("/maintenance/clean-media", tags=["Mantenimiento"])
async def clean_media_maintenance(
    payload: Dict[str, Any], user_data=Depends(verify_admin_token)
):
    """
    Elimina archivos de medios antiguos en el servidor.
    Payload: { 'days': int } (e.g. 180 para 6 meses)
    Solo CEO.
    """
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Acceso denegado.")

    days = payload.get("days", 180)
    cutoff = datetime.now() - timedelta(days=days)

    # Directorio de medios (ajustar según despliegue)
    media_dir = os.getenv("MEDIA_DIR", "/public/media")
    if not os.path.exists(media_dir):
        return {"message": "Dircetorio de medios no encontrado.", "deleted": 0}

    deleted_count = 0
    # Buscar archivos recursivamente
    for filepath in glob.glob(os.path.join(media_dir, "**", "*"), recursive=True):
        if os.path.isfile(filepath):
            mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
            if mtime < cutoff:
                try:
                    os.remove(filepath)
                    deleted_count += 1
                except Exception as e:
                    logger.error(f"Failed to delete {filepath}: {e}")

    return {"message": f"Limpieza completada.", "deleted": deleted_count}


# ==================== ENDPOINTS CREDENCIALES (SEGURIDAD) ====================


@router.get(
    "/credentials", dependencies=[Depends(verify_admin_token)], tags=["Credenciales"]
)
async def list_credentials(
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
    user_data=Depends(verify_admin_token),
):
    """
    Lista credenciales.
    - Globales: Solo visibles por CEO.
    - Tenant: Visibles por CEO y Admin de esa sede.
    Las valores se devuelven enmascarados (Sanitized).
    """
    try:
        # Consulta base
        query = "SELECT id, name, value, category, description, scope, tenant_id FROM credentials"
        params = []
        conditions = []

        if user_data.role != "ceo":
            # No CEO: solo scope='tenant' y tenant_id en allowed_ids (o tenant_id NULL para algunos casos?)
            # Asumimos que secretarias solo ven credenciales de SU tenant.
            conditions.append("scope = 'tenant'")
            conditions.append(f"tenant_id = ANY(${len(params) + 1}::int[])")
            params.append(allowed_ids)
        else:
            # CEO: ve todo global + todo tenant
            pass

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY category, name"

        rows = await db.pool.fetch(query, *params)

        results = []
        for r in rows:
            val = r["value"]
            # Enmascarar valor para UI
            masked = "••••••••"
            if val and len(str(val)) > 4:
                masked = f"••••••••{str(val)[-4:]}"

            # Ocultar valor real
            results.append(
                {
                    "id": r["id"],
                    "name": r["name"],
                    "value": masked,  # Frontend espera string, no enviamos el real
                    "category": r["category"],
                    "description": r["description"],
                    "scope": r["scope"],
                    "tenant_id": r["tenant_id"],
                }
            )
        return results
    except Exception as e:
        logger.error(f"Error listing credentials: {e}")
        raise HTTPException(status_code=500, detail="Error al listar credenciales")


@router.post(
    "/credentials", dependencies=[Depends(verify_admin_token)], tags=["Credenciales"]
)
async def create_or_update_credential(
    payload: CredentialPayload,
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
    user_data=Depends(verify_admin_token),
):
    """
    Crea o actualiza una credencial.
    - SIEMPRE encripta el valor (Fernet) antes de guardar.
    - Validación de permisos por scope/tenant.
    """
    # 1. Validaciones de Seguridad
    if payload.scope == "global" and user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar credenciales globales."
        )

    target_tenant_id = payload.tenant_id
    if payload.scope == "tenant":
        if target_tenant_id is None:
            # Si no viene, usamos el resuelto (propia clínica)
            target_tenant_id = resolved_tenant_id

        # Verificar acceso
        if target_tenant_id not in allowed_ids:
            raise HTTPException(
                status_code=403,
                detail="No tienes permiso para gestionar credenciales de esta clínica.",
            )
    else:
        target_tenant_id = None  # Global

    # 2. Encriptación
    encrypted_value = encrypt_value(payload.value)
    if not encrypted_value:
        raise HTTPException(
            status_code=503,
            detail="Error de encriptación. Verificar CREDENTIALS_FERNET_KEY.",
        )

    try:
        # 3. Upsert vs Update vs Insert
        # La UI envia ID si edita.
        if payload.id:
            # Update existente
            # Verificar ownership antes de update
            existing = await db.pool.fetchrow(
                "SELECT id, scope, tenant_id FROM credentials WHERE id = $1", payload.id
            )
            if not existing:
                raise HTTPException(status_code=404, detail="Credencial no encontrada")

            # Si era global y user no es CEO -> ya validado arriba, pero por si acaso cambio de scope
            if existing["scope"] == "global" and user_data.role != "ceo":
                raise HTTPException(
                    status_code=403, detail="No puedes editar credenciales globales."
                )
            if (
                existing["scope"] == "tenant"
                and existing["tenant_id"] not in allowed_ids
            ):
                raise HTTPException(
                    status_code=403, detail="No tienes acceso a esta credencial."
                )

            await db.pool.execute(
                """
                UPDATE credentials 
                SET name = $1, value = $2, category = $3, description = $4, scope = $5, tenant_id = $6, updated_at = NOW()
                WHERE id = $7
            """,
                payload.name,
                encrypted_value,
                payload.category,
                payload.description,
                payload.scope,
                target_tenant_id,
                payload.id,
            )
            action = "updated"
        else:
            # Insert
            await db.pool.execute(
                """
                INSERT INTO credentials (name, value, category, description, scope, tenant_id, created_at, updated_at)
                VALUES ($1, $2, $3, $4, $5, $6, NOW(), NOW())
            """,
                payload.name,
                encrypted_value,
                payload.category,
                payload.description,
                payload.scope,
                target_tenant_id,
            )
            action = "created"

        logger.info(
            f"Credential {action}: name={payload.name} scope={payload.scope} tenant={target_tenant_id} by {user_data.email}"
        )
        return {"status": "ok", "action": action}

    except Exception as e:
        logger.error(f"Error saving credential: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.delete(
    "/credentials/{id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Credenciales"],
)
async def delete_credential(
    id: int,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
    user_data=Depends(verify_admin_token),
):
    try:
        existing = await db.pool.fetchrow(
            "SELECT scope, tenant_id FROM credentials WHERE id = $1", id
        )
        if not existing:
            return {"status": "deleted"}  # Idempotente

        if existing["scope"] == "global" and user_data.role != "ceo":
            raise HTTPException(
                status_code=403, detail="Solo CEO puede eliminar credenciales globales."
            )

        if existing["scope"] == "tenant" and existing["tenant_id"] not in allowed_ids:
            raise HTTPException(
                status_code=403, detail="No tienes acceso a esta clínica."
            )

        await db.pool.execute("DELETE FROM credentials WHERE id = $1", id)
        return {"status": "deleted"}
    except Exception as e:
        logger.error(f"Error deleting credential: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.post(
    "/faqs/sync-embeddings", tags=["FAQs"], summary="Sincronizar embeddings de FAQs"
)
async def sync_faq_embeddings(
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Sincroniza manualmente los embeddings de FAQs (útil después de inserts masivos por SQL)."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede sync embeddings."
        )

    try:
        from services.embedding_service import sync_all_tenants_faq_embeddings
        import asyncio

        count = await sync_all_tenants_faq_embeddings()
        return {"status": "completed", "embeddings_synced": count}
    except Exception as e:
        logger.error(f"FAQ embedding sync failed: {e}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting credential: {e}")
        raise HTTPException(status_code=500, detail="Error al eliminar credencial")


@router.put(
    "/credentials/{cred_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Credenciales"],
)
async def update_credential(
    cred_id: int,
    payload: CredentialPayload,
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
    user_data=Depends(verify_admin_token),
):
    """
    Actualiza una credencial existente.
    - Verifica ownership antes de modificar.
    - Si el valor llega enmascarado (••••), conserva el valor encriptado anterior.
    - Solo CEO puede editar credenciales globales.
    """
    try:
        existing = await db.pool.fetchrow(
            "SELECT id, scope, tenant_id, value FROM credentials WHERE id = $1", cred_id
        )
        if not existing:
            raise HTTPException(status_code=404, detail="Credencial no encontrada.")

        if existing["scope"] == "global" and user_data.role != "ceo":
            raise HTTPException(
                status_code=403, detail="Solo CEO puede editar credenciales globales."
            )

        if existing["scope"] == "tenant" and existing["tenant_id"] not in allowed_ids:
            raise HTTPException(
                status_code=403, detail="No tienes acceso a esta credencial."
            )

        if payload.scope == "global" and user_data.role != "ceo":
            raise HTTPException(
                status_code=403, detail="Solo CEO puede asignar scope global."
            )

        target_tenant_id = payload.tenant_id
        if payload.scope == "tenant" and target_tenant_id is None:
            target_tenant_id = resolved_tenant_id

        # Conservar valor encriptado anterior si el frontend envía el valor enmascarado
        if payload.value and not payload.value.startswith("••••"):
            new_value = encrypt_value(payload.value)
            if not new_value:
                raise HTTPException(
                    status_code=503,
                    detail="Error de encriptación. Verificar CREDENTIALS_FERNET_KEY.",
                )
        else:
            new_value = existing["value"]

        await db.pool.execute(
            """
            UPDATE credentials
            SET name = $1, value = $2, category = $3, description = $4, scope = $5, tenant_id = $6, updated_at = NOW()
            WHERE id = $7
        """,
            payload.name,
            new_value,
            payload.category,
            payload.description,
            payload.scope,
            target_tenant_id,
            cred_id,
        )

        logger.info(
            f"Credential updated: id={cred_id} name={payload.name} scope={payload.scope} by {user_data.email}"
        )
        return {"status": "ok", "action": "updated"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating credential {cred_id}: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.post("/chat/send", dependencies=[Depends(verify_admin_token)], tags=["Chat"])
async def send_chat_message(
    payload: ChatSendMessage,
    request: Request,
    background_tasks: BackgroundTasks,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Envía un mensaje manual por WhatsApp; guarda en BD con tenant_id. Ventana 24h por clínica."""
    if payload.tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    try:
        has_tenant = await db.pool.fetchval(
            "SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public' AND table_name='chat_messages' AND column_name='tenant_id')"
        )
        if has_tenant:
            last_user_msg = await db.pool.fetchval(
                """
                SELECT created_at FROM chat_messages
                WHERE from_number = $1 AND role = 'user' AND tenant_id = $2
                ORDER BY created_at DESC LIMIT 1
            """,
                payload.phone,
                payload.tenant_id,
            )
        else:
            last_user_msg = await db.pool.fetchval(
                """
                SELECT created_at FROM chat_messages
                WHERE from_number = $1 AND role = 'user'
                ORDER BY created_at DESC LIMIT 1
            """,
                payload.phone,
            )
        if not last_user_msg:
            raise HTTPException(
                status_code=403,
                detail="No se puede enviar un mensaje si el usuario nunca ha escrito.",
            )
        # 24h delta: timezone choice does not affect the result, but we still use the
        # tenant tz for consistency with the rest of the booking flow.
        try:
            from services.tz_resolver import get_tenant_tz

            _fallback_tz = await get_tenant_tz(payload.tenant_id)
        except Exception:
            _fallback_tz = ARG_TZ
        now_localized = datetime.now(
            last_user_msg.tzinfo if last_user_msg.tzinfo else _fallback_tz
        )
        if (now_localized - last_user_msg) > timedelta(hours=24):
            raise HTTPException(
                status_code=403,
                detail="La ventana de 24hs de WhatsApp ha expirado. El paciente debe escribir primero.",
            )
        correlation_id = str(uuid.uuid4())
        await db.append_chat_message(
            from_number=payload.phone,
            role="assistant",
            content=payload.message,
            correlation_id=correlation_id,
            tenant_id=payload.tenant_id,
        )
        if hasattr(request.app.state, "emit_appointment_event"):
            await request.app.state.emit_appointment_event(
                "NEW_MESSAGE",
                {
                    "phone_number": payload.phone,
                    "tenant_id": payload.tenant_id,
                    "message": payload.message,
                    "role": "assistant",
                },
            )
        business_number = (
            os.getenv("YCLOUD_Phone_Number_ID")
            or os.getenv("BOT_PHONE_NUMBER")
            or "default"
        )
        background_tasks.add_task(
            send_to_whatsapp_task, payload.phone, payload.message, business_number
        )
        return {"status": "sent", "correlation_id": correlation_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error sending manual message: {str(e)}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ==================== ENDPOINTS DASHBOARD ====================


@router.get(
    "/dashboard/metrics",
    tags=["Dashboard"],
    summary="Métricas de tokens y agente IA (alias para frontend)",
)
@limiter.limit("10/minute")
async def get_dashboard_token_metrics(
    request: Request,
    days: int = Query(30, ge=1, le=365),
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    API de métricas de tokens del agente IA. Consumida por DashboardStatusView.
    Alias bajo /admin para garantizar que el proxy envíe al backend.
    """
    try:
        from dashboard.status_page import fetch_dashboard_metrics

        return await fetch_dashboard_metrics(days=days, tenant_id=tenant_id)
    except Exception as e:
        err_msg = str(e).lower()
        if "jinja2" in err_msg or "import" in err_msg or isinstance(e, ImportError):
            logger.warning(f"Dashboard metrics: dependencia no disponible ({e})")
            return {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "status": "modules_not_available",
                "message": "Módulos del dashboard no disponibles (jinja2 u otra dependencia faltante)",
                "token_metrics": {
                    "totals": {
                        "total_cost_usd": 0,
                        "total_tokens": 0,
                        "total_conversations": 0,
                    },
                    "today": {},
                    "current_month": {},
                },
                "daily_usage": [],
                "model_usage": [],
                "current_config": {},
                "system_metrics": {},
                "db_stats": {},
                "projections": {},
            }
        logger.error(f"Error obteniendo métricas dashboard: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.get(
    "/stats/summary",
    tags=["Estadísticas"],
    summary="Obtener resumen de métricas del dashboard",
)
@limiter.limit("10/minute")
async def get_dashboard_stats(
    request: Request,
    range: str = "all",
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Devuelve métricas avanzadas filtradas por rango temporal. Aislado por tenant_id (Regla de Oro)."""
    try:
        # Normalizar el mapeo de intervalos para Postgres
        interval_map = {
            "weekly": "INTERVAL '7 days'",
            "monthly": "INTERVAL '30 days'",
            "yearly": "INTERVAL '1 year'",
            "all": "INTERVAL '100 years'",
        }
        interval_expr = interval_map.get(range, "INTERVAL '100 years'")

        # 1. IA Conversations (Unificado de chat_conversations para todos los canales e hilos)
        ia_conversations = (
            await db.pool.fetchval(
                f"""
            SELECT COUNT(*) FROM chat_conversations
            WHERE tenant_id = $1 AND created_at >= CURRENT_DATE - {interval_expr}
        """,
                tenant_id,
            )
            or 0
        )

        # 2. IA Appointments (Turnos de IA en el rango seleccionado)
        ia_count = (
            await db.pool.fetchval(
                f"""
            SELECT COUNT(*) FROM appointments 
            WHERE tenant_id = $1 AND source = 'ai' AND appointment_datetime >= CURRENT_DATE - {interval_expr}
        """,
                tenant_id,
            )
            or 0
        )

        # 3. Urgencias activas (Pacientes marcados con urgencia)
        active_urgencies = (
            await db.pool.fetchval(
                f"""
            SELECT COUNT(*) FROM patients 
            WHERE tenant_id = $1 AND urgency_level IN ('high', 'emergency')
            AND updated_at >= CURRENT_DATE - {interval_expr}
        """,
                tenant_id,
            )
            or 0
        )

        # 4. Revenue Dual (Real vs Estimado)
        # 4a. Real Revenue (Pagos confirmados)
        confirmed_revenue = (
            await db.pool.fetchval(
                f"""
            SELECT COALESCE(SUM(at.amount), 0) 
            FROM accounting_transactions at
            WHERE at.tenant_id = $1 
            AND at.transaction_type = 'payment' 
            AND at.status = 'completed'
            AND at.created_at >= CURRENT_DATE - {interval_expr}
        """,
                tenant_id,
            )
            or 0
        )

        # 4b. Estimated Revenue (Valor de agenda IA - Basado en precios de tratamientos)
        # Si un turno no tiene tipo de tratamiento, usamos un falback de 100? No, mejor 0 o valor por defecto de la clínica.
        # Buscamos el precio base asignado a cada tipo de tratamiento.
        estimated_revenue = (
            await db.pool.fetchval(
                f"""
            SELECT COALESCE(SUM(tt.base_price), 0)
            FROM appointments a
            JOIN treatment_types tt ON a.appointment_type = tt.code AND tt.tenant_id = a.tenant_id
            WHERE a.tenant_id = $1 AND a.source = 'ai'
            AND a.status NOT IN ('cancelled')
            AND a.appointment_datetime >= CURRENT_DATE - {interval_expr}
        """,
                tenant_id,
            )
            or 0
        )

        # 5. Datos de crecimiento (Agrupación Inteligente: DÍA para Weekly/Monthly, MES para Anual/All)
        if range in ("yearly", "all"):
            growth_rows = await db.pool.fetch(
                f"""
                SELECT 
                    DATE_TRUNC('month', appointment_datetime) as period,
                    COUNT(*) FILTER (WHERE source = 'ai') as ia_referrals,
                    COUNT(*) FILTER (WHERE status IN ('completed', 'attended')) as completed_appointments
                FROM appointments
                WHERE tenant_id = $1 AND appointment_datetime >= CURRENT_DATE - {interval_expr}
                GROUP BY period
                ORDER BY period ASC
            """,
                tenant_id,
            )
        else:
            growth_rows = await db.pool.fetch(
                f"""
                SELECT 
                    DATE(appointment_datetime) as period,
                    COUNT(*) FILTER (WHERE source = 'ai') as ia_referrals,
                    COUNT(*) FILTER (WHERE status IN ('completed', 'attended')) as completed_appointments
                FROM appointments
                WHERE tenant_id = $1 AND appointment_datetime >= CURRENT_DATE - {interval_expr}
                GROUP BY period
                ORDER BY period ASC
            """,
                tenant_id,
            )

        growth_data = []
        for row in growth_rows:
            p = row["period"]
            date_str = (
                p.strftime("%Y-%m")
                if range in ("yearly", "all")
                else p.strftime("%Y-%m-%d")
            )
            growth_data.append(
                {
                    "date": date_str,
                    "ia_referrals": row["ia_referrals"],
                    "completed_appointments": row["completed_appointments"],
                }
            )

        if not growth_data:
            growth_data = [
                {
                    "date": date.today().isoformat(),
                    "ia_referrals": 0,
                    "completed_appointments": 0,
                }
            ]

        # 6. Calcular Rangos Previos para Tendencias (Crecimiento Real)
        prev_interval_expr = interval_expr  # El mismo intervalo pero desplazado

        # Consultas Previas
        prev_ia_conversations = (
            await db.pool.fetchval(
                f"""
            SELECT COUNT(*) FROM chat_conversations
            WHERE tenant_id = $1 
            AND created_at < CURRENT_DATE - {interval_expr}
            AND created_at >= CURRENT_DATE - {interval_expr} * 2
        """,
                tenant_id,
            )
            or 0
        )

        prev_ia_count = (
            await db.pool.fetchval(
                f"""
            SELECT COUNT(*) FROM appointments 
            WHERE tenant_id = $1 AND source = 'ai' 
            AND appointment_datetime < CURRENT_DATE - {interval_expr}
            AND appointment_datetime >= CURRENT_DATE - {interval_expr} * 2
        """,
                tenant_id,
            )
            or 0
        )

        prev_confirmed_revenue = (
            await db.pool.fetchval(
                f"""
            SELECT COALESCE(SUM(at.amount), 0) 
            FROM accounting_transactions at
            WHERE at.tenant_id = $1 
            AND at.transaction_type = 'payment' 
            AND at.status = 'completed'
            AND at.created_at < CURRENT_DATE - {interval_expr}
            AND at.created_at >= CURRENT_DATE - {interval_expr} * 2
        """,
                tenant_id,
            )
            or 0
        )

        # 7. Pending payments (non-cancelled appointments with pending/partial payment, excluding those linked to plans)
        # plus pending balance from approved/in_progress treatment plans
        legacy_appointments_pending = (
            await db.pool.fetchval(
                """
            SELECT COALESCE(SUM(billing_amount), 0)
            FROM appointments
            WHERE tenant_id = $1 
              AND payment_status IN ('pending', 'partial') 
              AND status NOT IN ('cancelled')
              AND plan_item_id IS NULL
        """,
                tenant_id,
            )
            or 0
        )

        plans_pending = (
            await db.pool.fetchval(
                """
            SELECT COALESCE(SUM(tp.approved_total - COALESCE(payments.total_paid, 0)), 0)
            FROM treatment_plans tp
            LEFT JOIN (
                SELECT plan_id, SUM(amount) as total_paid
                FROM treatment_plan_payments
                WHERE tenant_id = $1
                GROUP BY plan_id
            ) payments ON tp.id = payments.plan_id
            WHERE tp.tenant_id = $1
              AND tp.status IN ('approved', 'in_progress')
              AND tp.approved_total IS NOT NULL
              AND (payments.total_paid IS NULL OR payments.total_paid < tp.approved_total)
        """,
                tenant_id,
            )
            or 0
        )

        pending_payments = legacy_appointments_pending + plans_pending

        # 8. Today's revenue (paid appointments for today in Argentina timezone, excluding those linked to plans)
        # plus treatment plan payments recorded today
        legacy_appointments_today = (
            await db.pool.fetchval(
                """
            SELECT COALESCE(SUM(billing_amount), 0)
            FROM appointments
            WHERE tenant_id = $1 AND payment_status = 'paid'
              AND DATE(appointment_datetime AT TIME ZONE 'America/Argentina/Buenos_Aires') = CURRENT_DATE
              AND plan_item_id IS NULL
        """,
                tenant_id,
            )
            or 0
        )

        plan_payments_today = (
            await db.pool.fetchval(
                """
            SELECT COALESCE(SUM(amount), 0)
            FROM treatment_plan_payments
            WHERE tenant_id = $1 AND DATE(payment_date AT TIME ZONE 'America/Argentina/Buenos_Aires') = CURRENT_DATE
        """,
                tenant_id,
            )
            or 0
        )

        today_revenue = legacy_appointments_today + plan_payments_today

        def calc_trend(current, previous):
            if previous == 0:
                return "+100%" if current > 0 else "0%"
            change = ((current - previous) / previous) * 100
            return f"{'+' if change >= 0 else ''}{change:.1f}%"

        return {
            "ia_conversations": ia_conversations,
            "ia_conversations_trend": calc_trend(
                ia_conversations, prev_ia_conversations
            ),
            "ia_appointments": ia_count,
            "ia_appointments_trend": calc_trend(ia_count, prev_ia_count),
            "active_urgencies": active_urgencies,
            "total_revenue": float(confirmed_revenue),
            "total_revenue_trend": calc_trend(
                float(confirmed_revenue), float(prev_confirmed_revenue)
            ),
            "estimated_revenue": float(estimated_revenue),
            "pending_payments": float(pending_payments),
            "today_revenue": float(today_revenue),
            "growth_data": growth_data,
        }
    except Exception as e:
        logger.error(f"Error en get_dashboard_stats: {e}")
        raise HTTPException(status_code=500, detail="Error al cargar estadísticas.")


# --- CLÍNICAS (TENANTS) - CEO ONLY ---
# Tratamos "Tenant" como "Clínica". config (JSONB) incluye calendar_provider: 'local' | 'google'.


@router.get("/tenants", tags=["Sedes"], summary="Listar todas las sedes (clínicas)")
async def get_tenants(
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Lista las clínicas (tenants) accesibles por el usuario. Solo CEO."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar clínicas."
        )
    rows = await db.pool.fetch(
        "SELECT id, clinic_name, bot_phone_number, config, address, google_maps_url, working_hours, consultation_price, bank_cbu, bank_alias, bank_holder_name, derivation_email, logo_url, max_chairs, country_code, system_prompt_template, bot_name, payment_methods, financing_available, max_installments, installments_interest_free, financing_provider, financing_notes, cash_discount_percent, accepts_crypto, accepts_pregnant_patients, pregnancy_restricted_treatments, pregnancy_notes, accepts_pediatric, min_pediatric_age_years, pediatric_notes, high_risk_protocols, requires_anamnesis_before_booking, created_at, updated_at FROM tenants WHERE id = ANY($1::int[]) ORDER BY id ASC",
        allowed_ids,
    )
    result = []
    for r in rows:
        d = dict(r)
        # asyncpg puede devolver JSONB como string; asegurar que sean objetos
        for jfield in (
            "config",
            "working_hours",
            "payment_methods",
            "pregnancy_restricted_treatments",
            "high_risk_protocols",
        ):
            if isinstance(d.get(jfield), str):
                try:
                    d[jfield] = json.loads(d[jfield])
                except (json.JSONDecodeError, TypeError):
                    pass
        result.append(d)
    return result


@router.post("/tenants", tags=["Sedes"], summary="Crear una nueva sede")
async def create_tenant(data: Dict[str, Any], user_data=Depends(verify_admin_token)):
    """Crea una nueva clínica. config.calendar_provider obligatorio: 'local' o 'google'."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar clínicas."
        )
    calendar_provider = (
        data.get("calendar_provider")
        or data.get("config", {}).get("calendar_provider")
        or "local"
    ).lower()
    if calendar_provider not in ("local", "google"):
        calendar_provider = "local"
    config_json = json.dumps({"calendar_provider": calendar_provider})
    wh = data.get("working_hours")
    wh_json = json.dumps(wh) if wh else "{}"
    address = data.get("address") or None
    maps_url = data.get("google_maps_url") or None
    cp_raw = data.get("consultation_price")
    consultation_price = float(cp_raw) if cp_raw is not None and cp_raw != "" else None
    query = """
    INSERT INTO tenants (clinic_name, bot_phone_number, config, address, google_maps_url, working_hours, consultation_price, created_at)
    VALUES ($1, $2, $3::jsonb, $4, $5, $6::jsonb, $7, NOW())
    RETURNING id
    """
    try:
        new_id = await db.pool.fetchval(
            query,
            data.get("clinic_name"),
            data.get("bot_phone_number"),
            config_json,
            address,
            maps_url,
            wh_json,
            consultation_price,
        )
        return {"id": new_id, "status": "created"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put(
    "/tenants/{tenant_id}", tags=["Sedes"], summary="Actualizar datos de una sede"
)
async def update_tenant(
    tenant_id: int, data: Dict[str, Any], user_data=Depends(verify_admin_token)
):
    """Actualiza datos de una clínica. Incluye config.calendar_provider si se envía."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar clínicas."
        )
    updates = []
    params = []
    if "clinic_name" in data and data["clinic_name"] is not None:
        params.append(data["clinic_name"])
        updates.append(f"clinic_name = ${len(params)}")
    if "bot_phone_number" in data and data["bot_phone_number"] is not None:
        params.append(data["bot_phone_number"])
        updates.append(f"bot_phone_number = ${len(params)}")
    if "calendar_provider" in data and data["calendar_provider"] is not None:
        cp = (
            str(data["calendar_provider"]).lower()
            if data["calendar_provider"]
            else "local"
        )
        if cp not in ("local", "google"):
            cp = "local"
        # Persistir config como JSON explícito para evitar problemas de tipo en PostgreSQL
        config_patch = json.dumps({"calendar_provider": cp})
        params.append(config_patch)
        updates.append(
            f"config = COALESCE(config, '{{}}')::jsonb || ${len(params)}::jsonb"
        )
    if "address" in data:
        params.append(data.get("address") or None)
        updates.append(f"address = ${len(params)}")
    if "google_maps_url" in data:
        params.append(data.get("google_maps_url") or None)
        updates.append(f"google_maps_url = ${len(params)}")
    if "working_hours" in data and data["working_hours"] is not None:
        params.append(json.dumps(data["working_hours"]))
        updates.append(f"working_hours = ${len(params)}::jsonb")
    if "consultation_price" in data:
        val = data.get("consultation_price")
        params.append(float(val) if val is not None and val != "" else None)
        updates.append(f"consultation_price = ${len(params)}")
    if "bank_cbu" in data:
        params.append(data.get("bank_cbu") or None)
        updates.append(f"bank_cbu = ${len(params)}")
    if "bank_alias" in data:
        params.append(data.get("bank_alias") or None)
        updates.append(f"bank_alias = ${len(params)}")
    if "bank_holder_name" in data:
        params.append(data.get("bank_holder_name") or None)
        updates.append(f"bank_holder_name = ${len(params)}")
    if "derivation_email" in data:
        params.append(data.get("derivation_email") or None)
        updates.append(f"derivation_email = ${len(params)}")
    if "max_chairs" in data and data["max_chairs"] is not None:
        val = data.get("max_chairs")
        params.append(int(val) if val is not None and str(val).strip() != "" else 2)
        updates.append(f"max_chairs = ${len(params)}")
    if "country_code" in data and data["country_code"] is not None:
        cc = str(data["country_code"]).upper().strip()[:2]
        if len(cc) == 2:
            params.append(cc)
            updates.append(f"country_code = ${len(params)}")
    if "system_prompt_template" in data:
        # Normalize whitespace on save: collapse single newlines into spaces,
        # preserve double-newlines as paragraph separators, kill double spaces.
        # This mirrors what buffer_task does at runtime, so the value stored
        # in the DB is already clean and not just clean at render time.
        _raw = data.get("system_prompt_template")
        if _raw is None or _raw == "":
            params.append(None)
        else:
            import re as _re_spt

            _paragraphs = _re_spt.split(r"\n\s*\n", str(_raw))
            _clean = []
            for _p in _paragraphs:
                _flat = " ".join(
                    line.strip() for line in _p.splitlines() if line.strip()
                )
                _flat = _re_spt.sub(r"[ \t]{2,}", " ", _flat).strip()
                if _flat:
                    _clean.append(_flat)
            params.append("\n\n".join(_clean) if _clean else None)
        updates.append(f"system_prompt_template = ${len(params)}")
    if "bot_name" in data:
        # Editable bot display name (migration 033). NULL → fallback to "TORA"
        # at injection time in buffer_task. Manual validation since the
        # endpoint uses Dict[str, Any] instead of a Pydantic schema.
        _raw_bn = data.get("bot_name")
        _normalized_bn = (
            _raw_bn.strip() if isinstance(_raw_bn, str) and _raw_bn.strip() else None
        )
        if _normalized_bn is not None:
            import re as _re_bn

            if len(_normalized_bn) > 50 or not _re_bn.match(
                r"^[A-Za-z0-9 _-]+$", _normalized_bn
            ):
                raise HTTPException(
                    status_code=422,
                    detail="bot_name: máximo 50 caracteres, solo letras, números, espacios, guiones y guiones bajos.",
                )
        params.append(_normalized_bn)
        updates.append(f"bot_name = ${len(params)}")
    # === Payment & Financing (migration 035) ===
    if "payment_methods" in data:
        _raw_pm = data.get("payment_methods")
        if _raw_pm is None:
            params.append(None)
            updates.append(f"payment_methods = ${len(params)}")
        elif isinstance(_raw_pm, list):
            ALLOWED_METHODS = {
                "cash",
                "credit_card",
                "debit_card",
                "transfer",
                "mercado_pago",
                "rapipago",
                "pagofacil",
                "modo",
                "uala",
                "naranja",
                "crypto",
                "other",
            }
            invalid = [m for m in _raw_pm if m not in ALLOWED_METHODS]
            if invalid:
                raise HTTPException(
                    status_code=422,
                    detail=f"payment_methods contains invalid values: {invalid}",
                )
            params.append(json.dumps(_raw_pm))
            updates.append(f"payment_methods = ${len(params)}::jsonb")
    if "financing_available" in data:
        _raw_fa = data.get("financing_available")
        _val_fa = bool(_raw_fa) if _raw_fa is not None else False
        params.append(_val_fa)
        updates.append(f"financing_available = ${len(params)}")
    if "max_installments" in data:
        _raw_mi = data.get("max_installments")
        if _raw_mi is not None:
            try:
                _val_mi = int(str(_raw_mi).strip())
                if not (1 <= _val_mi <= 24):
                    raise HTTPException(
                        status_code=422,
                        detail="max_installments must be between 1 and 24",
                    )
                params.append(_val_mi)
                updates.append(f"max_installments = ${len(params)}")
            except (ValueError, TypeError):
                raise HTTPException(
                    status_code=422,
                    detail="max_installments must be a valid integer",
                )
    if "installments_interest_free" in data:
        _raw_iif = data.get("installments_interest_free")
        _val_iif = bool(_raw_iif) if _raw_iif is not None else True
        params.append(_val_iif)
        updates.append(f"installments_interest_free = ${len(params)}")
    if "financing_provider" in data:
        _raw_fp = data.get("financing_provider")
        _val_fp = (
            _raw_fp.strip() if isinstance(_raw_fp, str) and _raw_fp.strip() else None
        )
        params.append(_val_fp)
        updates.append(f"financing_provider = ${len(params)}")
    if "financing_notes" in data:
        _raw_fn = data.get("financing_notes")
        _val_fn = (
            _raw_fn.strip() if isinstance(_raw_fn, str) and _raw_fn.strip() else None
        )
        params.append(_val_fn)
        updates.append(f"financing_notes = ${len(params)}")
    if "cash_discount_percent" in data:
        _raw_cdp = data.get("cash_discount_percent")
        if _raw_cdp is not None:
            try:
                _val_cdp = float(str(_raw_cdp).strip())
                if not (0 <= _val_cdp <= 100):
                    raise HTTPException(
                        status_code=422,
                        detail="cash_discount_percent must be between 0 and 100",
                    )
                params.append(_val_cdp)
                updates.append(f"cash_discount_percent = ${len(params)}")
            except (ValueError, TypeError):
                raise HTTPException(
                    status_code=422,
                    detail="cash_discount_percent must be a valid number",
                )
    if "accepts_crypto" in data:
        _raw_ac = data.get("accepts_crypto")
        _val_ac = bool(_raw_ac) if _raw_ac is not None else False
        params.append(_val_ac)
        updates.append(f"accepts_crypto = ${len(params)}")

    # ----- Clinic special conditions (migration 036) -----
    if "accepts_pregnant_patients" in data and data["accepts_pregnant_patients"] is not None:
        params.append(bool(data["accepts_pregnant_patients"]))
        updates.append(f"accepts_pregnant_patients = ${len(params)}")

    if "pregnancy_restricted_treatments" in data:
        _val_prt = data.get("pregnancy_restricted_treatments")
        if _val_prt is None:
            params.append(None)
        else:
            if not isinstance(_val_prt, list) or not all(
                isinstance(c, str) for c in _val_prt
            ):
                raise HTTPException(
                    status_code=422,
                    detail="pregnancy_restricted_treatments must be a list of strings",
                )
            params.append(json.dumps(_val_prt))
        updates.append(f"pregnancy_restricted_treatments = ${len(params)}::jsonb")

    if "pregnancy_notes" in data:
        _val_pn = data.get("pregnancy_notes")
        params.append(_val_pn.strip() if isinstance(_val_pn, str) and _val_pn.strip() else None)
        updates.append(f"pregnancy_notes = ${len(params)}")

    if "accepts_pediatric" in data and data["accepts_pediatric"] is not None:
        params.append(bool(data["accepts_pediatric"]))
        updates.append(f"accepts_pediatric = ${len(params)}")

    if "min_pediatric_age_years" in data:
        _val_age = data.get("min_pediatric_age_years")
        if _val_age is None or _val_age == "":
            params.append(None)
        else:
            try:
                _ival_age = int(_val_age)
            except (TypeError, ValueError):
                raise HTTPException(
                    status_code=422,
                    detail="min_pediatric_age_years must be an integer",
                )
            if _ival_age < 0:
                raise HTTPException(
                    status_code=422,
                    detail="min_pediatric_age_years must be >= 0",
                )
            params.append(_ival_age)
        updates.append(f"min_pediatric_age_years = ${len(params)}")

    if "pediatric_notes" in data:
        _val_pedn = data.get("pediatric_notes")
        params.append(_val_pedn.strip() if isinstance(_val_pedn, str) and _val_pedn.strip() else None)
        updates.append(f"pediatric_notes = ${len(params)}")

    if "high_risk_protocols" in data:
        _val_hrp = data.get("high_risk_protocols")
        if _val_hrp is None:
            params.append(None)
        else:
            if not isinstance(_val_hrp, dict):
                raise HTTPException(
                    status_code=422,
                    detail="high_risk_protocols must be an object",
                )
            from pydantic import ValidationError

            validated: dict = {}
            for _cond_key, _cond_val in _val_hrp.items():
                if not isinstance(_cond_key, str) or not _cond_key.strip():
                    raise HTTPException(
                        status_code=422,
                        detail="high_risk_protocols keys must be non-empty strings",
                    )
                if not isinstance(_cond_val, dict):
                    raise HTTPException(
                        status_code=422,
                        detail=f"high_risk_protocols.{_cond_key} must be an object",
                    )
                try:
                    validated[_cond_key] = HighRiskProtocol(
                        **_cond_val
                    ).model_dump()
                except ValidationError as _ve:
                    raise HTTPException(
                        status_code=422,
                        detail=f"high_risk_protocols.{_cond_key}: {_ve.errors()}",
                    )
            params.append(json.dumps(validated))
        updates.append(f"high_risk_protocols = ${len(params)}::jsonb")

    if "requires_anamnesis_before_booking" in data and data["requires_anamnesis_before_booking"] is not None:
        params.append(bool(data["requires_anamnesis_before_booking"]))
        updates.append(f"requires_anamnesis_before_booking = ${len(params)}")

    if not updates:
        return {"status": "updated"}
    params.append(tenant_id)
    updates.append("updated_at = NOW()")
    query = f"UPDATE tenants SET {', '.join(updates)} WHERE id = ${len(params)}"
    await db.pool.execute(query, *params)
    # If country_code was in the payload, drop the cached tz so the new value
    # is picked up immediately on the next AI tool call.
    if "country_code" in data and data["country_code"] is not None:
        try:
            from services.tz_resolver import invalidate_tenant_tz_cache

            invalidate_tenant_tz_cache(tenant_id)
        except Exception:
            pass
    logger.info(
        f"Tenant {tenant_id} updated: calendar_provider={data.get('calendar_provider')} (persisted)"
    )
    result = {"status": "updated"}
    # Warn if country_code is not in supported list
    if "country_code" in data and data["country_code"] is not None:
        from services.holiday_service import SUPPORTED_COUNTRIES

        cc = str(data["country_code"]).upper().strip()[:2]
        if cc not in SUPPORTED_COUNTRIES:
            result["warning"] = (
                f"Country code '{cc}' is not in the supported holidays list. Custom holidays will still work, but national holidays won't be auto-detected."
            )
    return result


# ==================== HOLIDAYS ====================


@router.get(
    "/holidays", tags=["Feriados"], summary="Listar feriados (nacionales + custom)"
)
async def list_holidays(
    user_data=Depends(verify_admin_token), days: int = Query(365, ge=1, le=730)
):
    """Lista feriados nacionales (librería) + custom del tenant para los próximos N días."""
    from services.holiday_service import get_upcoming_holidays

    tenant_id = user_data.tenant_id
    holidays = await get_upcoming_holidays(db.pool, tenant_id, days_ahead=days)
    # Enrich upcoming entries with custom_hours if present
    for h in holidays:
        if "custom_hours_start" in h and h["custom_hours_start"] is not None:
            h["custom_hours"] = {
                "start": h["custom_hours_start"].strftime("%H:%M")
                if hasattr(h["custom_hours_start"], "strftime")
                else str(h["custom_hours_start"])[:5],
                "end": h["custom_hours_end"].strftime("%H:%M")
                if hasattr(h.get("custom_hours_end"), "strftime")
                else str(h.get("custom_hours_end", ""))[:5],
            }
        else:
            h["custom_hours"] = None
    # Also fetch custom-only rows for management
    custom_rows = await db.pool.fetch(
        "SELECT id, date, name, holiday_type, is_recurring, created_at, custom_hours_start, custom_hours_end FROM tenant_holidays WHERE tenant_id = $1 ORDER BY date ASC",
        tenant_id,
    )
    custom_list = []
    for r in custom_rows:
        row = dict(r)
        start = row.pop("custom_hours_start", None)
        end = row.pop("custom_hours_end", None)
        row["custom_hours_start"] = start.strftime("%H:%M") if start else None
        row["custom_hours_end"] = end.strftime("%H:%M") if end else None
        custom_list.append(row)
    return {"upcoming": holidays, "custom": custom_list}


@router.get(
    "/holidays/upcoming", tags=["Feriados"], summary="Próximos 30 días de feriados"
)
async def upcoming_holidays(user_data=Depends(verify_admin_token)):
    """Feriados en los próximos 30 días, ordenados por fecha."""
    from services.holiday_service import get_upcoming_holidays

    tenant_id = user_data.tenant_id
    return await get_upcoming_holidays(db.pool, tenant_id, days_ahead=30)


@router.post("/holidays", tags=["Feriados"], summary="Crear feriado custom")
async def create_holiday(data: Dict[str, Any], user_data=Depends(verify_admin_token)):
    """Crea un feriado custom (closure o override_open) para el tenant."""
    from datetime import date as date_type

    tenant_id = user_data.tenant_id
    holiday_date = data.get("date")
    name = data.get("name")
    holiday_type = data.get("holiday_type", "closure")
    is_recurring = data.get("is_recurring", False)

    if not holiday_date or not name:
        raise HTTPException(status_code=422, detail="date y name son obligatorios")
    if holiday_type not in ("closure", "override_open"):
        raise HTTPException(
            status_code=422, detail="holiday_type debe ser 'closure' o 'override_open'"
        )

    try:
        if isinstance(holiday_date, str):
            parsed_date = date_type.fromisoformat(holiday_date)
        else:
            parsed_date = holiday_date
    except (ValueError, TypeError):
        raise HTTPException(
            status_code=422, detail="Formato de fecha inválido. Usar YYYY-MM-DD"
        )

    try:
        new_id = await db.pool.fetchval(
            """INSERT INTO tenant_holidays (tenant_id, date, name, holiday_type, is_recurring)
               VALUES ($1, $2, $3, $4, $5) RETURNING id""",
            tenant_id,
            parsed_date,
            name.strip(),
            holiday_type,
            bool(is_recurring),
        )
        return {"id": new_id, "status": "created"}
    except Exception as e:
        if "uq_tenant_holidays_tenant_date_type" in str(e):
            raise HTTPException(
                status_code=409,
                detail="Ya existe un feriado con ese tipo para esa fecha",
            )
        raise HTTPException(status_code=400, detail=str(e))


@router.post(
    "/holidays/toggle", tags=["Feriados"], summary="Toggle feriado override_open"
)
async def toggle_holiday(
    data: Dict[str, Any],
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Alterna un feriado override_open para una fecha: si existe lo elimina, si no lo crea."""
    from datetime import date as date_type

    holiday_date = data.get("date")
    name = data.get("name", "Feriado con atención")

    if not holiday_date:
        raise HTTPException(status_code=422, detail="date es obligatorio")

    try:
        if isinstance(holiday_date, str):
            parsed_date = date_type.fromisoformat(holiday_date)
        else:
            parsed_date = holiday_date
    except (ValueError, TypeError):
        raise HTTPException(
            status_code=422, detail="Formato de fecha inválido. Usar YYYY-MM-DD"
        )

    existing = await db.pool.fetchrow(
        "SELECT id FROM tenant_holidays WHERE tenant_id = $1 AND date = $2 AND holiday_type = 'override_open'",
        tenant_id,
        parsed_date,
    )

    if existing:
        await db.pool.execute(
            "DELETE FROM tenant_holidays WHERE id = $1 AND tenant_id = $2",
            existing["id"],
            tenant_id,
        )
        return {"status": "toggled_closed", "holiday_id": None}
    else:
        new_id = await db.pool.fetchval(
            """INSERT INTO tenant_holidays (tenant_id, date, name, holiday_type, is_recurring)
               VALUES ($1, $2, $3, 'override_open', false) RETURNING id""",
            tenant_id,
            parsed_date,
            name.strip() if isinstance(name, str) else str(name),
        )
        return {"status": "toggled_open", "holiday_id": new_id}


@router.put(
    "/holidays/{holiday_id}", tags=["Feriados"], summary="Actualizar feriado custom"
)
async def update_holiday(
    holiday_id: int, data: Dict[str, Any], user_data=Depends(verify_admin_token)
):
    """Actualiza un feriado custom del tenant."""
    tenant_id = user_data.tenant_id
    existing = await db.pool.fetchrow(
        "SELECT id FROM tenant_holidays WHERE id = $1 AND tenant_id = $2",
        holiday_id,
        tenant_id,
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Feriado no encontrado")

    updates = []
    params = []
    if "name" in data:
        params.append(data["name"].strip())
        updates.append(f"name = ${len(params)}")

    new_holiday_type = data.get("holiday_type")
    if new_holiday_type is not None:
        if new_holiday_type not in ("closure", "override_open"):
            raise HTTPException(
                status_code=422,
                detail="holiday_type debe ser 'closure' o 'override_open'",
            )
        params.append(new_holiday_type)
        updates.append(f"holiday_type = ${len(params)}")

    if "is_recurring" in data:
        params.append(bool(data["is_recurring"]))
        updates.append(f"is_recurring = ${len(params)}")

    # Custom hours handling
    has_start = "custom_hours_start" in data
    has_end = "custom_hours_end" in data

    # If holiday_type is being changed to 'closure', NULL out custom hours
    if new_holiday_type == "closure":
        params.append(None)
        updates.append(f"custom_hours_start = ${len(params)}")
        params.append(None)
        updates.append(f"custom_hours_end = ${len(params)}")
    elif has_start or has_end:
        # If only one is provided, raise 422
        if has_start != has_end:
            raise HTTPException(
                status_code=422,
                detail="Ambos horarios (inicio y fin) son necesarios",
            )
        # Both provided — parse and validate
        raw_start = data["custom_hours_start"]
        raw_end = data["custom_hours_end"]
        if raw_start is None and raw_end is None:
            # Explicit NULL — clear both
            params.append(None)
            updates.append(f"custom_hours_start = ${len(params)}")
            params.append(None)
            updates.append(f"custom_hours_end = ${len(params)}")
        else:
            try:
                parsed_start = datetime.strptime(str(raw_start), "%H:%M").time()
                parsed_end = datetime.strptime(str(raw_end), "%H:%M").time()
            except (ValueError, TypeError):
                raise HTTPException(
                    status_code=422,
                    detail="Formato de horario inválido. Usar HH:MM",
                )
            if parsed_start >= parsed_end:
                raise HTTPException(
                    status_code=422,
                    detail="El horario de inicio debe ser anterior al de fin",
                )
            params.append(parsed_start)
            updates.append(f"custom_hours_start = ${len(params)}")
            params.append(parsed_end)
            updates.append(f"custom_hours_end = ${len(params)}")

    if not updates:
        return {"status": "updated"}

    params.append(holiday_id)
    params.append(tenant_id)
    query = f"UPDATE tenant_holidays SET {', '.join(updates)} WHERE id = ${len(params) - 1} AND tenant_id = ${len(params)}"
    await db.pool.execute(query, *params)
    return {"status": "updated"}


@router.delete(
    "/holidays/{holiday_id}", tags=["Feriados"], summary="Eliminar feriado custom"
)
async def delete_holiday(holiday_id: int, user_data=Depends(verify_admin_token)):
    """Elimina un feriado custom del tenant."""
    tenant_id = user_data.tenant_id
    result = await db.pool.execute(
        "DELETE FROM tenant_holidays WHERE id = $1 AND tenant_id = $2",
        holiday_id,
        tenant_id,
    )
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Feriado no encontrado")
    return {"status": "deleted"}


@router.post(
    "/tenants/{tenant_id}/logo", tags=["Sedes"], summary="Subir logo de la clínica"
)
async def upload_tenant_logo(
    tenant_id: int, file: UploadFile = File(...), user_data=Depends(verify_admin_token)
):
    """Sube un logo para la clínica. Solo CEO. Acepta PNG, JPG, SVG, WebP. Max 2MB."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar clínicas."
        )
    if not file.content_type or not file.content_type.startswith(
        ("image/", "image/svg")
    ):
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos de imagen (PNG, JPG, SVG, WebP).",
        )
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no puede superar 2MB.")
    ext = (
        file.filename.rsplit(".", 1)[-1].lower()
        if file.filename and "." in file.filename
        else "png"
    )
    if ext not in ("png", "jpg", "jpeg", "svg", "webp", "ico"):
        ext = "png"
    upload_dir = Path(f"/app/uploads/tenants/{tenant_id}")
    upload_dir.mkdir(parents=True, exist_ok=True)
    # Remove old logos
    for old in upload_dir.glob("logo.*"):
        old.unlink(missing_ok=True)
    file_path = upload_dir / f"logo.{ext}"
    file_path.write_bytes(content)
    logo_url = f"/public/tenant-logo/{tenant_id}"
    await db.pool.execute(
        "UPDATE tenants SET logo_url = $1 WHERE id = $2", logo_url, tenant_id
    )
    # Clear frontend cache
    return {"status": "ok", "logo_url": logo_url}


@router.delete("/tenants/{tenant_id}", tags=["Sedes"], summary="Eliminar una sede")
async def delete_tenant(
    tenant_id: int,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Elimina una clínica. Solo CEO, solo sus propias sedes."""
    if user_data.role != "ceo":
        raise HTTPException(
            status_code=403, detail="Solo el CEO puede gestionar clínicas."
        )
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tenés acceso a este tenant.")
    await db.pool.execute(
        "DELETE FROM tenants WHERE id = $1 AND id = ANY($2::int[])",
        tenant_id,
        allowed_ids,
    )
    return {"status": "deleted"}


# ─── FAQs por Clínica (Spec 2026-03-16) ───


@router.get(
    "/tenants/{tenant_id}/faqs", tags=["FAQs"], summary="Listar FAQs de una clínica"
)
async def get_tenant_faqs(
    tenant_id: int,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Retorna todas las FAQs de un tenant ordenadas por sort_order."""
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tenés acceso a este tenant.")
    rows = await db.pool.fetch(
        "SELECT id, tenant_id, category, question, answer, sort_order, created_at, updated_at FROM clinic_faqs WHERE tenant_id = $1 ORDER BY category, sort_order ASC, id ASC",
        tenant_id,
    )
    return [dict(r) for r in rows]


@router.post("/tenants/{tenant_id}/faqs", tags=["FAQs"], summary="Crear una FAQ")
async def create_tenant_faq(
    tenant_id: int,
    data: Dict[str, Any],
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Crea una nueva FAQ para un tenant."""
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Solo el CEO puede gestionar FAQs.")
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tenés acceso a este tenant.")
    question = _strip_html((data.get("question") or "").strip())
    answer = _strip_html((data.get("answer") or "").strip())
    if not question or not answer:
        raise HTTPException(
            status_code=400, detail="question y answer son obligatorios."
        )
    category = _strip_html((data.get("category") or "General").strip())
    sort_order = int(data.get("sort_order", 0))
    new_id = await db.pool.fetchval(
        "INSERT INTO clinic_faqs (tenant_id, category, question, answer, sort_order) VALUES ($1, $2, $3, $4, $5) RETURNING id",
        tenant_id,
        category,
        question,
        answer,
        sort_order,
    )

    # Generate embedding SYNCHRONOUSLY so the FAQ is immediately searchable.
    # If embedding fails, we still return success but include a warning so the
    # frontend can show it. The FAQ exists, just won't appear in semantic search.
    embedding_status = "skipped"
    embedding_error = None
    try:
        from services.embedding_service import (
            upsert_faq_embedding,
            check_pgvector_available,
        )

        if await check_pgvector_available():
            ok = await upsert_faq_embedding(tenant_id, new_id, question, answer)
            embedding_status = "ok" if ok else "failed"
            if not ok:
                embedding_error = "upsert_faq_embedding returned False (check logs)"
            logger.info(
                f"📚 FAQ {new_id} created and embedded: status={embedding_status}"
            )
        else:
            embedding_status = "no_pgvector"
            logger.warning(f"📚 FAQ {new_id} created but pgvector not available")
    except Exception as e:
        embedding_status = "error"
        embedding_error = str(e)
        logger.error(f"📚 FAQ {new_id} embedding failed: {e}", exc_info=True)

    return {
        "id": new_id,
        "status": "created",
        "embedding_status": embedding_status,
        "embedding_error": embedding_error,
    }


@router.put("/faqs/{faq_id}", tags=["FAQs"], summary="Actualizar una FAQ")
async def update_faq(
    faq_id: int,
    data: Dict[str, Any],
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Actualiza una FAQ existente."""
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Solo el CEO puede gestionar FAQs.")
    row = await db.pool.fetchrow(
        "SELECT tenant_id FROM clinic_faqs WHERE id = $1", faq_id
    )
    if not row:
        raise HTTPException(status_code=404, detail="FAQ no encontrada.")
    if row["tenant_id"] not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tenés acceso a esta FAQ.")
    updates = []
    params = []
    for field in ("category", "question", "answer"):
        if field in data and data[field] is not None:
            params.append(_strip_html(str(data[field]).strip()))
            updates.append(f"{field} = ${len(params)}")
    if "sort_order" in data:
        params.append(int(data["sort_order"]))
        updates.append(f"sort_order = ${len(params)}")
    if not updates:
        return {"status": "no_changes"}
    updates.append("updated_at = NOW()")
    params.append(faq_id)
    params.append(row["tenant_id"])
    query = f"UPDATE clinic_faqs SET {', '.join(updates)} WHERE id = ${len(params) - 1} AND tenant_id = ${len(params)}"
    await db.pool.execute(query, *params)

    # Re-generate embedding SYNCHRONOUSLY so the updated FAQ is immediately searchable
    embedding_status = "skipped"
    embedding_error = None
    try:
        from services.embedding_service import (
            upsert_faq_embedding,
            check_pgvector_available,
        )

        if await check_pgvector_available():
            updated_row = await db.pool.fetchrow(
                "SELECT tenant_id, question, answer FROM clinic_faqs WHERE id = $1",
                faq_id,
            )
            if updated_row:
                ok = await upsert_faq_embedding(
                    updated_row["tenant_id"],
                    faq_id,
                    updated_row["question"],
                    updated_row["answer"],
                )
                embedding_status = "ok" if ok else "failed"
                if not ok:
                    embedding_error = "upsert_faq_embedding returned False (check logs)"
                logger.info(
                    f"📚 FAQ {faq_id} updated and re-embedded: status={embedding_status}"
                )
        else:
            embedding_status = "no_pgvector"
    except Exception as e:
        embedding_status = "error"
        embedding_error = str(e)
        logger.error(f"📚 FAQ {faq_id} re-embedding failed: {e}", exc_info=True)

    return {
        "status": "updated",
        "embedding_status": embedding_status,
        "embedding_error": embedding_error,
    }


@router.delete("/faqs/{faq_id}", tags=["FAQs"], summary="Eliminar una FAQ")
async def delete_faq(
    faq_id: int,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Elimina una FAQ."""
    if user_data.role != "ceo":
        raise HTTPException(status_code=403, detail="Solo el CEO puede gestionar FAQs.")
    row = await db.pool.fetchrow(
        "SELECT tenant_id FROM clinic_faqs WHERE id = $1", faq_id
    )
    if not row:
        raise HTTPException(status_code=404, detail="FAQ no encontrada.")
    if row["tenant_id"] not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tenés acceso a esta FAQ.")
    # Delete embedding first (cascade should handle it, but be explicit)
    try:
        from services.embedding_service import delete_faq_embedding

        await delete_faq_embedding(faq_id)
        logger.info(f"📚 FAQ {faq_id} embedding deleted")
    except Exception as e:
        logger.warning(
            f"📚 FAQ {faq_id} embedding deletion failed (cascade should handle it): {e}"
        )
    await db.pool.execute(
        "DELETE FROM clinic_faqs WHERE id = $1 AND tenant_id = ANY($2::int[])",
        faq_id,
        allowed_ids,
    )
    return {"status": "deleted"}


@router.get(
    "/chat/urgencies",
    dependencies=[Depends(verify_admin_token)],
    tags=["Chat"],
    summary="Listar urgencias recientes detectadas",
)
async def get_recent_urgencies(
    limit: int = 10, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Retorna los últimos casos de urgencia. Aislado por tenant_id (Regla de Oro)."""
    try:
        rows = await db.pool.fetch(
            """
            SELECT 
                a.id,
                p.first_name || ' ' || COALESCE(p.last_name, '') as patient_name,
                p.phone_number as phone,
                UPPER(a.urgency_level) as urgency_level,
                COALESCE(a.urgency_reason, 'Consulta IA detectada') as reason,
                a.appointment_datetime as timestamp
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            WHERE a.tenant_id = $1 AND a.urgency_level IN ('high', 'emergency')
            ORDER BY a.created_at DESC
            LIMIT $2
        """,
            tenant_id,
            limit,
        )

        return [
            {
                "id": str(row["id"]),
                "patient_name": row["patient_name"],
                "phone": row["phone"],
                "urgency_level": "CRITICAL"
                if row["urgency_level"] == "EMERGENCY"
                else row["urgency_level"],
                "reason": row["reason"],
                "timestamp": row["timestamp"].strftime("%d/%m %H:%M")
                if hasattr(row["timestamp"], "strftime")
                else str(row["timestamp"]),
            }
            for row in rows
        ]
    except Exception as e:
        logger.error(f"Error fetching urgencies: {e}")
        return []


@router.get(
    "/config/deployment",
    dependencies=[Depends(verify_admin_token)],
    tags=["Configuración"],
    summary="Obtener configuración de despliegue",
)
async def get_deployment_config(request: Request):
    """Retorna configuración dinámica de despliegue (Webhooks, URLs)."""
    # Detectamos la URL base actual si no está forzada en ENV
    host = request.headers.get("host", "localhost:8000")
    protocol = (
        "https" if request.headers.get("x-forwarded-proto") == "https" else "http"
    )
    base_url = f"{protocol}://{host}"

    return {
        "webhook_ycloud_url": f"{base_url}/admin/ycloud/webhook",
        "webhook_ycloud_internal_port": os.getenv("WHATSAPP_SERVICE_PORT", "8002"),
        "orchestrator_url": base_url,
        "environment": os.getenv("ENVIRONMENT", "development"),
    }


@router.get(
    "/settings/clinic",
    dependencies=[Depends(verify_admin_token)],
    tags=["Configuración"],
    summary="Obtener configuración operativa de la clínica",
)
async def get_clinic_settings(
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Retorna la configuración operativa de la clínica (nombre, horarios, ui_language) desde el tenant."""
    try:
        row = await db.pool.fetchrow(
            "SELECT clinic_name, config FROM tenants WHERE id = $1", resolved_tenant_id
        )
        if not row:
            return _fallback_clinic_settings()
        config = row["config"] or {}
        ui_lang = (
            (config.get("ui_language") or "es") if isinstance(config, dict) else "es"
        )
        return {
            "name": row["clinic_name"] or os.getenv("CLINIC_NAME", "Clínica Dental"),
            "location": os.getenv("CLINIC_LOCATION", ""),
            "hours_start": os.getenv("CLINIC_HOURS_START", "08:00"),
            "hours_end": os.getenv("CLINIC_HOURS_END", "19:00"),
            "working_days": [0, 1, 2, 3, 4, 5],
            "time_zone": "America/Argentina/Buenos_Aires",
            "ui_language": ui_lang,
            "ai_engine_mode": row.get("ai_engine_mode") or "solo",
        }
    except Exception as e:
        logger.warning(f"get_clinic_settings failed: {e}")
        return _fallback_clinic_settings()


def _fallback_clinic_settings():
    """Config por defecto cuando no hay tenant o falla la consulta."""
    return {
        "name": os.getenv("CLINIC_NAME", "Clínica Dental"),
        "location": os.getenv("CLINIC_LOCATION", ""),
        "hours_start": os.getenv("CLINIC_HOURS_START", "08:00"),
        "hours_end": os.getenv("CLINIC_HOURS_END", "19:00"),
        "working_days": [0, 1, 2, 3, 4, 5],
        "time_zone": "America/Argentina/Buenos_Aires",
        "ui_language": "es",
        "ai_engine_mode": "solo",
    }


class ClinicSettingsUpdate(BaseModel):
    ui_language: Optional[str] = None  # "es" | "en" | "fr"
    ai_engine_mode: Optional[Literal["solo", "multi"]] = None  # dual-engine system


@router.patch(
    "/settings/clinic",
    dependencies=[Depends(verify_admin_token)],
    tags=["Configuración"],
    summary="Actualizar configuración operativa de la clínica",
)
async def update_clinic_settings(
    payload: ClinicSettingsUpdate,
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualiza configuración de la clínica (ej. idioma de la UI). Solo campos enviados."""
    if payload.ui_language is not None:
        if payload.ui_language not in ("es", "en", "fr"):
            raise HTTPException(
                status_code=400, detail="ui_language debe ser 'es', 'en' o 'fr'."
            )
        try:
            await db.pool.execute(
                """
                UPDATE tenants
                SET config = jsonb_set(COALESCE(config, '{}'), '{ui_language}', to_jsonb($1::text))
                WHERE id = $2
                """,
                payload.ui_language,
                resolved_tenant_id,
            )
        except Exception as e:
            logger.error(f"update_clinic_settings failed: {e}")
            raise HTTPException(
                status_code=500, detail="Error al guardar la configuración."
            )

    # Handle ai_engine_mode (dual-engine system)
    if payload.ai_engine_mode is not None:
        if payload.ai_engine_mode not in ("solo", "multi"):
            raise HTTPException(
                status_code=422, detail="ai_engine_mode debe ser 'solo' o 'multi'."
            )

        # If switching to multi, run health check probe first
        if payload.ai_engine_mode == "multi":
            try:
                from services.engine_router import MultiAgentEngine

                probe_result = await MultiAgentEngine().probe()
                if not probe_result.ok:
                    raise HTTPException(
                        status_code=422,
                        detail=f"Motor multi-agente no disponible: {probe_result.detail}",
                    )
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Multi-agent probe failed: {e}")
                raise HTTPException(
                    status_code=422,
                    detail="No se pudo verificar el motor multi-agente.",
                )

        # Update the database
        try:
            await db.pool.execute(
                """
                UPDATE tenants
                SET ai_engine_mode = $1, updated_at = now()
                WHERE id = $2
                """,
                payload.ai_engine_mode,
                resolved_tenant_id,
            )

            # Invalidate cache
            try:
                from services.engine_router import invalidate_cache

                invalidate_cache(resolved_tenant_id)
            except Exception as e:
                logger.warning(f"Failed to invalidate cache: {e}")

        except Exception as e:
            logger.error(f"update_clinic_settings ai_engine_mode failed: {e}")
            raise HTTPException(
                status_code=500, detail="Error al guardar la configuración del motor."
            )

    return {"status": "ok", "ui_language": getattr(payload, "ui_language", None)}


# ==================== ENDPOINTS BÚSQUEDA SEMÁNTICA ====================


@router.get(
    "/patients/search-semantic",
    tags=["Pacientes"],
    summary="Búsqueda semántica de pacientes por síntomas",
)
async def search_patients_by_symptoms(
    query: str,
    limit: int = 20,
    user_data=Depends(verify_admin_token),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Búsqueda semántica de pacientes por síntomas. Aislado por tenant_id (Regla de Oro).
    - CEO: retorna pacientes de todas sus sedes (allowed_ids).
    - Staff/Profesional: retorna pacientes solo de su sede activa.
    """
    try:
        is_ceo = user_data.role == "ceo"

        if is_ceo:
            # CEO: buscar en chat_messages sin filtro de sede (los patient_ids luego se filtran)
            chat_matches = await db.pool.fetch(
                """
                SELECT DISTINCT patient_id
                FROM chat_messages
                WHERE content ILIKE $1
                LIMIT $2
            """,
                f"%{query}%",
                limit,
            )
        else:
            # Staff: filtrar chat_messages por tenant_id de la sede activa
            chat_matches = await db.pool.fetch(
                """
                SELECT DISTINCT cm.patient_id
                FROM chat_messages cm
                JOIN patients p ON cm.patient_id = p.id
                WHERE cm.content ILIKE $1
                  AND p.tenant_id = $2
                LIMIT $3
            """,
                f"%{query}%",
                tenant_id,
                limit,
            )

        patient_ids = [
            row["patient_id"] for row in chat_matches if row["patient_id"] is not None
        ]

        if not patient_ids:
            return []

        if is_ceo:
            # CEO: filtrar por sus sedes permitidas
            patients = await db.pool.fetch(
                """
                SELECT id, first_name, last_name, phone_number, email, insurance_provider, dni, created_at
                FROM patients
                WHERE id = ANY($1::int[])
                  AND tenant_id = ANY($2::int[])
                  AND status = 'active'
            """,
                patient_ids,
                allowed_ids,
            )
        else:
            # Staff: filtrar por su sede activa
            patients = await db.pool.fetch(
                """
                SELECT id, first_name, last_name, phone_number, email, insurance_provider, dni, created_at
                FROM patients
                WHERE id = ANY($1::int[])
                  AND tenant_id = $2
                  AND status = 'active'
            """,
                patient_ids,
                tenant_id,
            )

        return [dict(row) for row in patients]

    except Exception as e:
        logger.error(f"Error en búsqueda semántica: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ==================== ENDPOINTS SEGURO / OBRAS SOCIALES ====================


@router.get(
    "/patients/{patient_id}/insurance-status",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Verificar estado de obra social del paciente",
)
async def get_patient_insurance_status(
    patient_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Verifica el estado de la credencial de obra social de un paciente.

    Returns:
        - status: 'ok' | 'warning' | 'expired'
        - requires_token: boolean (ej: OSDE requiere token)
        - message: string con descripción del estado
        - expiration_days: días hasta el vencimiento (negativo = vencido)
        - insurance_provider: nombre de la obra social
    """
    try:
        # Obtener datos del paciente y su obra social
        patient = await db.pool.fetchrow(
            """
            SELECT id, first_name, last_name, phone_number, insurance_provider,
                   insurance_id, insurance_valid_until, status
            FROM patients
            WHERE id = $1 AND tenant_id = $2 AND status = 'active'
        """,
            patient_id,
            tenant_id,
        )

        if not patient:
            raise HTTPException(status_code=404, detail="Paciente no encontrado")

        insurance_provider = patient.get("insurance_provider") or ""
        expiry_date = patient.get("insurance_valid_until")

        # Si no tiene obra social configurada
        if not insurance_provider:
            return {
                "status": "ok",
                "requires_token": False,
                "message": "Sin obra social configurada",
                "expiration_days": None,
                "insurance_provider": None,
            }

        # Verificar si requiere token (algunas obras sociales específicas)
        requires_token = insurance_provider.upper() in [
            "OSDE",
            "SWISS MEDICAL",
            "GALENO",
            "MEDICINA PREPAGA",
        ]

        # Calcular días hasta vencimiento
        expiration_days = None
        if expiry_date:
            # Asegurar que sea objeto date para la resta
            if isinstance(expiry_date, str):
                expiry_dt = date.fromisoformat(expiry_date.split("T")[0])
            elif isinstance(expiry_date, datetime):
                expiry_dt = expiry_date.date()
            else:
                expiry_dt = expiry_date  # Asumimos que ya es date

            today = date.today()
            delta = expiry_dt - today
            expiration_days = delta.days

        # Determinar estado
        if expiration_days is not None and expiration_days < 0:
            # Credencial vencida
            return {
                "status": "expired",
                "requires_token": requires_token,
                "message": f"Credencial vencida hace {abs(expiration_days)} días. Requiere renovación.",
                "expiration_days": expiration_days,
                "insurance_provider": insurance_provider,
            }
        elif expiration_days is not None and expiration_days <= 30:
            # Credencial próxima a vencer (30 días)
            return {
                "status": "warning",
                "requires_token": requires_token,
                "message": f"Credencial vence en {expiration_days} días. Considera renovar.",
                "expiration_days": expiration_days,
                "insurance_provider": insurance_provider,
            }
        else:
            # Todo bien
            message = "Credencial vigente"
            if requires_token:
                message += ". Requiere validación de token"

            return {
                "status": "ok",
                "requires_token": requires_token,
                "message": message,
                "expiration_days": expiration_days,
                "insurance_provider": insurance_provider,
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verificando seguro: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ==================== ENDPOINTS PACIENTES ====================


@router.post(
    "/patients",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Crear un nuevo paciente",
)
async def create_patient(
    p: PatientCreate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Crear un paciente nuevo en la sede actual. Aislado por tenant_id (Regla de Oro)."""
    try:
        row = await db.pool.fetchrow(
            """
            INSERT INTO patients (tenant_id, first_name, last_name, phone_number, email, dni, insurance_provider, city, birth_date, status, created_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'active', NOW())
            RETURNING id
        """,
            tenant_id,
            (p.first_name or "").strip() or "Sin nombre",
            (p.last_name or "").strip() or "",
            (p.phone_number or "").strip(),
            (p.email or "").strip() or None,
            (p.dni or "").strip() or None,
            (p.insurance or "").strip() or None,
            (p.city or "").strip() or None,
            p.birth_date,
        )
        return {"id": row["id"]}
    except asyncpg.UniqueViolationError as e:
        if (
            "patients_tenant_id_phone_number_key" in str(e)
            or "tenant_id" in str(e).lower()
            and "phone" in str(e).lower()
        ):
            raise HTTPException(
                status_code=409,
                detail="Ya existe un paciente con ese número de teléfono en esta sede. Podés buscarlo en la lista o usar otro teléfono.",
            )
        raise HTTPException(
            status_code=409,
            detail="Paciente duplicado (mismo DNI o teléfono en esta sede).",
        )
    except Exception as e:
        logger.error(f"Error creating patient: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== IMPORTACIÓN MASIVA DE PACIENTES ====================

# Mapeo de aliases de columnas (case-insensitive, sin tildes)
_COLUMN_ALIASES: Dict[str, str] = {}
for _field, _aliases in {
    "first_name": ["nombre", "nombres", "first_name", "name", "primer_nombre"],
    "last_name": ["apellido", "apellidos", "last_name", "surname", "segundo_nombre"],
    "_full_name": [
        "apellido y nombre",
        "nombre y apellido",
        "nombre completo",
        "full_name",
        "nombre_completo",
        "paciente",
    ],
    "phone_number": [
        "telefono",
        "teléfono",
        "phone",
        "phone_number",
        "celular",
        "tel",
        "whatsapp",
        "tel_celular",
        "numero_celular",
        "nro_telefono",
        "móvil",
        "movil",
    ],
    "dni": [
        "dni",
        "documento",
        "document",
        "id_number",
        "cedula",
        "cédula",
        "nro_doc",
        "numero_documento",
        "nro_documento",
    ],
    "email": ["email", "correo", "mail", "e-mail"],
    "birth_date": [
        "fecha_nacimiento",
        "birth_date",
        "nacimiento",
        "birthday",
        "fecha_nac",
    ],
    "insurance": ["obra_social", "insurance", "seguro", "prepaga", "mutual"],
    "city": ["ciudad", "city", "localidad", "domicilio"],
    "notes": ["notas", "notes", "observaciones", "comentarios"],
}.items():
    for _a in _aliases:
        _COLUMN_ALIASES[_a.lower().strip()] = _field


def _normalize_header(h: str) -> str:
    """Normaliza un header de CSV/XLSX a un campo conocido."""
    import unicodedata

    cleaned = (
        unicodedata.normalize("NFKD", h.lower().strip())
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    return _COLUMN_ALIASES.get(h.lower().strip()) or _COLUMN_ALIASES.get(cleaned) or ""


def _parse_birth_date(val: str):
    """Parsea fecha de nacimiento en DD/MM/AAAA o AAAA-MM-DD."""
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_import_file(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parsea un CSV o XLSX y devuelve lista de dicts normalizados."""
    rows = []
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not raw_rows:
            return []
        headers = [str(c or "").strip() for c in raw_rows[0]]
        for vals in raw_rows[1:]:
            row_dict = {}
            for i, h in enumerate(headers):
                v = vals[i] if i < len(vals) else None
                row_dict[h] = str(v).strip() if v is not None else ""
            rows.append(row_dict)
    elif ext == "csv":
        # Try UTF-8 first, fallback to latin-1
        text = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = file_bytes.decode(enc)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        if text is None:
            raise ValueError(
                "No se pudo decodificar el archivo. Probá guardándolo como UTF-8."
            )
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({k: (v or "").strip() for k, v in r.items()})
    else:
        raise ValueError(f"Formato no soportado: .{ext}. Usá .csv o .xlsx")

    # Normalizar headers
    normalized = []
    for row in rows:
        mapped = {}
        for raw_key, val in row.items():
            field = _normalize_header(raw_key)
            if field and val:
                mapped[field] = val
        # Split _full_name into first_name + last_name
        if "_full_name" in mapped and not mapped.get("first_name"):
            full = mapped.pop("_full_name").strip()
            parts = full.split(None, 1)  # split on first whitespace
            if len(parts) == 2:
                # "Apellido Nombre(s)" format (most common in Argentine CSVs)
                mapped["last_name"] = parts[0]
                mapped["first_name"] = parts[1]
            else:
                mapped["first_name"] = full
        elif "_full_name" in mapped:
            mapped.pop("_full_name")  # already have first_name, discard
        normalized.append(mapped)
    return normalized


@router.post(
    "/patients/import/preview",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Previsualizar importación de pacientes desde CSV/XLSX",
)
async def import_patients_preview(
    file: UploadFile = File(...),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Parsea un archivo CSV/XLSX y devuelve una previsualización sin insertar datos. Aislado por tenant_id (Regla de Oro)."""
    # Validar extensión
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx"):
        raise HTTPException(
            status_code=400, detail="Formato no soportado. Usá archivos .csv o .xlsx"
        )

    file_bytes = await file.read()
    try:
        rows = _parse_import_file(file_bytes, file.filename or "file.csv")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if len(rows) > 1000:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo tiene {len(rows)} filas. El máximo permitido es 1000.",
        )
    if not rows:
        raise HTTPException(
            status_code=400, detail="El archivo está vacío o no tiene datos."
        )

    # Generar placeholders y validar
    tel_counter = 1
    dni_counter = 1
    # Get existing placeholder counters from DB to avoid collisions
    existing_tel = (
        await db.pool.fetchval(
            "SELECT COUNT(*) FROM patients WHERE tenant_id = $1 AND phone_number LIKE 'SIN-TEL-%'",
            tenant_id,
        )
        or 0
    )
    existing_dni = (
        await db.pool.fetchval(
            "SELECT COUNT(*) FROM patients WHERE tenant_id = $1 AND dni LIKE 'SIN-DNI-%'",
            tenant_id,
        )
        or 0
    )
    tel_counter = existing_tel + 1
    dni_counter = existing_dni + 1

    preview_rows = []
    error_details = []
    phones_in_file: set = set()

    for i, row in enumerate(rows, start=2):  # Row 2 = first data row (1 is header)
        first_name = (row.get("first_name") or "").strip()
        if not first_name:
            error_details.append(
                {"row": i, "reason": "Falta el campo 'nombre' (obligatorio)"}
            )
            continue

        phone = re.sub(r"[^\d+]", "", row.get("phone_number") or "")
        if not phone:
            phone = f"SIN-TEL-{tel_counter:03d}"
            tel_counter += 1

        # Dedup within file
        if phone in phones_in_file and not phone.startswith("SIN-TEL-"):
            error_details.append(
                {"row": i, "reason": f"Teléfono duplicado dentro del archivo: {phone}"}
            )
            continue
        phones_in_file.add(phone)

        dni = re.sub(r"\D", "", row.get("dni") or "")
        if not dni:
            dni = f"SIN-DNI-{dni_counter:03d}"
            dni_counter += 1

        preview_rows.append(
            {
                "row": i,
                "first_name": first_name,
                "last_name": (row.get("last_name") or "").strip(),
                "phone_number": phone,
                "dni": dni,
                "email": (row.get("email") or "").strip() or None,
                "birth_date": str(_parse_birth_date(row.get("birth_date") or ""))
                if _parse_birth_date(row.get("birth_date") or "")
                else None,
                "insurance": (row.get("insurance") or "").strip() or None,
                "city": (row.get("city") or "").strip() or None,
                "notes": (row.get("notes") or "").strip() or None,
                "status": "new",  # will be updated below if duplicate
            }
        )

    # Check duplicates against DB (batch query by phone)
    phones = [
        r["phone_number"]
        for r in preview_rows
        if not r["phone_number"].startswith("SIN-TEL-")
    ]
    existing_patients = {}
    if phones:
        existing_rows = await db.pool.fetch(
            "SELECT id, first_name, last_name, phone_number, dni, email, insurance_provider, city FROM patients WHERE tenant_id = $1 AND phone_number = ANY($2)",
            tenant_id,
            phones,
        )
        for ep in existing_rows:
            existing_patients[ep["phone_number"]] = dict(ep)

    duplicate_details = []
    valid_new = 0
    for pr in preview_rows:
        existing = existing_patients.get(pr["phone_number"])
        if existing:
            pr["status"] = "duplicate"
            # Check if CSV has data that DB doesn't
            has_new = False
            for csv_field, db_field in [
                ("last_name", "last_name"),
                ("email", "email"),
                ("dni", "dni"),
                ("insurance", "insurance_provider"),
                ("city", "city"),
            ]:
                if pr.get(csv_field) and not existing.get(db_field):
                    has_new = True
                    break
            duplicate_details.append(
                {
                    "row": pr["row"],
                    "csv_name": f"{pr['first_name']} {pr.get('last_name', '')}".strip(),
                    "csv_phone": pr["phone_number"],
                    "existing_id": existing["id"],
                    "existing_name": f"{existing['first_name']} {existing.get('last_name', '')}".strip(),
                    "existing_phone": existing["phone_number"],
                    "has_new_data": has_new,
                }
            )
        else:
            valid_new += 1

    return {
        "total_rows": len(rows),
        "valid_new": valid_new,
        "duplicates": len(duplicate_details),
        "errors": len(error_details),
        "duplicate_details": duplicate_details,
        "error_details": error_details,
        "preview_rows": preview_rows,
    }


class ImportExecuteRequest(BaseModel):
    duplicate_action: str = "skip"  # "skip" or "update"
    rows: List[Dict[str, Any]]


@router.post(
    "/patients/import/execute",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Ejecutar importación de pacientes",
)
async def import_patients_execute(
    body: ImportExecuteRequest,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Ejecuta la importación de pacientes previsualizados. Aislado por tenant_id (Regla de Oro)."""
    imported = 0
    updated = 0
    skipped = 0
    errors = 0

    for row in body.rows:
        status = row.get("status", "new")
        first_name = (row.get("first_name") or "").strip()
        if not first_name:
            errors += 1
            continue

        if status == "new":
            try:
                birth = None
                if row.get("birth_date"):
                    try:
                        birth = date.fromisoformat(row["birth_date"])
                    except (ValueError, TypeError):
                        birth = _parse_birth_date(row["birth_date"])
                await db.pool.execute(
                    """
                    INSERT INTO patients (tenant_id, first_name, last_name, phone_number, email, dni, insurance_provider, city, birth_date, notes, status, created_at)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, 'active', NOW())
                """,
                    tenant_id,
                    first_name,
                    (row.get("last_name") or "").strip(),
                    (row.get("phone_number") or "").strip(),
                    row.get("email") or None,
                    (row.get("dni") or "").strip() or None,
                    row.get("insurance") or None,
                    row.get("city") or None,
                    birth,
                    row.get("notes") or None,
                )
                imported += 1
            except asyncpg.UniqueViolationError:
                skipped += 1
            except Exception as e:
                logger.error(f"Error importing patient row: {e}")
                errors += 1

        elif status == "duplicate":
            if body.duplicate_action == "skip":
                skipped += 1
            elif body.duplicate_action == "update":
                try:
                    phone = (row.get("phone_number") or "").strip()
                    birth = None
                    if row.get("birth_date"):
                        try:
                            birth = date.fromisoformat(row["birth_date"])
                        except (ValueError, TypeError):
                            birth = _parse_birth_date(row["birth_date"])
                    # COALESCE: only fill empty fields, never overwrite existing data
                    await db.pool.execute(
                        """
                        UPDATE patients SET
                            last_name = COALESCE(NULLIF(last_name, ''), $2),
                            email = COALESCE(email, $3),
                            dni = COALESCE(dni, $4),
                            insurance_provider = COALESCE(insurance_provider, $5),
                            city = COALESCE(city, $6),
                            birth_date = COALESCE(birth_date, $7),
                            notes = COALESCE(notes, $8),
                            updated_at = NOW()
                        WHERE tenant_id = $1 AND phone_number = $9
                    """,
                        tenant_id,
                        (row.get("last_name") or "").strip() or None,
                        row.get("email") or None,
                        (row.get("dni") or "").strip() or None,
                        row.get("insurance") or None,
                        row.get("city") or None,
                        birth,
                        row.get("notes") or None,
                        phone,
                    )
                    updated += 1
                except Exception as e:
                    logger.error(f"Error updating patient during import: {e}")
                    errors += 1
        else:
            errors += 1

    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


@router.get(
    "/patients",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Listar pacientes",
)
async def list_patients(
    search: str = None,
    limit: int = 200,
    professional_id: Optional[int] = None,
    assigned_professional_id: Optional[int] = None,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Listar todos los pacientes del tenant (incluye importados sin turnos).
    Aislado por tenant_id (Regla de Oro).
    Cuando professional_id se proporciona, filtra solo pacientes con al menos un turno con ese profesional (RBAC).
    Cuando assigned_professional_id se proporciona, filtra por profesional asignado.
    """
    query = """
        SELECT p.id, p.first_name, p.last_name, p.phone_number, p.email,
               p.insurance_provider as obra_social, p.dni, p.city, p.birth_date, p.created_at, p.status,
               p.assigned_professional_id,
               ap.first_name as assigned_professional_name,
               EXISTS (SELECT 1 FROM appointments a WHERE a.patient_id = p.id AND a.tenant_id = p.tenant_id) as has_appointments,
               lt.treatment_name as last_treatment
        FROM patients p
        LEFT JOIN professionals ap ON ap.id = p.assigned_professional_id
        LEFT JOIN LATERAL (
            SELECT tt.name as treatment_name
            FROM appointments a
            JOIN treatment_types tt ON tt.code = a.appointment_type AND tt.tenant_id = a.tenant_id
            WHERE a.patient_id = p.id AND a.tenant_id = p.tenant_id
            ORDER BY a.appointment_datetime DESC
            LIMIT 1
        ) lt ON true
        WHERE p.tenant_id = $1 AND p.status != 'deleted'
    """
    params: List[Any] = [tenant_id]

    # RBAC: filter patients by professional when professional_id is provided
    if professional_id is not None:
        params.append(professional_id)
        query += f" AND EXISTS (SELECT 1 FROM appointments a WHERE a.patient_id = p.id AND a.professional_id = ${len(params)})"

    # Filter by assigned professional
    if assigned_professional_id is not None:
        params.append(assigned_professional_id)
        query += f" AND p.assigned_professional_id = ${len(params)}"

    if search:
        params.append(f"%{search}%")
        query += f" AND (p.first_name ILIKE ${len(params)} OR p.last_name ILIKE ${len(params)} OR p.phone_number ILIKE ${len(params)} OR p.dni ILIKE ${len(params)})"
        query += f""" ORDER BY
            CASE WHEN EXISTS (SELECT 1 FROM appointments a WHERE a.patient_id = p.id AND a.tenant_id = p.tenant_id)
                 THEN 0 ELSE 1 END,
            p.created_at DESC LIMIT ${len(params) + 1}"""
        params.append(limit)
    else:
        query += f""" ORDER BY
            CASE WHEN EXISTS (SELECT 1 FROM appointments a WHERE a.patient_id = p.id AND a.tenant_id = p.tenant_id)
                 THEN 0 ELSE 1 END,
            p.created_at DESC LIMIT ${len(params) + 1}"""
        params.append(limit)
    rows = await db.pool.fetch(query, *params)
    patients = [dict(row) for row in rows]

    if patients:
        patient_ids = [p["id"] for p in patients]

        # Batch query: next appointment per patient
        next_apts = await db.pool.fetch(
            """
            SELECT DISTINCT ON (patient_id) patient_id, appointment_datetime
            FROM appointments
            WHERE tenant_id = $1 AND status IN ('scheduled','confirmed') AND appointment_datetime > NOW()
              AND patient_id = ANY($2::int[])
            ORDER BY patient_id, appointment_datetime ASC
        """,
            tenant_id,
            patient_ids,
        )

        # Batch query: pending balance per patient
        balances = await db.pool.fetch(
            """
            SELECT patient_id, COALESCE(SUM(billing_amount), 0) as pending
            FROM appointments
            WHERE tenant_id = $1 AND payment_status IN ('pending','partial') AND status NOT IN ('cancelled')
              AND patient_id = ANY($2::int[])
            GROUP BY patient_id
        """,
            tenant_id,
            patient_ids,
        )

        next_apt_map = {r["patient_id"]: r["appointment_datetime"] for r in next_apts}
        balance_map = {r["patient_id"]: float(r["pending"]) for r in balances}

        for p in patients:
            p["next_appointment_date"] = next_apt_map.get(p["id"])
            p["pending_balance"] = balance_map.get(p["id"], 0.0)

    return patients


@router.patch(
    "/patients/{patient_id}/assign-professional",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Asignar profesional a un paciente",
)
async def assign_professional_to_patient(
    patient_id: int,
    body: dict,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Asignar o desasignar un profesional a un paciente."""
    professional_id = body.get("professional_id")  # null to unassign

    # Validate patient belongs to tenant
    patient = await db.pool.fetchrow(
        "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Validate professional belongs to tenant (if assigning)
    if professional_id is not None:
        prof = await db.pool.fetchrow(
            "SELECT id, first_name, last_name FROM professionals WHERE id = $1 AND tenant_id = $2 AND is_active = true",
            professional_id,
            tenant_id,
        )
        if not prof:
            raise HTTPException(
                status_code=404, detail="Profesional no encontrado o inactivo"
            )

    await db.pool.execute(
        "UPDATE patients SET assigned_professional_id = $1 WHERE id = $2 AND tenant_id = $3",
        professional_id,
        patient_id,
        tenant_id,
    )

    # Return updated info
    prof_name = None
    if professional_id is not None:
        prof_row = await db.pool.fetchrow(
            "SELECT first_name, last_name FROM professionals WHERE id = $1",
            professional_id,
        )
        if prof_row:
            prof_name = f"{prof_row['first_name']} {prof_row.get('last_name', '') or ''}".strip()

    return {
        "patient_id": patient_id,
        "assigned_professional_id": professional_id,
        "assigned_professional_name": prof_name,
    }


@router.patch(
    "/patients/bulk-assign-professional",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Asignar profesional a múltiples pacientes",
)
async def bulk_assign_professional(
    body: dict,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Asignar o desasignar un profesional a múltiples pacientes (max 200)."""
    patient_ids = body.get("patient_ids", [])
    professional_id = body.get("professional_id")  # null to unassign

    if not patient_ids or len(patient_ids) > 200:
        raise HTTPException(
            status_code=400, detail="Se requieren entre 1 y 200 pacientes"
        )

    # Validate professional belongs to tenant (if assigning)
    if professional_id is not None:
        prof = await db.pool.fetchrow(
            "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2 AND is_active = true",
            professional_id,
            tenant_id,
        )
        if not prof:
            raise HTTPException(
                status_code=404, detail="Profesional no encontrado o inactivo"
            )

    # Bulk update — only patients belonging to this tenant
    result = await db.pool.execute(
        "UPDATE patients SET assigned_professional_id = $1 WHERE id = ANY($2::int[]) AND tenant_id = $3",
        professional_id,
        patient_ids,
        tenant_id,
    )

    # Parse "UPDATE N" to get count
    updated_count = int(result.split()[-1]) if result else 0

    return {
        "updated_count": updated_count,
        "professional_id": professional_id,
    }


@router.post(
    "/professionals",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Crear un nuevo profesional de la salud",
)
async def create_professional(
    professional: ProfessionalCreate,
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Crear un nuevo profesional. Soberanía: se asocia a la clínica elegida.
    Si el body trae tenant_id y el usuario tiene acceso a esa clínica, se usa; si no, la del contexto.
    Así el agente y el sistema saben a qué clínica pertenece el profesional.
    """
    try:
        tid_raw = professional.tenant_id
        tid_int = int(tid_raw) if tid_raw is not None else None
        tenant_id = (
            tid_int
            if tid_int is not None and tid_int in allowed_ids
            else resolved_tenant_id
        )
    except (TypeError, ValueError):
        tenant_id = resolved_tenant_id

    email = (
        professional.email or ""
    ).strip() or f"prof_{uuid.uuid4().hex[:8]}@dentalogic.local"
    name_part = (professional.name or "").strip()
    first_name = name_part.split(maxsplit=1)[0] if name_part else "Profesional"
    last_name = (
        name_part.split(maxsplit=1)[1]
        if name_part and len(name_part.split(maxsplit=1)) > 1
        else " "
    )

    # 1. Usuario: crear nuevo o reutilizar si el email ya existe (vincular a esta sede)
    existing_user = await db.pool.fetchrow(
        "SELECT id FROM users WHERE email = $1", email
    )
    if existing_user:
        user_id = existing_user["id"]
        # ¿Ya tiene fila en professionals para esta sede? Entonces es duplicado real.
        already_linked = await db.pool.fetchval(
            "SELECT 1 FROM professionals WHERE user_id = $1 AND tenant_id = $2",
            user_id,
            tenant_id,
        )
        if already_linked:
            raise HTTPException(
                status_code=409,
                detail="Ese profesional ya está vinculado a esta sede. Buscalo en la lista de Profesionales.",
            )
        # Si no está vinculado a esta sede: creamos la fila en professionals (aparece como activo)
    else:
        user_id = uuid.uuid4()
        try:
            await db.pool.execute(
                """
                INSERT INTO users (id, email, password_hash, role, first_name, status, created_at)
                VALUES ($1, $2, $3, 'professional', $4, 'active', NOW())
            """,
                user_id,
                email,
                "hash_placeholder",
                first_name,
            )
        except asyncpg.UndefinedColumnError:
            await db.pool.execute(
                """
                INSERT INTO users (id, email, password_hash, role, status, created_at)
                VALUES ($1, $2, $3, 'professional', 'active', NOW())
            """,
                user_id,
                email,
                "hash_placeholder",
            )

    try:
        # 2. Crear profesional (tenant_id = clínica elegida)
        wh = professional.working_hours or generate_default_working_hours()
        if not isinstance(wh, dict):
            wh = generate_default_working_hours()
        try:
            wh_json = json.dumps(wh)
        except (TypeError, ValueError):
            wh_json = json.dumps(generate_default_working_hours())
        matricula = professional.license_number or None
        phone_val = professional.phone or None
        specialty_val = professional.specialty or None

        params = [
            tenant_id,
            user_id,
            first_name,
            last_name,
            email,
            phone_val,
            specialty_val,
            matricula,
            professional.is_active,
            wh_json,
        ]
        try:
            await db.pool.execute(
                """
                INSERT INTO professionals (
                    tenant_id, user_id, first_name, last_name, email, phone_number,
                    specialty, registration_id, is_active, working_hours, created_at, updated_at
                ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb, NOW(), NOW())
            """,
                *params,
            )
        except asyncpg.UndefinedColumnError as e:
            err_str = str(e).lower()
            if "registration_id" in err_str:
                await db.pool.execute(
                    """
                    INSERT INTO professionals (
                        tenant_id, user_id, first_name, last_name, email, phone_number,
                        specialty, license_number, is_active, working_hours, created_at, updated_at
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb, NOW(), NOW())
                """,
                    *params,
                )
            elif "updated_at" in err_str:
                await db.pool.execute(
                    """
                    INSERT INTO professionals (
                        tenant_id, user_id, first_name, last_name, email, phone_number,
                        specialty, registration_id, is_active, working_hours, created_at
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb, NOW())
                """,
                    *params,
                )
            elif "phone_number" in err_str:
                # BD antigua sin columna phone_number (solo columnas típicas en esquemas viejos)
                params_no_phone = [
                    tenant_id,
                    user_id,
                    first_name,
                    last_name,
                    email,
                    specialty_val,
                    matricula,
                    professional.is_active,
                    wh_json,
                ]
                try:
                    await db.pool.execute(
                        """
                        INSERT INTO professionals (
                            tenant_id, user_id, first_name, last_name, email,
                            specialty, registration_id, is_active, working_hours, created_at, updated_at
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, NOW(), NOW())
                    """,
                        *params_no_phone,
                    )
                except asyncpg.UndefinedColumnError as e2:
                    if "updated_at" in str(e2).lower():
                        await db.pool.execute(
                            """
                            INSERT INTO professionals (
                                tenant_id, user_id, first_name, last_name, email,
                                specialty, registration_id, is_active, working_hours, created_at
                            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, NOW())
                        """,
                            *params_no_phone,
                        )
                    else:
                        raise
            elif "email" in err_str:
                params_no_email = [
                    tenant_id,
                    user_id,
                    first_name,
                    last_name,
                    phone_val,
                    specialty_val,
                    matricula,
                    professional.is_active,
                    wh_json,
                ]
                await db.pool.execute(
                    """
                    INSERT INTO professionals (
                        tenant_id, user_id, first_name, last_name, phone_number,
                        specialty, registration_id, is_active, working_hours, created_at, updated_at
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, NOW(), NOW())
                """,
                    *params_no_email,
                )
            elif "specialty" in err_str:
                # BD sin columna specialty: INSERT sin especialidad
                params_no_spec = [
                    tenant_id,
                    user_id,
                    first_name,
                    last_name,
                    email,
                    phone_val,
                    matricula,
                    professional.is_active,
                    wh_json,
                ]
                try:
                    await db.pool.execute(
                        """
                        INSERT INTO professionals (
                            tenant_id, user_id, first_name, last_name, email, phone_number,
                            registration_id, is_active, working_hours, created_at, updated_at
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, NOW(), NOW())
                    """,
                        *params_no_spec,
                    )
                except asyncpg.UndefinedColumnError as e2:
                    err2 = str(e2).lower()
                    if "phone_number" in err2:
                        params_no_spec_phone = [
                            tenant_id,
                            user_id,
                            first_name,
                            last_name,
                            email,
                            matricula,
                            professional.is_active,
                            wh_json,
                        ]
                        await db.pool.execute(
                            """
                            INSERT INTO professionals (
                                tenant_id, user_id, first_name, last_name, email,
                                registration_id, is_active, working_hours, created_at, updated_at
                            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8::jsonb, NOW(), NOW())
                        """,
                            *params_no_spec_phone,
                        )
                    elif "updated_at" in err2:
                        await db.pool.execute(
                            """
                            INSERT INTO professionals (
                                tenant_id, user_id, first_name, last_name, email, phone_number,
                                registration_id, is_active, working_hours, created_at
                            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, NOW())
                        """,
                            *params_no_spec,
                        )
                    else:
                        raise
            else:
                logger.exception("create_professional INSERT column error")
                raise HTTPException(
                    status_code=500,
                    detail=f"Columna no reconocida en professionals: {e}",
                )
        except asyncpg.UniqueViolationError as e:
            logger.warning(f"create_professional duplicate: {e}")
            raise HTTPException(
                status_code=409,
                detail="Ya existe un usuario o profesional con ese email o datos.",
            )
        except asyncpg.ForeignKeyViolationError as e:
            logger.warning(f"create_professional FK: {e}")
            raise HTTPException(
                status_code=400,
                detail="La clínica elegida no existe. Creá una sede primero en Sedes (Clínicas).",
            )

        # Enviar email de bienvenida si el profesional está activo
        if professional.is_active:
            asyncio.create_task(
                send_welcome_email(tenant_id, str(user_id), "professional", db.pool)
            )
            logger.info(f"📧 Email de bienvenida programado para usuario {user_id}")

        return {"status": "created", "user_id": str(user_id)}
    except HTTPException:
        raise
    except Exception as e:
        err_msg = str(e).lower()
        if "unique" in err_msg or "duplicate" in err_msg:
            raise HTTPException(
                status_code=409,
                detail="Ya existe un usuario o profesional con ese email o datos.",
            )
        if (
            "foreign key" in err_msg
            or "violates foreign key" in err_msg
            or "tenant" in err_msg
        ):
            raise HTTPException(
                status_code=400,
                detail="La clínica elegida no existe. Creá una sede primero en Sedes (Clínicas).",
            )
        logger.exception("Error creating professional")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.put(
    "/professionals/{id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Actualizar datos de un profesional",
)
async def update_professional(
    id: int,
    payload: ProfessionalUpdate,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """Actualizar datos de un profesional por su ID numérico."""
    try:
        # Verificar existencia Y ownership multi-tenant (Regla de Oro)
        prof_row = await db.pool.fetchrow(
            "SELECT tenant_id FROM professionals WHERE id = $1", id
        )
        if not prof_row:
            raise HTTPException(status_code=404, detail="Profesional no encontrado")
        if prof_row["tenant_id"] not in allowed_ids:
            raise HTTPException(status_code=403, detail="No tienes acceso a esta sede")

        # Actualizar datos básicos, working_hours, google_calendar_id, consultation_price e is_priority_professional
        current_wh = payload.working_hours
        gcal_id = (payload.google_calendar_id or "").strip() or None
        cp = payload.consultation_price
        is_priority = bool(payload.is_priority_professional)

        # Split full name into first_name + last_name to avoid the "Delgado Delgado Delgado" bug
        # where assigning the full name to first_name caused last_name to be re-appended on every edit.
        full_name = (payload.name or "").strip()
        name_parts = full_name.split(maxsplit=1)
        first_name_val = name_parts[0] if name_parts else ""
        last_name_val = name_parts[1] if len(name_parts) > 1 else ""

        if current_wh:
            sql_update = """
                UPDATE professionals SET
                    first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                    phone_number = $5, email = $6, is_active = $7,
                    working_hours = $8::jsonb,
                    google_calendar_id = $9, consultation_price = $10,
                    is_priority_professional = $11,
                    updated_at = NOW()
                WHERE id = $12
            """
            params = [
                first_name_val,
                last_name_val,
                payload.specialty,
                payload.license_number,
                payload.phone,
                payload.email,
                payload.is_active,
                json.dumps(current_wh),
                gcal_id,
                cp,
                is_priority,
                id,
            ]
        else:
            sql_update = """
                UPDATE professionals SET
                    first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                    phone_number = $5, email = $6, is_active = $7,
                    google_calendar_id = $8, consultation_price = $9,
                    is_priority_professional = $10,
                    updated_at = NOW()
                WHERE id = $11
            """
            params = [
                first_name_val,
                last_name_val,
                payload.specialty,
                payload.license_number,
                payload.phone,
                payload.email,
                payload.is_active,
                gcal_id,
                cp,
                is_priority,
                id,
            ]

        try:
            await db.pool.execute(sql_update, *params)
        except asyncpg.UndefinedColumnError as e:
            err_str = str(e).lower()
            if "consultation_price" in err_str:
                # DB without consultation_price column yet (migration not run)
                if current_wh:
                    await db.pool.execute(
                        """
                        UPDATE professionals SET
                            first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                            phone_number = $5, email = $6, is_active = $7,
                            working_hours = $8::jsonb, google_calendar_id = $9,
                            updated_at = NOW()
                        WHERE id = $10
                    """,
                        first_name_val,
                        last_name_val,
                        payload.specialty,
                        payload.license_number,
                        payload.phone,
                        payload.email,
                        payload.is_active,
                        json.dumps(current_wh),
                        gcal_id,
                        id,
                    )
                else:
                    await db.pool.execute(
                        """
                        UPDATE professionals SET
                            first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                            phone_number = $5, email = $6, is_active = $7,
                            google_calendar_id = $8,
                            updated_at = NOW()
                        WHERE id = $9
                    """,
                        first_name_val,
                        last_name_val,
                        payload.specialty,
                        payload.license_number,
                        payload.phone,
                        payload.email,
                        payload.is_active,
                        gcal_id,
                        id,
                    )
            elif "google_calendar_id" in err_str:
                # DB without google_calendar_id column
                if current_wh:
                    await db.pool.execute(
                        """
                        UPDATE professionals SET
                            first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                            phone_number = $5, email = $6, is_active = $7,
                            working_hours = $8::jsonb,
                            updated_at = NOW()
                        WHERE id = $9
                    """,
                        first_name_val,
                        last_name_val,
                        payload.specialty,
                        payload.license_number,
                        payload.phone,
                        payload.email,
                        payload.is_active,
                        json.dumps(current_wh),
                        id,
                    )
                else:
                    await db.pool.execute(
                        """
                        UPDATE professionals SET
                            first_name = $1, last_name = $2, specialty = $3, registration_id = $4,
                            phone_number = $5, email = $6, is_active = $7,
                            updated_at = NOW()
                        WHERE id = $8
                    """,
                        first_name_val,
                        last_name_val,
                        payload.specialty,
                        payload.license_number,
                        payload.phone,
                        payload.email,
                        payload.is_active,
                        id,
                    )
            else:
                raise

        return {"id": id, "status": "updated"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating professional {id}: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.get(
    "/patients/{id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Obtener detalle de un paciente específico",
)
async def get_patient(id: int, tenant_id: int = Depends(get_resolved_tenant_id)):
    """Obtener un paciente por ID. Aislado por tenant_id (Regla de Oro)."""
    row = await db.pool.fetchrow(
        """
        SELECT id, first_name, last_name, phone_number, email, insurance_provider as obra_social,
               dni, birth_date, city, first_touch_source as acquisition_source,
               meta_ad_id, meta_adset_id, meta_adset_name, meta_ad_name, meta_campaign_name,
               medical_history, created_at, status, notes, anamnesis_token
        FROM patients
        WHERE id = $1 AND tenant_id = $2
    """,
        id,
        tenant_id,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    patient_dict = dict(row)
    # Generar anamnesis_token si no existe (lazy generation)
    if not patient_dict.get("anamnesis_token"):
        new_token = str(uuid.uuid4())
        await db.pool.execute(
            "UPDATE patients SET anamnesis_token = $1 WHERE id = $2 AND tenant_id = $3",
            new_token,
            id,
            tenant_id,
        )
        patient_dict["anamnesis_token"] = new_token
    else:
        patient_dict["anamnesis_token"] = str(patient_dict["anamnesis_token"])
    # Asegurar que medical_history se devuelve como dict
    if patient_dict.get("medical_history") and isinstance(
        patient_dict["medical_history"], str
    ):
        try:
            patient_dict["medical_history"] = json.loads(
                patient_dict["medical_history"]
            )
        except:
            pass
    return patient_dict


@router.put(
    "/patients/{id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Actualizar datos de un paciente",
)
async def update_patient(
    id: int, p: PatientCreate, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Actualizar datos de un paciente. Aislado por tenant_id (Regla de Oro)."""
    try:
        result = await db.pool.execute(
            """
            UPDATE patients
            SET first_name = COALESCE($1, first_name),
                last_name = COALESCE($2, last_name),
                phone_number = COALESCE($3, phone_number),
                email = $4,
                dni = $5,
                insurance_provider = $6,
                city = $7,
                birth_date = $8,
                updated_at = NOW()
            WHERE id = $9 AND tenant_id = $10
        """,
            p.first_name,
            p.last_name,
            p.phone_number,
            p.email,
            p.dni,
            p.insurance,
            p.city,
            p.birth_date,
            id,
            tenant_id,
        )
        if result == "UPDATE 0":
            raise HTTPException(status_code=404, detail="Paciente no encontrado")
        return {"id": id, "status": "updated"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating patient {id}: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.delete(
    "/patients/{id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Eliminar (soft-delete) un paciente",
)
async def delete_patient(id: int, tenant_id: int = Depends(get_resolved_tenant_id)):
    """Marcar paciente como eliminado. Aislado por tenant_id (Regla de Oro)."""
    try:
        result = await db.pool.execute(
            "UPDATE patients SET status = 'deleted', updated_at = NOW() WHERE id = $1 AND tenant_id = $2",
            id,
            tenant_id,
        )
        if result == "UPDATE 0":
            raise HTTPException(status_code=404, detail="Paciente no encontrado")
        return {"status": "deleted", "id": id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting patient {id}: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.get(
    "/patients/{id}/records",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Obtener historia clínica del paciente",
)
async def get_clinical_records(
    id: int, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Obtener historia clínica de un paciente. Aislado por tenant_id (Regla de Oro)."""
    rows = await db.pool.fetch(
        """
        SELECT cr.id, cr.diagnosis, cr.treatment_plan, cr.odontogram_data, cr.created_at 
        FROM clinical_records cr
        JOIN patients p ON cr.patient_id = p.id
        WHERE cr.patient_id = $1 AND p.tenant_id = $2
        ORDER BY cr.created_at DESC
    """,
        id,
        tenant_id,
    )

    records = []
    for row in rows:
        record_dict = dict(row)

        # Parse JSONB string outputs back into native Python dictionaries so React receives JSON hierarchy
        if isinstance(record_dict.get("odontogram_data"), str):
            try:
                record_dict["odontogram_data"] = json.loads(
                    record_dict["odontogram_data"]
                )
            except:
                record_dict["odontogram_data"] = {}

        if isinstance(record_dict.get("treatment_plan"), str):
            try:
                record_dict["treatment_plan"] = json.loads(
                    record_dict["treatment_plan"]
                )
            except:
                record_dict["treatment_plan"] = {}

        records.append(record_dict)

    return records


@router.post(
    "/patients/{id}/records",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Agregar nota a la historia clínica",
)
async def add_clinical_note(
    id: int, note: ClinicalNote, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Agregar una evolución/nota a la historia clínica. Aislado por tenant_id (Regla de Oro)."""
    patient = await db.pool.fetchrow(
        "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2", id, tenant_id
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Incluir odontogram_data solo si tiene datos reales (no vacío)
    odontogram_json = None
    if note.odontogram_data and isinstance(note.odontogram_data, dict):
        # Check it actually has teeth data, not just an empty shell
        has_data = (
            note.odontogram_data.get("teeth")
            or note.odontogram_data.get("permanent")
            or any(str(k).isdigit() for k in note.odontogram_data.keys())
        )
        if has_data:
            odontogram_json = json.dumps(note.odontogram_data)

    await db.pool.execute(
        """
        INSERT INTO clinical_records (id, tenant_id, patient_id, diagnosis, treatment_plan, odontogram_data, created_at)
        VALUES ($1, $2, $3, $4, $5, $6, NOW())
    """,
        str(uuid.uuid4()),
        tenant_id,
        id,
        note.content,
        "{}",
        odontogram_json,
    )
    return {"status": "ok"}


# ==================== ENDPOINTS ODONTOGRAMA Y DOCUMENTOS (FICHA MÉDICA) ====================


@router.put(
    "/patients/{patient_id}/records/{record_id}/odontogram",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Actualizar odontograma en registro clínico",
)
async def update_odontogram(
    patient_id: int,
    record_id: str,
    odontogram_data: OdontogramUpdate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Actualiza el odontograma de un registro clínico específico.
    Aislado por tenant_id (Regla de Oro).
    """
    # Verificar que el registro pertenece al paciente y tenant
    record = await db.pool.fetchrow(
        """
        SELECT cr.id FROM clinical_records cr
        JOIN patients p ON cr.patient_id = p.id
        WHERE cr.id = $1 AND cr.patient_id = $2 AND p.tenant_id = $3
    """,
        record_id,
        patient_id,
        tenant_id,
    )

    if not record:
        raise HTTPException(status_code=404, detail="Registro clínico no encontrado")

    # Actualizar odontograma - Asegurar filtro tenant_id
    await db.pool.execute(
        """
        UPDATE clinical_records cr
        SET odontogram_data = $1
        FROM patients p
        WHERE cr.id = $2 
        AND cr.patient_id = p.id 
        AND p.tenant_id = $3
    """,
        json.dumps(normalize_to_v3(odontogram_data.odontogram_data)),
        record_id,
        tenant_id,
    )

    # Emit Socket.IO event for real-time UI sync
    try:
        from main import sio, to_json_safe

        v3_normalized = normalize_to_v3(odontogram_data.odontogram_data)
        await sio.emit(
            "ODONTOGRAM_UPDATED",
            to_json_safe(
                {
                    "patient_id": patient_id,
                    "record_id": record_id,
                    "tenant_id": tenant_id,
                    "odontogram_data": v3_normalized,
                }
            ),
        )
    except Exception:
        pass  # Non-critical — UI will still get the REST response

    return {"status": "ok", "message": "Odontograma actualizado"}


@router.get(
    "/patients/{patient_id}/records/{record_id}/odontogram",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Obtener odontograma de registro clínico",
)
async def get_odontogram(
    patient_id: int, record_id: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """
    Obtiene el odontograma de un registro clínico específico.
    Aislado por tenant_id (Regla de Oro).
    """
    record = await db.pool.fetchrow(
        """
        SELECT cr.odontogram_data FROM clinical_records cr
        JOIN patients p ON cr.patient_id = p.id
        WHERE cr.id = $1 AND cr.patient_id = $2 AND p.tenant_id = $3
    """,
        record_id,
        patient_id,
        tenant_id,
    )

    if not record:
        raise HTTPException(status_code=404, detail="Registro clínico no encontrado")

    return {"odontogram_data": normalize_to_v3(record["odontogram_data"] or {})}


@router.post(
    "/patients/{id}/documents",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Subir documento para paciente",
)
async def upload_patient_document(
    id: int,
    file: UploadFile = File(...),
    document_type: str = Form("clinical"),
    tenant_id: int = Depends(get_resolved_tenant_id),
    user_data: dict = Depends(verify_admin_token),
):
    """
    Sube un documento para un paciente específico.
    Aislado por tenant_id (Regla de Oro).
    """
    # Verificar que el paciente existe y pertenece al tenant
    patient = await db.pool.fetchrow(
        """
        SELECT id FROM patients WHERE id = $1 AND tenant_id = $2
    """,
        id,
        tenant_id,
    )

    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Validar tipo de archivo
    allowed_mime_types = [
        "application/pdf",
        "image/jpeg",
        "image/png",
        "image/gif",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ]

    if file.content_type not in allowed_mime_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido")

    # Validar tamaño (10MB máximo)
    file_size = 0
    content = await file.read()
    file_size = len(content)
    await file.seek(0)

    if file_size > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(
            status_code=400, detail="Archivo demasiado grande (máximo 10MB)"
        )

    # Crear directorio seguro
    import os
    from pathlib import Path

    upload_dir = Path(f"/app/uploads/patients/{tenant_id}/{id}")
    upload_dir.mkdir(parents=True, exist_ok=True)

    # Generar nombre de archivo seguro
    import uuid

    safe_filename = f"{uuid.uuid4()}_{file.filename.replace(' ', '_')}"
    file_path = upload_dir / safe_filename

    # Guardar archivo
    with open(file_path, "wb") as f:
        f.write(content)

    # Guardar en base de datos
    doc_id = await db.pool.fetchval(
        """
        INSERT INTO patient_documents 
        (tenant_id, patient_id, file_name, file_path, file_size, mime_type, document_type, uploaded_by, source, uploaded_at)
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, 'manual', NOW())
        RETURNING id
    """,
        tenant_id,
        id,
        file.filename,
        str(file_path),
        file_size,
        file.content_type,
        document_type,
        user_data.user_id,
    )

    return {
        "id": doc_id,
        "file_name": file.filename,
        "filename": file.filename,  # Compatibility
        "file_path": str(file_path),
        "file_size": file_size,
        "mime_type": file.content_type,
        "document_type": document_type,
        "uploaded_by": user_data.user_id,
        "proxy_url": f"/admin/patients/{id}/documents/{doc_id}/proxy",
    }


@router.get(
    "/patients/{id}/documents",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Listar documentos del paciente",
)
async def list_patient_documents(
    id: int, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """
    Lista todos los documentos de un paciente.
    Aislado por tenant_id (Regla de Oro).
    """
    # Verificar que el paciente existe y pertenece al tenant
    patient = await db.pool.fetchrow(
        """
        SELECT id FROM patients WHERE id = $1 AND tenant_id = $2
    """,
        id,
        tenant_id,
    )

    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    documents = await db.pool.fetch(
        """
        SELECT id, file_name as filename, file_name, file_path, file_size, mime_type, document_type, 
               uploaded_by, created_at, source, source_details, uploaded_at
        FROM patient_documents
        WHERE patient_id = $1 AND tenant_id = $2
        ORDER BY created_at DESC
    """,
        id,
        tenant_id,
    )

    from urllib.parse import urlencode

    out = []
    for doc in documents:
        d = dict(doc)
        file_path = d.get("file_path", "")
        if file_path:
            # ✅ FIX: Generar URL firmada para previsualización (Spec 19)
            signature, expires = generate_signed_url(file_path, tenant_id)
            proxy_params = {
                "url": file_path,
                "tenant_id": tenant_id,
                "signature": signature,
                "expires": expires,
            }
            # Usamos el proxy unificado definido en chat_api
            d["url"] = f"/admin/chat/media/proxy?{urlencode(proxy_params)}"
        out.append(d)

    return out


@router.get(
    "/patients/{patient_id}/documents/{doc_id}/proxy",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Proxy seguro para descargar documento",
)
async def download_patient_document_proxy(
    patient_id: int, doc_id: int, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """
    Proxy seguro para descargar documentos de pacientes.
    Aislado por tenant_id (Regla de Oro).
    """
    logger.info(
        f"📥 Document proxy request: patient={patient_id} doc={doc_id} tenant={tenant_id}"
    )
    # Verificar que el documento existe y pertenece al paciente/tenant
    document = await db.pool.fetchrow(
        """
        SELECT file_path, mime_type, file_name
        FROM patient_documents
        WHERE id = $1 AND patient_id = $2 AND tenant_id = $3
    """,
        doc_id,
        patient_id,
        tenant_id,
    )

    if not document:
        logger.warning(
            f"Document {doc_id} not found in DB for patient {patient_id}, tenant {tenant_id}"
        )
        raise HTTPException(
            status_code=404, detail="El documento no existe en la base de datos."
        )

    file_path = document["file_path"]

    # Si es una URL externa (ej: YCloud media), descargar a disco + actualizar DB + devolver
    if file_path.startswith("http://") or file_path.startswith("https://"):
        import httpx
        import uuid as _uuid

        try:
            headers_dl = {}
            ycloud_key = os.getenv("YCLOUD_API_KEY")
            if ycloud_key and "ycloud" in file_path.lower():
                headers_dl["X-API-Key"] = ycloud_key
            async with httpx.AsyncClient(
                timeout=30.0, follow_redirects=True, headers=headers_dl
            ) as client:
                resp = await client.get(file_path)
                if resp.status_code != 200:
                    raise HTTPException(
                        status_code=502,
                        detail=f"Error descargando archivo externo: {resp.status_code}",
                    )
                if len(resp.content) == 0:
                    raise HTTPException(status_code=502, detail="Archivo externo vacío")
                content_type = document["mime_type"] or resp.headers.get(
                    "content-type", "application/octet-stream"
                )
                file_name = document["file_name"] or "document"

                # Persistir a disco para que no se pierda
                ext = os.path.splitext(file_name)[1] or ".jpg"
                new_filename = f"{_uuid.uuid4()}{ext}"
                uploads_dir = os.getenv(
                    "UPLOADS_DIR", os.path.join(os.getcwd(), "uploads")
                )
                media_dir = os.path.join(uploads_dir, str(tenant_id))
                os.makedirs(media_dir, exist_ok=True)
                local_path = os.path.join(media_dir, new_filename)
                with open(local_path, "wb") as f:
                    f.write(resp.content)
                new_file_path = f"/uploads/{tenant_id}/{new_filename}"
                logger.info(
                    f"✅ External media persisted: {new_file_path} ({len(resp.content)} bytes)"
                )

                # Actualizar DB para que la próxima vez se sirva desde disco
                await db.pool.execute(
                    "UPDATE patient_documents SET file_path = $1 WHERE id = $2 AND tenant_id = $3",
                    new_file_path,
                    doc_id,
                    tenant_id,
                )

                return Response(
                    content=resp.content,
                    media_type=content_type,
                    headers={"Content-Disposition": f'inline; filename="{file_name}"'},
                )
        except httpx.HTTPError as e:
            logger.error(f"Error downloading external media: {e}")
            raise HTTPException(
                status_code=502,
                detail="No se pudo descargar el archivo desde el proveedor externo.",
            )

    # Limpiar path de posibles parámetros (HMAC legacy)
    clean_path = file_path.split("?")[0]

    # Intentar múltiples paths para encontrar el archivo en disco
    uploads_dir = os.getenv("UPLOADS_DIR", os.path.join(os.getcwd(), "uploads"))
    # Strip leading /media/ or /uploads/ prefix for recombination
    stripped = clean_path
    for prefix in ("/media/", "/uploads/"):
        if stripped.startswith(prefix):
            stripped = stripped[len(prefix) :]
            break
    candidates = [
        os.path.join(
            os.getcwd(), clean_path.lstrip("/")
        ),  # /app/media/1/file.jpg or /app/uploads/1/file.jpg
        clean_path,  # absolute path as-is
        os.path.join(uploads_dir, stripped),  # UPLOADS_DIR/1/file.jpg
        os.path.join(
            os.getcwd(), "media", stripped
        ),  # /app/media/1/file.jpg (explicit)
        os.path.join(
            os.getcwd(), "uploads", stripped
        ),  # /app/uploads/1/file.jpg (explicit)
        os.path.join("/media", stripped),  # /media/1/file.jpg (volume mount)
        os.path.join("/uploads", stripped),  # /uploads/1/file.jpg (volume mount)
    ]

    full_path = None
    for candidate in candidates:
        if os.path.exists(candidate):
            full_path = candidate
            break

    if not full_path:
        logger.error(f"File not found on disk. Tried: {candidates}")
        raise HTTPException(
            status_code=404,
            detail="Falta el archivo en el servidor. (Posible reinicio del contenedor sin volumen persistente)",
        )

    # Path traversal protection: ensure resolved path is within allowed roots
    import pathlib

    resolved = pathlib.Path(full_path).resolve()
    allowed_roots = [
        pathlib.Path("/app/uploads").resolve(),
        pathlib.Path("/app/media").resolve(),
        pathlib.Path(os.getcwd(), "uploads").resolve(),
        pathlib.Path(os.getcwd(), "media").resolve(),
    ]
    if not any(str(resolved).startswith(str(root)) for root in allowed_roots):
        raise HTTPException(status_code=403, detail="Acceso denegado al archivo")

    download_filename = document["file_name"] or os.path.basename(clean_path)

    return FileResponse(
        path=full_path,
        filename=download_filename,
        media_type=document["mime_type"] or "application/octet-stream",
    )


@router.delete(
    "/patients/{patient_id}/documents/{doc_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Eliminar documento del paciente",
)
async def delete_patient_document(
    patient_id: int, doc_id: int, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """
    Elimina un documento de un paciente.
    Aislado por tenant_id (Regla de Oro).
    """
    # Verify patient belongs to tenant
    patient_exists = await db.pool.fetchval(
        "SELECT 1 FROM patients WHERE id = $1 AND tenant_id = $2", patient_id, tenant_id
    )
    if not patient_exists:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Verificar que el documento existe y pertenece al paciente/tenant
    document = await db.pool.fetchrow(
        """
        SELECT file_path FROM patient_documents
        WHERE id = $1 AND patient_id = $2 AND tenant_id = $3
    """,
        doc_id,
        patient_id,
        tenant_id,
    )

    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    # Eliminar archivo físico
    import os

    file_path = document["file_path"]
    if os.path.exists(file_path):
        os.remove(file_path)

    # Eliminar registro de base de datos
    await db.pool.execute(
        """
        DELETE FROM patient_documents
        WHERE id = $1 AND patient_id = $2 AND tenant_id = $3
    """,
        doc_id,
        patient_id,
        tenant_id,
    )

    return {"status": "ok", "message": "Documento eliminado"}


@router.get(
    "/patients/{patient_id}/attachments-summary",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Obtener el resumen más reciente de adjuntos del paciente",
)
async def get_attachment_summary(
    patient_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Get the most recent attachment summary for a patient."""
    # Verify patient belongs to tenant
    patient = await db.pool.fetchrow(
        "SELECT id, tenant_id FROM patients WHERE id = $1 AND tenant_id = $2",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Get latest summary
    summary = await db.pool.fetchrow(
        """
        SELECT summary_text, attachments_count, attachments_types, created_at
        FROM clinical_record_summaries
        WHERE tenant_id = $1 AND patient_id = $2
        ORDER BY created_at DESC
        LIMIT 1
    """,
        tenant_id,
        patient_id,
    )

    if not summary:
        return {
            "summary_text": None,
            "attachments_count": 0,
            "attachments_types": [],
            "created_at": None,
        }

    return {
        "summary_text": summary["summary_text"],
        "attachments_count": summary["attachments_count"],
        "attachments_types": summary["attachments_types"] or [],
        "created_at": summary["created_at"].isoformat()
        if summary["created_at"]
        else None,
    }


# ==================== ENDPOINTS TRATAMIENTOS ====================


@router.get(
    "/treatment_types",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Listar tipos de tratamiento disponibles",
)
async def list_treatment_types_legacy(
    tenant_id: int = Depends(get_resolved_tenant_id),
    only_active: bool = Query(True, description="Solo mostrar tratamientos activos"),
    only_booking: bool = Query(
        False, description="Solo mostrar tratamientos disponibles para reserva"
    ),
):
    """Listar tipos de tratamiento por clínica. Aislado por tenant_id (Regla de Oro)."""
    query = "SELECT * FROM treatment_types WHERE tenant_id = $1"
    params = [tenant_id]

    if only_active:
        query += " AND is_active = true"
    if only_booking:
        query += " AND is_available_for_booking = true"

    query += " ORDER BY name ASC"
    rows = await db.pool.fetch(query, *params)
    from decimal import Decimal as _Dec2

    treatments = [
        {k: float(v) if isinstance(v, _Dec2) else v for k, v in dict(row).items()}
        for row in rows
    ]
    # Batch-fetch professional assignments
    if treatments:
        tt_ids = [t["id"] for t in treatments]
        ttp_rows = await db.pool.fetch(
            "SELECT treatment_type_id, professional_id FROM treatment_type_professionals WHERE tenant_id = $1 AND treatment_type_id = ANY($2)",
            tenant_id,
            tt_ids,
        )
        prof_map: dict = {}
        for r in ttp_rows:
            prof_map.setdefault(r["treatment_type_id"], []).append(r["professional_id"])
        for t in treatments:
            t["professional_ids"] = prof_map.get(t["id"], [])
    return treatments


# ==================== ENDPOINTS TURNOS (AGENDA) ====================


@router.get(
    "/appointments",
    dependencies=[Depends(verify_admin_token)],
    tags=["Turnos"],
    summary="Listar turnos en un rango de fechas",
    response_model=List[AppointmentResponse],
)
async def list_appointments(
    start_date: str,
    end_date: str,
    professional_id: Optional[int] = None,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Obtener turnos del calendario. Aislado por tenant_id (Regla de Oro)."""
    query = """
        SELECT a.id, a.patient_id, a.appointment_datetime, a.duration_minutes, a.status, a.urgency_level,
               a.source, a.appointment_type, a.notes,
               a.billing_amount, a.billing_installments, a.billing_notes, a.payment_status, a.payment_receipt_data,
               (a.appointment_datetime + (COALESCE(a.duration_minutes, 30) || ' minutes')::interval) AS end_datetime,
               (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name,
               p.phone_number as patient_phone,
               prof.first_name as professional_name, prof.id as professional_id,
               COALESCE(tt.name, a.appointment_type, 'Consulta') as appointment_name,
               CASE WHEN LOWER(COALESCE(p.notes, '') || ' ' || COALESCE(p.medical_history::text, '')) ~ '(diabetes|hipertension|cardiopatia|hemofilia|alergia penicilina|embarazo|anticoagulacion|vih|hepatitis|asma severa)' THEN true ELSE false END as has_medical_alerts
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        LEFT JOIN professionals prof ON a.professional_id = prof.id
        LEFT JOIN treatment_types tt ON a.appointment_type = tt.code AND a.tenant_id = tt.tenant_id
        WHERE a.tenant_id = $1 AND a.appointment_datetime BETWEEN $2 AND $3
    """
    params = [
        tenant_id,
        datetime.fromisoformat(start_date),
        datetime.fromisoformat(end_date),
    ]
    if professional_id:
        query += f" AND a.professional_id = ${len(params) + 1}"
        params.append(professional_id)
    query += " ORDER BY a.appointment_datetime ASC"
    rows = await db.pool.fetch(query, *params)
    # Serialize all fields properly (UUID, Decimal, JSONB)
    from decimal import Decimal as _Decimal

    result = []
    for row in rows:
        d = {}
        for key, val in dict(row).items():
            if isinstance(val, _Decimal):
                d[key] = float(val)
            elif isinstance(val, uuid.UUID):
                d[key] = str(val)
            elif isinstance(val, str) and key.endswith("_data"):
                # Try to parse JSONB string fields
                try:
                    d[key] = json.loads(val)
                except (json.JSONDecodeError, TypeError):
                    d[key] = val
            else:
                d[key] = val
        result.append(d)
    return result


# ==================== ENDPOINT: GET SINGLE APPOINTMENT ====================


@router.get(
    "/appointments/{appointment_id}/audit",
    tags=["Turnos"],
    summary="Audit log de un turno (TIER 3) — solo CEO/admin",
)
async def get_appointment_audit_log(
    appointment_id: str,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Return chronological audit history for an appointment.

    Multi-tenant isolation: returns 404 if the appointment is not in the caller's tenant.
    Role guard: only users with role 'ceo' or 'admin' can read audit logs.
    """
    role = (user_data or {}).get("role") or ""
    if role not in ("ceo", "admin"):
        raise HTTPException(
            status_code=403, detail="Solo CEO/admin pueden leer audit logs"
        )

    try:
        apt_uuid = uuid.UUID(appointment_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de turno inválido")

    # Verify the appointment belongs to caller's tenant (multi-tenant isolation)
    apt_check = await db.pool.fetchrow(
        "SELECT id FROM appointments WHERE id = $1 AND tenant_id = $2",
        apt_uuid,
        tenant_id,
    )
    if not apt_check:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    rows = await db.pool.fetch(
        """
        SELECT al.id, al.action, al.actor_type, al.actor_id,
               al.before_values, al.after_values, al.source_channel, al.reason,
               al.created_at,
               u.email AS actor_email
        FROM appointment_audit_log al
        LEFT JOIN users u ON u.id::text = al.actor_id AND al.actor_type = 'staff_user'
        WHERE al.tenant_id = $1 AND al.appointment_id = $2
        ORDER BY al.created_at ASC
        """,
        tenant_id,
        apt_uuid,
    )

    return [
        {
            "id": r["id"],
            "action": r["action"],
            "actor_type": r["actor_type"],
            "actor_id": r["actor_id"],
            "actor_display": r["actor_email"] or r["actor_id"],
            "before_values": r["before_values"],
            "after_values": r["after_values"],
            "source_channel": r["source_channel"],
            "reason": r["reason"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
        }
        for r in rows
    ]


@router.get(
    "/appointments/{appointment_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Turnos"],
    summary="Obtener un turno por ID con todos los campos",
)
async def get_appointment_by_id(
    appointment_id: str,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Get a single appointment with ALL fields including billing and receipt data."""
    from decimal import Decimal as _Decimal

    try:
        apt_uuid = uuid.UUID(appointment_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de turno inválido")

    row = await db.pool.fetchrow(
        """
        SELECT a.id, a.patient_id, a.appointment_datetime, a.duration_minutes, a.status, a.urgency_level,
               a.source, a.appointment_type, a.notes,
               a.billing_amount, a.billing_installments, a.billing_notes, a.payment_status, a.payment_receipt_data,
               a.plan_item_id,
               (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name,
               p.phone_number as patient_phone,
               prof.first_name as professional_name, prof.id as professional_id,
               COALESCE(tt.name, a.appointment_type, 'Consulta') as appointment_name,
               tp.name as plan_name, tp.id as plan_id
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        LEFT JOIN professionals prof ON a.professional_id = prof.id
        LEFT JOIN treatment_types tt ON a.appointment_type = tt.code AND a.tenant_id = tt.tenant_id
        LEFT JOIN treatment_plan_items tpi ON a.plan_item_id = tpi.id AND a.tenant_id = tpi.tenant_id
        LEFT JOIN treatment_plans tp ON tpi.plan_id = tp.id AND tpi.tenant_id = tp.tenant_id
        WHERE a.id = $1 AND a.tenant_id = $2
    """,
        apt_uuid,
        tenant_id,
    )

    if not row:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    d = {}
    for key, val in dict(row).items():
        if isinstance(val, _Decimal):
            d[key] = float(val)
        elif isinstance(val, uuid.UUID):
            d[key] = str(val)
        elif isinstance(val, str) and key.endswith("_data"):
            try:
                d[key] = json.loads(val)
            except (json.JSONDecodeError, TypeError):
                d[key] = val
        else:
            d[key] = val
    return d


# ==================== ENDPOINT COLISION DETECTION ====================


@router.get(
    "/appointments/check-collisions",
    dependencies=[Depends(verify_admin_token)],
    tags=["Turnos"],
    summary="Verificar colisiones de horario para un turno",
)
async def check_collisions(
    professional_id: int,
    datetime_str: str,
    duration_minutes: int = 60,
    exclude_appointment_id: str = None,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Verificar colisiones de horario. Aislado por tenant_id (Regla de Oro)."""
    target_datetime = datetime.fromisoformat(datetime_str)
    target_end = target_datetime + timedelta(minutes=duration_minutes)
    overlap_query = """
        SELECT id, appointment_datetime, duration_minutes, status, source
        FROM appointments
        WHERE tenant_id = $1 AND professional_id = $2
        AND status NOT IN ('cancelled', 'no-show')
        AND appointment_datetime < $4
        AND appointment_datetime + (duration_minutes || ' minutes')::interval > $3
    """
    params = [tenant_id, professional_id, target_datetime, target_end]
    if exclude_appointment_id:
        overlap_query += " AND id != $5"
        params.append(exclude_appointment_id)
    overlapping = await db.pool.fetch(overlap_query, *params)
    gcal_blocks = await db.pool.fetch(
        """
        SELECT id, title, start_datetime, end_datetime
        FROM google_calendar_blocks
        WHERE tenant_id = $1 AND (professional_id = $2 OR professional_id IS NULL)
        AND start_datetime < $4
        AND end_datetime > $3
    """,
        tenant_id,
        professional_id,
        target_datetime,
        target_end,
    )

    has_collisions = len(overlapping) > 0 or len(gcal_blocks) > 0

    return {
        "has_collisions": has_collisions,
        "conflicting_appointments": [dict(row) for row in overlapping],
        "conflicting_blocks": [dict(row) for row in gcal_blocks],
    }


@router.post(
    "/appointments",
    tags=["Turnos"],
    summary="Agendar un turno manualmente",
)
async def create_appointment_manual(
    apt: AppointmentCreate,
    request: Request,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Agendar turno manualmente. Aislado por tenant_id (Regla de Oro)."""
    try:
        if apt.check_collisions:
            collision_response = await check_collisions(
                apt.professional_id,
                apt.appointment_datetime.isoformat(),
                apt.duration_minutes or 60,  # Use duration_minutes from model
                None,
                tenant_id=tenant_id,
            )
            if collision_response["has_collisions"]:
                conflicts = []
                for apt_conflict in collision_response["conflicting_appointments"]:
                    conflicts.append(
                        f"Turno existente: {apt_conflict['appointment_datetime']}"
                    )
                for block in collision_response["conflicting_blocks"]:
                    conflicts.append(
                        f"Bloqueo GCalendar: {block['title']} ({block['start_datetime']})"
                    )
                raise HTTPException(
                    status_code=409,
                    detail=f"Hay colisiones de horario: {'; '.join(conflicts)}",
                )

        # 1. Validar que professional_id existe y pertenece al tenant
        prof_exists = await db.pool.fetchval(
            "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2 AND is_active = true",
            apt.professional_id,
            tenant_id,
        )
        if not prof_exists:
            raise HTTPException(
                status_code=400, detail="Profesional inválido o inactivo"
            )

        # 2. Resolver patient_id (solo dentro del tenant)
        pid = apt.patient_id
        if not pid and apt.patient_phone:
            exist = await db.pool.fetchrow(
                "SELECT id FROM patients WHERE tenant_id = $1 AND phone_number = $2",
                tenant_id,
                apt.patient_phone,
            )
            if exist:
                pid = exist["id"]
            else:
                new_p = await db.pool.fetchrow(
                    "INSERT INTO patients (tenant_id, phone_number, first_name, created_at) VALUES ($1, $2, $3, NOW()) RETURNING id",
                    tenant_id,
                    apt.patient_phone,
                    "Paciente Manual",
                )
                pid = new_p["id"]
        if not pid:
            raise HTTPException(
                status_code=400, detail="Se requiere ID de paciente o teléfono válido."
            )
        patient_exists = await db.pool.fetchval(
            "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2", pid, tenant_id
        )
        if not patient_exists:
            raise HTTPException(status_code=400, detail="Paciente no encontrado")
        # 4. Crear turno (source='manual')
        new_id = str(uuid.uuid4())
        duration = apt.duration_minutes or 30
        await db.pool.execute(
            """
            INSERT INTO appointments (
                id, tenant_id, patient_id, professional_id, appointment_datetime,
                duration_minutes, appointment_type, status, urgency_level, source, notes, created_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, 'confirmed', 'normal', 'manual', $8, NOW())
        """,
            new_id,
            tenant_id,
            pid,
            apt.professional_id,
            apt.appointment_datetime,
            duration,
            apt.appointment_type,
            apt.notes,
        )
        # Audit log (TIER 3 cap.3 Phase B)
        try:
            from services.audit_log import log_appointment_mutation as _audit

            await _audit(
                pool=db.pool,
                tenant_id=tenant_id,
                appointment_id=new_id,
                action="created",
                actor_type="staff_user",
                actor_id=getattr(user_data, "user_id", None),
                before_values=None,
                after_values={
                    "patient_id": pid,
                    "professional_id": apt.professional_id,
                    "appointment_datetime": apt.appointment_datetime.isoformat()
                    if hasattr(apt.appointment_datetime, "isoformat")
                    else str(apt.appointment_datetime),
                    "duration_minutes": duration,
                    "appointment_type": apt.appointment_type,
                    "status": "confirmed",
                    "source": "manual",
                    "notes": apt.notes,
                },
                source_channel="web_admin",
                reason=None,
            )
        except Exception as _audit_err:
            logger.warning(
                f"audit_log create_appointment_manual failed (non-blocking): {_audit_err}"
            )
        # 5. Obtener datos completos del turno para evento y GCal (incluye source y tenant_id para notificación)
        appointment_data = await db.pool.fetchrow(
            """
            SELECT a.id, a.patient_id, a.professional_id, a.appointment_datetime, 
                   a.appointment_type, a.status, a.urgency_level, a.notes, a.duration_minutes,
                   a.source, a.tenant_id,
                   (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name, 
                   p.phone_number as patient_phone,
                   p.first_name, p.last_name, -- Para el summary de GCal
                   prof.first_name as professional_name
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            JOIN professionals prof ON a.professional_id = prof.id
            WHERE a.id = $1
        """,
            new_id,
        )

        # 6. Sincronizar con Google Calendar
        try:
            # Obtener google_calendar_id del profesional
            google_calendar_id = await db.pool.fetchval(
                "SELECT google_calendar_id FROM professionals WHERE id = $1",
                apt.professional_id,
            )

            if google_calendar_id and appointment_data:
                summary = f"Cita Dental: {appointment_data['first_name']} {appointment_data['last_name'] or ''} - {apt.appointment_type}"
                start_time = apt.appointment_datetime.isoformat()
                end_time = (
                    apt.appointment_datetime + timedelta(minutes=60)
                ).isoformat()

                gcal_event = gcal_service.create_event(
                    calendar_id=google_calendar_id,
                    summary=summary,
                    start_time=start_time,
                    end_time=end_time,
                    description=f"Paciente: {appointment_data['first_name']}\nTel: {appointment_data['patient_phone']}\nNotas: {apt.notes or ''}",
                )

                if gcal_event:
                    await db.pool.execute(
                        "UPDATE appointments SET google_calendar_event_id = $1, google_calendar_sync_status = 'synced' WHERE id = $2",
                        gcal_event["id"],
                        new_id,
                    )
        except Exception as ge:
            logger.warning(f"GCal sync failed for appointment {new_id}: {ge}")

        # 7. Emitir evento de Socket.IO para actualización en tiempo real (no fallar la respuesta si falla el emit)
        if appointment_data:
            try:
                await emit_appointment_event(
                    "NEW_APPOINTMENT", dict(appointment_data), request
                )
            except Exception as emit_err:
                logger.warning(
                    f"Socket emit failed for NEW_APPOINTMENT {new_id}: {emit_err}"
                )

        return {
            "id": new_id,
            "status": "confirmed",
            "patient_id": pid,
            "source": "manual",
        }

    except asyncpg.ForeignKeyViolationError:
        raise HTTPException(
            status_code=400, detail="ID de profesional o paciente no válido"
        )
    except asyncpg.UniqueViolationError:
        raise HTTPException(status_code=409, detail="Turno duplicado para ese horario")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creando turno: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.put(
    "/appointments/{id}/status",
    tags=["Turnos"],
    summary="Actualizar el estado de un turno (confirmed, cancelled, etc.)",
)
@router.patch(
    "/appointments/{id}/status",
    tags=["Turnos"],
    summary="Actualizar el estado de un turno (vía PATCH)",
)
async def update_appointment_status(
    id: str,
    payload: StatusUpdate,
    request: Request,
    background_tasks: BackgroundTasks,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Cambiar estado: confirmed, cancelled, attended, no_show. Aislado por tenant_id (Regla de Oro)."""
    # Capture previous status for audit log
    _prev_status = await db.pool.fetchval(
        "SELECT status FROM appointments WHERE id = $1 AND tenant_id = $2",
        id,
        tenant_id,
    )
    # 2. Actualizar en BD
    update_fields = "status = $1, updated_at = NOW()"
    params = [payload.status, id, tenant_id]

    if payload.notes is not None:
        update_fields = "status = $1, notes = $4, updated_at = NOW()"
        params.append(payload.notes)

    result = await db.pool.execute(
        f"UPDATE appointments SET {update_fields} WHERE id = $2 AND tenant_id = $3",
        *params,
    )
    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail="Turno no encontrado o sin acceso")

    # Audit log (TIER 3 cap.3 Phase B)
    try:
        from services.audit_log import log_appointment_mutation as _audit

        _action = "cancelled" if payload.status == "cancelled" else "status_changed"
        await _audit(
            pool=db.pool,
            tenant_id=tenant_id,
            appointment_id=id,
            action=_action,
            actor_type="staff_user",
            actor_id=getattr(user_data, "user_id", None),
            before_values={"status": _prev_status} if _prev_status else None,
            after_values={"status": payload.status, "notes": payload.notes}
            if payload.notes is not None
            else {"status": payload.status},
            source_channel="web_admin",
            reason=None,
        )
    except Exception as _audit_err:
        logger.warning(
            f"audit_log update_appointment_status failed (non-blocking): {_audit_err}"
        )

    # Obtener datos actuales del turno para emitir evento y lógica de feedback
    appointment_data = await db.pool.fetchrow(
        """
        SELECT a.id, a.patient_id, a.professional_id, a.appointment_datetime, 
               a.appointment_type, a.status, a.urgency_level,
               (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name, 
               p.first_name, p.last_name, p.phone_number,
               prof.first_name as professional_name,
               a.google_calendar_event_id, a.google_calendar_sync_status
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        JOIN professionals prof ON a.professional_id = prof.id
        WHERE a.id = $1
    """,
        id,
    )

    if not appointment_data:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    if (
        payload.status == "completed"
        and appointment_data["status"] != "completed"
        and appointment_data["phone_number"]
    ):
        # Disparar feedback automático en background
        try:
            background_tasks.add_task(
                trigger_feedback_after_delay,
                appointment_id=int(id) if str(id).isdigit() else id,
                tenant_id=tenant_id,
                patient_name=f"{appointment_data['first_name'] or ''} {appointment_data['last_name'] or ''}".strip()
                or "paciente",
                phone_number=appointment_data["phone_number"],
                delay_minutes=45,
            )
            logger.info(f"⏰ Feedback programado para turno {id}")
        except Exception as e:
            logger.error(f"Error programando feedback: {e}")

    if appointment_data:
        # 1. Sincronizar cancelación con Google Calendar
        if (
            payload.status == "cancelled"
            and appointment_data["google_calendar_event_id"]
        ):
            try:
                # Need to fetch professional's calendar ID
                google_calendar_id = await db.pool.fetchval(
                    "SELECT google_calendar_id FROM professionals WHERE id = $1",
                    appointment_data["professional_id"],
                )

                if google_calendar_id:
                    gcal_service.delete_event(
                        calendar_id=google_calendar_id,
                        event_id=appointment_data["google_calendar_event_id"],
                    )
                    await db.pool.execute(
                        "UPDATE appointments SET google_calendar_sync_status = 'cancelled' WHERE id = $1",
                        id,
                    )
            except Exception as ge:
                logger.warning(f"Error deleting GCal event: {ge}")

        if (
            payload.status == "completed"
            and apt["status"] != "completed"
            and apt["phone_number"]
        ):
            # Disparar feedback automático en background
            try:
                background_tasks.add_task(
                    trigger_feedback_after_delay,
                    appointment_id=int(id) if str(id).isdigit() else id,
                    tenant_id=tenant_id,
                    patient_name=f"{apt['first_name'] or ''} {apt['last_name'] or ''}".strip()
                    or "paciente",
                    phone_number=apt["phone_number"],
                    delay_minutes=45,
                )
                logger.info(f"⏰ Feedback programado para turno {id}")
            except Exception as e:
                logger.error(f"Error programando feedback: {e}")

        # 2. Emitir evento según el nuevo estado
        try:
            if payload.status == "cancelled":
                await emit_appointment_event("APPOINTMENT_DELETED", id, request)
            else:
                await emit_appointment_event(
                    "APPOINTMENT_UPDATED", appointment_data_dict, request
                )
        except Exception as emit_err:
            logger.warning(f"Socket emit failed for status update {id}: {emit_err}")

    return {
        "status": "updated",
        "appointment_id": str(id),
        "new_status": payload.status,
    }


@router.put(
    "/appointments/{id}",
    tags=["Turnos"],
    summary="Actualizar datos completos de un turno",
)
async def update_appointment(
    id: str,
    apt: AppointmentCreate,
    request: Request,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualizar datos de un turno (fecha, profesional, tipo, notas)."""
    try:
        # 1. Obtener datos actuales (solo del tenant del usuario)
        old_apt = await db.pool.fetchrow(
            """
            SELECT id, professional_id, appointment_datetime, google_calendar_event_id, status
            FROM appointments WHERE id = $1 AND tenant_id = $2
        """,
            id,
            tenant_id,
        )

        if not old_apt:
            raise HTTPException(status_code=404, detail="Turno no encontrado")

        # 2. Verificar colisiones si la fecha o el profesional cambiaron
        date_changed = old_apt["appointment_datetime"] != apt.appointment_datetime
        prof_changed = old_apt["professional_id"] != apt.professional_id

        if apt.check_collisions and (date_changed or prof_changed):
            collision_response = await check_collisions(
                apt.professional_id,
                apt.appointment_datetime.isoformat(),
                60,
                id,  # Excluir este mismo turno de la búsqueda de colisiones
                tenant_id=tenant_id,
            )
            if collision_response["has_collisions"]:
                raise HTTPException(
                    status_code=409,
                    detail="Hay colisiones de horario en la nueva fecha/profesional",
                )

        # 3. Actualizar en Base de Datos (solo si pertenece al tenant)
        await db.pool.execute(
            """
            UPDATE appointments SET
                patient_id = $1,
                professional_id = $2,
                appointment_datetime = $3,
                appointment_type = $4,
                notes = $5,
                updated_at = NOW()
            WHERE id = $6 AND tenant_id = $7
        """,
            apt.patient_id,
            apt.professional_id,
            apt.appointment_datetime,
            apt.appointment_type,
            apt.notes,
            id,
            tenant_id,
        )
        # Audit log (TIER 3 cap.3 Phase B)
        try:
            from services.audit_log import log_appointment_mutation as _audit

            _old_dt = old_apt["appointment_datetime"] if old_apt else None
            _date_changed = (
                _old_dt is not None
                and apt.appointment_datetime is not None
                and _old_dt != apt.appointment_datetime
            )
            await _audit(
                pool=db.pool,
                tenant_id=tenant_id,
                appointment_id=id,
                action="rescheduled" if _date_changed else "status_changed",
                actor_type="staff_user",
                actor_id=getattr(user_data, "user_id", None),
                before_values={
                    "professional_id": old_apt["professional_id"] if old_apt else None,
                    "appointment_datetime": _old_dt.isoformat()
                    if _old_dt and hasattr(_old_dt, "isoformat")
                    else str(_old_dt),
                    "status": old_apt["status"] if old_apt else None,
                }
                if old_apt
                else None,
                after_values={
                    "patient_id": apt.patient_id,
                    "professional_id": apt.professional_id,
                    "appointment_datetime": apt.appointment_datetime.isoformat()
                    if hasattr(apt.appointment_datetime, "isoformat")
                    else str(apt.appointment_datetime),
                    "appointment_type": apt.appointment_type,
                    "notes": apt.notes,
                },
                source_channel="web_admin",
                reason=None,
            )
        except Exception as _audit_err:
            logger.warning(
                f"audit_log update_appointment failed (non-blocking): {_audit_err}"
            )

        # 4. Sincronizar con Google Calendar
        try:
            # Obtener datos completos para GCal
            appointment_data = await db.pool.fetchrow(
                """
                SELECT a.id, p.first_name, p.last_name, p.phone_number as patient_phone,
                       prof.first_name as professional_name, prof.google_calendar_id
                FROM appointments a
                JOIN patients p ON a.patient_id = p.id
                JOIN professionals prof ON a.professional_id = prof.id
                WHERE a.id = $1
            """,
                id,
            )

            if appointment_data:
                # Si cambió el profesional, hay que borrar el evento del calendario viejo y crear uno en el nuevo
                if prof_changed and old_apt["google_calendar_event_id"]:
                    old_prof_gcal = await db.pool.fetchval(
                        "SELECT google_calendar_id FROM professionals WHERE id = $1",
                        old_apt["professional_id"],
                    )
                    if old_prof_gcal:
                        gcal_service.delete_event(
                            calendar_id=old_prof_gcal,
                            event_id=old_apt["google_calendar_event_id"],
                        )

                    # Crear nuevo evento en el nuevo calendario
                    if appointment_data["google_calendar_id"]:
                        summary = f"Cita Dental: {appointment_data['first_name']} {appointment_data['last_name'] or ''} - {apt.appointment_type}"
                        new_gcal = gcal_service.create_event(
                            calendar_id=appointment_data["google_calendar_id"],
                            summary=summary,
                            start_time=apt.appointment_datetime.isoformat(),
                            end_time=(
                                apt.appointment_datetime + timedelta(minutes=60)
                            ).isoformat(),
                            description=f"Paciente: {appointment_data['first_name']}\nTel: {appointment_data['patient_phone']}\nNotas: {apt.notes or ''}",
                        )
                        if new_gcal:
                            await db.pool.execute(
                                "UPDATE appointments SET google_calendar_event_id = $1 WHERE id = $2",
                                new_gcal["id"],
                                id,
                            )

                # Si no cambió el profesional pero sí la fecha u otros datos, intentar actualizar el evento existente
                elif (
                    old_apt["google_calendar_event_id"]
                    and appointment_data["google_calendar_id"]
                ):
                    # Por ahora el gcal_service solo tiene create y delete, así que borramos y creamos
                    # TODO: Implementar update_event en gcal_service para mayor eficiencia
                    gcal_service.delete_event(
                        calendar_id=appointment_data["google_calendar_id"],
                        event_id=old_apt["google_calendar_event_id"],
                    )
                    summary = f"Cita Dental: {appointment_data['first_name']} {appointment_data['last_name'] or ''} - {apt.appointment_type}"
                    new_gcal = gcal_service.create_event(
                        calendar_id=appointment_data["google_calendar_id"],
                        summary=summary,
                        start_time=apt.appointment_datetime.isoformat(),
                        end_time=(
                            apt.appointment_datetime + timedelta(minutes=60)
                        ).isoformat(),
                        description=f"Paciente: {appointment_data['first_name']}\nTel: {appointment_data['patient_phone']}\nNotas: {apt.notes or ''}",
                    )
                    if new_gcal:
                        await db.pool.execute(
                            "UPDATE appointments SET google_calendar_event_id = $1 WHERE id = $2",
                            new_gcal["id"],
                            id,
                        )

        except Exception as ge:
            logger.error(f"Error syncing GCal on update: {ge}")

        # 5. Emitir evento Socket.IO
        full_data = await db.pool.fetchrow(
            """
            SELECT a.id, a.patient_id, a.professional_id, a.appointment_datetime, 
                   a.appointment_type, a.status, a.urgency_level,
                   (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name, 
                   p.phone_number as patient_phone, prof.first_name as professional_name
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            JOIN professionals prof ON a.professional_id = prof.id
            WHERE a.id = $1
        """,
            id,
        )
        if full_data:
            full_data_dict = dict(full_data)
            if full_data_dict.get("appointment_datetime"):
                full_data_dict["appointment_datetime"] = full_data_dict[
                    "appointment_datetime"
                ].isoformat()

            try:
                await emit_appointment_event(
                    "APPOINTMENT_UPDATED", full_data_dict, request
                )
            except Exception as emit_err:
                logger.warning(
                    f"Socket emit failed for appointment update {id}: {emit_err}"
                )

        return {"status": "updated", "id": id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating appointment: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.post(
    "/appointments/{id}/approve-payment",
    dependencies=[Depends(verify_admin_token)],
    tags=["Turnos"],
    summary="Aprobar pago manualmente (cuando la verificación automática falló)",
)
async def approve_payment_manually(
    id: str,
    tenant_id: int = Depends(get_resolved_tenant_id),
    user_data=Depends(verify_admin_token),
):
    """Aprueba manualmente un pago que la IA no pudo verificar. Actualiza status y payment_receipt_data."""
    apt = await db.pool.fetchrow(
        "SELECT id, payment_receipt_data, status, payment_status FROM appointments WHERE id = $1 AND tenant_id = $2",
        id,
        tenant_id,
    )
    if not apt:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    # Update receipt data to mark as manually verified
    receipt = apt.get("payment_receipt_data")
    if isinstance(receipt, str):
        receipt = json.loads(receipt) if receipt else {}
    if not isinstance(receipt, dict):
        receipt = {}

    receipt["status"] = "verified_manual"
    receipt["manually_approved_by"] = user_data.email or user_data.role or "admin"
    receipt["manually_approved_at"] = datetime.now(timezone.utc).isoformat()
    receipt.pop("failure_reason", None)

    await db.pool.execute(
        """UPDATE appointments SET
            status = 'confirmed',
            payment_status = 'paid',
            payment_receipt_data = $1::jsonb,
            updated_at = NOW()
        WHERE id = $2 AND tenant_id = $3""",
        json.dumps(receipt),
        id,
        tenant_id,
    )
    logger.info(
        f"Payment manually approved: apt={id} by={receipt.get('manually_approved_by')}"
    )
    return {"status": "approved", "appointment_id": id}


@router.put(
    "/appointments/{id}/billing",
    tags=["Turnos"],
    summary="Actualizar facturación de un turno",
)
async def update_appointment_billing(
    id: str,
    data: Dict[str, Any],
    request: Request,
    background_tasks: BackgroundTasks,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualizar datos de facturación de un turno (monto, cuotas, notas, estado de pago)."""
    # 1. Obtener datos actuales del turno, paciente y clínica
    current = await db.pool.fetchrow(
        """
        SELECT a.payment_status as current_payment_status,
               p.email as patient_email,
               t.country_code, t.clinic_name, t.address, t.bot_phone_number,
               a.billing_amount, a.appointment_datetime, a.appointment_type,
               p.first_name, p.last_name, prof.first_name as professional_name
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        JOIN tenants t ON a.tenant_id = t.id
        LEFT JOIN professionals prof ON a.professional_id = prof.id
        WHERE a.id = $1 AND a.tenant_id = $2
        """,
        id,
        tenant_id,
    )
    if not current:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    updates = []
    params = []
    new_payment_status = None
    if "billing_amount" in data:
        val = data.get("billing_amount")
        params.append(float(val) if val is not None and val != "" else None)
        updates.append(f"billing_amount = ${len(params)}")
    if "billing_installments" in data:
        val = data.get("billing_installments")
        params.append(int(val) if val is not None and val != "" else None)
        updates.append(f"billing_installments = ${len(params)}")
    if "billing_notes" in data:
        params.append(data.get("billing_notes") or None)
        updates.append(f"billing_notes = ${len(params)}")
    if "payment_status" in data:
        ps = data.get("payment_status") or "pending"
        if ps not in ("pending", "partial", "paid"):
            ps = "pending"
        params.append(ps)
        updates.append(f"payment_status = ${len(params)}")
        new_payment_status = ps
    if not updates:
        return {"status": "no_changes"}
    updates.append("updated_at = NOW()")
    params.append(id)
    params.append(tenant_id)
    query = f"UPDATE appointments SET {', '.join(updates)} WHERE id = ${len(params) - 1} AND tenant_id = ${len(params)}"
    result = await db.pool.execute(query, *params)
    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    # Audit log (TIER 3 cap.3 Phase B)
    try:
        from services.audit_log import log_appointment_mutation as _audit

        _before = {
            "billing_amount": float(current["billing_amount"])
            if current and current.get("billing_amount") is not None
            else None,
            "payment_status": current["current_payment_status"] if current else None,
        }
        _after = {
            k: data.get(k)
            for k in (
                "billing_amount",
                "billing_installments",
                "billing_notes",
                "payment_status",
            )
            if k in data
        }
        await _audit(
            pool=db.pool,
            tenant_id=tenant_id,
            appointment_id=id,
            action="payment_updated",
            actor_type="staff_user",
            actor_id=getattr(user_data, "user_id", None),
            before_values=_before,
            after_values=_after,
            source_channel="web_admin",
            reason=None,
        )
    except Exception as _audit_err:
        logger.warning(
            f"audit_log update_appointment_billing failed (non-blocking): {_audit_err}"
        )

    # Emit update event
    full_data = await db.pool.fetchrow(
        """
        SELECT a.id, a.patient_id, a.professional_id, a.appointment_datetime,
               a.appointment_type, a.status, a.urgency_level, a.billing_amount,
               a.billing_installments, a.billing_notes, a.payment_status,
               (p.first_name || ' ' || COALESCE(p.last_name, '')) as patient_name,
               p.phone_number as patient_phone, prof.first_name as professional_name
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        JOIN professionals prof ON a.professional_id = prof.id
        WHERE a.id = $1
    """,
        id,
    )
    if full_data:
        full_data_dict = dict(full_data)
        if full_data_dict.get("appointment_datetime"):
            full_data_dict["appointment_datetime"] = full_data_dict[
                "appointment_datetime"
            ].isoformat()
        try:
            await emit_appointment_event("APPOINTMENT_UPDATED", full_data_dict, request)
        except Exception as emit_err:
            logger.warning(f"Socket emit failed for billing update {id}: {emit_err}")

    # 2. Enviar email de confirmación si se pagó (y no estaba pagado antes)
    if new_payment_status == "paid" and current.get("current_payment_status") != "paid":
        patient_email = current.get("patient_email")
        if patient_email:
            # Preparar datos para el email
            country_code = current.get("country_code") or "AR"
            clinic_name = current.get("clinic_name") or ""
            clinic_address = current.get("address") or ""
            clinic_phone = current.get("bot_phone_number") or ""
            patient_name = f"{current.get('first_name', '')} {current.get('last_name', '')}".strip()
            amount = str(current.get("billing_amount", ""))
            appointment_datetime = current.get("appointment_datetime")
            appointment_date = (
                appointment_datetime.strftime("%d/%m/%Y")
                if appointment_datetime
                else ""
            )
            appointment_time = (
                appointment_datetime.strftime("%H:%M") if appointment_datetime else ""
            )
            treatment = current.get("appointment_type") or "Tratamiento"
            professional_name = current.get("professional_name") or ""

            # Enviar email en background
            def send_payment_email_task():
                email_service.send_payment_email(
                    to_email=patient_email,
                    country_code=country_code,
                    patient_name=patient_name,
                    clinic_name=clinic_name,
                    appointment_date=appointment_date,
                    appointment_time=appointment_time,
                    treatment=treatment,
                    amount=amount,
                    payment_method="Pago registrado",
                    clinic_address=clinic_address,
                    clinic_phone=clinic_phone,
                )

            background_tasks.add_task(send_payment_email_task)
            logger.info(f"📧 Email de pago programado para {patient_email}")

    return {"status": "updated", "id": id}


@router.delete(
    "/appointments/{id}",
    tags=["Turnos"],
    summary="Eliminar físicamente un turno de la agenda",
)
async def delete_appointment(
    id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Eliminar turno físicamente de la base de datos y de GCal. Aislado por tenant_id (Regla de Oro)."""
    try:
        # 1. Obtener datos antes de borrar — verificar ownership
        apt = await db.pool.fetchrow(
            """
            SELECT google_calendar_event_id, professional_id
            FROM appointments WHERE id = $1 AND tenant_id = $2
        """,
            id,
            tenant_id,
        )

        if not apt:
            raise HTTPException(status_code=404, detail="Turno no encontrado")

        # 2. Borrar de Google Calendar si existe
        if apt["google_calendar_event_id"]:
            try:
                google_calendar_id = await db.pool.fetchval(
                    "SELECT google_calendar_id FROM professionals WHERE id = $1",
                    apt["professional_id"],
                )
                if google_calendar_id:
                    gcal_service.delete_event(
                        calendar_id=google_calendar_id,
                        event_id=apt["google_calendar_event_id"],
                    )
            except Exception as ge:
                logger.error(f"Error borrando de GCal: {ge}")

        # Audit log (TIER 3 cap.3 Phase B) — MUST run BEFORE the DELETE so the FK to
        # appointments(id) still resolves at INSERT time. After the DELETE, the FK
        # constraint ON DELETE SET NULL will null out appointment_id in the audit row,
        # preserving the historical record without a dangling reference.
        try:
            from services.audit_log import log_appointment_mutation as _audit

            await _audit(
                pool=db.pool,
                tenant_id=tenant_id,
                appointment_id=id,
                action="cancelled",
                actor_type="staff_user",
                actor_id=getattr(user_data, "user_id", None),
                before_values={
                    "professional_id": apt["professional_id"] if apt else None,
                    "google_calendar_event_id": apt["google_calendar_event_id"]
                    if apt
                    else None,
                },
                after_values=None,
                source_channel="web_admin",
                reason="Hard-deleted from admin UI",
            )
        except Exception as _audit_err:
            logger.warning(
                f"audit_log delete_appointment failed (non-blocking): {_audit_err}"
            )

        # 3. Borrar de la base de datos — con doble verificación de tenant_id
        await db.pool.execute(
            "DELETE FROM appointments WHERE id = $1 AND tenant_id = $2", id, tenant_id
        )

        # 4. Notificar a la UI
        await emit_appointment_event("APPOINTMENT_DELETED", id, request)

        return {"status": "deleted", "id": id}
    except Exception as e:
        logger.error(f"Error deleting appointment: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ==================== ENDPOINTS SLOTS DISPONIBLES ====================


class NextSlotsResponse(BaseModel):
    slot_start: str
    slot_end: str
    duration_minutes: int
    professional_id: int
    professional_name: str


@router.get(
    "/appointments/next-slots",
    response_model=List[NextSlotsResponse],
    dependencies=[Depends(verify_admin_token)],
    tags=["Turnos"],
    summary="Consultar próximos huecos disponibles para turnos",
)
async def get_next_available_slots(
    days_ahead: int = 3,
    slot_duration_minutes: int = 20,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Próximos huecos disponibles para urgencias. Aislado por tenant_id (Regla de Oro).
    """
    professionals = await db.pool.fetch(
        """
        SELECT id, first_name, last_name 
        FROM professionals 
        WHERE tenant_id = $1 AND is_active = true
    """,
        tenant_id,
    )
    if not professionals:
        return []
    available_slots: List[Dict[str, Any]] = []
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for day_offset in range(days_ahead + 1):
        current_date = today + timedelta(days=day_offset)
        if current_date.weekday() >= 5:
            continue
        day_start = current_date.replace(hour=9, minute=0, second=0, microsecond=0)
        day_end = current_date.replace(hour=18, minute=0, second=0, microsecond=0)
        for prof in professionals:
            prof_id = prof["id"]
            prof_name = f"{prof['first_name']} {prof.get('last_name', '')}".strip()
            appointments = await db.pool.fetch(
                """
                SELECT appointment_datetime, duration_minutes
                FROM appointments
                WHERE tenant_id = $1 AND professional_id = $2
                AND DATE(appointment_datetime) = $3
                AND status NOT IN ('cancelled', 'no-show')
                ORDER BY appointment_datetime ASC
            """,
                tenant_id,
                prof_id,
                current_date.date(),
            )
            gcal_blocks = await db.pool.fetch(
                """
                SELECT start_datetime, end_datetime
                FROM google_calendar_blocks
                WHERE tenant_id = $1 AND (professional_id = $2 OR professional_id IS NULL)
                AND DATE(start_datetime) = $3
                ORDER BY start_datetime ASC
            """,
                tenant_id,
                prof_id,
                current_date.date(),
            )

            # Crear lista de busy periods (turnos + bloques)
            busy_periods = []

            for apt in appointments:
                apt_start = apt["appointment_datetime"]
                apt_end = apt_start + timedelta(minutes=apt.get("duration_minutes", 60))
                busy_periods.append((apt_start, apt_end))

            for block in gcal_blocks:
                busy_periods.append((block["start_datetime"], block["end_datetime"]))

            # Ordenar por inicio
            busy_periods.sort(key=lambda x: x[0])

            # Encontrar gaps
            current_time = day_start

            for busy_start, busy_end in busy_periods:
                # Verificar gap antes del primer ocupado
                if current_time < busy_start:
                    gap_duration = int((busy_start - current_time).total_seconds() / 60)
                    if (
                        gap_duration >= slot_duration_minutes - 5
                    ):  # Allow 5 min tolerance
                        slot_start = current_time
                        slot_end = slot_start + timedelta(minutes=slot_duration_minutes)
                        available_slots.append(
                            {
                                "slot_start": slot_start.isoformat(),
                                "slot_end": slot_end.isoformat(),
                                "duration_minutes": slot_duration_minutes,
                                "professional_id": prof_id,
                                "professional_name": prof_name,
                            }
                        )

                current_time = max(current_time, busy_end)

            # Verificar gap después del último ocupado hasta fin del día
            if current_time < day_end:
                gap_duration = int((day_end - current_time).total_seconds() / 60)
                if gap_duration >= slot_duration_minutes - 5:
                    slot_start = current_time
                    slot_end = slot_start + timedelta(minutes=slot_duration_minutes)
                    available_slots.append(
                        {
                            "slot_start": slot_start.isoformat(),
                            "slot_end": slot_end.isoformat(),
                            "duration_minutes": slot_duration_minutes,
                            "professional_id": prof_id,
                            "professional_name": prof_name,
                        }
                    )

    # Ordenar por fecha y hora, tomar los primeros 5
    available_slots.sort(key=lambda x: x["slot_start"])
    return available_slots[:5]


# ==================== ENDPOINTS PROFESIONALES ====================


@router.get(
    "/professionals",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Listar profesionales activos y aprobados",
)
async def list_professionals(
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Lista profesionales aprobados y activos. CEO ve todos los de sus sedes; secretary/professional solo los de su clínica.
    Solo se incluyen profesionales cuyo usuario tiene status = 'active' (aprobados por el CEO).
    """
    # Solo listar profesionales dentales aprobados (u.role IN ('professional', 'ceo') y u.status = 'active').
    base_join = "FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active'"
    # CEO (varias sedes): listar profesionales de todas las sedes permitidas
    if len(allowed_ids) > 1:
        try:
            rows = await db.pool.fetch(
                f"SELECT p.id, p.first_name, p.last_name, p.specialty, p.is_active, p.tenant_id, p.is_priority_professional {base_join} WHERE p.tenant_id = ANY($1::int[])",
                allowed_ids,
            )
            return [dict(row) for row in rows]
        except Exception as e:
            err_str = str(e).lower()
            if (
                "last_name" in err_str
                or "tenant_id" in err_str
                or "is_priority_professional" in err_str
            ):
                try:
                    rows = await db.pool.fetch(
                        f"SELECT p.id, p.first_name, p.specialty, p.is_active, p.tenant_id {base_join} WHERE p.tenant_id = ANY($1::int[])",
                        allowed_ids,
                    )
                    return [
                        dict(r) | {"last_name": "", "is_priority_professional": False}
                        for r in rows
                    ]
                except Exception:
                    pass
            try:
                rows = await db.pool.fetch(
                    "SELECT p.id, p.first_name, p.last_name, p.specialty, p.is_active, p.is_priority_professional FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active' WHERE p.tenant_id = ANY($1::int[])",
                    allowed_ids,
                )
                return [dict(row) for row in rows]
            except Exception:
                pass
        try:
            rows = await db.pool.fetch(
                "SELECT p.id, p.first_name, p.last_name, p.specialty, p.is_active FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active' WHERE p.tenant_id = ANY($1::int[])",
                allowed_ids,
            )
            return [dict(row) | {"is_priority_professional": False} for row in rows]
        except Exception as e2:
            logger.warning(f"list_professionals CEO fallback failed: {e2}")
            return []

    tenant_id = resolved_tenant_id
    try:
        rows = await db.pool.fetch(
            f"SELECT p.id, p.first_name, p.last_name, p.specialty, p.is_active, p.is_priority_professional {base_join} WHERE p.tenant_id = $1",
            tenant_id,
        )
        return [dict(row) for row in rows]
    except Exception as e:
        logger.warning(f"list_professionals primary query failed: {e}")

    try:
        rows = await db.pool.fetch(
            "SELECT p.id, p.first_name, p.specialty, p.is_active FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active' WHERE p.tenant_id = $1",
            tenant_id,
        )
        return [
            dict(r) | {"last_name": "", "is_priority_professional": False} for r in rows
        ]
    except Exception as e:
        logger.warning(f"list_professionals fallback (no last_name) failed: {e}")

    try:
        rows = await db.pool.fetch(
            "SELECT p.id, p.first_name, p.last_name, p.specialty, p.is_active FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active' WHERE p.tenant_id = $1",
            tenant_id,
        )
        return [dict(row) | {"is_priority_professional": False} for row in rows]
    except Exception as e:
        logger.warning(f"list_professionals fallback (no tenant) failed: {e}")
    try:
        rows = await db.pool.fetch(
            "SELECT p.id, p.first_name, p.specialty, p.is_active FROM professionals p INNER JOIN users u ON p.user_id = u.id AND u.role IN ('professional', 'ceo') AND u.status = 'active' WHERE p.tenant_id = $1",
            tenant_id,
        )
        return [
            dict(r) | {"last_name": "", "is_priority_professional": False} for r in rows
        ]
    except Exception as e:
        logger.error(f"list_professionals all fallbacks failed: {e}", exc_info=True)
        return []


@router.get(
    "/professionals/by-user/{user_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Obtener sedes de un profesional por su user_id",
)
async def get_professionals_by_user(
    user_id: str,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Devuelve las filas de professionals asociadas a un usuario (por user_id).
    Usado por el modal de detalle al hacer clic en un miembro de Personal Activo.
    Solo se devuelven sedes a las que el requestor tiene acceso.
    """
    try:
        uid = uuid.UUID(user_id)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="user_id inválido")
    try:
        rows = await db.pool.fetch(
            """
            SELECT id, tenant_id, user_id, first_name, last_name, email, specialty,
                   is_active, working_hours, created_at, phone_number, registration_id, google_calendar_id, consultation_price, is_priority_professional
            FROM professionals
            WHERE user_id = $1 AND tenant_id = ANY($2::int[])
            ORDER BY tenant_id
            """,
            uid,
            allowed_ids,
        )
        return [dict(r) for r in rows]
    except Exception as e:
        err_str = str(e).lower()
        if (
            "working_hours" in err_str
            or "tenant_id" in err_str
            or "phone_number" in err_str
            or "registration_id" in err_str
            or "google_calendar_id" in err_str
            or "is_priority_professional" in err_str
        ):
            try:
                rows = await db.pool.fetch(
                    "SELECT id, tenant_id, user_id, first_name, last_name, email, specialty, is_active, working_hours, phone_number, registration_id FROM professionals WHERE user_id = $1 AND tenant_id = ANY($2::int[]) ORDER BY tenant_id",
                    uid,
                    allowed_ids,
                )
                return [dict(r) | {"is_priority_professional": False} for r in rows]
            except Exception:
                pass
        try:
            rows = await db.pool.fetch(
                "SELECT id, first_name, last_name, email, specialty, is_active FROM professionals WHERE user_id = $1",
                uid,
            )
            return [dict(r) | {"is_priority_professional": False} for r in rows]
        except Exception as e2:
            logger.warning(f"get_professionals_by_user failed: {e2}")
            return []


@router.get(
    "/professionals/{id}/analytics",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Obtener analíticas detalladas de un profesional",
)
async def get_professional_analytics(
    id: int,
    tenant_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Métricas de un solo profesional para un tenant y rango de fechas.
    Usado por el modal de datos del profesional (acordeón). Solo sedes permitidas.
    """
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta sede.")
    if not start_date or not end_date:
        today = datetime.now()
        start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:
        start = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
        end = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
    result = await analytics_service.get_professional_summary(id, start, end, tenant_id)
    if result is None:
        raise HTTPException(
            status_code=404, detail="Profesional no encontrado en esta sede."
        )
    return result


@router.post(
    "/professionals/{id}/conversation-insights",
    dependencies=[Depends(verify_admin_token)],
    tags=["Profesionales"],
    summary="Generar insights de conversaciones con IA (beta)",
)
async def get_professional_conversation_insights(
    id: int,
    tenant_id: int,
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Analiza conversaciones recientes de pacientes del profesional usando IA.
    Extrae sentimiento, temas comunes y calidad de atención. Beta/experimental.
    """
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta sede.")

    # Fetch recent patients for this professional
    patient_ids = await db.pool.fetch(
        """
        SELECT DISTINCT patient_id FROM appointments
        WHERE professional_id = $1 AND tenant_id = $2
        AND created_at > NOW() - INTERVAL '60 days'
        ORDER BY patient_id
        LIMIT 20
    """,
        id,
        tenant_id,
    )

    if not patient_ids:
        return {
            "insights": "No hay suficientes datos de pacientes para generar insights."
        }

    pids = [r["patient_id"] for r in patient_ids]

    # Get conversation snippets (first 200 chars of last messages)
    snippets = []
    for pid in pids[:10]:
        phone = await db.pool.fetchval(
            "SELECT phone_number FROM patients WHERE id = $1 AND tenant_id = $2",
            pid,
            tenant_id,
        )
        if not phone:
            continue
        messages = await db.pool.fetch(
            """
            SELECT role, LEFT(content, 200) as content FROM chat_messages
            WHERE from_number = $1 AND tenant_id = $2
            ORDER BY created_at DESC LIMIT 6
        """,
            phone,
            tenant_id,
        )
        if messages:
            conv_text = " | ".join(
                [f"{m['role']}: {m['content']}" for m in reversed(messages)]
            )
            snippets.append(conv_text)

    if not snippets:
        return {"insights": "No se encontraron conversaciones recientes para analizar."}

    # Call OpenAI for analysis
    try:
        import openai

        client = openai.AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        all_snippets = "\n---\n".join(snippets[:10])
        # Get configurable model for insights (from system_config)
        model_name = "gpt-4o-mini"
        try:
            model_row = await db.pool.fetchrow(
                "SELECT value FROM system_config WHERE key = 'MODEL_INSIGHTS' AND tenant_id = $1",
                tenant_id,
            )
            if model_row and model_row.get("value"):
                model_name = str(model_row["value"]).strip()
            else:
                # Fallback to main OPENAI_MODEL
                model_row = await db.pool.fetchrow(
                    "SELECT value FROM system_config WHERE key = 'OPENAI_MODEL' AND tenant_id = $1",
                    tenant_id,
                )
                if model_row and model_row.get("value"):
                    model_name = str(model_row["value"]).strip()
        except Exception:
            pass
        messages_payload = [
            {
                "role": "system",
                "content": "Sos un analista de experiencia del paciente en una clínica dental. Analizá los fragmentos de conversaciones y generá un resumen conciso en español con: 1) Sentimiento general (positivo/negativo/neutro), 2) Temas positivos recurrentes, 3) Quejas o preocupaciones frecuentes, 4) Puntuación de calidad de atención (1-10), 5) Recomendaciones breves. Respondé en formato estructurado, máximo 300 palabras.",
            },
            {
                "role": "user",
                "content": f"Fragmentos de {len(snippets)} conversaciones recientes:\n\n{all_snippets}",
            },
        ]
        # Newer OpenAI models (gpt-5, o-series) require `max_completion_tokens`
        # while older ones (gpt-4o, gpt-4o-mini) accept `max_tokens`. Try the
        # modern parameter first and fallback if the model rejects it.
        try:
            response = await client.chat.completions.create(
                model=model_name,
                messages=messages_payload,
                max_completion_tokens=500,
                temperature=0.3,
            )
        except Exception as param_err:
            err_str = str(param_err).lower()
            if (
                "max_completion_tokens" in err_str
                or "unsupported" in err_str
                or "unexpected keyword" in err_str
            ):
                response = await client.chat.completions.create(
                    model=model_name,
                    messages=messages_payload,
                    max_tokens=500,
                    temperature=0.3,
                )
            else:
                raise
        insights_text = response.choices[0].message.content

        # Track token usage
        usage = response.usage
        input_tokens = usage.prompt_tokens if usage else 0
        output_tokens = usage.completion_tokens if usage else 0
        total_tokens = input_tokens + output_tokens
        cost_usd = 0.0

        if total_tokens > 0:
            try:
                from dashboard.token_tracker import token_tracker, TokenUsage
                from datetime import timezone as tz

                if token_tracker:
                    cost_usd = float(
                        token_tracker.calculate_cost(
                            model_name, input_tokens, output_tokens
                        )
                    )
                    token_usage = TokenUsage(
                        conversation_id=f"insights-prof-{id}",
                        patient_phone="system-insights",
                        model=model_name,
                        input_tokens=input_tokens,
                        output_tokens=output_tokens,
                        total_tokens=total_tokens,
                        cost_usd=token_tracker.calculate_cost(
                            model_name, input_tokens, output_tokens
                        ),
                        timestamp=datetime.now(tz.utc),
                        tenant_id=tenant_id,
                    )
                    await token_tracker.track_usage(token_usage)
                # Update tenant totals
                await db.pool.execute(
                    "UPDATE tenants SET total_tokens_used = COALESCE(total_tokens_used, 0) + $1, total_tool_calls = COALESCE(total_tool_calls, 0) + 1 WHERE id = $2",
                    total_tokens,
                    tenant_id,
                )
                logger.info(
                    f"📊 Insights tokens tracked: {total_tokens} (in={input_tokens}, out={output_tokens}) cost=${cost_usd:.4f} tenant={tenant_id}"
                )
            except Exception as tk_err:
                logger.warning(f"⚠️ Insights token tracking error (non-fatal): {tk_err}")

        return {
            "insights": insights_text,
            "conversations_analyzed": len(snippets),
            "tokens_used": total_tokens,
            "cost_usd": round(cost_usd, 6),
        }
    except Exception as e:
        logger.error(f"Error generating conversation insights: {e}")
        return {
            "insights": f"Error al generar insights: {str(e)}",
            "conversations_analyzed": 0,
        }


# ==================== ENDPOINTS GOOGLE CALENDAR ====================


@router.post(
    "/calendar/connect-sovereign",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Conectar calendario de Google vía Auth0 (Soberano)",
)
async def connect_sovereign_calendar(
    payload: ConnectSovereignPayload,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    allowed_ids: List[int] = Depends(get_allowed_tenant_ids),
):
    """
    Guarda el token de Auth0 cifrado (Fernet) en credentials para la clínica
    y cambia calendar_provider a 'google'. Preparación para integración Auth0.
    """
    tenant_id = (
        payload.tenant_id
        if (payload.tenant_id is not None and payload.tenant_id in allowed_ids)
        else resolved_tenant_id
    )
    if tenant_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta clínica.")
    encrypted = encrypt_value(payload.access_token)
    if not encrypted:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Cifrado no configurado. Definí CREDENTIALS_FERNET_KEY en el entorno.",
        )
    try:
        existing = await db.pool.fetchrow(
            """
            SELECT id FROM credentials
            WHERE tenant_id = $1 AND category = 'google_calendar' AND name = 'access_token'
            LIMIT 1
            """,
            tenant_id,
        )
        if existing:
            await db.pool.execute(
                """
                UPDATE credentials SET value = $1
                WHERE tenant_id = $2 AND category = 'google_calendar' AND name = 'access_token'
                """,
                encrypted,
                tenant_id,
            )
        else:
            await db.pool.execute(
                """
                INSERT INTO credentials (name, value, category, scope, tenant_id, description)
                VALUES ('access_token', $1, 'google_calendar', 'tenant', $2, 'Auth0/Google Calendar token (encrypted)')
                """,
                encrypted,
                tenant_id,
            )
        await db.pool.execute(
            """
            UPDATE tenants
            SET config = COALESCE(config, '{}')::jsonb || jsonb_build_object('calendar_provider', 'google'),
                updated_at = NOW()
            WHERE id = $1
            """,
            tenant_id,
        )
        logger.info(
            f"Calendar connect-sovereign: tenant_id={tenant_id}, calendar_provider=google"
        )
        return {
            "status": "connected",
            "tenant_id": tenant_id,
            "calendar_provider": "google",
        }
    except Exception as e:
        logger.error(f"connect-sovereign failed: {e}")
        raise HTTPException(
            status_code=500, detail="Error al guardar el token o actualizar la clínica."
        )


@router.get(
    "/calendar/blocks",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Listar bloqueos manuales o de GCalendar",
)
async def get_calendar_blocks(
    start_date: str,
    end_date: str,
    professional_id: Optional[int] = None,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Obtener bloques de calendario. Aislado por tenant_id (Regla de Oro)."""
    try:
        query = """
            SELECT id, google_event_id, title, description,
                   start_datetime, end_datetime, all_day,
                   professional_id, sync_status
            FROM google_calendar_blocks
            WHERE tenant_id = $1 AND start_datetime < $3 AND end_datetime > $2
        """
        params = [
            tenant_id,
            datetime.fromisoformat(start_date.replace("Z", "+00:00")),
            datetime.fromisoformat(end_date.replace("Z", "+00:00")),
        ]
        if professional_id:
            query += " AND professional_id = $4"
            params.append(professional_id)
        query += " ORDER BY start_datetime ASC"
        rows = await db.pool.fetch(query, *params)
        return [dict(row) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching calendar blocks: {e}")
        return []


@router.post(
    "/calendar/blocks",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Crear un nuevo bloqueo en el calendario",
)
async def create_calendar_block(
    block: GCalendarBlockCreate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Crear un bloque de calendario. Aislado por tenant_id (Regla de Oro)."""
    try:
        new_id = str(uuid.uuid4())
        await db.pool.execute(
            """
            INSERT INTO google_calendar_blocks (
                id, tenant_id, google_event_id, title, description,
                start_datetime, end_datetime, all_day, professional_id, sync_status, created_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'synced', NOW())
        """,
            new_id,
            tenant_id,
            block.google_event_id,
            block.title,
            block.description,
            block.start_datetime,
            block.end_datetime,
            block.all_day,
            block.professional_id,
        )
        return {"id": new_id, "status": "created"}
    except Exception as e:
        return {"id": str(uuid.uuid4()), "status": "simulated", "message": str(e)}


@router.delete(
    "/calendar/blocks/{block_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Eliminar un bloqueo del calendario",
)
async def delete_calendar_block(
    block_id: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Eliminar un bloque de calendario. Aislado por tenant_id (Regla de Oro)."""
    await db.pool.execute(
        "DELETE FROM google_calendar_blocks WHERE id = $1 AND tenant_id = $2",
        block_id,
        tenant_id,
    )
    return {"status": "deleted"}


@router.post(
    "/sync/calendar",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Disparar sincronización manual de GCalendar (Alias)",
)
@router.post(
    "/calendar/sync",
    dependencies=[Depends(verify_admin_token)],
    tags=["Calendario"],
    summary="Disparar sincronización manual de GCalendar",
)
async def trigger_sync(tenant_id: int = Depends(get_resolved_tenant_id)):
    """
    Sincronización con Google Calendar para profesionales activos del tenant.
    Aislado por tenant_id (Regla de Oro).
    """
    try:
        professionals = await db.pool.fetch(
            """
            SELECT id, first_name, google_calendar_id 
            FROM professionals 
            WHERE tenant_id = $1 AND is_active = true AND google_calendar_id IS NOT NULL
        """,
            tenant_id,
        )

        if not professionals:
            return {
                "status": "warning",
                "message": "No hay profesionales con calendario configurado.",
            }

        appointment_google_ids = await db.pool.fetch(
            "SELECT google_calendar_event_id FROM appointments WHERE tenant_id = $1 AND google_calendar_event_id IS NOT NULL",
            tenant_id,
        )
        apt_ids_set = {
            row["google_calendar_event_id"] for row in appointment_google_ids
        }

        total_created = 0
        total_updated = 0
        total_processed = 0

        # Sync from yesterday to ensure we catch today's earlier events
        time_min = (
            (datetime.now(timezone.utc) - timedelta(days=1))
            .isoformat()
            .replace("+00:00", "Z")
        )
        time_max = (
            (datetime.now(timezone.utc) + timedelta(days=30))
            .isoformat()
            .replace("+00:00", "Z")
        )

        for prof in professionals:
            prof_id = prof["id"]
            cal_id = prof["google_calendar_id"]

            logger.info(
                f"🔄 Syncing GCal for {prof['first_name']} (ID: {prof_id}) on {cal_id}"
            )
            logger.info(f"   Time Range: {time_min} to {time_max}")

            events = gcal_service.list_events(
                calendar_id=cal_id, time_min=time_min, time_max=time_max
            )
            logger.info(f"   Found {len(events)} events in GCal.")
            total_processed += len(events)

            # Obtener IDs ya existentes para este profesional
            existing_blocks = await db.pool.fetch(
                "SELECT google_event_id FROM google_calendar_blocks WHERE professional_id = $1",
                prof_id,
            )
            existing_ids_set = {row["google_event_id"] for row in existing_blocks}

            for event in events:
                g_id = event["id"]
                summary = event.get("summary", "Sin Título")

                # Ignorar si es un turno ya registrado
                if g_id in apt_ids_set:
                    logger.info(
                        f"   Skipping event {g_id} ({summary}) - Already linked to appointment"
                    )
                    continue

                description = event.get("description", "")
                start = event["start"].get("dateTime") or event["start"].get("date")
                end = event["end"].get("dateTime") or event["end"].get("date")
                all_day = "date" in event["start"]

                # Parsing dates with safety
                try:
                    dt_start = datetime.fromisoformat(start.replace("Z", "+00:00"))
                    dt_end = datetime.fromisoformat(end.replace("Z", "+00:00"))
                except Exception as de:
                    logger.warning(
                        f"   Error parsing date {start}/{end} for event {g_id}: {de}"
                    )
                    continue

                if g_id in existing_ids_set:
                    await db.pool.execute(
                        """
                        UPDATE google_calendar_blocks SET
                            title = $1, description = $2, start_datetime = $3, end_datetime = $4,
                            all_day = $5, updated_at = NOW()
                        WHERE google_event_id = $6 AND professional_id = $7
                    """,
                        summary,
                        description,
                        dt_start,
                        dt_end,
                        all_day,
                        g_id,
                        prof_id,
                    )
                    total_updated += 1
                else:
                    await db.pool.execute(
                        """
                        INSERT INTO google_calendar_blocks (
                            tenant_id, google_event_id, title, description,
                            start_datetime, end_datetime, all_day, professional_id
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                    """,
                        tenant_id,
                        g_id,
                        summary,
                        description,
                        dt_start,
                        dt_end,
                        all_day,
                        prof_id,
                    )
                    total_created += 1

        # Registrar sync global en log
        await db.pool.execute(
            """
            INSERT INTO calendar_sync_log (
                tenant_id, sync_type, direction, events_processed, 
                events_created, events_updated, completed_at
            ) VALUES ($1, 'manual', 'inbound', $2, $3, $4, NOW())
        """,
            tenant_id,
            total_processed,
            total_created,
            total_updated,
        )

        return {
            "status": "success",
            "professionals_synced": len(professionals),
            "events_processed": total_processed,
            "created": total_created,
            "updated": total_updated,
            "message": f"Sincronización completada para {len(professionals)} profesionales.",
        }
    except Exception as e:
        logger.error(f"Error en trigger_sync: {e}")
        return {"status": "error", "message": f"Error en sincronización: {str(e)}"}


# --- Función Helper de Entorno (Legacy support) ---
async def sync_environment():
    """Crea la clínica por defecto si no existe (startup main.py)."""
    exists = await db.pool.fetchval("SELECT id FROM tenants LIMIT 1")
    if not exists:
        await db.pool.execute("""
            INSERT INTO tenants (clinic_name, bot_phone_number, config)
            VALUES ('Clínica Dental', '5491100000000', '{"calendar_provider": "local"}'::jsonb)
        """)


# ==================== SEGUROS / OBRAS SOCIALES ====================

VALID_INSURANCE_STATUSES = ("accepted", "restricted", "external_derivation", "rejected")


class TreatmentCoverage(BaseModel):
    """Structured coverage entry for a single treatment code under an insurance
    provider. Migration 034 replaces the flat `restrictions` array with a dict
    keyed by treatment_types.code where each value is a TreatmentCoverage."""

    covered: bool = True
    copay_percent: float = 0.0
    requires_pre_authorization: bool = False
    pre_auth_leadtime_days: int = 0
    waiting_period_days: int = 0
    max_annual_coverage: Optional[float] = None
    notes: str = ""


class InsuranceProviderCreate(BaseModel):
    provider_name: str
    status: str  # 'accepted','restricted','external_derivation','rejected'
    coverage_by_treatment: Optional[Dict[str, TreatmentCoverage]] = None
    is_prepaid: bool = False
    employee_discount_percent: Optional[float] = None
    default_copay_percent: Optional[float] = None
    external_target: Optional[str] = None
    requires_copay: bool = True
    copay_notes: Optional[str] = None
    ai_response_template: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True


class InsuranceProviderUpdate(BaseModel):
    provider_name: str
    status: str
    coverage_by_treatment: Optional[Dict[str, TreatmentCoverage]] = None
    is_prepaid: bool = False
    employee_discount_percent: Optional[float] = None
    default_copay_percent: Optional[float] = None
    external_target: Optional[str] = None
    requires_copay: bool = True
    copay_notes: Optional[str] = None
    ai_response_template: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True


def _validate_insurance_provider(data) -> None:
    """Valida campos de obra social / seguro (migration 034 shape)."""
    name = data.provider_name.strip() if data.provider_name else ""
    if not name:
        raise HTTPException(status_code=422, detail="provider_name es obligatorio")
    if len(name) > 100:
        raise HTTPException(
            status_code=422, detail="provider_name no puede superar 100 caracteres"
        )
    if data.status not in VALID_INSURANCE_STATUSES:
        raise HTTPException(
            status_code=422,
            detail=f"status inválido. Debe ser uno de: {', '.join(VALID_INSURANCE_STATUSES)}",
        )
    # Per-treatment coverage validation (migration 034)
    if data.coverage_by_treatment:
        for code, cov in data.coverage_by_treatment.items():
            if not code or not code.strip():
                raise HTTPException(
                    status_code=422,
                    detail="Los códigos de tratamiento no pueden estar vacíos",
                )
            if not (0 <= cov.copay_percent <= 100):
                raise HTTPException(
                    status_code=422,
                    detail=f"copay_percent para '{code}' debe estar entre 0 y 100",
                )
            if cov.pre_auth_leadtime_days < 0:
                raise HTTPException(
                    status_code=422,
                    detail=f"pre_auth_leadtime_days para '{code}' no puede ser negativo",
                )
            if cov.waiting_period_days < 0:
                raise HTTPException(
                    status_code=422,
                    detail=f"waiting_period_days para '{code}' no puede ser negativo",
                )
            if cov.max_annual_coverage is not None and cov.max_annual_coverage <= 0:
                raise HTTPException(
                    status_code=422,
                    detail=f"max_annual_coverage para '{code}' debe ser positivo",
                )
            if len(cov.notes) > 500:
                raise HTTPException(
                    status_code=422,
                    detail=f"notes para '{code}' no puede superar 500 caracteres",
                )
    if data.employee_discount_percent is not None and not (
        0 <= data.employee_discount_percent <= 100
    ):
        raise HTTPException(
            status_code=422,
            detail="employee_discount_percent debe estar entre 0 y 100",
        )
    if data.default_copay_percent is not None and not (
        0 <= data.default_copay_percent <= 100
    ):
        raise HTTPException(
            status_code=422,
            detail="default_copay_percent debe estar entre 0 y 100",
        )
    if data.status == "external_derivation" and not data.external_target:
        raise HTTPException(
            status_code=422,
            detail="external_target es obligatorio cuando status='external_derivation'",
        )
    if data.ai_response_template and len(data.ai_response_template) > 1000:
        raise HTTPException(
            status_code=422,
            detail="ai_response_template no puede superar 1000 caracteres",
        )
    if data.copay_notes and len(data.copay_notes) > 500:
        raise HTTPException(
            status_code=422, detail="copay_notes no puede superar 500 caracteres"
        )


def _serialize_coverage_for_db(coverage: Optional[Dict[str, TreatmentCoverage]]) -> str:
    """Serialize a coverage_by_treatment dict for asyncpg JSONB binding.
    Returns '{}' if coverage is None or empty so the column always holds a
    valid JSON object (DB default is the same)."""
    if not coverage:
        return "{}"
    return json.dumps({k: v.dict() for k, v in coverage.items()})


@router.get(
    "/insurance-providers",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Listar obras sociales / seguros de la clínica",
)
async def list_insurance_providers(tenant_id: int = Depends(get_resolved_tenant_id)):
    """Listar obras sociales configuradas. Aislado por tenant_id (Regla de Oro)."""
    rows = await db.pool.fetch(
        """
        SELECT id, provider_name, status, coverage_by_treatment, is_prepaid,
               employee_discount_percent, default_copay_percent, external_target,
               requires_copay, copay_notes, ai_response_template, sort_order, is_active,
               created_at, updated_at
        FROM tenant_insurance_providers
        WHERE tenant_id = $1
        ORDER BY sort_order, provider_name
        """,
        tenant_id,
    )
    # asyncpg may return JSONB as a string in some versions — defensively parse
    # so the API response always carries a dict, not a JSON-encoded string.
    result = []
    for row in rows:
        d = dict(row)
        if isinstance(d.get("coverage_by_treatment"), str):
            try:
                d["coverage_by_treatment"] = json.loads(d["coverage_by_treatment"])
            except (json.JSONDecodeError, TypeError):
                d["coverage_by_treatment"] = {}
        result.append(d)
    return result


@router.post(
    "/insurance-providers",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Crear una obra social / seguro",
)
async def create_insurance_provider(
    data: InsuranceProviderCreate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Crear obra social. Aislado por tenant_id (Regla de Oro)."""
    _validate_insurance_provider(data)
    provider_name = data.provider_name.strip()

    # Duplicate check (case-insensitive)
    existing = await db.pool.fetchval(
        "SELECT id FROM tenant_insurance_providers WHERE tenant_id = $1 AND provider_name ILIKE $2",
        tenant_id,
        provider_name,
    )
    if existing:
        raise HTTPException(
            status_code=409, detail="Ya existe una obra social con ese nombre"
        )

    coverage_json = _serialize_coverage_for_db(data.coverage_by_treatment)
    row = await db.pool.fetchrow(
        """
        INSERT INTO tenant_insurance_providers (
            tenant_id, provider_name, status, coverage_by_treatment, is_prepaid,
            employee_discount_percent, default_copay_percent, external_target,
            requires_copay, copay_notes, ai_response_template, sort_order, is_active,
            created_at, updated_at
        ) VALUES ($1, $2, $3, $4::jsonb, $5, $6, $7, $8, $9, $10, $11, $12, $13, NOW(), NOW())
        RETURNING id
        """,
        tenant_id,
        provider_name,
        data.status,
        coverage_json,
        data.is_prepaid,
        data.employee_discount_percent,
        data.default_copay_percent,
        data.external_target,
        data.requires_copay,
        data.copay_notes,
        data.ai_response_template,
        data.sort_order,
        data.is_active,
    )
    # Sync embedding (background, non-blocking). NOTE: upsert_insurance_embedding
    # does not currently exist in embedding_service — the except below swallows
    # the ImportError. When implemented, it should receive the coverage JSON so
    # the RAG index reflects per-treatment details.
    try:
        from services.embedding_service import upsert_insurance_embedding
        import asyncio

        asyncio.create_task(
            upsert_insurance_embedding(
                tenant_id,
                row["id"],
                data.provider_name,
                data.status,
                coverage_json,
                data.copay_notes or "",
                data.ai_response_template or "",
            )
        )
    except Exception:
        pass
    return {"status": "created", "id": row["id"]}


@router.put(
    "/insurance-providers/{provider_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Actualizar una obra social / seguro",
)
async def update_insurance_provider(
    provider_id: int,
    data: InsuranceProviderUpdate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualizar obra social. Aislado por tenant_id (Regla de Oro)."""
    _validate_insurance_provider(data)
    provider_name = data.provider_name.strip()

    coverage_json = _serialize_coverage_for_db(data.coverage_by_treatment)
    result = await db.pool.execute(
        """
        UPDATE tenant_insurance_providers SET
            provider_name = $1, status = $2, coverage_by_treatment = $3::jsonb,
            is_prepaid = $4, employee_discount_percent = $5, default_copay_percent = $6,
            external_target = $7, requires_copay = $8, copay_notes = $9,
            ai_response_template = $10, sort_order = $11, is_active = $12,
            updated_at = NOW()
        WHERE id = $13 AND tenant_id = $14
        """,
        provider_name,
        data.status,
        coverage_json,
        data.is_prepaid,
        data.employee_discount_percent,
        data.default_copay_percent,
        data.external_target,
        data.requires_copay,
        data.copay_notes,
        data.ai_response_template,
        data.sort_order,
        data.is_active,
        provider_id,
        tenant_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail="Obra social no encontrada")
    # Sync embedding (background, non-blocking). See note in create_insurance_provider.
    try:
        from services.embedding_service import upsert_insurance_embedding
        import asyncio

        asyncio.create_task(
            upsert_insurance_embedding(
                tenant_id,
                provider_id,
                data.provider_name,
                data.status,
                coverage_json,
                data.copay_notes or "",
                data.ai_response_template or "",
            )
        )
    except Exception:
        pass
    return {"status": "updated", "id": provider_id}


@router.delete(
    "/insurance-providers/{provider_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Eliminar una obra social / seguro",
)
async def delete_insurance_provider(
    provider_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Eliminar obra social. Aislado por tenant_id (Regla de Oro)."""
    result = await db.pool.execute(
        "DELETE FROM tenant_insurance_providers WHERE id = $1 AND tenant_id = $2",
        provider_id,
        tenant_id,
    )
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Obra social no encontrada")
    try:
        from services.embedding_service import delete_insurance_embedding
        import asyncio

        asyncio.create_task(delete_insurance_embedding(provider_id, tenant_id))
    except Exception:
        pass
    return {"status": "deleted", "id": provider_id}


@router.patch(
    "/insurance-providers/{provider_id}/toggle-active",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Activar / desactivar una obra social",
)
async def toggle_insurance_provider_active(
    provider_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Flip is_active de obra social. Aislado por tenant_id (Regla de Oro)."""
    row = await db.pool.fetchrow(
        "SELECT id, is_active FROM tenant_insurance_providers WHERE id = $1 AND tenant_id = $2",
        provider_id,
        tenant_id,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Obra social no encontrada")
    new_value = not row["is_active"]
    await db.pool.execute(
        "UPDATE tenant_insurance_providers SET is_active = $1, updated_at = NOW() WHERE id = $2 AND tenant_id = $3",
        new_value,
        provider_id,
        tenant_id,
    )
    return {"status": "updated", "id": provider_id, "is_active": new_value}


class InsuranceReorderItem(BaseModel):
    id: int
    sort_order: int


class InsuranceReorderBody(BaseModel):
    order: List[InsuranceReorderItem]


@router.put(
    "/insurance-providers/reorder",
    dependencies=[Depends(verify_admin_token)],
    tags=["Seguros"],
    summary="Reordenar obras sociales",
)
async def reorder_insurance_providers(
    body: InsuranceReorderBody,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Batch update sort_order de obras sociales. Aislado por tenant_id (Regla de Oro)."""
    for item in body.order:
        await db.pool.execute(
            "UPDATE tenant_insurance_providers SET sort_order = $1, updated_at = NOW() WHERE id = $2 AND tenant_id = $3",
            item.sort_order,
            item.id,
            tenant_id,
        )
    return {"status": "reordered", "count": len(body.order)}


# ==================== REGLAS DE DERIVACIÓN ====================

VALID_PATIENT_CONDITIONS = ("new_patient", "existing_patient", "any")
VALID_TARGET_TYPES = ("specific_professional", "priority_professional", "team")
MAX_DERIVATION_RULES = 20


class DerivationRuleCreate(BaseModel):
    rule_name: str
    patient_condition: str  # 'new_patient','existing_patient','any'
    treatment_categories: List[str]  # ['cirugia','implantes'] or ['*']
    target_type: str  # 'specific_professional','priority_professional','team'
    target_professional_id: Optional[int] = None
    priority_order: int = 0
    is_active: bool = True
    description: Optional[str] = None
    # Escalation fallback (migration 038)
    enable_escalation: bool = False
    fallback_professional_id: Optional[int] = None
    fallback_team_mode: bool = False
    max_wait_days_before_escalation: int = Field(default=7, ge=1, le=30)
    escalation_message_template: Optional[str] = None
    criteria_custom: Optional[dict] = None


class DerivationRuleUpdate(BaseModel):
    rule_name: str
    patient_condition: str
    treatment_categories: List[str]
    target_type: str
    target_professional_id: Optional[int] = None
    priority_order: int = 0
    is_active: bool = True
    description: Optional[str] = None
    # Escalation fallback (migration 038)
    enable_escalation: bool = False
    fallback_professional_id: Optional[int] = None
    fallback_team_mode: bool = False
    max_wait_days_before_escalation: int = Field(default=7, ge=1, le=30)
    escalation_message_template: Optional[str] = None
    criteria_custom: Optional[dict] = None


async def _validate_derivation_rule(data, tenant_id: int) -> None:
    """Valida campos de regla de derivación (migration 038 escalation-aware)."""
    if data.patient_condition not in VALID_PATIENT_CONDITIONS:
        raise HTTPException(
            status_code=422,
            detail=f"patient_condition inválido. Debe ser uno de: {', '.join(VALID_PATIENT_CONDITIONS)}",
        )
    if data.target_type not in VALID_TARGET_TYPES:
        raise HTTPException(
            status_code=422,
            detail=f"target_type inválido. Debe ser uno de: {', '.join(VALID_TARGET_TYPES)}",
        )
    if (
        data.target_type == "specific_professional"
        and data.target_professional_id is None
    ):
        raise HTTPException(
            status_code=422,
            detail="target_professional_id es obligatorio cuando target_type='specific_professional'",
        )
    if not data.treatment_categories:
        raise HTTPException(
            status_code=422, detail="treatment_categories no puede estar vacío"
        )
    # Verify professional belongs to tenant
    if data.target_professional_id is not None:
        prof_exists = await db.pool.fetchval(
            "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2",
            data.target_professional_id,
            tenant_id,
        )
        if not prof_exists:
            raise HTTPException(
                status_code=422,
                detail="El profesional especificado no pertenece a esta clínica",
            )

    # --- Escalation-specific validation (migration 038) ---
    # Conflict guard: can't specify both fallback_professional_id AND fallback_team_mode
    if data.fallback_professional_id is not None and data.fallback_team_mode:
        raise HTTPException(
            status_code=422,
            detail="No se puede especificar fallback_professional_id cuando fallback_team_mode es true",
        )
    # Tenant isolation for fallback professional
    if data.fallback_professional_id is not None:
        fb_exists = await db.pool.fetchval(
            "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2",
            data.fallback_professional_id,
            tenant_id,
        )
        if not fb_exists:
            raise HTTPException(
                status_code=422,
                detail="El profesional de fallback no pertenece a esta clínica",
            )
    # Implicit team mode when escalation is enabled but no fallback configured.
    # Mutate data in-place so the INSERT/UPDATE sees the corrected value.
    if (
        data.enable_escalation
        and data.fallback_professional_id is None
        and not data.fallback_team_mode
    ):
        data.fallback_team_mode = True


@router.get(
    "/derivation-rules",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Listar reglas de derivación",
)
async def list_derivation_rules(tenant_id: int = Depends(get_resolved_tenant_id)):
    """Listar reglas de derivación. Aislado por tenant_id (Regla de Oro)."""
    rows = await db.pool.fetch(
        """
        SELECT dr.id, dr.rule_name, dr.patient_condition, dr.treatment_categories,
               dr.target_type, dr.target_professional_id, dr.priority_order,
               dr.is_active, dr.description, dr.created_at, dr.updated_at,
               dr.enable_escalation, dr.fallback_professional_id,
               dr.fallback_team_mode, dr.max_wait_days_before_escalation,
               dr.escalation_message_template, dr.criteria_custom,
               p.first_name AS professional_first_name, p.last_name AS professional_last_name,
               fp.first_name AS fallback_first_name, fp.last_name AS fallback_last_name
        FROM professional_derivation_rules dr
        LEFT JOIN professionals p ON dr.target_professional_id = p.id
        LEFT JOIN professionals fp ON dr.fallback_professional_id = fp.id
        WHERE dr.tenant_id = $1
        ORDER BY dr.priority_order, dr.id
        """,
        tenant_id,
    )
    result = []
    for r in rows:
        item = dict(r)
        if item.get("professional_first_name") is not None:
            item["professional_name"] = (
                f"{item.pop('professional_first_name')} {item.pop('professional_last_name') or ''}".strip()
            )
        else:
            item.pop("professional_first_name", None)
            item.pop("professional_last_name", None)
            item["professional_name"] = None
        # Migration 038: same treatment for fallback professional name
        if item.get("fallback_first_name") is not None:
            item["fallback_professional_name"] = (
                f"{item.pop('fallback_first_name')} {item.pop('fallback_last_name') or ''}".strip()
            )
        else:
            item.pop("fallback_first_name", None)
            item.pop("fallback_last_name", None)
            item["fallback_professional_name"] = None
        # treatment_categories may be returned as a list or JSON string by asyncpg
        cats = item.get("treatment_categories")
        if isinstance(cats, str):
            try:
                item["treatment_categories"] = json.loads(cats)
            except Exception:
                pass
        # criteria_custom is JSONB — asyncpg may return it as string
        cc = item.get("criteria_custom")
        if isinstance(cc, str):
            try:
                item["criteria_custom"] = json.loads(cc)
            except Exception:
                item["criteria_custom"] = None
        result.append(item)
    return result


@router.post(
    "/derivation-rules",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Crear una regla de derivación",
)
async def create_derivation_rule(
    data: DerivationRuleCreate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Crear regla de derivación. Aislado por tenant_id (Regla de Oro)."""
    await _validate_derivation_rule(data, tenant_id)

    # Max rules check
    count = await db.pool.fetchval(
        "SELECT COUNT(*) FROM professional_derivation_rules WHERE tenant_id = $1",
        tenant_id,
    )
    if count >= MAX_DERIVATION_RULES:
        raise HTTPException(
            status_code=422,
            detail=f"Se alcanzó el límite máximo de {MAX_DERIVATION_RULES} reglas de derivación por clínica",
        )

    row = await db.pool.fetchrow(
        """
        INSERT INTO professional_derivation_rules (
            tenant_id, rule_name, patient_condition, treatment_categories,
            target_type, target_professional_id, priority_order, is_active,
            description,
            enable_escalation, fallback_professional_id, fallback_team_mode,
            max_wait_days_before_escalation, escalation_message_template,
            criteria_custom,
            created_at, updated_at
        ) VALUES ($1, $2, $3, $4::text[], $5, $6, $7, $8, $9,
                  $10, $11, $12, $13, $14, $15::jsonb,
                  NOW(), NOW())
        RETURNING id
        """,
        tenant_id,
        data.rule_name,
        data.patient_condition,
        list(data.treatment_categories or []),
        data.target_type,
        data.target_professional_id,
        data.priority_order,
        data.is_active,
        data.description,
        data.enable_escalation,
        data.fallback_professional_id,
        data.fallback_team_mode,
        data.max_wait_days_before_escalation,
        data.escalation_message_template,
        json.dumps(data.criteria_custom) if data.criteria_custom is not None else None,
    )
    # Sync embedding (background, non-blocking)
    try:
        from services.embedding_service import upsert_derivation_embedding
        import asyncio

        # Need to resolve professional name
        prof_name = ""
        if data.target_professional_id:
            prof_row = await db.pool.fetchrow(
                "SELECT first_name FROM professionals WHERE id = $1 AND tenant_id = $2",
                data.target_professional_id,
                tenant_id,
            )
            prof_name = prof_row["first_name"] if prof_row else ""
        asyncio.create_task(
            upsert_derivation_embedding(
                tenant_id,
                row["id"],
                data.rule_name,
                data.patient_condition,
                data.treatment_categories,
                data.target_type,
                prof_name,
            )
        )
    except Exception:
        pass
    return {"status": "created", "id": row["id"]}


@router.put(
    "/derivation-rules/{rule_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Actualizar una regla de derivación",
)
async def update_derivation_rule(
    rule_id: int,
    data: DerivationRuleUpdate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualizar regla de derivación. Aislado por tenant_id (Regla de Oro)."""
    await _validate_derivation_rule(data, tenant_id)

    result = await db.pool.execute(
        """
        UPDATE professional_derivation_rules SET
            rule_name = $1, patient_condition = $2, treatment_categories = $3::text[],
            target_type = $4, target_professional_id = $5, priority_order = $6,
            is_active = $7, description = $8,
            enable_escalation = $9, fallback_professional_id = $10,
            fallback_team_mode = $11, max_wait_days_before_escalation = $12,
            escalation_message_template = $13, criteria_custom = $14::jsonb,
            updated_at = NOW()
        WHERE id = $15 AND tenant_id = $16
        """,
        data.rule_name,
        data.patient_condition,
        list(data.treatment_categories or []),
        data.target_type,
        data.target_professional_id,
        data.priority_order,
        data.is_active,
        data.description,
        data.enable_escalation,
        data.fallback_professional_id,
        data.fallback_team_mode,
        data.max_wait_days_before_escalation,
        data.escalation_message_template,
        json.dumps(data.criteria_custom) if data.criteria_custom is not None else None,
        rule_id,
        tenant_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail="Regla de derivación no encontrada")
    # Sync embedding (background, non-blocking)
    try:
        from services.embedding_service import upsert_derivation_embedding
        import asyncio

        # Need to resolve professional name
        prof_name = ""
        if data.target_professional_id:
            prof_row = await db.pool.fetchrow(
                "SELECT first_name FROM professionals WHERE id = $1 AND tenant_id = $2",
                data.target_professional_id,
                tenant_id,
            )
            prof_name = prof_row["first_name"] if prof_row else ""
        asyncio.create_task(
            upsert_derivation_embedding(
                tenant_id,
                rule_id,
                data.rule_name,
                data.patient_condition,
                data.treatment_categories,
                data.target_type,
                prof_name,
            )
        )
    except Exception:
        pass
    return {"status": "updated", "id": rule_id}


@router.delete(
    "/derivation-rules/{rule_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Eliminar una regla de derivación",
)
async def delete_derivation_rule(
    rule_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Eliminar regla de derivación. Aislado por tenant_id (Regla de Oro)."""
    result = await db.pool.execute(
        "DELETE FROM professional_derivation_rules WHERE id = $1 AND tenant_id = $2",
        rule_id,
        tenant_id,
    )
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Regla de derivación no encontrada")
    try:
        from services.embedding_service import delete_derivation_embedding
        import asyncio

        asyncio.create_task(delete_derivation_embedding(rule_id, tenant_id))
    except Exception:
        pass
    return {"status": "deleted", "id": rule_id}


@router.patch(
    "/derivation-rules/{rule_id}/toggle-active",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Activar / desactivar una regla de derivación",
)
async def toggle_derivation_rule_active(
    rule_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Flip is_active de regla de derivación. Aislado por tenant_id (Regla de Oro)."""
    row = await db.pool.fetchrow(
        "SELECT id, is_active FROM professional_derivation_rules WHERE id = $1 AND tenant_id = $2",
        rule_id,
        tenant_id,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Regla de derivación no encontrada")
    new_value = not row["is_active"]
    await db.pool.execute(
        "UPDATE professional_derivation_rules SET is_active = $1, updated_at = NOW() WHERE id = $2 AND tenant_id = $3",
        new_value,
        rule_id,
        tenant_id,
    )
    return {"status": "updated", "id": rule_id, "is_active": new_value}


class DerivationReorderItem(BaseModel):
    id: int
    priority_order: int


class DerivationReorderBody(BaseModel):
    order: List[DerivationReorderItem]


@router.put(
    "/derivation-rules/reorder",
    dependencies=[Depends(verify_admin_token)],
    tags=["Derivación"],
    summary="Reordenar reglas de derivación",
)
async def reorder_derivation_rules(
    body: DerivationReorderBody,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Batch update priority_order de reglas de derivación. Aislado por tenant_id (Regla de Oro)."""
    for item in body.order:
        await db.pool.execute(
            "UPDATE professional_derivation_rules SET priority_order = $1, updated_at = NOW() WHERE id = $2 AND tenant_id = $3",
            item.priority_order,
            item.id,
            tenant_id,
        )
    return {"status": "reordered", "count": len(body.order)}


@router.get(
    "/patients/{patient_id}/attachments-summary",
    dependencies=[Depends(verify_admin_token)],
    tags=["Pacientes"],
    summary="Obtener el resumen más reciente de adjuntos del paciente",
)
async def get_attachment_summary(
    patient_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Get the most recent attachment summary for a patient."""
    # Verify patient belongs to tenant
    patient = await db.pool.fetchrow(
        "SELECT id, tenant_id FROM patients WHERE id = $1 AND tenant_id = $2",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    # Get latest summary
    summary = await db.pool.fetchrow(
        """
        SELECT summary_text, attachments_count, attachments_types, created_at
        FROM clinical_record_summaries
        WHERE tenant_id = $1 AND patient_id = $2
        ORDER BY created_at DESC
        LIMIT 1
    """,
        tenant_id,
        patient_id,
    )

    if not summary:
        return {
            "summary_text": None,
            "attachments_count": 0,
            "attachments_types": [],
            "created_at": None,
        }

    return {
        "summary_text": summary["summary_text"],
        "attachments_count": summary["attachments_count"],
        "attachments_types": summary["attachments_types"] or [],
        "created_at": summary["created_at"].isoformat()
        if summary["created_at"]
        else None,
    }


# ==================== ENDPOINTS TRATAMIENTOS ====================


class TreatmentTypeUpdate(BaseModel):
    name: str
    description: Optional[str] = None
    default_duration_minutes: int
    min_duration_minutes: int
    max_duration_minutes: int
    complexity_level: str
    category: str
    requires_multiple_sessions: bool = False
    session_gap_days: int = 0
    is_active: bool = True
    is_available_for_booking: bool = True
    internal_notes: Optional[str] = None
    base_price: Optional[float] = 0
    priority: Optional[str] = "medium"
    # Migration 036: structured pre/post instructions. See PreInstructions /
    # PostInstructions Pydantic models near the top of this file.
    pre_instructions: Optional[Union[PreInstructions, dict, str]] = None
    post_instructions: Optional[Union[PostInstructions, list, dict, str]] = None
    followup_template: Optional[Any] = None
    confirm_unusual_price: bool = False


async def _validate_treatment_price_scale(
    base_price: Optional[float], confirm_unusual_price: bool, tenant_id: int
) -> None:
    """Valida que base_price no tenga escala incorrecta respecto al precio de consulta.

    Si base_price > consultation_price * 100 y confirm_unusual_price es False,
    lanza HTTP 422 para que el frontend muestre confirmación al operador.

    Si consultation_price es NULL o 0, la validación se omite silenciosamente
    (no hay referencia con qué comparar).
    """
    if not base_price or base_price <= 0:
        return
    row = await db.pool.fetchrow(
        "SELECT consultation_price FROM tenants WHERE id = $1", tenant_id
    )
    if not row:
        return
    consultation_price = row["consultation_price"]
    # Si no hay precio de consulta configurado, no podemos comparar — skip silencioso
    if consultation_price is None or consultation_price == 0:
        return
    threshold = float(consultation_price) * 100
    if float(base_price) > threshold and not confirm_unusual_price:
        raise HTTPException(
            status_code=422,
            detail=(
                "Precio inusualmente alto. ¿Estás seguro? "
                "Si es correcto, incluí confirm_unusual_price=true en el cuerpo."
            ),
        )


def _coerce_pre_instructions(value):
    """Normalize pre_instructions to a JSON-serializable dict (or None).

    Migration 036 stores pre_instructions as JSONB. This helper converts
    every accepted input shape into the canonical dict form before
    json.dumps + INSERT/UPDATE.
    - None → None
    - PreInstructions (Pydantic) → .dict()
    - dict → returned as-is (assumed already in the right shape)
    - str (legacy) → wrapped as {"general_notes": str}
    - anything else → None (defensive)
    """
    if value is None:
        return None
    if isinstance(value, PreInstructions):
        return value.dict(exclude_none=False)
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        return {"general_notes": value}
    return None


def _coerce_post_instructions(value):
    """Normalize post_instructions to dict, list, or None.

    Migration 036 keeps post_instructions JSONB but allows two shapes:
    - new structured PostInstructions dict (recovery protocol)
    - legacy timed-sequence list (preserved as-is)
    """
    if value is None:
        return None
    if isinstance(value, PostInstructions):
        return value.dict(exclude_none=False)
    if isinstance(value, list):
        return value  # legacy timed sequence — store as-is
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        # Defensive: try to parse JSON; on failure wrap as general_notes
        try:
            parsed = json.loads(value)
            return parsed
        except (json.JSONDecodeError, TypeError):
            return {"general_notes": value}
    return None


def _validate_treatment_instruction_fields(treatment) -> None:
    """Valida los campos de instrucciones de un tipo de tratamiento (migration 036).

    Acepta tanto la forma legacy (string para pre, lista timed para post) como
    la nueva forma estructurada (PreInstructions / PostInstructions dicts).
    """
    # pre_instructions: string ≤ 2000 chars OR dict with general_notes ≤ 2000
    pre = treatment.pre_instructions
    if pre is not None:
        if isinstance(pre, str):
            if len(pre) > 2000:
                raise HTTPException(
                    status_code=422,
                    detail="pre_instructions no puede superar 2000 caracteres",
                )
        elif isinstance(pre, PreInstructions):
            notes = pre.general_notes or ""
            if len(notes) > 2000:
                raise HTTPException(
                    status_code=422,
                    detail="pre_instructions.general_notes no puede superar 2000 caracteres",
                )
            if pre.fasting_hours is not None and pre.fasting_hours < 0:
                raise HTTPException(
                    status_code=422,
                    detail="fasting_hours no puede ser negativo",
                )
            if (
                pre.preparation_days_before is not None
                and pre.preparation_days_before < 0
            ):
                raise HTTPException(
                    status_code=422,
                    detail="preparation_days_before no puede ser negativo",
                )
        elif isinstance(pre, dict):
            notes = pre.get("general_notes") or ""
            if isinstance(notes, str) and len(notes) > 2000:
                raise HTTPException(
                    status_code=422,
                    detail="pre_instructions.general_notes no puede superar 2000 caracteres",
                )
        else:
            raise HTTPException(
                status_code=422,
                detail="pre_instructions debe ser un objeto o un string",
            )

    # post_instructions: PostInstructions dict OR legacy list of timed entries
    post = treatment.post_instructions
    if post is not None:
        if isinstance(post, PostInstructions):
            if (
                post.care_duration_days is not None
                and post.care_duration_days <= 0
            ):
                raise HTTPException(
                    status_code=422,
                    detail="care_duration_days debe ser mayor a 0",
                )
            if (
                post.sutures_removal_day is not None
                and post.sutures_removal_day < 0
            ):
                raise HTTPException(
                    status_code=422,
                    detail="sutures_removal_day no puede ser negativo",
                )
            esc = post.escalation_message or ""
            if len(esc) > 500:
                raise HTTPException(
                    status_code=422,
                    detail="escalation_message no puede superar 500 caracteres",
                )
        elif isinstance(post, list):
            # Legacy timed-sequence shape — preserved for backwards compat.
            valid_timings = (
                "before",
                "after",
                "day_of",
                "same_day",
                "immediate",
                "24h",
                "48h",
                "72h",
                "1w",
                "stitch_removal",
                "custom",
            )
            for item in post:
                if not isinstance(item, dict):
                    raise HTTPException(
                        status_code=422,
                        detail="Cada elemento de post_instructions debe ser un objeto",
                    )
                timing = item.get("timing")
                if timing is not None and timing not in valid_timings:
                    raise HTTPException(
                        status_code=422,
                        detail=(
                            f"Valor de timing inválido: '{timing}'. Debe ser uno de: "
                            f"{', '.join(valid_timings)}"
                        ),
                    )
        elif isinstance(post, dict):
            # Raw dict (no Pydantic coercion) — accept without strict
            # validation to keep the API flexible for future keys.
            cd = post.get("care_duration_days")
            if cd is not None and isinstance(cd, (int, float)) and cd <= 0:
                raise HTTPException(
                    status_code=422,
                    detail="care_duration_days debe ser mayor a 0",
                )
        else:
            raise HTTPException(
                status_code=422,
                detail="post_instructions debe ser un objeto, lista o string",
            )
    if treatment.followup_template is not None:
        if not isinstance(treatment.followup_template, list):
            raise HTTPException(
                status_code=422,
                detail="followup_template debe ser una lista de objetos",
            )
        for item in treatment.followup_template:
            if not isinstance(item, dict):
                raise HTTPException(
                    status_code=422,
                    detail="Cada elemento de followup_template debe ser un objeto",
                )
            delay = item.get("delay_hours")
            if delay is not None:
                if not isinstance(delay, (int, float)) or not (1 <= delay <= 720):
                    raise HTTPException(
                        status_code=422,
                        detail=f"delay_hours debe ser un número entre 1 y 720, recibido: {delay}",
                    )
            if "message_template" not in item:
                raise HTTPException(
                    status_code=422,
                    detail="Cada elemento de followup_template debe incluir 'message_template'",
                )


@router.get(
    "/treatment-types",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Listar tipos de tratamientos ofrecidos",
    response_model=List[TreatmentTypeResponse],
)
async def list_treatment_types(tenant_id: int = Depends(get_resolved_tenant_id)):
    """Listar tipos de tratamiento de la clínica. Aislado por tenant_id (Regla de Oro)."""
    rows = await db.pool.fetch(
        """
        SELECT id, code, name, description, default_duration_minutes,
               min_duration_minutes, max_duration_minutes, complexity_level,
               category, requires_multiple_sessions, session_gap_days,
               is_active, is_available_for_booking, internal_notes, base_price, priority,
               pre_instructions, post_instructions, followup_template
        FROM treatment_types
        WHERE tenant_id = $1
        ORDER BY category, name
    """,
        tenant_id,
    )
    from decimal import Decimal as _Dec

    treatments = [
        {k: float(v) if isinstance(v, _Dec) else v for k, v in dict(row).items()}
        for row in rows
    ]
    # Batch-fetch professional assignments
    if treatments:
        tt_ids = [t["id"] for t in treatments]
        ttp_rows = await db.pool.fetch(
            "SELECT treatment_type_id, professional_id FROM treatment_type_professionals WHERE tenant_id = $1 AND treatment_type_id = ANY($2)",
            tenant_id,
            tt_ids,
        )
        prof_map: dict = {}
        for r in ttp_rows:
            prof_map.setdefault(r["treatment_type_id"], []).append(r["professional_id"])
        for t in treatments:
            t["professional_ids"] = prof_map.get(t["id"], [])
    return treatments


@router.get(
    "/treatments/price-audit",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Auditar precios con posible escala incorrecta",
)
async def price_audit(tenant_id: int = Depends(get_resolved_tenant_id)):
    """Retorna los tratamientos cuyo base_price excede 100x el consultation_price del tenant.

    Criterio de sospecha: base_price > consultation_price * 100
    Ejemplo: consultation_price=$15.000, threshold=$1.500.000.
    Un tratamiento con base_price=$7.000.000 aparecería en la lista (ratio=466x).

    Si consultation_price no está configurado, retorna lista vacía con un aviso.
    El operador decide manualmente si corregir los precios marcados — el sistema
    NO auto-corrige datos existentes.

    Tenant_id siempre viene del contexto de autenticación (Regla de Oro).
    """
    from decimal import Decimal as _Dec

    tenant_row = await db.pool.fetchrow(
        "SELECT consultation_price FROM tenants WHERE id = $1", tenant_id
    )
    consultation_price = tenant_row["consultation_price"] if tenant_row else None

    treatment_rows = await db.pool.fetch(
        "SELECT code, name, base_price FROM treatment_types WHERE tenant_id = $1 AND is_active = true",
        tenant_id,
    )
    total_treatments = len(treatment_rows)

    if consultation_price is None or consultation_price == 0:
        return {
            "tenant_id": tenant_id,
            "consultation_price": None,
            "threshold": None,
            "suspicious": [],
            "total_treatments": total_treatments,
            "total_suspicious": 0,
            "warning": "consultation_price no configurado — no es posible detectar escala incorrecta",
        }

    cp = float(consultation_price)
    threshold = cp * 100
    suspicious = []
    for row in treatment_rows:
        bp = float(row["base_price"]) if row["base_price"] is not None else 0
        if bp > threshold:
            suspicious.append(
                {
                    "code": row["code"],
                    "name": row["name"],
                    "base_price": bp,
                    "ratio": round(bp / cp, 1),
                }
            )

    return {
        "tenant_id": tenant_id,
        "consultation_price": cp,
        "threshold": threshold,
        "suspicious": suspicious,
        "total_treatments": total_treatments,
        "total_suspicious": len(suspicious),
    }


@router.get(
    "/treatment-types/{code}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Obtener detalle de un tratamiento por su código",
)
async def get_treatment_type(
    code: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Obtener un tipo de tratamiento. Aislado por tenant_id (Regla de Oro)."""
    row = await db.pool.fetchrow(
        """
        SELECT id, code, name, description, default_duration_minutes,
               min_duration_minutes, max_duration_minutes, complexity_level,
               category, requires_multiple_sessions, session_gap_days,
               is_active, is_available_for_booking, internal_notes, priority,
               pre_instructions, post_instructions, followup_template
        FROM treatment_types
        WHERE tenant_id = $1 AND code = $2
    """,
        tenant_id,
        code,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Tipo de tratamiento no encontrado")
    result = dict(row)
    ttp_rows = await db.pool.fetch(
        "SELECT professional_id FROM treatment_type_professionals WHERE tenant_id = $1 AND treatment_type_id = $2",
        tenant_id,
        row["id"],
    )
    result["professional_ids"] = [r["professional_id"] for r in ttp_rows]
    return result


@router.post(
    "/treatment-types",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Crear un nuevo tipo de tratamiento",
)
async def create_treatment_type(
    treatment: TreatmentTypeCreate, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Crear un nuevo tipo de tratamiento. Aislado por tenant_id (Regla de Oro)."""
    try:
        # Validate new instruction fields
        _validate_treatment_instruction_fields(treatment)
        # Validate price scale (Bug #3): prevenir escala incorrecta (e.g. 7000000 en vez de 70000)
        await _validate_treatment_price_scale(
            treatment.base_price, treatment.confirm_unusual_price, tenant_id
        )

        row = await db.pool.fetchrow(
            """
            INSERT INTO treatment_types (
                tenant_id, code, name, description, default_duration_minutes,
                min_duration_minutes, max_duration_minutes, complexity_level,
                category, requires_multiple_sessions, session_gap_days,
                is_active, is_available_for_booking, internal_notes, base_price, priority,
                pre_instructions, post_instructions, followup_template, created_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, NOW())
            RETURNING id
        """,
            tenant_id,
            treatment.code,
            treatment.name,
            treatment.description,
            treatment.default_duration_minutes,
            treatment.min_duration_minutes,
            treatment.max_duration_minutes,
            treatment.complexity_level,
            treatment.category,
            treatment.requires_multiple_sessions,
            treatment.session_gap_days,
            treatment.is_active,
            treatment.is_available_for_booking,
            treatment.internal_notes,
            treatment.base_price or 0,
            treatment.priority
            if treatment.priority in ("high", "medium-high", "medium", "low")
            else "medium",
            # Migration 036: pre_instructions is now JSONB. Coerce string /
            # PreInstructions / dict into a canonical dict before serializing.
            json.dumps(_coerce_pre_instructions(treatment.pre_instructions))
            if treatment.pre_instructions is not None
            else None,
            json.dumps(_coerce_post_instructions(treatment.post_instructions))
            if treatment.post_instructions is not None
            else None,
            json.dumps(treatment.followup_template)
            if treatment.followup_template is not None
            else None,
        )
        # Insert professional assignments if provided
        if treatment.professional_ids and row:
            tt_id = row["id"]
            for pid in treatment.professional_ids:
                await db.pool.execute(
                    "INSERT INTO treatment_type_professionals (tenant_id, treatment_type_id, professional_id) VALUES ($1, $2, $3) ON CONFLICT DO NOTHING",
                    tenant_id,
                    tt_id,
                    pid,
                )
        return {"status": "created", "code": treatment.code}
    except asyncpg.UniqueViolationError:
        raise HTTPException(
            status_code=400,
            detail=f"El código de tratamiento '{treatment.code}' ya existe.",
        )
    except Exception as e:
        logger.error(f"Error creating treatment type: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.put(
    "/treatment-types/{code}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Actualizar datos de un tipo de tratamiento",
)
async def update_treatment_type(
    code: str,
    treatment: TreatmentTypeUpdate,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualizar tipo de tratamiento. Aislado por tenant_id (Regla de Oro)."""
    # Validate new instruction fields
    _validate_treatment_instruction_fields(treatment)
    # Validate price scale (Bug #3): prevenir escala incorrecta (e.g. 7000000 en vez de 70000)
    await _validate_treatment_price_scale(
        treatment.base_price, treatment.confirm_unusual_price, tenant_id
    )

    result = await db.pool.execute(
        """
        UPDATE treatment_types SET
            name = $1, description = $2, default_duration_minutes = $3,
            min_duration_minutes = $4, max_duration_minutes = $5,
            complexity_level = $6, category = $7, requires_multiple_sessions = $8,
            session_gap_days = $9, is_active = $10, is_available_for_booking = $11,
            internal_notes = $12, base_price = $13, priority = $14,
            pre_instructions = $15, post_instructions = $16, followup_template = $17,
            updated_at = NOW()
        WHERE tenant_id = $18 AND code = $19
    """,
        treatment.name,
        treatment.description,
        treatment.default_duration_minutes,
        treatment.min_duration_minutes,
        treatment.max_duration_minutes,
        treatment.complexity_level,
        treatment.category,
        treatment.requires_multiple_sessions,
        treatment.session_gap_days,
        treatment.is_active,
        treatment.is_available_for_booking,
        treatment.internal_notes,
        treatment.base_price or 0,
        treatment.priority
        if treatment.priority in ("high", "medium-high", "medium", "low")
        else "medium",
        # Migration 036: pre_instructions is now JSONB. Coerce string /
        # PreInstructions / dict into a canonical dict before serializing.
        json.dumps(_coerce_pre_instructions(treatment.pre_instructions))
        if treatment.pre_instructions is not None
        else None,
        json.dumps(_coerce_post_instructions(treatment.post_instructions))
        if treatment.post_instructions is not None
        else None,
        json.dumps(treatment.followup_template)
        if treatment.followup_template is not None
        else None,
        tenant_id,
        code,
    )
    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail="Tipo de tratamiento no encontrado")
    # Sync treatment instruction embedding if instructions were provided
    if treatment.pre_instructions or treatment.post_instructions:
        try:
            from services.embedding_service import (
                upsert_treatment_instruction_embedding,
            )
            import asyncio

            post = treatment.post_instructions
            if isinstance(post, str):
                post = json.loads(post)
            # Need treatment ID from DB
            tt_row = await db.pool.fetchrow(
                "SELECT id FROM treatment_types WHERE tenant_id = $1 AND code = $2",
                tenant_id,
                code,
            )
            if tt_row:
                asyncio.create_task(
                    upsert_treatment_instruction_embedding(
                        tenant_id,
                        tt_row["id"],
                        treatment.name,
                        treatment.pre_instructions or "",
                        post or [],
                    )
                )
        except Exception:
            pass
    return {"status": "updated", "code": code}


@router.delete(
    "/treatment-types/{code}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Eliminar o desactivar un tipo de tratamiento",
)
async def delete_treatment_type(
    code: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Eliminar o desactivar tipo de tratamiento. Aislado por tenant_id (Regla de Oro)."""
    has_appointments = await db.pool.fetchval(
        "SELECT EXISTS(SELECT 1 FROM appointments WHERE tenant_id = $1 AND appointment_type = $2)",
        tenant_id,
        code,
    )
    if has_appointments:
        await db.pool.execute(
            "UPDATE treatment_types SET is_active = false, is_available_for_booking = false WHERE tenant_id = $1 AND code = $2",
            tenant_id,
            code,
        )
        return {
            "status": "deactivated",
            "code": code,
            "message": "Tratamiento desactivado por tener citas asociadas.",
        }

    # Eliminar físicas las imágenes primero (el cascade borra la DB, pero hay que borrar de disco)
    try:
        images = await db.pool.fetch(
            "SELECT file_path FROM treatment_images WHERE tenant_id = $1 AND treatment_code = $2",
            tenant_id,
            code,
        )
        for img in images:
            if os.path.exists(img["file_path"]):
                try:
                    os.remove(img["file_path"])
                except Exception as e:
                    logger.error(
                        f"Error borrando imagen huérfana ({img['file_path']}): {e}"
                    )
    except Exception as e:
        logger.error(f"Error obteniendo imagenes para borrar: {e}")

    result = await db.pool.execute(
        "DELETE FROM treatment_types WHERE tenant_id = $1 AND code = $2",
        tenant_id,
        code,
    )
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Tipo de tratamiento no encontrado")
    return {"status": "deleted", "code": code}


@router.get(
    "/treatment-types/{code}/professionals",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Listar profesionales asignados a un tratamiento",
)
async def get_treatment_professionals(
    code: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Listar profesionales asignados a un tratamiento. Aislado por tenant_id (Regla de Oro)."""
    tt = await db.pool.fetchrow(
        "SELECT id FROM treatment_types WHERE tenant_id = $1 AND code = $2",
        tenant_id,
        code,
    )
    if not tt:
        raise HTTPException(status_code=404, detail="Tipo de tratamiento no encontrado")
    rows = await db.pool.fetch(
        """
        SELECT p.id, p.first_name, p.last_name, p.specialty
        FROM treatment_type_professionals ttp
        JOIN professionals p ON ttp.professional_id = p.id
        WHERE ttp.tenant_id = $1 AND ttp.treatment_type_id = $2
        ORDER BY p.first_name
    """,
        tenant_id,
        tt["id"],
    )
    return [dict(r) for r in rows]


@router.put(
    "/treatment-types/{code}/professionals",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Reemplazar profesionales asignados a un tratamiento",
)
async def set_treatment_professionals(
    code: str, body: dict, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Reemplazar asignaciones de profesionales para un tratamiento. Aislado por tenant_id (Regla de Oro)."""
    professional_ids = body.get("professional_ids", [])
    tt = await db.pool.fetchrow(
        "SELECT id FROM treatment_types WHERE tenant_id = $1 AND code = $2",
        tenant_id,
        code,
    )
    if not tt:
        raise HTTPException(status_code=404, detail="Tipo de tratamiento no encontrado")
    tt_id = tt["id"]
    # Delete all current assignments, then insert new ones
    await db.pool.execute(
        "DELETE FROM treatment_type_professionals WHERE tenant_id = $1 AND treatment_type_id = $2",
        tenant_id,
        tt_id,
    )
    for pid in professional_ids:
        await db.pool.execute(
            "INSERT INTO treatment_type_professionals (tenant_id, treatment_type_id, professional_id) VALUES ($1, $2, $3) ON CONFLICT DO NOTHING",
            tenant_id,
            tt_id,
            pid,
        )
    return {"status": "updated", "code": code, "professional_ids": professional_ids}


@router.get(
    "/treatment-types/{code}/duration",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Obtener duración estimada según nivel de urgencia",
)
async def get_treatment_duration(
    code: str,
    urgency_level: str = "normal",
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Obtener duración calculada para un tratamiento según urgencia.

    Args:
        code: Código del tratamiento (ej: 'root_canal', 'cleaning')
        urgency_level: 'low', 'normal', 'high', 'emergency'

    Returns:
        duration_minutes: Duración calculada en minutos
    """
    row = await db.pool.fetchrow(
        """
        SELECT default_duration_minutes, min_duration_minutes, max_duration_minutes
        FROM treatment_types
        WHERE tenant_id = $1 AND code = $2 AND is_active = TRUE AND is_available_for_booking = TRUE
    """,
        tenant_id,
        code,
    )

    if not row:
        # Return default duration if treatment not found
        return {"duration_minutes": 30, "source": "default"}

    default_duration = row["default_duration_minutes"]
    min_duration = row["min_duration_minutes"]
    max_duration = row["max_duration_minutes"]

    # Calculate duration based on urgency
    if urgency_level == "emergency":
        calculated_duration = min(min_duration, default_duration)
    elif urgency_level == "high":
        calculated_duration = default_duration
    elif urgency_level == "normal":
        calculated_duration = default_duration
    else:  # low
        calculated_duration = max(default_duration, max_duration)

    return {
        "duration_minutes": calculated_duration,
        "source": "calculated",
        "treatment_code": code,
        "urgency_level": urgency_level,
        "details": {
            "default": default_duration,
            "min": min_duration,
            "max": max_duration,
        },
    }


# --- TRATAMIENTO IMÁGENES ---


@router.get(
    "/treatment-types/{code}/images",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Listar imágenes de un tratamiento",
)
async def list_treatment_images(
    code: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Lista las imágenes físicas asociadas a un tratamiento."""
    rows = await db.pool.fetch(
        """
        SELECT id, filename, file_path, mime_type, file_size, created_at
        FROM treatment_images
        WHERE tenant_id = $1 AND treatment_code = $2
        ORDER BY created_at ASC
    """,
        tenant_id,
        code,
    )
    return [dict(row) for row in rows]


@router.post(
    "/treatment-types/{code}/images",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Subir imagen a un tratamiento",
)
async def upload_treatment_image(
    code: str,
    file: UploadFile = File(...),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Sube una imagen física y la asocia a un tratamiento."""
    # Verificar que el tratamiento existe
    exists = await db.pool.fetchval(
        "SELECT 1 FROM treatment_types WHERE tenant_id = $1 AND code = $2",
        tenant_id,
        code,
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Tratamiento no encontrado")

    upload_dir = f"uploads/treatments/{tenant_id}/{code}"
    os.makedirs(upload_dir, exist_ok=True)

    # Validate file extension
    ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".pdf"}
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de archivo no permitido. Permitidos: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    safe_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(upload_dir, safe_filename)

    try:
        content = await file.read()
        # Check file size (max 10MB)
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(
                status_code=400, detail="Archivo demasiado grande. Máximo 10MB."
            )
        with open(file_path, "wb") as f:
            f.write(content)

        file_size = len(content)
        mime_type = file.content_type

        new_id = await db.pool.fetchval(
            """
            INSERT INTO treatment_images (tenant_id, treatment_code, filename, file_path, mime_type, file_size)
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING id
        """,
            tenant_id,
            code,
            file.filename,
            file_path,
            mime_type,
            file_size,
        )

        return {
            "id": new_id,
            "filename": file.filename,
            "file_path": file_path,
            "status": "uploaded",
        }
    except Exception as e:
        logger.error(f"Error uploading treatment image: {e}")
        raise HTTPException(status_code=500, detail="Error al guardar la imagen")


@router.delete(
    "/treatment-types/{code}/images/{image_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Tratamientos"],
    summary="Borrar imagen de un tratamiento",
)
async def delete_treatment_image(
    code: str, image_id: str, tenant_id: int = Depends(get_resolved_tenant_id)
):
    """Elimina una imagen física de un tratamiento."""
    row = await db.pool.fetchrow(
        """
        SELECT file_path FROM treatment_images WHERE id = $1 AND tenant_id = $2 AND treatment_code = $3
    """,
        image_id,
        tenant_id,
        code,
    )

    if not row:
        raise HTTPException(status_code=404, detail="Imagen no encontrada")

    file_path = row["file_path"]
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            logger.error(f"Error borrando archivo físico {file_path}: {e}")

    await db.pool.execute(
        "DELETE FROM treatment_images WHERE id = $1 AND tenant_id = $2",
        image_id,
        tenant_id,
    )
    return {"status": "deleted"}


@router.get(
    "/public/media/{image_id}", tags=["Media"], summary="Servidor público de imágenes"
)
async def get_public_media(image_id: str):
    """Endpoint público para servir las imágenes temporalmente a YCloud/WhatsApp a través de la URL"""
    try:
        row = await db.pool.fetchrow(
            "SELECT file_path, mime_type FROM treatment_images WHERE id = $1", image_id
        )
        if not row or not os.path.exists(row["file_path"]):
            raise HTTPException(status_code=404, detail="Imagen no encontrada")

        return FileResponse(row["file_path"], media_type=row["mime_type"])
    except Exception as e:
        logger.error(f"Error fetching media: {e}")
        raise HTTPException(status_code=400, detail="ID Invalido o error interno")


@router.get(
    "/public/tenant-logo/{tenant_id}",
    tags=["Media"],
    summary="Logo público de la clínica",
)
async def get_tenant_logo(tenant_id: int):
    """Endpoint público para servir el logo de la clínica (usado como favicon y en sidebar)."""
    upload_dir = Path(f"/app/uploads/tenants/{tenant_id}")
    if upload_dir.exists():
        for ext in ("png", "jpg", "jpeg", "svg", "webp", "ico"):
            logo_path = upload_dir / f"logo.{ext}"
            if logo_path.exists():
                mime_map = {
                    "png": "image/png",
                    "jpg": "image/jpeg",
                    "jpeg": "image/jpeg",
                    "svg": "image/svg+xml",
                    "webp": "image/webp",
                    "ico": "image/x-icon",
                }
                return FileResponse(
                    str(logo_path), media_type=mime_map.get(ext, "image/png")
                )
    raise HTTPException(status_code=404, detail="Logo no encontrado")


# ==================== ENDPOINTS ANALYTICS ====================


@router.get(
    "/analytics/professionals/summary",
    tags=["Analítica"],
    summary="Obtener reporte analítico de profesionales para el CEO",
)
async def get_professionals_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Retorna métricas estratégicas de los profesionales para el dashboard del CEO.
    """
    try:
        # Default to current month if not specified
        if not start_date or not end_date:
            today = datetime.now()
            start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # End of month roughly
            if today.month == 12:
                end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(
                    days=1
                )
            else:
                end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        else:
            start = datetime.fromisoformat(start_date)
            end = datetime.fromisoformat(end_date)

        data = await analytics_service.get_professionals_summary(
            start, end, tenant_id=tenant_id
        )
        return data

    except Exception as e:
        logger.error(f"Error in analytics endpoint: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


@router.get(
    "/analytics/professionals/liquidation",
    dependencies=[Depends(verify_admin_token)],
    tags=["Analítica"],
    summary="Liquidación por profesional con trazabilidad completa",
)
async def get_liquidation(
    start_date: str = Query(...),
    end_date: str = Query(...),
    professional_id: Optional[int] = Query(None),
    payment_status: Optional[str] = Query("all"),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Retorna el detalle completo de facturación por profesional para liquidaciones.
    Agrupa por (paciente, tratamiento) con sesiones individuales y totales.
    """
    # --- Validation ---
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de fecha inválido. Usar YYYY-MM-DD.",
        )

    if start > end:
        raise HTTPException(
            status_code=400,
            detail="start_date debe ser menor o igual a end_date.",
        )

    if (end - start).days > 366:
        raise HTTPException(
            status_code=400,
            detail="El rango no puede superar 366 días.",
        )

    valid_statuses = ("all", "pending", "partial", "paid")
    if payment_status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"payment_status inválido. Valores permitidos: {', '.join(valid_statuses)}.",
        )

    try:
        data = await analytics_service.get_professionals_liquidation(
            pool=db.pool,
            tenant_id=tenant_id,
            start_date=start,
            end_date=end,
            professional_id=professional_id,
            payment_status=payment_status,
        )
        return data
    except Exception as e:
        logger.error(f"Error en liquidación: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# ==================== END OF ANALYTICS ====================


# =============================================
# Spec 12: HEALTH CHECK — INTEGRACIONES
# =============================================


@router.get(
    "/health/integrations",
    tags=["Mantenimiento"],
    summary="Health-Check de integraciones externas (Facebook/Meta Ads)",
)
async def health_check_integrations(
    user_data=Depends(verify_admin_token),
):
    """
    Spec 12: Verifica el estado de las integraciones externas (Meta Ads).
    Retorna status de conexión, validez de token y cuentas activas.
    """
    if user_data.role not in ["ceo", "secretary"]:
        raise HTTPException(
            status_code=403,
            detail="Solo administradores pueden ver estado de integraciones.",
        )

    from scripts.check_meta_health import check_meta_health

    result = await check_meta_health()

    status_code = 200 if result.get("status") == "ok" else 503
    return JSONResponse(content=result, status_code=status_code)


@router.get(
    "/marketing/automation-logs",
    tags=["Marketing Analytics"],
    summary="Consultar logs de automatización de marketing",
)
async def get_automation_logs(
    limit: int = Query(50, ge=1, le=200),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
    user_data=Depends(verify_admin_token),
):
    """
    Spec 08: Auditoría de logs de automatización (UI Transparente).
    Muestra el historial de recordatorios y seguimientos enviados.
    """
    if user_data.role not in ["ceo", "secretary"]:
        raise HTTPException(status_code=403, detail="Acceso denegado.")

    try:
        rows = await db.pool.fetch(
            """
            SELECT l.id, l.patient_id, p.first_name || ' ' || COALESCE(p.last_name, '') as patient_name,
                   l.trigger_type, l.status, l.created_at, l.error_details, l.meta
            FROM automation_logs l
            JOIN patients p ON l.patient_id = p.id
            WHERE l.tenant_id = $1
            ORDER BY l.created_at DESC
            LIMIT $2
        """,
            resolved_tenant_id,
            limit,
        )

        return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"Error fetching automation logs: {e}")
        raise HTTPException(
            status_code=500, detail="Error obteniendo logs de automatización."
        )


# ==================== ENDPOINTS ANAMNESIS (Spec 2026-03-11) ====================


class AnamnesisUpdate(BaseModel):
    base_diseases: Optional[str] = None
    habitual_medication: Optional[str] = None
    allergies: Optional[str] = None
    previous_surgeries: Optional[str] = None
    is_smoker: Optional[str] = None
    smoker_amount: Optional[str] = None
    pregnancy_lactation: Optional[str] = None
    negative_experiences: Optional[str] = None
    specific_fears: Optional[str] = None


@router.get("/patients/by-phone/{phone}", tags=["Pacientes"])
async def get_patient_by_phone(
    phone: str,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Devuelve el paciente (incluyendo medical_history JSONB) buscado por número de teléfono.
    Usado por el panel de anamnesis en ChatsView y EditAppointment modal.
    Filtrado obligatorio por tenant_id extraído del JWT.
    """
    tenant_id = resolved_tenant_id
    normalized = normalize_phone(phone)
    row = await db.pool.fetchrow(
        """SELECT id, first_name, last_name, phone_number, email, dni, birth_date,
                  city, acquisition_source, insurance_provider, status, medical_history,
                  created_at
           FROM patients
           WHERE tenant_id = $1 AND (phone_number = $2 OR phone_number = $3)
             AND status != 'deleted'
           LIMIT 1""",
        tenant_id,
        normalized,
        phone,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Paciente no encontrado.")

    patient_dict = dict(row)
    # Asegurar que medical_history se devuelve como dict (o None)
    if patient_dict.get("medical_history") and isinstance(
        patient_dict["medical_history"], str
    ):
        try:
            patient_dict["medical_history"] = json.loads(
                patient_dict["medical_history"]
            )
        except Exception:
            pass
    return patient_dict


@router.patch("/patients/{patient_id}/anamnesis", tags=["Pacientes"])
async def update_patient_anamnesis(
    patient_id: int,
    payload: AnamnesisUpdate,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Actualiza (merge) la anamnesis de un paciente.
    Permisos:
    - CEO: puede editar cualquier paciente de su/sus clínica(s).
    - Profesional: solo puede editar pacientes asignados a él (tienen al menos 1 turno con ese profesional).
    - Secretaria / otros: 403.
    Siempre filtra por tenant_id del JWT.
    """
    tenant_id = resolved_tenant_id

    # 1. Verificar que el paciente existe en el tenant (incluir phone_number para el socket)
    patient = await db.pool.fetchrow(
        "SELECT id, phone_number, medical_history FROM patients WHERE id = $1 AND tenant_id = $2 AND status != 'deleted'",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado.")

    # 2. Control de rol
    role = user_data.role
    if role == "ceo":
        pass  # CEO puede editar siempre
    elif role == "professional":
        # Verificar que tiene al menos 1 turno con este paciente
        uid = uuid.UUID(user_data.user_id)
        prof = await db.pool.fetchrow(
            "SELECT id FROM professionals WHERE user_id = $1 AND tenant_id = $2 AND is_active = true",
            uid,
            tenant_id,
        )
        if not prof:
            raise HTTPException(
                status_code=403, detail="Profesional no encontrado en esta sede."
            )
        has_appointment = await db.pool.fetchval(
            """SELECT 1 FROM appointments
               WHERE tenant_id = $1 AND patient_id = $2 AND professional_id = $3
               LIMIT 1""",
            tenant_id,
            patient_id,
            prof["id"],
        )
        if not has_appointment:
            raise HTTPException(
                status_code=403, detail="No tenés turnos asignados con este paciente."
            )
    else:
        raise HTTPException(
            status_code=403, detail="No tenés permisos para editar la anamnesis."
        )

    # 3. Merge: conservar campos existentes, sobreescribir solo los enviados
    current = {}
    if patient["medical_history"]:
        if isinstance(patient["medical_history"], str):
            try:
                current = json.loads(patient["medical_history"])
            except Exception:
                current = {}
        elif isinstance(patient["medical_history"], dict):
            current = patient["medical_history"]

    updates = {k: v for k, v in payload.dict().items() if v is not None}
    current.update(updates)
    current["anamnesis_last_edited_by"] = user_data.user_id
    current["anamnesis_last_edited_at"] = datetime.now(timezone.utc).isoformat()

    await db.pool.execute(
        "UPDATE patients SET medical_history = $1::jsonb WHERE id = $2 AND tenant_id = $3",
        json.dumps(current),
        patient_id,
        tenant_id,
    )

    # Notificar via Socket.IO para refrescar UI
    try:
        from main import sio, to_json_safe

        await sio.emit(
            "PATIENT_UPDATED",
            to_json_safe(
                {
                    "phone_number": patient["phone_number"],
                    "patient_id": patient_id,
                    "tenant_id": tenant_id,
                    "update_type": "anamnesis_manual_update",
                }
            ),
        )
    except:
        pass

    return {
        "message": "Anamnesis actualizada correctamente.",
        "medical_history": current,
    }


class EmailUpdate(BaseModel):
    email: str


@router.put("/patients/{patient_id}/email", tags=["Pacientes"])
async def update_patient_email(
    patient_id: int,
    payload: EmailUpdate,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Actualiza el email de un paciente.
    Valida formato de email y filtra por tenant_id.
    """
    import re

    tenant_id = resolved_tenant_id
    email = payload.email.strip()

    # Validar formato de email
    email_pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    if not re.match(email_pattern, email):
        raise HTTPException(status_code=400, detail="Formato de email inválido.")

    # Verificar que el paciente existe
    patient = await db.pool.fetchrow(
        "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2 AND status != 'deleted'",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado.")

    # Actualizar email
    await db.pool.execute(
        "UPDATE patients SET email = $1, updated_at = NOW() WHERE id = $2 AND tenant_id = $3",
        email,
        patient_id,
        tenant_id,
    )

    # Notificar via Socket.IO
    try:
        from main import sio, to_json_safe

        await sio.emit(
            "PATIENT_UPDATED",
            to_json_safe(
                {
                    "patient_id": patient_id,
                    "tenant_id": tenant_id,
                    "update_type": "email_update",
                }
            ),
        )
    except:
        pass

    return {"message": "Email actualizado correctamente.", "email": email}


# ============================================
# AUTOMATION ENGINE v2 — Motor de Reglas
# ============================================


class AutomationRuleCreate(BaseModel):
    name: str
    trigger_type: str
    condition_json: dict = {}
    message_type: str = "free_text"
    free_text_message: Optional[str] = None
    ycloud_template_name: Optional[str] = None
    ycloud_template_lang: Optional[str] = "es"
    ycloud_template_vars: dict = {}
    channels: List[str] = ["whatsapp"]
    send_hour_min: int = 8
    send_hour_max: int = 20
    is_active: bool = True


class AutomationRuleUpdate(BaseModel):
    name: Optional[str] = None
    trigger_type: Optional[str] = None
    condition_json: Optional[dict] = None
    message_type: Optional[str] = None
    free_text_message: Optional[str] = None
    ycloud_template_name: Optional[str] = None
    ycloud_template_lang: Optional[str] = None
    ycloud_template_vars: Optional[dict] = None
    channels: Optional[List[str]] = None
    send_hour_min: Optional[int] = None
    send_hour_max: Optional[int] = None
    is_active: Optional[bool] = None


async def write_automation_log(
    tenant_id: int,
    trigger_type: str,
    status: str,
    patient_name: str = None,
    phone_number: str = None,
    channel: str = "whatsapp",
    message_type: str = "free_text",
    message_preview: str = None,
    template_name: str = None,
    rule_id: int = None,
    rule_name: str = None,
    skip_reason: str = None,
    ycloud_message_id: str = None,
    patient_id: int = None,
    error_detail: str = None,
):
    """Helper centralizado para escribir en automation_logs desde cualquier job o endpoint."""
    try:
        await db.pool.execute(
            """
            INSERT INTO automation_logs (
                tenant_id, automation_rule_id, rule_name, trigger_type,
                patient_id, patient_name, phone_number, channel,
                message_type, message_preview, template_name,
                status, skip_reason, ycloud_message_id, error_details,
                triggered_at, sent_at
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,NOW(),
                CASE WHEN $12 = 'sent' THEN NOW() ELSE NULL END
            )
        """,
            tenant_id,
            rule_id,
            rule_name,
            trigger_type,
            patient_id,
            patient_name,
            phone_number,
            channel,
            message_type,
            (message_preview[:200] if message_preview else None),
            template_name,
            status,
            skip_reason,
            ycloud_message_id,
            error_detail,
        )
    except Exception as e:
        logger.warning(f"⚠️ No se pudo escribir automation_log: {e}")


@router.get("/automations/rules")
async def get_automation_rules(
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Lista todas las reglas de automatización del tenant (sistema + personalizadas)."""
    rows = await db.pool.fetch(
        """
        SELECT id, name, is_active, is_system, trigger_type, condition_json,
               message_type, free_text_message, ycloud_template_name, ycloud_template_lang,
               ycloud_template_vars, channels, send_hour_min, send_hour_max, created_at, updated_at
        FROM automation_rules
        WHERE tenant_id = $1
        ORDER BY is_system DESC, created_at ASC
    """,
        tenant_id,
    )
    return {"rules": [dict(r) for r in rows]}


@router.post("/automations/rules")
async def create_automation_rule(
    payload: AutomationRuleCreate,
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
    user_data=Depends(verify_admin_token),
):
    """Crea una nueva regla de automatización personalizada."""
    row = await db.pool.fetchrow(
        """
        INSERT INTO automation_rules (
            tenant_id, name, trigger_type, condition_json, message_type,
            free_text_message, ycloud_template_name, ycloud_template_lang,
            ycloud_template_vars, channels, send_hour_min, send_hour_max,
            is_active, is_system, created_by, created_at, updated_at
        ) VALUES ($1,$2,$3,$4::jsonb,$5,$6,$7,$8,$9::jsonb,$10,$11,$12,$13,FALSE,$14,NOW(),NOW())
        RETURNING *
    """,
        tenant_id,
        payload.name,
        payload.trigger_type,
        json.dumps(payload.condition_json),
        payload.message_type,
        payload.free_text_message,
        payload.ycloud_template_name,
        payload.ycloud_template_lang,
        json.dumps(payload.ycloud_template_vars),
        payload.channels,
        payload.send_hour_min,
        payload.send_hour_max,
        payload.is_active,
        getattr(user_data, "user_id", None),
    )
    return {"rule": dict(row), "message": "Regla creada correctamente."}


@router.patch("/automations/rules/{rule_id}")
async def update_automation_rule(
    rule_id: int,
    payload: AutomationRuleUpdate,
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Edita una regla de automatización. No permite editar reglas de sistema."""
    rule = await db.pool.fetchrow(
        "SELECT * FROM automation_rules WHERE id = $1 AND tenant_id = $2",
        rule_id,
        tenant_id,
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Regla no encontrada.")
    if rule["is_system"]:
        # Las reglas de sistema permiten editar el MENSAJE (tipo + contenido) y el toggle is_active.
        # NO se puede cambiar trigger_type, channels ni horarios.
        system_updates = {}
        if payload.is_active is not None:
            system_updates["is_active"] = payload.is_active
        if payload.message_type is not None:
            system_updates["message_type"] = payload.message_type
        if payload.free_text_message is not None:
            system_updates["free_text_message"] = payload.free_text_message
        if payload.ycloud_template_name is not None:
            system_updates["ycloud_template_name"] = payload.ycloud_template_name
        if payload.ycloud_template_lang is not None:
            system_updates["ycloud_template_lang"] = payload.ycloud_template_lang

        if payload.ycloud_template_vars is not None:
            await db.pool.execute(
                "UPDATE automation_rules SET ycloud_template_vars=$1::jsonb, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
                json.dumps(payload.ycloud_template_vars),
                rule_id,
                tenant_id,
            )

        for field, value in system_updates.items():
            await db.pool.execute(
                f"UPDATE automation_rules SET {field}=$1, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
                value,
                rule_id,
                tenant_id,
            )

        updated = await db.pool.fetchrow(
            "SELECT * FROM automation_rules WHERE id=$1", rule_id
        )
        return {"rule": dict(updated), "message": "Regla de sistema actualizada."}

    # Construir UPDATE dinámico
    updates = {}
    if payload.name is not None:
        updates["name"] = payload.name
    if payload.trigger_type is not None:
        updates["trigger_type"] = payload.trigger_type
    if payload.message_type is not None:
        updates["message_type"] = payload.message_type
    if payload.free_text_message is not None:
        updates["free_text_message"] = payload.free_text_message
    if payload.ycloud_template_name is not None:
        updates["ycloud_template_name"] = payload.ycloud_template_name
    if payload.ycloud_template_lang is not None:
        updates["ycloud_template_lang"] = payload.ycloud_template_lang
    if payload.channels is not None:
        updates["channels"] = payload.channels
    if payload.send_hour_min is not None:
        updates["send_hour_min"] = payload.send_hour_min
    if payload.send_hour_max is not None:
        updates["send_hour_max"] = payload.send_hour_max
    if payload.is_active is not None:
        updates["is_active"] = payload.is_active

    if payload.condition_json is not None:
        await db.pool.execute(
            "UPDATE automation_rules SET condition_json=$1::jsonb, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
            json.dumps(payload.condition_json),
            rule_id,
            tenant_id,
        )
    if payload.ycloud_template_vars is not None:
        await db.pool.execute(
            "UPDATE automation_rules SET ycloud_template_vars=$1::jsonb, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
            json.dumps(payload.ycloud_template_vars),
            rule_id,
            tenant_id,
        )

    for field, value in updates.items():
        await db.pool.execute(
            f"UPDATE automation_rules SET {field}=$1, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
            value,
            rule_id,
            tenant_id,
        )

    updated = await db.pool.fetchrow(
        "SELECT * FROM automation_rules WHERE id=$1", rule_id
    )
    return {"rule": dict(updated), "message": "Regla actualizada."}


@router.patch("/automations/rules/{rule_id}/toggle")
async def toggle_automation_rule(
    rule_id: int,
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Activa o desactiva una regla (incluyendo las de sistema)."""
    rule = await db.pool.fetchrow(
        "SELECT is_active FROM automation_rules WHERE id=$1 AND tenant_id=$2",
        rule_id,
        tenant_id,
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Regla no encontrada.")
    new_state = not rule["is_active"]
    await db.pool.execute(
        "UPDATE automation_rules SET is_active=$1, updated_at=NOW() WHERE id=$2 AND tenant_id=$3",
        new_state,
        rule_id,
        tenant_id,
    )
    return {
        "is_active": new_state,
        "message": f"Regla {'activada' if new_state else 'desactivada'}.",
    }


@router.delete("/automations/rules/{rule_id}")
async def delete_automation_rule(
    rule_id: int,
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Elimina una regla personalizada. No se pueden eliminar las reglas de sistema."""
    rule = await db.pool.fetchrow(
        "SELECT is_system FROM automation_rules WHERE id=$1 AND tenant_id=$2",
        rule_id,
        tenant_id,
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Regla no encontrada.")
    if rule["is_system"]:
        raise HTTPException(
            status_code=403, detail="Las reglas de sistema no se pueden eliminar."
        )
    await db.pool.execute(
        "DELETE FROM automation_rules WHERE id=$1 AND tenant_id=$2", rule_id, tenant_id
    )
    return {"message": "Regla eliminada correctamente."}


@router.get("/automations/logs")
async def get_automation_logs(
    request: Request,
    trigger_type: Optional[str] = None,
    status: Optional[str] = None,
    channel: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Logs de automatización con filtros. Incluye logs de todos los jobs del sistema."""
    offset = (page - 1) * limit

    conditions = ["al.tenant_id = $1"]
    params = [tenant_id]
    idx = 2

    if trigger_type:
        conditions.append(f"al.trigger_type = ${idx}")
        params.append(trigger_type)
        idx += 1
    if status:
        conditions.append(f"al.status = ${idx}")
        params.append(status)
        idx += 1
    if channel:
        conditions.append(f"al.channel = ${idx}")
        params.append(channel)
        idx += 1
    if date_from:
        conditions.append(f"al.triggered_at >= ${idx}::date")
        params.append(date_from)
        idx += 1
    if date_to:
        conditions.append(f"al.triggered_at <= (${idx}::date + INTERVAL '1 day')")
        params.append(date_to)
        idx += 1

    where_clause = " AND ".join(conditions)

    rows = await db.pool.fetch(
        f"""
        SELECT al.id, al.automation_rule_id, al.rule_name, al.trigger_type,
               al.patient_name, al.phone_number, al.channel,
               al.message_type, al.message_preview, al.template_name,
               al.status, al.skip_reason, al.error_details, al.ycloud_message_id,
               al.triggered_at, al.sent_at, al.delivered_at
        FROM automation_logs al
        WHERE {where_clause}
        ORDER BY al.triggered_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
    """,
        *params,
        limit,
        offset,
    )

    total = await db.pool.fetchval(
        f"""
        SELECT COUNT(*) FROM automation_logs al WHERE {where_clause}
    """,
        *params,
    )

    return {
        "logs": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit if total else 0,
    }


@router.get("/automations/ycloud-templates")
async def list_ycloud_templates(
    request: Request,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Proxy: obtiene plantillas HSM aprobadas desde YCloud API. No expone la API key al frontend."""
    # Lee la API key del Vault de credenciales del tenant (Settings → YCloud)
    ycloud_api_key = await get_tenant_credential(tenant_id, "YCLOUD_API_KEY")
    if not ycloud_api_key:
        logger.warning(
            f"📋 YCloud templates: API key not configured for tenant {tenant_id}"
        )
        return {
            "templates": [],
            "warning": "YCloud API Key no configurada. Configurala en Configuración → YCloud.",
            "diagnostic": {"step": "credentials", "found": False},
        }

    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                "https://api.ycloud.com/v2/whatsapp/templates",
                headers={"X-API-Key": ycloud_api_key},
                params={"limit": 100},
            )
        logger.info(
            f"📋 YCloud templates API: status={response.status_code} for tenant {tenant_id}"
        )

        if response.status_code != 200:
            logger.warning(
                f"📋 YCloud templates API error {response.status_code}: {response.text[:500]}"
            )
            return {
                "templates": [],
                "warning": f"YCloud devolvió {response.status_code}: {response.text[:200]}",
                "diagnostic": {
                    "step": "ycloud_api",
                    "status_code": response.status_code,
                    "body": response.text[:500],
                },
            }

        data = response.json()

        # YCloud may return items in different keys depending on API version
        templates = (
            data.get("items")
            or data.get("data")
            or data.get("results")
            or (data if isinstance(data, list) else [])
        )

        # Log raw response for debugging
        logger.info(
            f"📋 YCloud raw response: {len(templates)} total templates, "
            f"keys={list(data.keys()) if isinstance(data, dict) else 'list'}"
        )
        if templates:
            sample_statuses = [t.get("status") for t in templates[:5]]
            sample_names = [t.get("name") for t in templates[:5]]
            logger.info(f"📋 Sample template statuses: {sample_statuses}")
            logger.info(f"📋 Sample template names: {sample_names}")

        # Filter approved templates — case-insensitive, accept multiple status values
        # Meta/YCloud uses uppercase "APPROVED" but we defensively accept variants
        accepted_statuses = {"approved", "active", "enabled"}
        approved = [
            t
            for t in templates
            if str(t.get("status", "")).strip().lower() in accepted_statuses
        ]

        logger.info(
            f"📋 YCloud templates: {len(approved)}/{len(templates)} approved for tenant {tenant_id}"
        )

        # Diagnostic info if filter excluded everything
        diagnostic = {
            "step": "ok",
            "raw_count": len(templates),
            "approved_count": len(approved),
        }
        if templates and not approved:
            unique_statuses = list({str(t.get("status", "")) for t in templates})
            diagnostic["unmatched_statuses"] = unique_statuses
            logger.warning(
                f"📋 YCloud filter matched 0 templates. Found statuses: {unique_statuses}. "
                f"Expected one of: {accepted_statuses}"
            )

        result = {
            "templates": approved,
            "total": len(approved),
            "diagnostic": diagnostic,
        }

        # If we got templates from YCloud but none approved, give a helpful warning
        if templates and not approved:
            result["warning"] = (
                f"YCloud devolvió {len(templates)} plantillas pero ninguna está aprobada. "
                f"Estados encontrados: {', '.join(diagnostic['unmatched_statuses'])}. "
                "Si las plantillas fueron creadas directamente en Meta, puede que YCloud aún no las haya sincronizado."
            )
        elif not templates:
            result["warning"] = (
                "YCloud devolvió una lista vacía. Verificá que la API key corresponda al WABA correcto "
                "y que las plantillas estén creadas/sincronizadas en YCloud (no solo en Meta)."
            )

        return result
    except Exception as e:
        logger.error(f"📋 Error obteniendo templates YCloud: {e}", exc_info=True)
        return {
            "templates": [],
            "warning": str(e),
            "diagnostic": {"step": "exception", "error": str(e)},
        }


# ============================================
# APPOINTMENTS — PATCH STATUS (Event-Driven)
# ============================================


class AppointmentStatusUpdate(BaseModel):
    status: str  # pending, confirmed, completed, cancelled, no_show
    notes: Optional[str] = None


async def trigger_feedback_after_delay(
    appointment_id: int,
    tenant_id: int,
    patient_name: str,
    phone_number: str,
    delay_minutes: int = 45,
):
    """
    Job event-driven: se ejecuta delay_minutes después de marcar un turno como completado.
    Envía el mensaje de feedback configurado en la regla 'post_appointment_completed'.
    """
    import asyncio

    try:
        await asyncio.sleep(delay_minutes * 60)

        # 1. Verificar que la regla de feedback esté activa
        rule = await db.pool.fetchrow(
            """
            SELECT * FROM automation_rules
            WHERE tenant_id=$1 AND trigger_type='post_appointment_completed' AND is_active=TRUE
            ORDER BY is_system DESC LIMIT 1
        """,
            tenant_id,
        )

        if not rule:
            logger.info(
                f"⏭️ Regla feedback inactiva para tenant {tenant_id}, no se envía."
            )
            return

        # 2. Verificar que no se haya enviado ya (idempotencia)
        existing = await db.pool.fetchval(
            """
            SELECT id FROM automation_logs
            WHERE automation_rule_id=$1 AND phone_number=$2 AND DATE(triggered_at)=CURRENT_DATE AND status='sent'
        """,
            rule["id"],
            phone_number,
        )
        if existing:
            logger.info(f"⏭️ Feedback ya enviado hoy a {phone_number}")
            return

        # 3. Construir mensaje con variables
        first_name = patient_name.split()[0] if patient_name else "paciente"
        message = (
            rule["free_text_message"] or "Hola! ¿Cómo te sentís después de la consulta?"
        ).replace("{{first_name}}", first_name)

        # 4. Enviar via WhatsApp Service
        internal_token = os.getenv("INTERNAL_API_TOKEN")
        whatsapp_url = os.getenv("WHATSAPP_SERVICE_URL", "http://whatsapp_service:8000")
        sent_ok = False
        ycloud_msg_id = None

        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(
                    f"{whatsapp_url}/send",
                    json={"to": phone_number, "text": message, "tenant_id": tenant_id},
                    headers={
                        "X-Internal-Token": internal_token,
                        "Content-Type": "application/json",
                    },
                )
            sent_ok = resp.status_code == 200
            if sent_ok:
                result = resp.json()
                ycloud_msg_id = result.get("messageId")
        except Exception as send_err:
            logger.error(f"❌ Error enviando feedback a {phone_number}: {send_err}")

        # 5. Registrar en automation_logs
        await write_automation_log(
            tenant_id=tenant_id,
            trigger_type="post_appointment_completed",
            status="sent" if sent_ok else "failed",
            patient_name=patient_name,
            phone_number=phone_number,
            channel="whatsapp",
            message_type="free_text",
            message_preview=message[:200],
            rule_id=rule["id"],
            rule_name=rule["name"],
            ycloud_message_id=ycloud_msg_id,
            error_detail=None if sent_ok else "WhatsApp service error",
        )

        # 6. Marcar followup_sent en el turno
        if sent_ok:
            await db.pool.execute(
                "UPDATE appointments SET followup_sent=TRUE, followup_sent_at=NOW() WHERE id=$1 AND tenant_id=$2",
                appointment_id,
                tenant_id,
            )
            logger.info(f"✅ Feedback post-completado enviado a {phone_number}")

    except Exception as e:
        logger.error(f"❌ Error en trigger_feedback_after_delay: {e}")


# ============================================
# TREATMENT PLAN BILLING - CRUD ENDPOINTS (EP-01 a EP-05)
# ============================================


async def emit_treatment_plan_event(
    event_type: str, data: Dict[str, Any], request: Request
):
    """Emite eventos de treatment plan via Socket.IO."""
    if hasattr(request.app.state, "emit_appointment_event"):
        await request.app.state.emit_appointment_event(event_type, data)


async def recalculate_plan_totals(
    pool, plan_id: str, tenant_id: int
) -> Dict[str, float]:
    """Recalcula estimated_total del plan basado en sus items."""
    result = await pool.fetchrow(
        """
        SELECT 
            COALESCE(SUM(estimated_price), 0) as estimated_total,
            COALESCE(SUM(approved_price), 0) as approved_total
        FROM treatment_plan_items
        WHERE plan_id = $1 AND tenant_id = $2 AND status != 'cancelled'
        """,
        plan_id,
        tenant_id,
    )
    return {
        "estimated_total": float(result["estimated_total"]) if result else 0,
        "approved_total": float(result["approved_total"]) if result else 0,
    }


async def get_professional_name(pool, professional_id: int) -> Optional[str]:
    """Obtiene el nombre del profesional."""
    row = await pool.fetchrow(
        "SELECT first_name FROM professionals WHERE id = $1", professional_id
    )
    return row["first_name"] if row else None


def validate_plan_status_transition(current: str, new: str) -> bool:
    """Valida transiciones de estado válidas para planes."""
    valid_transitions = {
        "draft": ["approved", "cancelled"],
        "approved": ["in_progress", "cancelled"],
        "in_progress": ["completed", "cancelled"],
        "completed": [],
        "cancelled": [],
    }
    return new in valid_transitions.get(current, [])


# BILLING-SUMMARY: GET /admin/patients/{patient_id}/billing-summary
@router.get(
    "/patients/{patient_id}/billing-summary",
    tags=["Planes de Tratamiento"],
    summary="Resumen completo de facturación del paciente: turnos + billing agrupados por tratamiento",
)
async def get_patient_billing_summary(
    patient_id: int,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Retorna resumen completo de facturación del paciente: turnos + billing agrupados por tratamiento."""
    tenant_id = resolved_tenant_id

    # 1. Fetch all appointments with billing data
    rows = await db.pool.fetch(
        """
        SELECT
            a.id as appointment_id,
            a.appointment_datetime,
            a.status as appointment_status,
            a.appointment_type,
            COALESCE(tt.name, a.appointment_type, 'Consulta') as treatment_name,
            tt.base_price,
            tt.code as treatment_code,
            COALESCE(a.billing_amount, tt.base_price, 0) as billing_amount,
            a.billing_installments,
            a.billing_notes,
            a.payment_status,
            a.payment_receipt_data,
            a.plan_item_id,
            prof.first_name || ' ' || COALESCE(prof.last_name, '') as professional_name,
            a.duration_minutes
        FROM appointments a
        LEFT JOIN treatment_types tt ON tt.code = a.appointment_type AND tt.tenant_id = $1
        LEFT JOIN professionals prof ON prof.id = a.professional_id AND prof.tenant_id = $1
        WHERE a.tenant_id = $1
          AND a.patient_id = $2
          AND a.status NOT IN ('cancelled', 'deleted')
        ORDER BY a.appointment_datetime DESC
        """,
        tenant_id,
        patient_id,
    )

    # 2. Build appointments list (keys match frontend AppointmentBilling interface)
    appointments = []
    for row in rows:
        receipt = row["payment_receipt_data"]
        if isinstance(receipt, str):
            try:
                receipt = json.loads(receipt)
            except Exception:
                receipt = None

        appointments.append(
            {
                "id": str(row["appointment_id"]),
                "scheduled_at": row["appointment_datetime"].isoformat()
                if row["appointment_datetime"]
                else None,
                "status": row["appointment_status"],
                "treatment_code": row["treatment_code"] or row["appointment_type"],
                "treatment_name": row["treatment_name"],
                "base_price": float(row["base_price"]) if row["base_price"] else 0,
                "billing_amount": float(row["billing_amount"] or 0),
                "billing_installments": row["billing_installments"],
                "billing_notes": row["billing_notes"],
                "payment_status": row["payment_status"] or "pending",
                "payment_receipt_data": receipt,
                "professional_name": (row["professional_name"] or "").strip(),
                "plan_item_id": str(row["plan_item_id"])
                if row["plan_item_id"]
                else None,
                "duration_minutes": row["duration_minutes"] or 30,
            }
        )

    # 3. Group by treatment — include full appointment objects (not just IDs)
    groups_map = {}
    for apt in appointments:
        code = apt["treatment_code"] or "sin_tipo"
        if code not in groups_map:
            groups_map[code] = {
                "treatment_code": code,
                "treatment_name": apt["treatment_name"],
                "base_price": apt["base_price"],
                "appointments": [],
                "total_billed": 0.0,
                "total_paid": 0.0,
                "total_pending": 0.0,
            }
        g = groups_map[code]
        g["appointments"].append(apt)
        g["total_billed"] += apt["billing_amount"]
        if apt["payment_status"] == "paid":
            g["total_paid"] += apt["billing_amount"]

    for g in groups_map.values():
        g["total_pending"] = g["total_billed"] - g["total_paid"]
        g["total_billed"] = round(g["total_billed"], 2)
        g["total_paid"] = round(g["total_paid"], 2)
        g["total_pending"] = round(g["total_pending"], 2)

    treatment_groups = sorted(
        groups_map.values(), key=lambda x: x["total_billed"], reverse=True
    )

    # 4. Totals
    total_estimated = sum(a["billing_amount"] for a in appointments)
    total_paid = sum(
        a["billing_amount"] for a in appointments if a["payment_status"] == "paid"
    )

    # 5. Check active plan
    active_plan = await db.pool.fetchrow(
        "SELECT id FROM treatment_plans WHERE tenant_id=$1 AND patient_id=$2 AND status IN ('draft','approved','in_progress') ORDER BY created_at DESC LIMIT 1",
        tenant_id,
        patient_id,
    )

    # 6. Patient email
    patient_email_row = await db.pool.fetchrow(
        "SELECT email FROM patients WHERE id=$1 AND tenant_id=$2",
        patient_id,
        tenant_id,
    )
    patient_email = (
        patient_email_row["email"]
        if patient_email_row and patient_email_row["email"]
        else None
    )

    return {
        "appointments": appointments,
        "treatment_groups": treatment_groups,
        "global_total_estimated": round(total_estimated, 2),
        "global_total_paid": round(total_paid, 2),
        "global_total_pending": round(total_estimated - total_paid, 2),
        "has_active_plan": active_plan is not None,
        "active_plan_id": str(active_plan["id"]) if active_plan else None,
        "patient_email": patient_email,
    }


# EP-NEW-02: POST /admin/patients/{patient_id}/generate-plan-from-appointments
class GeneratePlanFromAppointmentsBody(BaseModel):
    name: Optional[str] = None
    professional_id: Optional[int] = None
    treatment_codes: Optional[List[str]] = (
        None  # Filter: only include these treatment types
    )


@router.post(
    "/patients/{patient_id}/generate-plan-from-appointments",
    tags=["Planes de Tratamiento"],
    summary="EP-NEW-02: Genera un plan de tratamiento automáticamente desde los turnos del paciente",
    status_code=status.HTTP_201_CREATED,
)
async def generate_plan_from_appointments(
    patient_id: int,
    payload: GeneratePlanFromAppointmentsBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Genera un plan de tratamiento atómicamente a partir de los turnos existentes:
    - Crea treatment_plan (status='draft')
    - Agrupa por treatment_type → un treatment_plan_item por grupo
    - Vincula cada appointment a su item (SET plan_item_id)
    - Migra solo pagos verificados a treatment_plan_payments + accounting_transactions
    """
    tenant_id = resolved_tenant_id

    # 1. Fetch UNLINKED appointments (not yet in any plan)
    rows = await db.pool.fetch(
        """
        SELECT
            a.id as appointment_id,
            a.appointment_datetime,
            a.appointment_type,
            COALESCE(tt.name, a.appointment_type, 'Consulta') as treatment_name,
            tt.base_price,
            tt.code as treatment_code,
            COALESCE(a.billing_amount, tt.base_price, 0) as billing_amount,
            a.payment_status,
            a.payment_receipt_data
        FROM appointments a
        LEFT JOIN treatment_types tt ON tt.code = a.appointment_type AND tt.tenant_id = $1
        WHERE a.tenant_id = $1
          AND a.patient_id = $2
          AND a.status NOT IN ('cancelled', 'deleted')
          AND a.plan_item_id IS NULL
        ORDER BY a.appointment_datetime ASC
        """,
        tenant_id,
        patient_id,
    )

    if not rows:
        raise HTTPException(
            status_code=422,
            detail="No hay turnos sin presupuesto asignado. Todos los turnos ya están vinculados a un presupuesto existente.",
        )

    # 3. Preparar appointments parseando JSONB defensivamente
    appointments = []
    for row in rows:
        receipt = row["payment_receipt_data"]
        if isinstance(receipt, str):
            try:
                receipt = json.loads(receipt)
            except Exception:
                receipt = None

        appointments.append(
            {
                "id": str(row["appointment_id"]),
                "treatment_code": row["treatment_code"]
                or row["appointment_type"]
                or "sin_tipo",
                "treatment_name": row["treatment_name"],
                "base_price": float(row["base_price"]) if row["base_price"] else 0.0,
                "billing_amount": float(row["billing_amount"] or 0),
                "payment_status": row["payment_status"] or "pending",
                "payment_receipt": receipt,
            }
        )

    # 3b. Filter by selected treatment_codes if provided
    if payload.treatment_codes:
        selected = set(payload.treatment_codes)
        appointments = [a for a in appointments if a["treatment_code"] in selected]
        if not appointments:
            raise HTTPException(
                status_code=422,
                detail="Ninguno de los tratamientos seleccionados tiene turnos sin asignar",
            )

    # 4. Agrupar por treatment_code
    groups_map: Dict[str, Dict] = {}
    for apt in appointments:
        code = apt["treatment_code"]
        if code not in groups_map:
            groups_map[code] = {
                "treatment_code": code,
                "treatment_name": apt["treatment_name"],
                "base_price": apt["base_price"],
                "appointments": [],
                "total_billed": 0.0,
            }
        groups_map[code]["appointments"].append(apt)
        groups_map[code]["total_billed"] += apt["billing_amount"]

    # 5. Auto-nombre si no se provee
    plan_name = payload.name
    if not plan_name:
        patient_row = await db.pool.fetchrow(
            "SELECT first_name, last_name FROM patients WHERE id=$1 AND tenant_id=$2",
            patient_id,
            tenant_id,
        )
        if patient_row:
            full_name = (
                f"{patient_row['first_name']} {patient_row['last_name'] or ''}".strip()
            )
            plan_name = f"Tratamiento de {full_name}"
        else:
            plan_name = f"Tratamiento generado"

    estimated_total = sum(a["billing_amount"] for a in appointments)
    plan_id = uuid.uuid4()
    items_created = 0
    payments_migrated = 0

    # 6. Transacción atómica: crear plan + items + vincular appointments + migrar pagos
    async with db.pool.acquire() as conn:
        async with conn.transaction():
            # 6a. Crear treatment_plan
            await conn.execute(
                """
                INSERT INTO treatment_plans
                (id, tenant_id, patient_id, professional_id, name, status, estimated_total, created_at, updated_at)
                VALUES ($1, $2, $3, $4, $5, 'draft', $6, NOW(), NOW())
                """,
                plan_id,
                tenant_id,
                patient_id,
                payload.professional_id,
                plan_name,
                estimated_total,
            )

            # 6b. Por cada grupo crear un treatment_plan_item y vincular sus appointments
            for sort_order, (code, group) in enumerate(groups_map.items()):
                item_id = uuid.uuid4()
                estimated_price = (
                    group["total_billed"]
                    if group["total_billed"] > 0
                    else group["base_price"]
                )

                await conn.execute(
                    """
                    INSERT INTO treatment_plan_items
                    (id, plan_id, tenant_id, treatment_type_code, custom_description,
                     estimated_price, approved_price, status, sort_order, created_at, updated_at)
                    VALUES ($1, $2, $3, $4, $5, $6, NULL, 'pending', $7, NOW(), NOW())
                    """,
                    item_id,
                    plan_id,
                    tenant_id,
                    code if code != "sin_tipo" else None,
                    group["treatment_name"],
                    estimated_price,
                    sort_order,
                )
                items_created += 1

                # 6c. Vincular cada appointment al item
                for apt in group["appointments"]:
                    await conn.execute(
                        "UPDATE appointments SET plan_item_id = $1 WHERE id = $2 AND tenant_id = $3",
                        item_id,
                        uuid.UUID(apt["id"]),
                        tenant_id,
                    )

                # 6d. Migrar pagos verificados (sin duplicar)
                for apt in group["appointments"]:
                    receipt = apt["payment_receipt"]
                    if not receipt:
                        continue

                    receipt_status = receipt.get("status", "")
                    if receipt_status not in ("verified", "verified_manual"):
                        continue

                    apt_id = apt["id"]

                    # Check duplicado por traceability note
                    existing_payment = await conn.fetchval(
                        "SELECT id FROM treatment_plan_payments WHERE plan_id=$1 AND notes LIKE $2",
                        plan_id,
                        f"%migrated:apt:{apt_id}%",
                    )
                    if existing_payment:
                        continue

                    # Monto: preferir amount_detected del comprobante, fallback a billing_amount
                    amount_detected = receipt.get("amount_detected") or receipt.get(
                        "amount"
                    )
                    payment_amount = (
                        float(amount_detected)
                        if amount_detected
                        else apt["billing_amount"]
                    )

                    if payment_amount <= 0:
                        continue

                    payment_id = uuid.uuid4()
                    await conn.execute(
                        """
                        INSERT INTO treatment_plan_payments
                        (id, plan_id, plan_item_id, tenant_id, amount, payment_method,
                         notes, created_at, updated_at)
                        VALUES ($1, $2, $3, $4, $5, 'transfer', $6, NOW(), NOW())
                        """,
                        payment_id,
                        plan_id,
                        item_id,
                        tenant_id,
                        payment_amount,
                        f"migrated:apt:{apt_id}",
                    )

                    # 6e. Sync a accounting_transactions (si no existe)
                    existing_tx = await conn.fetchval(
                        "SELECT id FROM accounting_transactions WHERE reference_id=$1 AND tenant_id=$2",
                        str(payment_id),
                        tenant_id,
                    )
                    if not existing_tx:
                        tx_id = uuid.uuid4()
                        await conn.execute(
                            """
                            INSERT INTO accounting_transactions
                            (id, tenant_id, patient_id, amount, type, category,
                             description, reference_id, transaction_date, created_at)
                            VALUES ($1, $2, $3, $4, 'income', 'treatment_payment',
                                    $5, $6, NOW(), NOW())
                            """,
                            tx_id,
                            tenant_id,
                            patient_id,
                            payment_amount,
                            f"Pago migrado desde turno {apt_id}",
                            str(payment_id),
                        )

                    payments_migrated += 1

    # 7. Emitir socket event (non-fatal)
    try:
        await emit_treatment_plan_event(
            "TREATMENT_PLAN_CREATED",
            {
                "plan_id": str(plan_id),
                "patient_id": patient_id,
                "tenant_id": tenant_id,
                "name": plan_name,
                "status": "draft",
                "source": "generate_from_appointments",
            },
            request,
        )
    except Exception as e:
        logger.warning(f"⚠️ Socket emit TREATMENT_PLAN_CREATED falló (non-fatal): {e}")

    return {
        "status": "created",
        "plan_id": str(plan_id),
        "items_created": items_created,
        "payments_migrated": payments_migrated,
        "estimated_total": round(estimated_total, 2),
    }


# EP-NEW-05: POST /admin/treatment-plans/{plan_id}/sync-appointments
@router.post(
    "/treatment-plans/{plan_id}/sync-appointments",
    tags=["Planes de Tratamiento"],
    summary="EP-NEW-05: Sincronizar turnos no vinculados al plan existente",
)
async def sync_appointments_to_plan(
    plan_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Detecta turnos del paciente que NO están vinculados a ningún plan y los agrega
    como items al plan existente. Migra pagos verificados de esos turnos.
    """
    tenant_id = resolved_tenant_id

    try:
        plan_uuid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de plan inválido")

    # 1. Get plan + patient_id
    plan = await db.pool.fetchrow(
        "SELECT id, patient_id, status FROM treatment_plans WHERE id=$1 AND tenant_id=$2",
        plan_uuid,
        tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    if plan["status"] in ("cancelled", "completed"):
        raise HTTPException(
            status_code=422, detail="No se puede sincronizar un plan finalizado"
        )

    patient_id = plan["patient_id"]

    # 2. Fetch unlinked appointments (no plan_item_id)
    rows = await db.pool.fetch(
        """
        SELECT
            a.id as appointment_id,
            a.appointment_type,
            COALESCE(tt.name, a.appointment_type, 'Consulta') as treatment_name,
            tt.base_price,
            tt.code as treatment_code,
            COALESCE(a.billing_amount, tt.base_price, 0) as billing_amount,
            a.payment_status,
            a.payment_receipt_data
        FROM appointments a
        LEFT JOIN treatment_types tt ON tt.code = a.appointment_type AND tt.tenant_id = $1
        WHERE a.tenant_id = $1
          AND a.patient_id = $2
          AND a.status NOT IN ('cancelled', 'deleted')
          AND a.plan_item_id IS NULL
        ORDER BY a.appointment_datetime ASC
        """,
        tenant_id,
        patient_id,
    )

    if not rows:
        return {
            "status": "no_changes",
            "message": "Todos los turnos ya están vinculados al presupuesto",
        }

    # 3. Parse appointments
    appointments = []
    for row in rows:
        receipt = row["payment_receipt_data"]
        if isinstance(receipt, str):
            try:
                receipt = json.loads(receipt)
            except Exception:
                receipt = None
        appointments.append(
            {
                "id": str(row["appointment_id"]),
                "treatment_code": row["treatment_code"]
                or row["appointment_type"]
                or "sin_tipo",
                "treatment_name": row["treatment_name"],
                "base_price": float(row["base_price"]) if row["base_price"] else 0.0,
                "billing_amount": float(row["billing_amount"] or 0),
                "payment_status": row["payment_status"] or "pending",
                "payment_receipt": receipt,
            }
        )

    # 4. Group by treatment_code
    groups_map: Dict[str, Dict] = {}
    for apt in appointments:
        code = apt["treatment_code"]
        if code not in groups_map:
            groups_map[code] = {
                "treatment_code": code,
                "treatment_name": apt["treatment_name"],
                "base_price": apt["base_price"],
                "appointments": [],
                "total_billed": 0.0,
            }
        groups_map[code]["appointments"].append(apt)
        groups_map[code]["total_billed"] += apt["billing_amount"]

    # 5. Check which treatment codes already exist as items in this plan
    existing_items = await db.pool.fetch(
        "SELECT id, treatment_type_code FROM treatment_plan_items WHERE plan_id=$1 AND tenant_id=$2",
        plan_uuid,
        tenant_id,
    )
    existing_codes = {ei["treatment_type_code"]: ei["id"] for ei in existing_items}
    max_sort = await db.pool.fetchval(
        "SELECT COALESCE(MAX(sort_order), -1) FROM treatment_plan_items WHERE plan_id=$1",
        plan_uuid,
    )

    items_created = 0
    items_updated = 0
    payments_migrated = 0

    async with db.pool.acquire() as conn:
        async with conn.transaction():
            for code, group in groups_map.items():
                if code in existing_codes:
                    # Item already exists — just link appointments to it
                    item_id = existing_codes[code]
                    items_updated += 1
                else:
                    # Create new item
                    max_sort += 1
                    item_id = uuid.uuid4()
                    estimated_price = (
                        group["total_billed"]
                        if group["total_billed"] > 0
                        else group["base_price"]
                    )
                    await conn.execute(
                        """
                        INSERT INTO treatment_plan_items
                        (id, plan_id, tenant_id, treatment_type_code, custom_description,
                         estimated_price, approved_price, status, sort_order, created_at, updated_at)
                        VALUES ($1, $2, $3, $4, $5, $6, NULL, 'pending', $7, NOW(), NOW())
                        """,
                        item_id,
                        plan_uuid,
                        tenant_id,
                        code if code != "sin_tipo" else None,
                        group["treatment_name"],
                        estimated_price,
                        max_sort,
                    )
                    items_created += 1

                # Link appointments to item
                for apt in group["appointments"]:
                    await conn.execute(
                        "UPDATE appointments SET plan_item_id = $1 WHERE id = $2 AND tenant_id = $3",
                        item_id,
                        uuid.UUID(apt["id"]),
                        tenant_id,
                    )

                # Migrate verified payments
                for apt in group["appointments"]:
                    receipt = apt["payment_receipt"]
                    if not receipt:
                        continue
                    receipt_status = receipt.get("status", "")
                    if receipt_status not in ("verified", "verified_manual"):
                        continue

                    apt_id = apt["id"]
                    # Use full apt_id in both check and insert for reliable dedup
                    existing_payment = await conn.fetchval(
                        "SELECT id FROM treatment_plan_payments WHERE plan_id=$1 AND notes LIKE $2",
                        plan_uuid,
                        f"%apt:{apt_id}%",
                    )
                    if existing_payment:
                        continue

                    amount_detected = receipt.get("amount_detected") or receipt.get(
                        "amount"
                    )
                    payment_amount = (
                        float(amount_detected)
                        if amount_detected
                        else apt["billing_amount"]
                    )
                    if payment_amount <= 0:
                        continue

                    payment_id = uuid.uuid4()
                    await conn.execute(
                        """
                        INSERT INTO treatment_plan_payments
                        (id, plan_id, plan_item_id, tenant_id, amount, payment_method,
                         notes, created_at, updated_at)
                        VALUES ($1, $2, $3, $4, $5, 'transfer', $6, NOW(), NOW())
                        """,
                        payment_id,
                        plan_uuid,
                        item_id,
                        tenant_id,
                        payment_amount,
                        f"migrated:apt:{apt_id}",
                    )
                    payments_migrated += 1

            # Recalculate estimated_total
            new_total = await recalculate_plan_estimated_total(
                conn, plan_uuid, tenant_id
            )
            await conn.execute(
                "UPDATE treatment_plans SET estimated_total=$1, updated_at=NOW() WHERE id=$2",
                new_total,
                plan_uuid,
            )

    # Emit event
    try:
        await emit_appointment_event(
            "TREATMENT_PLAN_UPDATED",
            {"plan_id": str(plan_uuid), "tenant_id": tenant_id},
            request,
        )
    except Exception:
        pass

    return {
        "status": "synced",
        "items_created": items_created,
        "items_updated": items_updated,
        "payments_migrated": payments_migrated,
        "new_estimated_total": float(new_total),
    }


# EP-01: GET /admin/patients/{patient_id}/treatment-plans
@router.get(
    "/patients/{patient_id}/treatment-plans",
    tags=["Treatment Plans"],
    summary="EP-01: Lista planes de tratamiento de un paciente",
)
async def list_patient_treatment_plans(
    patient_id: int,
    request: Request,
    status: str = None,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Lista todos los planes de tratamiento de un paciente.
    Retorna cada plan con campos agregados: items_count, paid_total, pending_total.
    Acepta query param opcional ?status= para filtrar por estado.
    """
    tenant_id = resolved_tenant_id

    ALLOWED_PLAN_STATUSES = {
        "draft",
        "approved",
        "in_progress",
        "completed",
        "cancelled",
    }
    if status is not None and status not in ALLOWED_PLAN_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Estado inválido. Valores permitidos: {', '.join(sorted(ALLOWED_PLAN_STATUSES))}",
        )

    patient = await db.pool.fetchrow(
        "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    if status is not None:
        rows = await db.pool.fetch(
            """
            SELECT
                tp.id, tp.tenant_id, tp.patient_id, tp.professional_id, tp.name, tp.status,
                tp.estimated_total, tp.approved_total, tp.approved_by, tp.approved_at,
                tp.notes, tp.created_at, tp.updated_at,
                prof.first_name as professional_first_name,
                (SELECT COUNT(*) FROM treatment_plan_items tpi
                 WHERE tpi.plan_id = tp.id AND tpi.tenant_id = tp.tenant_id AND tpi.status != 'cancelled') as items_count,
                COALESCE((SELECT SUM(tp_pay.amount) FROM treatment_plan_payments tp_pay
                          WHERE tp_pay.plan_id = tp.id AND tp_pay.tenant_id = tp.tenant_id), 0) as paid_total
            FROM treatment_plans tp
            LEFT JOIN professionals prof ON tp.professional_id = prof.id
            WHERE tp.patient_id = $1 AND tp.tenant_id = $2 AND tp.status = $3
            ORDER BY tp.created_at DESC
            """,
            patient_id,
            tenant_id,
            status,
        )
    else:
        rows = await db.pool.fetch(
            """
            SELECT
                tp.id, tp.tenant_id, tp.patient_id, tp.professional_id, tp.name, tp.status,
                tp.estimated_total, tp.approved_total, tp.approved_by, tp.approved_at,
                tp.notes, tp.created_at, tp.updated_at,
                prof.first_name as professional_first_name,
                (SELECT COUNT(*) FROM treatment_plan_items tpi
                 WHERE tpi.plan_id = tp.id AND tpi.tenant_id = tp.tenant_id AND tpi.status != 'cancelled') as items_count,
                COALESCE((SELECT SUM(tp_pay.amount) FROM treatment_plan_payments tp_pay
                          WHERE tp_pay.plan_id = tp.id AND tp_pay.tenant_id = tp.tenant_id), 0) as paid_total
            FROM treatment_plans tp
            LEFT JOIN professionals prof ON tp.professional_id = prof.id
            WHERE tp.patient_id = $1 AND tp.tenant_id = $2
            ORDER BY tp.created_at DESC
            """,
            patient_id,
            tenant_id,
        )

    plans = []
    for row in rows:
        approved = float(row["approved_total"] or 0)
        paid = float(row["paid_total"]) if row["paid_total"] else 0
        pending = max(approved - paid, 0) if approved > 0 else 0

        plans.append(
            {
                "id": str(row["id"]),
                "tenant_id": row["tenant_id"],
                "patient_id": row["patient_id"],
                "professional_id": row["professional_id"],
                "name": row["name"],
                "status": row["status"],
                "estimated_total": float(row["estimated_total"]),
                "approved_total": row["approved_total"],
                "approved_by": row["approved_by"],
                "approved_at": row["approved_at"].isoformat()
                if row["approved_at"]
                else None,
                "notes": row["notes"],
                "created_at": row["created_at"].isoformat(),
                "updated_at": row["updated_at"].isoformat(),
                "items_count": row["items_count"],
                "paid_total": paid,
                "pending_total": pending,
                "professional_name": row["professional_first_name"],
            }
        )

    return plans


# EP-02: POST /admin/patients/{patient_id}/treatment-plans
@router.post(
    "/patients/{patient_id}/treatment-plans",
    tags=["Treatment Plans"],
    summary="EP-02: Crear nuevo plan de tratamiento",
    status_code=status.HTTP_201_CREATED,
)
async def create_treatment_plan(
    patient_id: int,
    payload: CreateTreatmentPlanBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Crea un nuevo plan de tratamiento para un paciente.
    Puede incluir items iniciales opcionales.
    """
    tenant_id = resolved_tenant_id

    patient = await db.pool.fetchrow(
        "SELECT id FROM patients WHERE id = $1 AND tenant_id = $2",
        patient_id,
        tenant_id,
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    if payload.professional_id:
        prof = await db.pool.fetchrow(
            "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2",
            payload.professional_id,
            tenant_id,
        )
        if not prof:
            raise HTTPException(status_code=400, detail="Profesional no válido")

    plan_id = uuid.uuid4()

    async with db.pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute(
                """
                INSERT INTO treatment_plans 
                (id, tenant_id, patient_id, professional_id, name, status, estimated_total, notes, created_at, updated_at)
                VALUES ($1, $2, $3, $4, $5, 'draft', 0, $6, NOW(), NOW())
                """,
                plan_id,
                tenant_id,
                patient_id,
                payload.professional_id,
                payload.name,
                payload.notes,
            )

            items_created = []
            if payload.items:
                max_order = await conn.fetchval(
                    "SELECT COALESCE(MAX(sort_order), -1) FROM treatment_plan_items WHERE plan_id = $1",
                    plan_id,
                )

                for idx, item in enumerate(payload.items):
                    item_id = uuid.uuid4()
                    sort_order = max_order + idx + 1

                    estimated_price = item.estimated_price
                    if item.treatment_type_code and estimated_price == 0:
                        base_price = await conn.fetchval(
                            "SELECT base_price FROM treatment_types WHERE code = $1 AND tenant_id = $2",
                            item.treatment_type_code,
                            tenant_id,
                        )
                        if base_price:
                            estimated_price = float(base_price)

                    await conn.execute(
                        """
                        INSERT INTO treatment_plan_items
                        (id, plan_id, tenant_id, treatment_type_code, custom_description, 
                         estimated_price, approved_price, status, sort_order, created_at, updated_at)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, NOW(), NOW())
                        """,
                        item_id,
                        plan_id,
                        tenant_id,
                        item.treatment_type_code,
                        item.custom_description,
                        estimated_price,
                        item.approved_price,
                        item.status or "pending",
                        sort_order,
                    )

                    items_created.append(
                        {
                            "id": str(item_id),
                            "plan_id": str(plan_id),
                            "tenant_id": tenant_id,
                            "treatment_type_code": item.treatment_type_code,
                            "custom_description": item.custom_description,
                            "estimated_price": estimated_price,
                            "approved_price": item.approved_price,
                            "status": item.status or "pending",
                            "sort_order": sort_order,
                        }
                    )

            totals = await recalculate_plan_totals(conn, str(plan_id), tenant_id)
            if totals["estimated_total"] > 0:
                await conn.execute(
                    "UPDATE treatment_plans SET estimated_total = $1 WHERE id = $2",
                    totals["estimated_total"],
                    plan_id,
                )

    plan = await db.pool.fetchrow(
        "SELECT * FROM treatment_plans WHERE id = $1", plan_id
    )
    prof_name = (
        await get_professional_name(db.pool, plan["professional_id"])
        if plan["professional_id"]
        else None
    )

    await emit_treatment_plan_event(
        "TREATMENT_PLAN_CREATED",
        {
            "plan_id": str(plan_id),
            "patient_id": patient_id,
            "tenant_id": tenant_id,
            "name": plan["name"],
            "status": plan["status"],
        },
        request,
    )

    return {
        "id": str(plan["id"]),
        "tenant_id": plan["tenant_id"],
        "patient_id": plan["patient_id"],
        "professional_id": plan["professional_id"],
        "name": plan["name"],
        "status": plan["status"],
        "estimated_total": float(plan["estimated_total"]),
        "approved_total": plan["approved_total"],
        "approved_by": plan["approved_by"],
        "approved_at": plan["approved_at"].isoformat() if plan["approved_at"] else None,
        "notes": plan["notes"],
        "created_at": plan["created_at"].isoformat(),
        "updated_at": plan["updated_at"].isoformat(),
        "items": items_created,
        "professional_name": prof_name,
    }


# EP-03: GET /admin/treatment-plans/{plan_id}
@router.get(
    "/treatment-plans/{plan_id}",
    tags=["Treatment Plans"],
    summary="EP-03: Detalle completo de plan de tratamiento",
)
async def get_treatment_plan_detail(
    plan_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Retorna el detalle completo de un plan: datos del plan + arrays de items y payments.
    """
    tenant_id = resolved_tenant_id

    try:
        plan_uuid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de plan inválido")

    plan = await db.pool.fetchrow(
        """
        SELECT tp.*, 
               p.first_name as patient_first_name, p.last_name as patient_last_name,
               prof.first_name as professional_first_name
        FROM treatment_plans tp
        LEFT JOIN patients p ON tp.patient_id = p.id
        LEFT JOIN professionals prof ON tp.professional_id = prof.id
        WHERE tp.id = $1 AND tp.tenant_id = $2
        """,
        plan_uuid,
        tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    items = await db.pool.fetch(
        """
        SELECT tpi.*, tt.name as treatment_name,
               (SELECT COUNT(*) FROM appointments a WHERE a.plan_item_id = tpi.id AND a.tenant_id = tpi.tenant_id) as appointments_count,
               (SELECT COALESCE(JSON_AGG(JSON_BUILD_OBJECT(
                   'id', a.id,
                   'datetime', a.appointment_datetime,
                   'status', a.status,
                   'type', COALESCE(a.appointment_type, 'Sin tipo')
               ) ORDER BY a.appointment_datetime), '[]'::json)
               FROM appointments a
               WHERE a.plan_item_id = tpi.id AND a.tenant_id = $2
               ) as appointments
        FROM treatment_plan_items tpi
        LEFT JOIN treatment_types tt ON tpi.treatment_type_code = tt.code AND tpi.tenant_id = tt.tenant_id
        WHERE tpi.plan_id = $1 AND tpi.tenant_id = $2
        ORDER BY tpi.sort_order
        """,
        plan_uuid,
        tenant_id,
    )

    items_list = []
    for item in items:
        raw_appointments = item["appointments"]
        if isinstance(raw_appointments, str):
            import json as _json

            appts = _json.loads(raw_appointments)
        else:
            appts = raw_appointments if raw_appointments is not None else []
        items_list.append(
            {
                "id": str(item["id"]),
                "plan_id": str(item["plan_id"]),
                "tenant_id": item["tenant_id"],
                "treatment_type_code": item["treatment_type_code"],
                "custom_description": item["custom_description"],
                "estimated_price": float(item["estimated_price"]),
                "approved_price": float(item["approved_price"])
                if item["approved_price"]
                else None,
                "status": item["status"],
                "sort_order": item["sort_order"],
                "created_at": item["created_at"].isoformat(),
                "updated_at": item["updated_at"].isoformat(),
                "treatment_name": item["treatment_name"],
                "appointments_count": item["appointments_count"] or 0,
                "appointments": appts,
            }
        )

    payments = await db.pool.fetch(
        """
        SELECT tpp.*, u.email as recorded_by_email
        FROM treatment_plan_payments tpp
        LEFT JOIN users u ON tpp.recorded_by = u.email
        WHERE tpp.plan_id = $1 AND tpp.tenant_id = $2
        ORDER BY tpp.payment_date DESC, tpp.created_at DESC
        """,
        plan_uuid,
        tenant_id,
    )

    payments_list = []
    for pay in payments:
        payments_list.append(
            {
                "id": str(pay["id"]),
                "plan_id": str(pay["plan_id"]),
                "tenant_id": pay["tenant_id"],
                "amount": float(pay["amount"]),
                "payment_method": pay["payment_method"],
                "payment_date": pay["payment_date"].isoformat()
                if pay["payment_date"]
                else None,
                "recorded_by": pay["recorded_by_email"] or pay["recorded_by"],
                "appointment_id": str(pay["appointment_id"])
                if pay["appointment_id"]
                else None,
                "receipt_data": pay["receipt_data"],
                "payment_receipt_data": pay["receipt_data"],
                "notes": pay["notes"],
                "created_at": pay["created_at"].isoformat(),
            }
        )

    approved = float(plan["approved_total"] or 0)
    paid = sum(float(p["amount"]) for p in payments_list)
    pending = max(approved - paid, 0) if approved > 0 else 0
    progress_pct = round(paid / approved * 100, 1) if approved > 0 else 0.0
    patient_name = (
        f"{plan['patient_first_name']} {plan['patient_last_name'] or ''}".strip()
    )

    return {
        "id": str(plan["id"]),
        "tenant_id": plan["tenant_id"],
        "patient_id": plan["patient_id"],
        "professional_id": plan["professional_id"],
        "name": plan["name"],
        "status": plan["status"],
        "estimated_total": float(plan["estimated_total"]),
        "approved_total": plan["approved_total"],
        "approved_by": plan["approved_by"],
        "approved_by_name": plan.get("approved_by") or None,
        "approved_at": plan["approved_at"].isoformat() if plan["approved_at"] else None,
        "notes": plan["notes"],
        "created_at": plan["created_at"].isoformat(),
        "updated_at": plan["updated_at"].isoformat(),
        "items": items_list,
        "payments": payments_list,
        "professional_name": plan["professional_first_name"],
        "patient_name": patient_name,
        "paid_total": paid,
        "pending_total": pending,
        "progress_pct": progress_pct,
    }


# EP-04: PATCH /admin/treatment-plans/{plan_id}  (also accepts PUT for backwards compat)
@router.patch(
    "/treatment-plans/{plan_id}",
    tags=["Treatment Plans"],
    summary="EP-04: Actualizar plan de tratamiento",
)
@router.put(
    "/treatment-plans/{plan_id}",
    tags=["Treatment Plans"],
    summary="EP-04: Actualizar plan de tratamiento (PUT alias)",
    include_in_schema=False,
)
async def update_treatment_plan(
    plan_id: str,
    payload: UpdateTreatmentPlanBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Actualiza un plan de tratamiento.
    Caso especial: status='approved' establece approved_by y approved_at.
    """
    tenant_id = resolved_tenant_id

    try:
        plan_uuid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de plan inválido")

    plan = await db.pool.fetchrow(
        "SELECT * FROM treatment_plans WHERE id = $1 AND tenant_id = $2",
        plan_uuid,
        tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    update_fields = []
    params = []
    idx = 1

    if payload.name is not None:
        update_fields.append(f"name = ${idx}")
        params.append(payload.name)
        idx += 1

    if payload.professional_id is not None:
        if payload.professional_id:
            prof = await db.pool.fetchrow(
                "SELECT id FROM professionals WHERE id = $1 AND tenant_id = $2",
                payload.professional_id,
                tenant_id,
            )
            if not prof:
                raise HTTPException(status_code=400, detail="Profesional no válido")
        update_fields.append(f"professional_id = ${idx}")
        params.append(payload.professional_id)
        idx += 1

    if payload.status is not None:
        if not validate_plan_status_transition(plan["status"], payload.status):
            raise HTTPException(
                status_code=422,
                detail=f"Transición de estado inválida: {plan['status']} -> {payload.status}",
            )

        if payload.status == "approved":
            update_fields.append(f"status = ${idx}")
            params.append(payload.status)
            idx += 1
            update_fields.append(f"approved_by = ${idx}")
            params.append(
                user_data.email
            )  # approved_by stores user email (VARCHAR column)
            idx += 1
            update_fields.append(f"approved_at = NOW()")
            if payload.approved_total is not None:
                update_fields.append(f"approved_total = ${idx}")
                params.append(payload.approved_total)
                idx += 1
            else:
                update_fields.append(f"approved_total = estimated_total")
        else:
            update_fields.append(f"status = ${idx}")
            params.append(payload.status)
            idx += 1

    if payload.approved_total is not None and payload.status != "approved":
        update_fields.append(f"approved_total = ${idx}")
        params.append(payload.approved_total)
        idx += 1

    # Merge budget config fields into notes when no new DB columns exist yet.
    # If `notes` was explicitly sent, use it as-is (caller already composed the string).
    # Otherwise, if any budget field is present, build a JSON metadata note.
    budget_fields_present = any(
        [
            payload.payment_conditions is not None,
            payload.discount_pct is not None,
            payload.discount_amount is not None,
            payload.installments is not None,
            payload.installments_amount is not None,
        ]
    )

    if payload.notes is not None:
        # Explicit notes wins; budget fields are ignored for storage (already in approved_total)
        update_fields.append(f"notes = ${idx}")
        params.append(payload.notes)
        idx += 1
    elif budget_fields_present:
        # Build a structured JSON note from budget config fields
        import json as _json

        budget_meta: dict = {}
        if payload.payment_conditions is not None:
            budget_meta["payment_conditions"] = payload.payment_conditions
        if payload.discount_pct is not None:
            budget_meta["discount_pct"] = payload.discount_pct
        if payload.discount_amount is not None:
            budget_meta["discount_amount"] = payload.discount_amount
        if payload.installments is not None:
            budget_meta["installments"] = payload.installments
        if payload.installments_amount is not None:
            budget_meta["installments_amount"] = payload.installments_amount
        # Merge with existing config if notes is already JSON
        existing_notes = plan["notes"] or ""
        try:
            existing_meta = (
                _json.loads(existing_notes)
                if existing_notes.strip().startswith("{")
                else {}
            )
        except Exception:
            existing_meta = {}
        existing_meta.update(budget_meta)
        merged_notes = _json.dumps(existing_meta, ensure_ascii=False)
        update_fields.append(f"notes = ${idx}")
        params.append(merged_notes)
        idx += 1

    if not update_fields:
        raise HTTPException(status_code=400, detail="No hay campos para actualizar")

    update_fields.append(f"updated_at = NOW()")
    params.append(plan_uuid)

    await db.pool.execute(
        f"UPDATE treatment_plans SET {', '.join(update_fields)} WHERE id = ${idx}",
        *params,
    )

    updated_plan = await db.pool.fetchrow(
        "SELECT * FROM treatment_plans WHERE id = $1", plan_uuid
    )
    prof_name = (
        await get_professional_name(db.pool, updated_plan["professional_id"])
        if updated_plan["professional_id"]
        else None
    )

    paid = await db.pool.fetchval(
        "SELECT COALESCE(SUM(amount), 0) FROM treatment_plan_payments WHERE plan_id = $1 AND tenant_id = $2",
        plan_uuid,
        tenant_id,
    )
    paid = float(paid) if paid else 0
    approved = float(updated_plan["approved_total"] or 0)
    pending = max(approved - paid, 0) if approved > 0 else 0
    items_count = await db.pool.fetchval(
        "SELECT COUNT(*) FROM treatment_plan_items WHERE plan_id = $1 AND status != 'cancelled'",
        plan_uuid,
    )

    await emit_treatment_plan_event(
        "TREATMENT_PLAN_UPDATED",
        {
            "plan_id": str(plan_uuid),
            "patient_id": updated_plan["patient_id"],
            "tenant_id": tenant_id,
            "status": updated_plan["status"],
            "changes": list(payload.model_dump(exclude_none=True).keys()),
        },
        request,
    )

    return {
        "id": str(updated_plan["id"]),
        "tenant_id": updated_plan["tenant_id"],
        "patient_id": updated_plan["patient_id"],
        "professional_id": updated_plan["professional_id"],
        "name": updated_plan["name"],
        "status": updated_plan["status"],
        "estimated_total": float(updated_plan["estimated_total"]),
        "approved_total": updated_plan["approved_total"],
        "approved_by": updated_plan["approved_by"],
        "approved_at": updated_plan["approved_at"].isoformat()
        if updated_plan["approved_at"]
        else None,
        "notes": updated_plan["notes"],
        "created_at": updated_plan["created_at"].isoformat(),
        "updated_at": updated_plan["updated_at"].isoformat(),
        "items_count": items_count,
        "paid_total": paid,
        "pending_total": pending,
        "professional_name": prof_name,
    }


# EP-05: DELETE /admin/treatment-plans/{plan_id}
@router.delete(
    "/treatment-plans/{plan_id}",
    tags=["Treatment Plans"],
    summary="EP-05: Soft-cancel plan de tratamiento",
)
async def delete_treatment_plan(
    plan_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Soft-cancel de un plan de tratamiento (status='cancelled').
    No permite cancelar planes con status='completed'.
    """
    tenant_id = resolved_tenant_id

    try:
        plan_uuid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de plan inválido")

    plan = await db.pool.fetchrow(
        "SELECT * FROM treatment_plans WHERE id = $1 AND tenant_id = $2",
        plan_uuid,
        tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    if plan["status"] == "completed":
        raise HTTPException(
            status_code=400, detail="No se puede cancelar un plan completado"
        )

    if plan["status"] == "cancelled":
        return {
            "status": "cancelled",
            "plan_id": str(plan_uuid),
            "message": "El plan ya estaba cancelado",
        }

    await db.pool.execute(
        "UPDATE treatment_plans SET status = 'cancelled', updated_at = NOW() WHERE id = $1",
        plan_uuid,
    )

    await emit_treatment_plan_event(
        "TREATMENT_PLAN_UPDATED",
        {
            "plan_id": str(plan_uuid),
            "patient_id": plan["patient_id"],
            "tenant_id": tenant_id,
            "name": plan["name"],
        },
        request,
    )

    return {
        "status": "cancelled",
        "plan_id": str(plan_uuid),
        "message": "Plan cancelado exitosamente",
    }


# ============================================
# TREATMENT PLAN ITEMS CRUD (EP-06 a EP-08)
# ============================================


async def recalculate_plan_estimated_total(
    pool, plan_id: uuid.UUID, tenant_id: int
) -> Decimal:
    """
    Recalcula el estimated_total del plan sumando COALESCE(approved_price, estimated_price)
    de todos los ítems no cancelados. Si existe precio aprobado, ese tiene prioridad.
    """
    total = await pool.fetchval(
        """
        SELECT COALESCE(SUM(COALESCE(approved_price, estimated_price)), 0)
        FROM treatment_plan_items
        WHERE plan_id = $1 AND tenant_id = $2 AND status != 'cancelled'
        """,
        plan_id,
        tenant_id,
    )
    return Decimal(str(total)) if total else Decimal("0")


@router.post(
    "/treatment-plans/{plan_id}/items",
    tags=["Planes de Tratamiento"],
    summary="EP-06: Agregar ítem a plan",
    status_code=status.HTTP_201_CREATED,
)
async def add_plan_item(
    plan_id: str,
    payload: AddPlanItemBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-06: Agrega un ítem a un plan de tratamiento existente.
    - Calcula sort_order automáticamente
    - Carga base_price de treatment_types si no se provee estimated_price
    - Recalcula estimated_total del plan
    - Valida que el plan no esté cancelled/completed
    """
    # Validar plan_id como UUID
    try:
        plan_uuid = uuid.UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de plan inválido")

    # Verificar que el plan existe y no está cancelled/completed
    plan = await db.pool.fetchrow(
        """
        SELECT id, tenant_id, status, name, patient_id
        FROM treatment_plans
        WHERE id = $1
        """,
        plan_uuid,
    )

    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    # Aislamiento por tenant
    if plan["tenant_id"] != resolved_tenant_id:
        raise HTTPException(status_code=403, detail="No tienes acceso a este plan")

    # Validar estado del plan
    if plan["status"] in ("cancelled", "completed"):
        raise HTTPException(
            status_code=422,
            detail=f"No se puede agregar ítems a un plan en estado '{plan['status']}'",
        )

    # Obtener base_price de treatment_types si no se provee
    estimated_price = payload.estimated_price
    if estimated_price is None:
        treatment_type = await db.pool.fetchrow(
            "SELECT base_price FROM treatment_types WHERE code = $1 AND tenant_id = $2",
            payload.treatment_type_code,
            resolved_tenant_id,
        )
        if treatment_type and treatment_type["base_price"]:
            estimated_price = Decimal(str(treatment_type["base_price"]))
        else:
            estimated_price = Decimal("0")

    # Calcular sort_order (siguiente número libre)
    max_sort_order = await db.pool.fetchval(
        """
        SELECT COALESCE(MAX(sort_order), 0)
        FROM treatment_plan_items
        WHERE plan_id = $1 AND tenant_id = $2
        """,
        plan_uuid,
        resolved_tenant_id,
    )
    new_sort_order = (max_sort_order or 0) + 1

    # Crear el ítem
    item_id = uuid.uuid4()
    await db.pool.execute(
        """
        INSERT INTO treatment_plan_items (
            id, plan_id, tenant_id, treatment_type_code, custom_description,
            estimated_price, status, sort_order, created_at, updated_at
        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, NOW(), NOW())
        """,
        item_id,
        plan_uuid,
        resolved_tenant_id,
        payload.treatment_type_code,
        payload.custom_description,
        estimated_price,
        "pending",
        new_sort_order,
    )

    # Recalcular estimated_total del plan
    new_total = await recalculate_plan_estimated_total(
        db.pool, plan_uuid, resolved_tenant_id
    )
    await db.pool.execute(
        """
        UPDATE treatment_plans SET estimated_total = $1, updated_at = NOW()
        WHERE id = $2 AND tenant_id = $3
        """,
        new_total,
        plan_uuid,
        resolved_tenant_id,
    )

    # Emitir evento Socket.IO
    await emit_appointment_event(
        "TREATMENT_PLAN_UPDATED",
        {"plan_id": str(plan_uuid), "tenant_id": resolved_tenant_id},
        request,
    )

    # Obtener el ítem creado para devolverlo
    item = await db.pool.fetchrow(
        """
        SELECT id, plan_id, tenant_id, treatment_type_code, custom_description,
               estimated_price, approved_price, status, sort_order, created_at, updated_at
        FROM treatment_plan_items WHERE id = $1
        """,
        item_id,
    )

    response = TreatmentPlanItemResponse(
        id=str(item["id"]),
        plan_id=str(item["plan_id"]),
        tenant_id=item["tenant_id"],
        treatment_type_code=item["treatment_type_code"],
        custom_description=item["custom_description"],
        estimated_price=item["estimated_price"],
        approved_price=item["approved_price"],
        status=ItemStatus(item["status"]),
        sort_order=item["sort_order"],
        appointments_count=0,
        created_at=item["created_at"],
        updated_at=item["updated_at"],
    )

    return {
        "status": "created",
        "item": response.model_dump(),
        "plan_estimated_total": float(new_total),
    }


@router.put(
    "/treatment-plan-items/{item_id}",
    tags=["Planes de Tratamiento"],
    summary="EP-07: Actualizar ítem de plan",
)
async def update_plan_item(
    item_id: str,
    payload: UpdatePlanItemBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-07: Actualiza un ítem de plan existente.
    - Update dinámico de campos
    - Recalcula estimated_total si cambia el precio
    """
    # Validar item_id como UUID
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de ítem inválido")

    # Verificar que el ítem existe
    item = await db.pool.fetchrow(
        """
        SELECT tpi.id, tpi.plan_id, tpi.tenant_id, tpi.status, tpi.estimated_price,
               tp.status as plan_status
        FROM treatment_plan_items tpi
        JOIN treatment_plans tp ON tpi.plan_id = tp.id
        WHERE tpi.id = $1
        """,
        item_uuid,
    )

    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")

    # Aislamiento por tenant
    if item["tenant_id"] != resolved_tenant_id:
        raise HTTPException(status_code=403, detail="No tienes acceso a este ítem")

    # Validar estado del ítem (no permite cambiar de completed a pending)
    if item["status"] == "completed" and payload.status == "pending":
        raise HTTPException(
            status_code=422,
            detail="No se puede revertir un ítem completado a pendiente",
        )

    # Validar estado del plan
    if item["plan_status"] in ("cancelled", "completed"):
        raise HTTPException(
            status_code=422,
            detail=f"No se puede modificar un ítem de un plan en estado '{item['plan_status']}'",
        )

    # Construir update dinámico
    update_fields = []
    update_values = []
    param_idx = 1

    if payload.treatment_type_code is not None:
        update_fields.append(f"treatment_type_code = ${param_idx}")
        update_values.append(payload.treatment_type_code)
        param_idx += 1

    if payload.custom_description is not None:
        update_fields.append(f"custom_description = ${param_idx}")
        update_values.append(payload.custom_description)
        param_idx += 1

    if payload.estimated_price is not None:
        update_fields.append(f"estimated_price = ${param_idx}")
        update_values.append(payload.estimated_price)
        param_idx += 1

    if payload.approved_price is not None:
        update_fields.append(f"approved_price = ${param_idx}")
        update_values.append(payload.approved_price)
        param_idx += 1

    if payload.status is not None:
        update_fields.append(f"status = ${param_idx}")
        update_values.append(payload.status.value)
        param_idx += 1

    if payload.sort_order is not None:
        update_fields.append(f"sort_order = ${param_idx}")
        update_values.append(payload.sort_order)
        param_idx += 1

    # Si no hay campos para actualizar, devolver el ítem actual
    if not update_fields:
        item_detail = await db.pool.fetchrow(
            """
            SELECT id, plan_id, tenant_id, treatment_type_code, custom_description,
                   estimated_price, approved_price, status, sort_order, created_at, updated_at
            FROM treatment_plan_items WHERE id = $1
            """,
            item_uuid,
        )
        response = TreatmentPlanItemResponse(
            id=str(item_detail["id"]),
            plan_id=str(item_detail["plan_id"]),
            tenant_id=item_detail["tenant_id"],
            treatment_type_code=item_detail["treatment_type_code"],
            custom_description=item_detail["custom_description"],
            estimated_price=item_detail["estimated_price"],
            approved_price=item_detail["approved_price"],
            status=ItemStatus(item_detail["status"]),
            sort_order=item_detail["sort_order"],
            appointments_count=0,
            created_at=item_detail["created_at"],
            updated_at=item_detail["updated_at"],
        )
        return {"item": response.model_dump()}

    # Agregar updated_at
    update_fields.append(f"updated_at = ${param_idx}")
    update_values.append(datetime.now())
    param_idx += 1

    # Ejecutar update
    update_values.append(item_uuid)
    await db.pool.execute(
        f"""
        UPDATE treatment_plan_items
        SET {", ".join(update_fields)}
        WHERE id = ${param_idx}
        """,
        *update_values,
    )

    # Recalcular estimated_total si cambió algún precio
    if payload.estimated_price is not None or payload.approved_price is not None:
        plan_uuid = item["plan_id"]
        new_total = await recalculate_plan_estimated_total(
            db.pool, plan_uuid, resolved_tenant_id
        )
        await db.pool.execute(
            """
            UPDATE treatment_plans SET estimated_total = $1, updated_at = NOW()
            WHERE id = $2 AND tenant_id = $3
            """,
            new_total,
            plan_uuid,
            resolved_tenant_id,
        )

    # Emitir evento Socket.IO
    await emit_appointment_event(
        "TREATMENT_PLAN_UPDATED",
        {"plan_id": str(item["plan_id"]), "tenant_id": resolved_tenant_id},
        request,
    )

    # Obtener ítem actualizado
    item_updated = await db.pool.fetchrow(
        """
        SELECT id, plan_id, tenant_id, treatment_type_code, custom_description,
               estimated_price, approved_price, status, sort_order, created_at, updated_at
        FROM treatment_plan_items WHERE id = $1
        """,
        item_uuid,
    )

    response = TreatmentPlanItemResponse(
        id=str(item_updated["id"]),
        plan_id=str(item_updated["plan_id"]),
        tenant_id=item_updated["tenant_id"],
        treatment_type_code=item_updated["treatment_type_code"],
        custom_description=item_updated["custom_description"],
        estimated_price=item_updated["estimated_price"],
        approved_price=item_updated["approved_price"],
        status=ItemStatus(item_updated["status"]),
        sort_order=item_updated["sort_order"],
        appointments_count=0,
        created_at=item_updated["created_at"],
        updated_at=item_updated["updated_at"],
    )

    # Recalcular total si hubo cambio
    plan_uuid = item["plan_id"]
    new_total = await recalculate_plan_estimated_total(
        db.pool, plan_uuid, resolved_tenant_id
    )

    return {
        "status": "updated",
        "item_id": str(item_uuid),
        "item": response.model_dump(),
        "plan_estimated_total": float(new_total),
    }


@router.delete(
    "/treatment-plan-items/{item_id}",
    tags=["Planes de Tratamiento"],
    summary="EP-08: Eliminar ítem de plan",
    status_code=status.HTTP_200_OK,
)
async def delete_plan_item(
    item_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-08: Elimina un ítem de plan (hard-delete).
    - No permite eliminar ítems completados (debe cancelarse primero)
    - Desvincula turnos asociados (SET NULL)
    - Recalcula estimated_total del plan
    """
    # Validar item_id como UUID
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de ítem inválido")

    # Verificar que el ítem existe
    item = await db.pool.fetchrow(
        """
        SELECT tpi.id, tpi.plan_id, tpi.tenant_id, tpi.status,
               tp.status as plan_status
        FROM treatment_plan_items tpi
        JOIN treatment_plans tp ON tpi.plan_id = tp.id
        WHERE tpi.id = $1
        """,
        item_uuid,
    )

    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")

    # Aislamiento por tenant
    if item["tenant_id"] != resolved_tenant_id:
        raise HTTPException(status_code=403, detail="No tienes acceso a este ítem")

    # No permite eliminar ítems completados
    if item["status"] == "completed":
        raise HTTPException(
            status_code=400,
            detail="No se puede eliminar un ítem completado. Cancélalo primero.",
        )

    plan_uuid = item["plan_id"]

    # Desvincular turnos asociados (SET NULL en appointments)
    await db.pool.execute(
        """
        UPDATE appointments
        SET plan_item_id = NULL
        WHERE plan_item_id = $1 AND tenant_id = $2
        """,
        item_uuid,
        resolved_tenant_id,
    )

    # Hard-delete del ítem
    await db.pool.execute(
        """
        DELETE FROM treatment_plan_items
        WHERE id = $1 AND tenant_id = $2
        """,
        item_uuid,
        resolved_tenant_id,
    )

    # Recalcular estimated_total del plan
    new_total = await recalculate_plan_estimated_total(
        db.pool, plan_uuid, resolved_tenant_id
    )
    await db.pool.execute(
        """
        UPDATE treatment_plans SET estimated_total = $1, updated_at = NOW()
        WHERE id = $2 AND tenant_id = $3
        """,
        new_total,
        plan_uuid,
        resolved_tenant_id,
    )

    # Emitir evento Socket.IO
    await emit_appointment_event(
        "TREATMENT_PLAN_UPDATED",
        {"plan_id": str(plan_uuid), "tenant_id": resolved_tenant_id},
        request,
    )

    return {
        "status": "deleted",
        "item_id": str(item_uuid),
        "plan_estimated_total": float(new_total),
    }


# =============================================================================
# EP-NEW-03: Generar PDF de presupuesto
# =============================================================================


@router.post(
    "/treatment-plans/{plan_id}/generate-pdf",
    tags=["Planes de Tratamiento"],
    summary="EP-NEW-03: Generar PDF de presupuesto",
)
async def generate_treatment_plan_pdf(
    plan_id: str,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Genera PDF de presupuesto y lo retorna para descarga."""
    from services.budget_service import generate_budget_pdf, gather_budget_data

    pdf_path = await generate_budget_pdf(db.pool, plan_id, resolved_tenant_id)
    if not pdf_path:
        raise HTTPException(
            status_code=404,
            detail="Plan no encontrado o error generando PDF",
        )

    data = await gather_budget_data(db.pool, plan_id, resolved_tenant_id)
    if data and data.get("patient", {}).get("name"):
        safe_name = data["patient"]["name"].replace(" ", "_")
        filename = f"Presupuesto_{safe_name}.pdf"
    else:
        filename = "Presupuesto.pdf"

    return FileResponse(pdf_path, media_type="application/pdf", filename=filename)


# =============================================================================
# EP-NEW-04: Enviar presupuesto por email
# =============================================================================


@router.post(
    "/treatment-plans/{plan_id}/send-email",
    tags=["Planes de Tratamiento"],
    summary="EP-NEW-04: Enviar presupuesto por email al paciente",
)
async def send_treatment_plan_email(
    plan_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Envía presupuesto por email al paciente con PDF adjunto."""
    import asyncio
    from email_service import email_service
    from services.budget_service import generate_budget_pdf, gather_budget_data

    body = await request.json()
    to_email = body.get("email") or body.get("to_email")
    if not to_email:
        raise HTTPException(status_code=422, detail="Email requerido")

    # Generate PDF (same pattern as digital records)
    try:
        pdf_path = await generate_budget_pdf(db.pool, plan_id, resolved_tenant_id)
    except Exception as e:
        logger.error(f"Error generating budget PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
    if not pdf_path:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    data = await gather_budget_data(db.pool, plan_id, resolved_tenant_id)
    if not data:
        raise HTTPException(status_code=404, detail="Datos del plan no encontrados")

    # Send email with PDF attachment (same pattern as digital records)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: email_service.send_budget_email(
                to_email=to_email,
                pdf_path=pdf_path,
                patient_name=data["patient"]["name"],
                clinic_name=data["clinic"]["name"],
            ),
        )
    except Exception as e:
        logger.error(f"Error sending budget email: {e}")
        raise HTTPException(status_code=500, detail=f"Error enviando email: {str(e)}")

    return {"success": True, "message": f"Presupuesto enviado a {to_email}"}


# =============================================================================
# EP-09: Registrar pago en plan de tratamiento (treatment-plan-billing)
# =============================================================================


@router.post(
    "/treatment-plans/{plan_id}/payments",
    tags=["Planes de Tratamiento"],
    summary="Registrar pago en plan de tratamiento",
    status_code=201,
)
async def register_plan_payment(
    plan_id: str,
    body: RegisterPaymentBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Registra un pago sobre el plan de tratamiento.
    Sincroniza con accounting_transactions.
    Auto-avanza el estado del plan de approved a in_progress.
    Envía email de confirmación al paciente (si tiene email).
    """
    # 1. Verificar plan existe y pertenece al tenant
    plan = await db.pool.fetchrow(
        """
        SELECT id, patient_id, status, approved_total, name
        FROM treatment_plans
        WHERE id = $1 AND tenant_id = $2
        """,
        plan_id,
        resolved_tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan de tratamiento no encontrado")

    # 2. Validar estado del plan (solo approved o in_progress)
    if plan["status"] not in ["approved", "in_progress"]:
        raise HTTPException(
            status_code=422,
            detail="Solo se pueden registrar pagos en planes aprobados o en progreso",
        )

    # 3. Si appointment_id se provee, validar que pertenece al mismo paciente
    if body.appointment_id:
        appointment = await db.pool.fetchrow(
            """
            SELECT id, patient_id FROM appointments
            WHERE id = $1 AND tenant_id = $2
            """,
            body.appointment_id,
            resolved_tenant_id,
        )
        if not appointment:
            raise HTTPException(status_code=404, detail="Turno no encontrado")
        if appointment["patient_id"] != plan["patient_id"]:
            raise HTTPException(
                status_code=422,
                detail="El turno no pertenece al paciente de este plan",
            )

    # 4. Generar payment_id
    payment_id = str(uuid.uuid4())

    # 5-8. Ejecutar en transacción atómica
    async with db.pool.acquire() as conn:
        async with conn.transaction():
            # 5. INSERT en treatment_plan_payments
            await conn.execute(
                """
                INSERT INTO treatment_plan_payments (
                    id, plan_id, tenant_id, amount, payment_method,
                    payment_date, recorded_by, appointment_id, receipt_data, notes,
                    created_at
                )
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, NOW())
                """,
                payment_id,
                plan_id,
                resolved_tenant_id,
                body.amount,
                body.payment_method.value
                if hasattr(body.payment_method, "value")
                else body.payment_method,
                body.payment_date or datetime.now(),
                body.recorded_by
                or getattr(user_data, "email", None)
                or str(getattr(user_data, "id", "")),
                body.appointment_id,
                json.dumps(body.receipt_data) if body.receipt_data else None,
                body.notes,
            )

            # 6. Sync con accounting_transactions
            accounting_tx_id = str(uuid.uuid4())
            await conn.execute(
                """
                INSERT INTO accounting_transactions (
                    id, tenant_id, patient_id, amount,
                    transaction_type, payment_method, status,
                    description, reference_id, reference_type,
                    created_at, updated_at
                )
                VALUES ($1, $2, $3, $4, 'payment', $5, 'completed', $6, $7, 'treatment_plan_payment', NOW(), NOW())
                """,
                accounting_tx_id,
                resolved_tenant_id,
                plan["patient_id"],
                body.amount,
                body.payment_method.value
                if hasattr(body.payment_method, "value")
                else body.payment_method,
                f"Pago plan: {plan['name']}",
                payment_id,
            )

            # 7. Actualizar accounting_transaction_id en el pago
            await conn.execute(
                """
                UPDATE treatment_plan_payments
                SET accounting_transaction_id = $1
                WHERE id = $2
                """,
                accounting_tx_id,
                payment_id,
            )

            # 8. Auto-avance de estado: approved → in_progress
            new_plan_status = plan["status"]
            if plan["status"] == "approved":
                new_plan_status = "in_progress"
                await conn.execute(
                    """
                    UPDATE treatment_plans
                    SET status = 'in_progress', updated_at = NOW()
                    WHERE id = $1 AND tenant_id = $2
                    """,
                    plan_id,
                    resolved_tenant_id,
                )

    # 9. Emitir evento Socket.IO
    await emit_appointment_event(
        "BILLING_UPDATED",
        {
            "plan_id": plan_id,
            "tenant_id": resolved_tenant_id,
            "payment_id": payment_id,
            "amount": float(body.amount),
            "plan_status": new_plan_status,
        },
        request,
    )

    # 10. Enviar email de confirmación (async, no rompe respuesta)
    try:
        asyncio.create_task(
            send_plan_payment_confirmation_email(
                tenant_id=resolved_tenant_id,
                patient_id=plan["patient_id"],
                payment_id=payment_id,
                db_pool=db.pool,
            )
        )
    except Exception as email_err:
        logger.warning(f"⚠️ Error al enviar email de confirmación de pago: {email_err}")

    # 11. Retornar respuesta
    return {
        "payment": {
            "id": payment_id,
            "plan_id": plan_id,
            "amount": float(body.amount),
            "payment_method": body.payment_method.value
            if hasattr(body.payment_method, "value")
            else body.payment_method,
            "payment_date": (body.payment_date or datetime.now()).isoformat()
            if hasattr((body.payment_date or datetime.now()), "isoformat")
            else str(body.payment_date or datetime.now()),
            "recorded_by": getattr(user_data, "email", None)
            or str(getattr(user_data, "id", "")),
            "appointment_id": body.appointment_id,
            "notes": body.notes,
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
        "plan_status": new_plan_status,
        "accounting_transaction_id": accounting_tx_id,
    }


# =============================================================================
# EP-10: Historial de pagos de un plan de tratamiento
# =============================================================================


@router.get(
    "/treatment-plans/{plan_id}/payments",
    tags=["Planes de Tratamiento"],
    summary="Historial de pagos de un plan de tratamiento",
)
async def get_plan_payments(
    plan_id: str,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Lista todos los pagos registrados para un plan de tratamiento.
    """
    # Verificar plan existe
    plan = await db.pool.fetchrow(
        "SELECT id FROM treatment_plans WHERE id = $1 AND tenant_id = $2",
        plan_id,
        resolved_tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan de tratamiento no encontrado")

    # Query pagos
    rows = await db.pool.fetch(
        """
        SELECT
            tpp.id,
            tpp.amount,
            tpp.payment_method,
            tpp.payment_date,
            tpp.notes,
            tpp.appointment_id,
            tpp.created_at,
            CONCAT(u.first_name, ' ', COALESCE(u.last_name, '')) AS recorded_by_name,
            a.appointment_type AS appointment_type
        FROM treatment_plan_payments tpp
        LEFT JOIN users u ON u.id = tpp.recorded_by
        LEFT JOIN appointments a ON a.id = tpp.appointment_id
        WHERE tpp.plan_id = $1 AND tpp.tenant_id = $2
        ORDER BY tpp.payment_date DESC
        """,
        plan_id,
        resolved_tenant_id,
    )

    return [dict(r) for r in rows]


# =============================================================================
# EP-11: Eliminar pago de plan de tratamiento
# =============================================================================


@router.delete(
    "/treatment-plan-payments/{payment_id}",
    tags=["Planes de Tratamiento"],
    summary="Eliminar pago de plan de tratamiento",
    status_code=204,
)
async def delete_plan_payment(
    payment_id: str,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Elimina un pago registrado de un plan de tratamiento.
    Solo CEO y secretary pueden eliminar pagos.
    Elimina la accounting_transaction relacionada.
    """
    # 1. Verificar rol (solo CEO o secretary)
    if user_data.role not in ["ceo", "secretary"]:
        raise HTTPException(
            status_code=403,
            detail="Solo CEO o secretary pueden eliminar pagos",
        )

    # 2. Verificar que el pago existe y pertenece al tenant
    payment = await db.pool.fetchrow(
        """
        SELECT id, plan_id, accounting_transaction_id
        FROM treatment_plan_payments
        WHERE id = $1 AND tenant_id = $2
        """,
        payment_id,
        resolved_tenant_id,
    )
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    # 3. Si tiene accounting_transaction_id, eliminar el registro de accounting
    if payment["accounting_transaction_id"]:
        await db.pool.execute(
            """
            DELETE FROM accounting_transactions
            WHERE id = $1 AND tenant_id = $2
            """,
            payment["accounting_transaction_id"],
            resolved_tenant_id,
        )

    # 4. Eliminar el pago
    await db.pool.execute(
        "DELETE FROM treatment_plan_payments WHERE id = $1",
        payment_id,
    )

    # 5. Emitir evento Socket.IO
    await emit_appointment_event(
        "BILLING_UPDATED",
        {
            "plan_id": payment["plan_id"],
            "tenant_id": resolved_tenant_id,
            "payment_id": payment_id,
            "action": "payment_deleted",
        },
        request,
    )

    return None


# =============================================================================
# EP-12: Vincular turno a ítem de plan de tratamiento
# =============================================================================


@router.put(
    "/appointments/{id}/link-plan-item",
    tags=["Planes de Tratamiento"],
    summary="Vincular turno a ítem de plan de tratamiento",
)
async def link_appointment_to_plan_item(
    id: str,
    body: LinkPlanItemBody,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Vincula o desvincula un turno a un ítem de plan de tratamiento.
    - Vincular: valida que el paciente del turno = paciente del plan, y el plan no está cancelado
    - Desvincular: set plan_item_id = NULL
    """
    appointment_uuid = id

    # 1. Verificar que el turno existe y pertenece al tenant
    appointment = await db.pool.fetchrow(
        """
        SELECT id, patient_id, plan_item_id
        FROM appointments
        WHERE id = $1 AND tenant_id = $2
        """,
        appointment_uuid,
        resolved_tenant_id,
    )
    if not appointment:
        raise HTTPException(status_code=404, detail="Turno no encontrado")

    # Caso 2: Desvincular (plan_item_id es None/null)
    if body.plan_item_id is None:
        # Solo desvincular si actualmente tiene un plan_item_id vinculado
        if appointment["plan_item_id"] is not None:
            await db.pool.execute(
                """
                UPDATE appointments
                SET plan_item_id = NULL, updated_at = NOW()
                WHERE id = $1 AND tenant_id = $2
                """,
                appointment_uuid,
                resolved_tenant_id,
            )

            # Emitir evento APPOINTMENT_UPDATED
            await emit_appointment_event(
                "APPOINTMENT_UPDATED",
                {
                    "id": appointment_uuid,
                    "plan_item_id": None,
                    "tenant_id": resolved_tenant_id,
                },
                request,
            )

        return {
            "status": "unlinked",
            "appointment_id": appointment_uuid,
            "plan_item_id": None,
        }

    # Caso 1: Vincular (plan_item_id no es None)
    plan_item_uuid = body.plan_item_id

    # 2. Verificar que el ítem existe y pertenece al tenant
    plan_item = await db.pool.fetchrow(
        """
        SELECT id, plan_id, treatment_type_code, custom_description
        FROM treatment_plan_items
        WHERE id = $1 AND tenant_id = $2
        """,
        plan_item_uuid,
        resolved_tenant_id,
    )
    if not plan_item:
        raise HTTPException(status_code=404, detail="Ítem de plan no encontrado")

    # 3. Verificar que el plan del ítem pertenece al mismo paciente que el turno
    plan = await db.pool.fetchrow(
        """
        SELECT id, patient_id, status, name
        FROM treatment_plans
        WHERE id = $1 AND tenant_id = $2
        """,
        plan_item["plan_id"],
        resolved_tenant_id,
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plan de tratamiento no encontrado")

    if plan["patient_id"] != appointment["patient_id"]:
        raise HTTPException(
            status_code=422,
            detail="El turno y el plan pertenecen a pacientes distintos",
        )

    # 4. Verificar que el plan no está cancelado
    if plan["status"] == "cancelled":
        raise HTTPException(
            status_code=422,
            detail="No se puede vincular a un ítem de plan cancelado",
        )

    # 5. UPDATE el turno con el plan_item_id
    await db.pool.execute(
        """
        UPDATE appointments
        SET plan_item_id = $1, updated_at = NOW()
        WHERE id = $2 AND tenant_id = $3
        """,
        plan_item_uuid,
        appointment_uuid,
        resolved_tenant_id,
    )

    # 6. Emitir APPOINTMENT_UPDATED
    await emit_appointment_event(
        "APPOINTMENT_UPDATED",
        {
            "id": appointment_uuid,
            "plan_item_id": plan_item_uuid,
            "tenant_id": resolved_tenant_id,
        },
        request,
    )

    return {
        "status": "linked",
        "appointment_id": appointment_uuid,
        "plan_item_id": plan_item_uuid,
    }


# =============================================
# Financial Command Center
# =============================================

# --- Pydantic Models for Financial Endpoints ---


class GenerateLiquidationBody(BaseModel):
    professional_id: int
    period_start: str
    period_end: str


class GenerateBulkLiquidationsBody(BaseModel):
    period_start: str
    period_end: str


class UpdateLiquidationStatusBody(BaseModel):
    status: str
    notes: Optional[str] = None


class CreatePayoutBody(BaseModel):
    amount: float
    payment_method: str
    payment_date: str
    reference_number: Optional[str] = None
    notes: Optional[str] = None


class UpsertCommissionBody(BaseModel):
    default_commission_pct: float
    per_treatment: Optional[List[dict]] = None


# --- EP-FC-01: POST /admin/liquidations/generate ---


@router.post(
    "/liquidations/generate",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Generate a liquidation for a specific professional and period",
)
async def generate_liquidation(
    body: GenerateLiquidationBody,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-01: Generates a liquidation snapshot for a professional in a given period.
    Idempotent: returns existing record if one already exists.
    Returns 201 for new records, 200 for existing.
    """
    # Validate dates
    try:
        period_start = date.fromisoformat(body.period_start)
        period_end = date.fromisoformat(body.period_end)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de fecha inválido. Usar YYYY-MM-DD.",
        )

    if period_start > period_end:
        raise HTTPException(
            status_code=400,
            detail="period_start debe ser menor o igual a period_end.",
        )

    if (period_end - period_start).days > 366:
        raise HTTPException(
            status_code=400,
            detail="El rango no puede superar 366 días.",
        )

    # Verify professional exists and belongs to tenant
    prof = await db.pool.fetchrow(
        """
        SELECT id, first_name, last_name FROM professionals
        WHERE id = $1 AND tenant_id = $2
        """,
        body.professional_id,
        tenant_id,
    )
    if not prof:
        raise HTTPException(
            status_code=404,
            detail=f"Profesional {body.professional_id} no encontrado en esta clínica.",
        )

    try:
        record = await liquidation_service.generate_liquidation(
            pool=db.pool,
            tenant_id=tenant_id,
            professional_id=body.professional_id,
            period_start=period_start,
            period_end=period_end,
            generated_by_email=user_data.email or "admin",
        )

        # Determine if it was newly created or existing
        created_at = record.get("created_at")
        is_new = False
        if created_at and isinstance(created_at, datetime):
            diff = (datetime.utcnow() - created_at).total_seconds()
            is_new = diff < 60

        status_code = 201 if is_new else 200
        return JSONResponse(content=record, status_code=status_code)

    except Exception as e:
        logger.error(f"Error generating liquidation: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-02: POST /admin/liquidations/generate-bulk ---


@router.post(
    "/liquidations/generate-bulk",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Generate liquidations for all active professionals in a period",
)
async def generate_bulk_liquidations(
    body: GenerateBulkLiquidationsBody,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-02: Generates liquidations for ALL active professionals in the period.
    Returns { generated_count, skipped_count, liquidations: [...] }.
    """
    # Validate dates
    try:
        period_start = date.fromisoformat(body.period_start)
        period_end = date.fromisoformat(body.period_end)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de fecha inválido. Usar YYYY-MM-DD.",
        )

    if period_start > period_end:
        raise HTTPException(
            status_code=400,
            detail="period_start debe ser menor o igual a period_end.",
        )

    if (period_end - period_start).days > 366:
        raise HTTPException(
            status_code=400,
            detail="El rango no puede superar 366 días.",
        )

    try:
        result = await liquidation_service.generate_bulk_liquidations(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=period_start,
            period_end=period_end,
            generated_by_email=user_data.email or "admin",
        )
        return JSONResponse(content=result, status_code=201)

    except Exception as e:
        logger.error(f"Error generating bulk liquidations: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-03: GET /admin/liquidations ---


@router.get(
    "/liquidations",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="List liquidations with filters and pagination",
)
async def list_liquidations(
    professional_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    period_start: Optional[str] = Query(None),
    period_end: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-03: Paginated list of liquidation records with optional filters.
    """
    # Parse optional dates
    parsed_start = None
    parsed_end = None
    if period_start:
        try:
            parsed_start = date.fromisoformat(period_start)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Formato de period_start inválido. Usar YYYY-MM-DD.",
            )
    if period_end:
        try:
            parsed_end = date.fromisoformat(period_end)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Formato de period_end inválido. Usar YYYY-MM-DD.",
            )

    limit = page_size
    offset = (page - 1) * page_size

    try:
        result = await liquidation_service.list_liquidations(
            pool=db.pool,
            tenant_id=tenant_id,
            professional_id=professional_id,
            status=status,
            period_start=parsed_start,
            period_end=parsed_end,
            limit=limit,
            offset=offset,
        )
        return result

    except Exception as e:
        logger.error(f"Error listing liquidations: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-04: GET /admin/liquidations/{liquidation_id} ---


@router.get(
    "/liquidations/{liquidation_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Get full liquidation detail with treatment groups and payouts",
)
async def get_liquidation_detail(
    liquidation_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-04: Returns full liquidation detail including treatment groups and payouts.
    """
    try:
        detail = await liquidation_service.get_liquidation_detail(
            pool=db.pool,
            tenant_id=tenant_id,
            liquidation_id=liquidation_id,
        )
        if not detail:
            raise HTTPException(
                status_code=404,
                detail=f"Liquidación {liquidation_id} no encontrada.",
            )
        return detail

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting liquidation detail: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-05: PATCH /admin/liquidations/{liquidation_id} ---


@router.patch(
    "/liquidations/{liquidation_id}",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Update liquidation status with audit trail",
)
async def update_liquidation_status(
    liquidation_id: int,
    body: UpdateLiquidationStatusBody,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-05: Updates liquidation status with validation of allowed transitions.
    Valid transitions: draft→generated→approved→paid
    """
    valid_statuses = ("draft", "generated", "approved", "paid")
    if body.status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Status inválido. Valores permitidos: {', '.join(valid_statuses)}.",
        )

    try:
        result = await liquidation_service.update_liquidation_status(
            pool=db.pool,
            tenant_id=tenant_id,
            liquidation_id=liquidation_id,
            new_status=body.status,
            user_email=user_data.email or "admin",
            notes=body.notes,
        )
        if not result:
            raise HTTPException(
                status_code=404,
                detail=f"Liquidación {liquidation_id} no encontrada.",
            )
        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating liquidation status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-06: POST /admin/liquidations/{liquidation_id}/payout ---


@router.post(
    "/liquidations/{liquidation_id}/payout",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Register a payout to a professional",
)
async def create_payout(
    liquidation_id: int,
    body: CreatePayoutBody,
    user_data=Depends(verify_admin_token),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-06: Creates a professional payout record.
    Auto-updates liquidation status to 'paid' if total payouts >= payout_amount.
    """
    # Validate amount
    if body.amount <= 0:
        raise HTTPException(
            status_code=400,
            detail="El monto debe ser mayor a 0.",
        )

    # Validate payment method
    valid_methods = ["transfer", "cash", "check"]
    if body.payment_method not in valid_methods:
        raise HTTPException(
            status_code=400,
            detail=f"Método de pago inválido. Valores permitidos: {', '.join(valid_methods)}.",
        )

    # Parse payment date
    try:
        payment_date = date.fromisoformat(body.payment_date)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de payment_date inválido. Usar YYYY-MM-DD.",
        )

    try:
        payout = await liquidation_service.create_payout(
            pool=db.pool,
            tenant_id=tenant_id,
            liquidation_id=liquidation_id,
            amount=body.amount,
            payment_method=body.payment_method,
            payment_date=payment_date,
            reference_number=body.reference_number,
            notes=body.notes,
            user_email=user_data.email or "admin",
        )
        return JSONResponse(content=payout, status_code=201)

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating payout: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-07: GET /admin/liquidations/{liquidation_id}/payouts ---


@router.get(
    "/liquidations/{liquidation_id}/payouts",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="List all payouts for a liquidation",
)
async def get_liquidation_payouts(
    liquidation_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-07: Returns all payouts associated with a liquidation.
    """
    try:
        # First verify the liquidation belongs to this tenant
        record = await db.pool.fetchrow(
            """
            SELECT id FROM liquidation_records
            WHERE id = $1 AND tenant_id = $2
            """,
            liquidation_id,
            tenant_id,
        )
        if not record:
            raise HTTPException(
                status_code=404,
                detail=f"Liquidación {liquidation_id} no encontrada.",
            )

        payouts = await db.pool.fetch(
            """
            SELECT id, liquidation_id, professional_id, amount,
                   payment_method, payment_date, reference_number,
                   notes, created_at
            FROM professional_payouts
            WHERE liquidation_id = $1 AND tenant_id = $2
            ORDER BY payment_date DESC
            """,
            liquidation_id,
            tenant_id,
        )

        result = []
        for p in payouts:
            result.append(
                {
                    "id": p["id"],
                    "liquidation_id": p["liquidation_id"],
                    "amount": float(p["amount"]),
                    "payment_method": p["payment_method"],
                    "payment_date": p["payment_date"].isoformat()
                    if p["payment_date"]
                    else None,
                    "reference_number": p["reference_number"],
                    "notes": p["notes"],
                    "created_at": p["created_at"].isoformat()
                    if p["created_at"]
                    else None,
                }
            )

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting payouts: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-08: GET /admin/financial-dashboard ---


@router.get(
    "/financial-dashboard",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Get comprehensive financial dashboard data",
)
async def get_financial_dashboard(
    period_start: str = Query(...),
    period_end: str = Query(...),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-08: Returns aggregated financial dashboard data including:
    - Summary (revenue, payouts, profit, pending)
    - Revenue by professional
    - Revenue by treatment
    - Daily cash flow
    - Month-over-month growth
    - Top treatments
    - Pending collections
    """
    # Validate dates
    try:
        p_start = date.fromisoformat(period_start)
        p_end = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de fecha inválido. Usar YYYY-MM-DD.",
        )

    if p_start > p_end:
        raise HTTPException(
            status_code=400,
            detail="period_start debe ser menor o igual a period_end.",
        )

    try:
        # Fetch all dashboard data in parallel would be ideal, but we do sequential for simplicity
        summary = await financial_dashboard_service.get_financial_summary(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=p_start,
            period_end=p_end,
        )
        revenue_by_professional = (
            await financial_dashboard_service.get_revenue_by_professional(
                pool=db.pool,
                tenant_id=tenant_id,
                period_start=p_start,
                period_end=p_end,
            )
        )
        revenue_by_treatment = (
            await financial_dashboard_service.get_revenue_by_treatment(
                pool=db.pool,
                tenant_id=tenant_id,
                period_start=p_start,
                period_end=p_end,
            )
        )
        daily_cash_flow = await financial_dashboard_service.get_daily_cash_flow(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=p_start,
            period_end=p_end,
        )
        mom_growth = await financial_dashboard_service.get_mom_growth(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=p_start,
            period_end=p_end,
        )
        top_treatments = await financial_dashboard_service.get_top_treatments(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=p_start,
            period_end=p_end,
        )
        pending_collections = await financial_dashboard_service.get_pending_collections(
            pool=db.pool,
            tenant_id=tenant_id,
            period_start=p_start,
            period_end=p_end,
        )

        return {
            "summary": summary,
            "revenue_by_professional": revenue_by_professional,
            "revenue_by_treatment": revenue_by_treatment,
            "daily_cash_flow": daily_cash_flow,
            "mom_growth": mom_growth,
            "top_treatments": top_treatments,
            "pending_collections": pending_collections,
        }

    except Exception as e:
        logger.error(f"Error getting financial dashboard: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-09: GET /admin/professionals/{professional_id}/commissions ---


@router.get(
    "/professionals/{professional_id}/commissions",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Get commission configuration for a professional",
)
async def get_professional_commissions(
    professional_id: int,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-09: Returns commission configuration for a professional.
    If no config exists, returns 0% with a warning field.
    """
    # Verify professional exists and belongs to tenant
    prof = await db.pool.fetchrow(
        """
        SELECT id, first_name, last_name FROM professionals
        WHERE id = $1 AND tenant_id = $2
        """,
        professional_id,
        tenant_id,
    )
    if not prof:
        raise HTTPException(
            status_code=404,
            detail=f"Profesional {professional_id} no encontrado en esta clínica.",
        )

    try:
        config = await liquidation_service.get_commission_config(
            pool=db.pool,
            tenant_id=tenant_id,
            professional_id=professional_id,
        )
        # Add professional name
        config["professional_name"] = (
            f"{prof['first_name']} {prof['last_name']}".strip()
        )
        return config

    except Exception as e:
        logger.error(f"Error getting commission config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-10: PUT /admin/professionals/{professional_id}/commissions ---


@router.put(
    "/professionals/{professional_id}/commissions",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Create or update commission configuration for a professional",
)
async def upsert_professional_commissions(
    professional_id: int,
    body: UpsertCommissionBody,
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-10: Creates or updates commission configuration for a professional.
    Validates percentages (0-100) and treatment codes.
    """
    # Validate default commission percentage
    if not (0 <= body.default_commission_pct <= 100):
        raise HTTPException(
            status_code=400,
            detail="default_commission_pct debe estar entre 0 y 100.",
        )

    # Validate per-treatment percentages
    if body.per_treatment:
        for entry in body.per_treatment:
            pct = entry.get("commission_pct", 0)
            if not (0 <= pct <= 100):
                raise HTTPException(
                    status_code=400,
                    detail=f"commission_pct para {entry.get('treatment_code', 'unknown')} debe estar entre 0 y 100.",
                )
            if not entry.get("treatment_code"):
                raise HTTPException(
                    status_code=400,
                    detail="Cada override de tratamiento requiere treatment_code.",
                )

        # Validate treatment codes exist
        treatment_codes = [e["treatment_code"] for e in body.per_treatment]
        existing_treatments = await db.pool.fetch(
            """
            SELECT code FROM treatment_types
            WHERE tenant_id = $1 AND code = ANY($2)
            """,
            tenant_id,
            treatment_codes,
        )
        existing_codes = {r["code"] for r in existing_treatments}
        missing_codes = set(treatment_codes) - existing_codes
        if missing_codes:
            raise HTTPException(
                status_code=400,
                detail=f"Tratamientos no encontrados: {', '.join(missing_codes)}",
            )

    # Verify professional exists and belongs to tenant
    prof = await db.pool.fetchrow(
        """
        SELECT id FROM professionals
        WHERE id = $1 AND tenant_id = $2
        """,
        professional_id,
        tenant_id,
    )
    if not prof:
        raise HTTPException(
            status_code=404,
            detail=f"Profesional {professional_id} no encontrado en esta clínica.",
        )

    try:
        result = await liquidation_service.upsert_commission_config(
            pool=db.pool,
            tenant_id=tenant_id,
            professional_id=professional_id,
            default_commission_pct=body.default_commission_pct,
            per_treatment=body.per_treatment or [],
        )
        return result

    except Exception as e:
        logger.error(f"Error upserting commission config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# --- EP-FC-11: GET /admin/reconciliation ---


@router.get(
    "/reconciliation",
    dependencies=[Depends(verify_admin_token)],
    tags=["Financial Command Center"],
    summary="Financial reconciliation report comparing patient payments vs professional payouts",
)
async def get_reconciliation(
    period_start: str = Query(...),
    period_end: str = Query(...),
    tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    EP-FC-11: Compares patient payments vs professional payouts and detects discrepancies.
    Discrepancies = paid appointments not included in any liquidation record.
    """
    # Validate dates
    try:
        p_start = date.fromisoformat(period_start)
        p_end = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de fecha inválido. Usar YYYY-MM-DD.",
        )

    if p_start > p_end:
        raise HTTPException(
            status_code=400,
            detail="period_start debe ser menor o igual a period_end.",
        )

    try:
        # 1. Total patient payments (paid appointments + treatment plan payments)
        total_patient_appt = await db.pool.fetchval(
            """
            SELECT COALESCE(SUM(billing_amount), 0)
            FROM appointments
            WHERE tenant_id = $1
              AND appointment_datetime >= $2
              AND appointment_datetime < ($3::date + INTERVAL '1 day')
              AND payment_status = 'paid'
              AND status NOT IN ('cancelled', 'deleted')
            """,
            tenant_id,
            p_start,
            p_end,
        )

        total_plan_payments = await db.pool.fetchval(
            """
            SELECT COALESCE(SUM(amount), 0)
            FROM treatment_plan_payments tpp
            JOIN treatment_plans tp ON tp.id = tpp.plan_id AND tp.tenant_id = $1
            WHERE tpp.tenant_id = $1
              AND tpp.payment_date >= $2
              AND tpp.payment_date < ($3::date + INTERVAL '1 day')
            """,
            tenant_id,
            p_start,
            p_end,
        )

        total_patient_payments = float(total_patient_appt or 0) + float(
            total_plan_payments or 0
        )

        # 2. Total professional payouts
        total_professional_payouts = await db.pool.fetchval(
            """
            SELECT COALESCE(SUM(amount), 0)
            FROM professional_payouts
            WHERE tenant_id = $1
              AND payment_date >= $2
              AND payment_date < ($3::date + INTERVAL '1 day')
            """,
            tenant_id,
            p_start,
            p_end,
        )
        total_professional_payouts = float(total_professional_payouts or 0)

        difference = total_patient_payments - total_professional_payouts

        # 3. Detect discrepancies: paid appointments not in any liquidation
        # Get all paid appointment IDs in the period
        paid_appts = await db.pool.fetch(
            """
            SELECT
                a.id AS appointment_id,
                a.appointment_datetime,
                a.billing_amount,
                a.payment_status,
                pat.first_name || ' ' || COALESCE(pat.last_name, '') AS patient_name,
                COALESCE(tt.name, a.appointment_type, 'Sin tratamiento') AS treatment_name,
                p.first_name || ' ' || COALESCE(p.last_name, '') AS professional_name
            FROM appointments a
            JOIN patients pat ON pat.id = a.patient_id AND pat.tenant_id = $1
            LEFT JOIN treatment_types tt ON tt.code = a.appointment_type AND tt.tenant_id = $1
            LEFT JOIN professionals p ON p.id = a.professional_id AND p.tenant_id = $1
            WHERE a.tenant_id = $1
              AND a.appointment_datetime >= $2
              AND a.appointment_datetime < ($3::date + INTERVAL '1 day')
              AND a.payment_status = 'paid'
              AND a.status NOT IN ('cancelled', 'deleted')
              AND a.billing_amount > 0
            """,
            tenant_id,
            p_start,
            p_end,
        )

        # Get all appointment IDs that ARE in liquidation records for the period
        liquidated_appt_ids = await db.pool.fetch(
            """
            SELECT DISTINCT a.id
            FROM appointments a
            JOIN liquidation_records lr ON lr.tenant_id = $1
                AND lr.professional_id = a.professional_id
                AND a.appointment_datetime >= lr.period_start
                AND a.appointment_datetime < (lr.period_end::date + INTERVAL '1 day')
            WHERE a.tenant_id = $1
              AND a.appointment_datetime >= $2
              AND a.appointment_datetime < ($3::date + INTERVAL '1 day')
              AND a.payment_status = 'paid'
              AND a.status NOT IN ('cancelled', 'deleted')
            """,
            tenant_id,
            p_start,
            p_end,
        )
        liquidated_ids = {r["id"] for r in liquidated_appt_ids}

        # Build discrepancies
        discrepancies = []
        for appt in paid_appts:
            if appt["appointment_id"] not in liquidated_ids:
                discrepancies.append(
                    {
                        "type": "payment_without_liquidation",
                        "appointment_id": appt["appointment_id"],
                        "patient_name": (appt["patient_name"] or "").strip(),
                        "treatment_name": appt["treatment_name"],
                        "amount": float(appt["billing_amount"]),
                        "appointment_date": appt["appointment_datetime"].isoformat()
                        if appt["appointment_datetime"]
                        else None,
                        "professional_name": (appt["professional_name"] or "").strip(),
                        "description": "Pago registrado sin liquidación asociada",
                    }
                )

        return {
            "period_start": period_start,
            "period_end": period_end,
            "total_patient_payments": round(total_patient_payments, 2),
            "total_professional_payouts": round(total_professional_payouts, 2),
            "difference": round(difference, 2),
            "discrepancies": discrepancies,
            "discrepancy_count": len(discrepancies),
        }

    except Exception as e:
        logger.error(f"Error in reconciliation: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")


# =============================================================================
# EP-FC-12: Generar PDF de liquidación profesional
# =============================================================================


@router.get(
    "/liquidations/{liquidation_id}/pdf",
    tags=["Liquidaciones"],
    summary="EP-FC-12: Generar PDF de liquidación profesional",
)
async def generate_liquidation_pdf_endpoint(
    liquidation_id: int,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Genera y sirve el PDF de una liquidación profesional con caché en disco."""
    from services.liquidation_pdf_service import (
        generate_liquidation_pdf,
        gather_liquidation_pdf_data,
    )

    # Verify the liquidation exists and belongs to tenant
    record = await db.pool.fetchrow(
        "SELECT id, status, professional_id FROM liquidation_records WHERE id = $1 AND tenant_id = $2",
        liquidation_id,
        resolved_tenant_id,
    )
    if not record:
        raise HTTPException(
            status_code=404,
            detail="Liquidación no encontrada",
        )

    # Check if cached PDF exists and is still valid
    pdf_path = f"/app/uploads/liquidations/{resolved_tenant_id}/{liquidation_id}.pdf"
    use_cache = False
    if os.path.exists(pdf_path):
        # Check if the PDF was generated after the last status change
        pdf_mtime = os.path.getmtime(pdf_path)
        # Compare with the record's created_at (or updated_at if available)
        record_ts = record.get("created_at")
        if record_ts:
            if hasattr(record_ts, "timestamp"):
                record_epoch = record_ts.timestamp()
            else:
                record_epoch = 0
            if pdf_mtime >= record_epoch:
                use_cache = True

    if not use_cache:
        # Generate fresh PDF
        pdf_path = await generate_liquidation_pdf(
            db.pool, liquidation_id, resolved_tenant_id
        )
        if not pdf_path:
            raise HTTPException(
                status_code=404,
                detail="Error generando PDF de liquidación",
            )

    # Build filename
    data = await gather_liquidation_pdf_data(
        db.pool, liquidation_id, resolved_tenant_id
    )
    if data:
        prof_name = data["professional"]["full_name"].replace(" ", "_")
        period_label = data["period"]["label"].replace(" ", "_")
        filename = f"Liquidacion_{prof_name}_{period_label}.pdf"
    else:
        filename = f"Liquidacion_{liquidation_id}.pdf"

    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=filename,
    )


# =============================================================================
# EP-FC-13: Enviar liquidación por email
# =============================================================================


class SendLiquidationEmailBody(BaseModel):
    to_email: Optional[str] = None
    message: Optional[str] = None


@router.post(
    "/liquidations/{liquidation_id}/send-email",
    tags=["Liquidaciones"],
    summary="EP-FC-13: Enviar liquidación por email al profesional",
)
async def send_liquidation_email_endpoint(
    liquidation_id: int,
    body: SendLiquidationEmailBody,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Envía el PDF de liquidación por email al profesional."""
    from services.liquidation_pdf_service import (
        generate_liquidation_pdf,
        gather_liquidation_pdf_data,
        render_liquidation_email_html,
    )

    # 1. Verify liquidation exists and belongs to tenant
    record = await db.pool.fetchrow(
        """
        SELECT lr.*, p.email AS professional_email
        FROM liquidation_records lr
        JOIN professionals p ON p.id = lr.professional_id AND p.tenant_id = lr.tenant_id
        WHERE lr.id = $1 AND lr.tenant_id = $2
        """,
        liquidation_id,
        resolved_tenant_id,
    )
    if not record:
        raise HTTPException(
            status_code=404,
            detail="Liquidación no encontrada",
        )

    # 2. Determine recipient email
    to_email = body.to_email
    if not to_email:
        to_email = record.get("professional_email")
        if not to_email:
            raise HTTPException(
                status_code=422,
                detail="No se encontró email del profesional. Proporcione to_email.",
            )

    # 3. Generate PDF (reuse cache if available)
    pdf_path = await generate_liquidation_pdf(
        db.pool, liquidation_id, resolved_tenant_id
    )
    if not pdf_path:
        raise HTTPException(
            status_code=500,
            detail="Error generando PDF para envío",
        )

    # 4. Gather data for email body
    data = await gather_liquidation_pdf_data(
        db.pool, liquidation_id, resolved_tenant_id
    )
    if not data:
        raise HTTPException(
            status_code=404,
            detail="Datos de liquidación no encontrados",
        )

    # Render email HTML body
    email_html = render_liquidation_email_html(data)

    # 5. Send email
    email_svc = EmailService()
    try:
        await asyncio.to_thread(
            lambda: email_svc.send_liquidation_email(
                to_email=to_email,
                pdf_path=pdf_path,
                professional_name=data["professional"]["full_name"],
                clinic_name=data["clinic"]["name"],
                period_label=data["period"]["label"],
                payout_amount=data["summary"]["payout_amount"],
                html_body=email_html,
            )
        )
    except Exception as e:
        logger.error(f"Error sending liquidation email: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error enviando email: {str(e)}",
        )

    # 6. Log email send in audit trail
    existing_notes = record.get("notes") or {}
    audit_trail = existing_notes.get("audit_trail", [])
    audit_trail.append(
        {
            "action": "email_sent",
            "to": to_email,
            "by": user_data.email if hasattr(user_data, "email") else "admin",
            "at": datetime.now(timezone.utc).isoformat(),
        }
    )
    existing_notes["audit_trail"] = audit_trail
    await db.pool.execute(
        """
        UPDATE liquidation_records
        SET notes = $1
        WHERE id = $2 AND tenant_id = $3
        """,
        existing_notes,
        liquidation_id,
        resolved_tenant_id,
    )

    return {
        "success": True,
        "message": f"Liquidación enviada a {to_email}",
    }


# ---------------------------------------------------------------------------
# T1.3 — Telegram Authorized Users CRUD
# ---------------------------------------------------------------------------


class TelegramAuthorizedUserCreate(BaseModel):
    display_name: str
    telegram_chat_id: str
    user_role: str = "ceo"


class TelegramAuthorizedUserUpdate(BaseModel):
    display_name: Optional[str] = None
    user_role: Optional[str] = None
    is_active: Optional[bool] = None


def _mask_chat_id(chat_id: str) -> str:
    """Return first 3 chars + *** + last 3 chars. Falls back to full masking for short values."""
    if len(chat_id) <= 6:
        return "***"
    return f"{chat_id[:3]}***{chat_id[-3:]}"


@router.get("/telegram/authorized-users", tags=["Telegram"])
async def list_telegram_authorized_users(
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Lista usuarios autorizados de Telegram para el tenant (chat_id enmascarado)."""
    rows = await db.fetch(
        """
        SELECT id, display_name, user_role, is_active, telegram_chat_id, created_at, updated_at
        FROM telegram_authorized_users
        WHERE tenant_id = $1
        ORDER BY created_at ASC
        """,
        resolved_tenant_id,
    )

    result = []
    for row in rows:
        decrypted = decrypt_value(row["telegram_chat_id"])
        result.append(
            {
                "id": row["id"],
                "display_name": row["display_name"],
                "user_role": row["user_role"],
                "is_active": row["is_active"],
                "telegram_chat_id": _mask_chat_id(decrypted),
                "created_at": row["created_at"].isoformat()
                if row["created_at"]
                else None,
                "updated_at": row["updated_at"].isoformat()
                if row["updated_at"]
                else None,
            }
        )

    return result


@router.post("/telegram/authorized-users", tags=["Telegram"])
async def create_telegram_authorized_user(
    payload: TelegramAuthorizedUserCreate,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Agrega un usuario autorizado de Telegram. Encripta el chat_id antes de guardar."""
    encrypted_chat_id = encrypt_value(payload.telegram_chat_id)

    try:
        row = await db.pool.fetchrow(
            """
            INSERT INTO telegram_authorized_users
                (tenant_id, display_name, telegram_chat_id, user_role, is_active, created_at, updated_at)
            VALUES ($1, $2, $3, $4, true, NOW(), NOW())
            RETURNING id, display_name, user_role, is_active, telegram_chat_id, created_at, updated_at
            """,
            resolved_tenant_id,
            payload.display_name,
            encrypted_chat_id,
            payload.user_role,
        )
    except asyncpg.UniqueViolationError:
        raise HTTPException(
            status_code=409,
            detail="Ya existe un usuario con ese Telegram chat ID para este tenant.",
        )

    decrypted = decrypt_value(row["telegram_chat_id"])
    return {
        "id": row["id"],
        "display_name": row["display_name"],
        "user_role": row["user_role"],
        "is_active": row["is_active"],
        "telegram_chat_id": _mask_chat_id(decrypted),
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


@router.put("/telegram/authorized-users/{user_id}", tags=["Telegram"])
async def update_telegram_authorized_user(
    user_id: int,
    payload: TelegramAuthorizedUserUpdate,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Actualiza display_name, user_role o is_active de un usuario autorizado."""
    # Build dynamic SET clause from non-None fields
    set_clauses = []
    values: List[Any] = []
    param_idx = 1

    if payload.display_name is not None:
        set_clauses.append(f"display_name = ${param_idx}")
        values.append(payload.display_name)
        param_idx += 1

    if payload.user_role is not None:
        set_clauses.append(f"user_role = ${param_idx}")
        values.append(payload.user_role)
        param_idx += 1

    if payload.is_active is not None:
        set_clauses.append(f"is_active = ${param_idx}")
        values.append(payload.is_active)
        param_idx += 1

    if not set_clauses:
        raise HTTPException(status_code=400, detail="No hay campos para actualizar.")

    set_clauses.append(f"updated_at = NOW()")

    values.append(user_id)  # $N   → WHERE id
    values.append(resolved_tenant_id)  # $N+1 → WHERE tenant_id

    query = f"""
        UPDATE telegram_authorized_users
        SET {", ".join(set_clauses)}
        WHERE id = ${param_idx} AND tenant_id = ${param_idx + 1}
        RETURNING id, display_name, user_role, is_active, telegram_chat_id, created_at, updated_at
    """

    row = await db.pool.fetchrow(query, *values)
    if not row:
        raise HTTPException(
            status_code=404,
            detail="Usuario no encontrado o no pertenece a este tenant.",
        )

    decrypted = decrypt_value(row["telegram_chat_id"])
    return {
        "id": row["id"],
        "display_name": row["display_name"],
        "user_role": row["user_role"],
        "is_active": row["is_active"],
        "telegram_chat_id": _mask_chat_id(decrypted),
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


@router.delete("/telegram/authorized-users/{user_id}", tags=["Telegram"])
async def delete_telegram_authorized_user(
    user_id: int,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """Elimina un usuario autorizado de Telegram verificando propiedad del tenant."""
    result = await db.pool.execute(
        """
        DELETE FROM telegram_authorized_users
        WHERE id = $1 AND tenant_id = $2
        """,
        user_id,
        resolved_tenant_id,
    )

    # asyncpg returns "DELETE N" — check count
    deleted = int(result.split()[-1])
    if deleted == 0:
        raise HTTPException(
            status_code=404,
            detail="Usuario no encontrado o no pertenece a este tenant.",
        )

    return {"success": True, "deleted_id": user_id}


# ---------------------------------------------------------------------------
# T1.4 — Telegram Integration Config
# ---------------------------------------------------------------------------


class TelegramConfigCreate(BaseModel):
    bot_token: str


@router.get("/telegram/config", tags=["Telegram"])
async def get_telegram_config(
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Devuelve configuración del bot de Telegram para el tenant.
    Si no está configurado: { configured: false }.
    Si está configurado: llama a getMe para obtener el username del bot.
    """
    token = await get_tenant_credential(resolved_tenant_id, TELEGRAM_BOT_TOKEN)
    if not token:
        return {"configured": False}

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(f"https://api.telegram.org/bot{token}/getMe")
            resp.raise_for_status()
            bot_info = resp.json().get("result", {})
    except Exception as e:
        logger.error(f"[Telegram] getMe failed for tenant {resolved_tenant_id}: {e}")
        raise HTTPException(
            status_code=502,
            detail="No se pudo verificar el bot con Telegram. Revisá el token.",
        )

    api_base = os.getenv("BASE_URL", "").rstrip("/")
    if not api_base:
        scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
        host = request.headers.get("x-forwarded-host", request.url.netloc)
        api_base = f"{scheme}://{host}"

    access_token = await get_tenant_credential(
        resolved_tenant_id, TELEGRAM_WEBHOOK_ACCESS_TOKEN
    )
    webhook_url = (
        f"{api_base}/telegram/webhook/{access_token}" if access_token else None
    )

    return {
        "configured": True,
        "bot_username": bot_info.get("username"),
        "webhook_url": webhook_url,
    }


@router.post("/telegram/config", tags=["Telegram"])
async def save_telegram_config(
    payload: TelegramConfigCreate,
    request: Request,
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Configura el bot de Telegram para el tenant:
    1. Valida el token llamando a getMe.
    2. Genera webhook_secret y access_token aleatorios.
    3. Guarda las tres credenciales encriptadas.
    4. Registra el webhook en Telegram.
    """
    # 1. Validate token first
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                f"https://api.telegram.org/bot{payload.bot_token}/getMe"
            )
            resp.raise_for_status()
            bot_info = resp.json().get("result", {})
    except Exception as e:
        logger.error(f"[Telegram] getMe validation failed: {e}")
        raise HTTPException(
            status_code=400,
            detail="Token inválido o Telegram no accesible. Verificá el bot token.",
        )

    # 2. Generate secrets
    webhook_secret = os.urandom(32).hex()  # 64 hex chars
    access_token = str(uuid.uuid4())

    # 3. Persist credentials (encrypted via save_tenant_credential)
    await save_tenant_credential(
        resolved_tenant_id, TELEGRAM_BOT_TOKEN, payload.bot_token, "telegram"
    )
    await save_tenant_credential(
        resolved_tenant_id, TELEGRAM_WEBHOOK_SECRET, webhook_secret, "telegram"
    )
    await save_tenant_credential(
        resolved_tenant_id, TELEGRAM_WEBHOOK_ACCESS_TOKEN, access_token, "telegram"
    )

    # 4. Build webhook URL
    api_base = os.getenv("BASE_URL", "").rstrip("/")
    if not api_base:
        scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
        host = request.headers.get("x-forwarded-host", request.url.netloc)
        api_base = f"{scheme}://{host}"

    webhook_url = f"{api_base}/telegram/webhook/{access_token}"

    # 5. Register webhook with Telegram
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                f"https://api.telegram.org/bot{payload.bot_token}/setWebhook",
                json={
                    "url": webhook_url,
                    "secret_token": webhook_secret,
                    "allowed_updates": ["message", "callback_query"],
                },
            )
            resp.raise_for_status()
            tg_result = resp.json()
    except Exception as e:
        logger.error(
            f"[Telegram] setWebhook failed for tenant {resolved_tenant_id}: {e}"
        )
        raise HTTPException(
            status_code=502,
            detail="Credenciales guardadas pero no se pudo registrar el webhook en Telegram.",
        )

    if not tg_result.get("ok"):
        raise HTTPException(
            status_code=502,
            detail=f"Telegram rechazó el webhook: {tg_result.get('description')}",
        )

    # 6. Reload bot polling in background
    try:
        from services.telegram_bot import reload_telegram_bot

        asyncio.ensure_future(reload_telegram_bot(resolved_tenant_id))
    except Exception as e:
        logger.warning(f"[Telegram] Bot reload after config: {e}")

    return {
        "configured": True,
        "bot_username": bot_info.get("username"),
        "webhook_url": webhook_url,
    }


@router.delete("/telegram/config", tags=["Telegram"])
async def delete_telegram_config(
    user_data=Depends(verify_admin_token),
    resolved_tenant_id: int = Depends(get_resolved_tenant_id),
):
    """
    Desactiva el bot de Telegram para el tenant:
    1. Llama a deleteWebhook en Telegram.
    2. Elimina las tres credenciales del vault.
    """
    token = await get_tenant_credential(resolved_tenant_id, TELEGRAM_BOT_TOKEN)

    if token:
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                resp = await client.post(
                    f"https://api.telegram.org/bot{token}/deleteWebhook"
                )
                resp.raise_for_status()
        except Exception as e:
            # Log but continue — we still remove credentials locally
            logger.warning(
                f"[Telegram] deleteWebhook failed for tenant {resolved_tenant_id}: {e}"
            )

    # Remove all three credentials
    await db.pool.execute(
        """
        DELETE FROM credentials
        WHERE tenant_id = $1 AND name = ANY($2::text[])
        """,
        resolved_tenant_id,
        [TELEGRAM_BOT_TOKEN, TELEGRAM_WEBHOOK_SECRET, TELEGRAM_WEBHOOK_ACCESS_TOKEN],
    )

    # Stop bot polling
    try:
        from services.telegram_bot import reload_telegram_bot

        asyncio.ensure_future(reload_telegram_bot(resolved_tenant_id))
    except Exception:
        pass

    return {"configured": False}
